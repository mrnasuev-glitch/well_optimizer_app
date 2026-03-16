import streamlit as st
import pandas as pd
import numpy as np
import math
from io import BytesIO
from datetime import datetime, timedelta
import pickle
import json
from pathlib import Path
import re

# Ленивые импорты - будут загружаться только когда нужны
_plotly_loaded = False
_openpyxl_loaded = False
_scipy_loaded = False
_stx_loaded = False

def _load_plotly():
    global _plotly_loaded
    if not _plotly_loaded:
        import plotly.graph_objects as go
        import plotly.express as px
        from plotly.subplots import make_subplots
        globals()['go'] = go
        globals()['px'] = px
        globals()['make_subplots'] = make_subplots
        _plotly_loaded = True

def _load_openpyxl():
    global _openpyxl_loaded
    if not _openpyxl_loaded:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        globals()['Workbook'] = Workbook
        globals()['PatternFill'] = PatternFill
        globals()['Font'] = Font
        globals()['Alignment'] = Alignment
        globals()['Border'] = Border
        globals()['Side'] = Side
        globals()['get_column_letter'] = get_column_letter
        _openpyxl_loaded = True

def _load_scipy():
    global _scipy_loaded
    if not _scipy_loaded:
        from scipy.optimize import differential_evolution, curve_fit
        globals()['differential_evolution'] = differential_evolution
        globals()['curve_fit'] = curve_fit
        _scipy_loaded = True

def _load_stx():
    global _stx_loaded
    if not _stx_loaded:
        import extra_streamlit_components as stx
        globals()['stx'] = stx
        _stx_loaded = True

def clean_excel_value(value):
    """Очищает значение от артефактов Excel"""
    if pd.isna(value):
        return None
    str_value = str(value).strip()
    str_value = str_value.replace('_x000D_', '')
    str_value = ' '.join(str_value.split())
    return str_value if str_value else None

def clean_numeric_value(value):
    """Очищает и конвертирует числовое значение"""
    if pd.isna(value):
        return None
    str_value = str(value).strip()
    str_value = str_value.replace('_x000D_', '')
    str_value = str_value.strip()
    if not str_value:
        return None
    try:
        # Удаляем все, кроме цифр, точки и минуса
        str_value = re.sub(r'[^\d\.\-]', '', str_value)
        return float(str_value) if str_value else None
    except:
        return None

# ============================================================
# ФУНКЦИИ ДЛЯ РАСЧЕТА СКОРОСТИ И ПЛОТНОСТИ (НОВЫЕ)
# ============================================================

import math

# Стандартные диаметры труб (наружный диаметр в мм)
STANDARD_PIPELINE_DIAMETERS = {
    '114': {'wall_thickness': 5, 'inner_diameter': 104},   # 114 - 2*5 = 104 мм
    '159': {'wall_thickness': 6, 'inner_diameter': 147},   # 159 - 2*6 = 147 мм  
    '219': {'wall_thickness': 8, 'inner_diameter': 203},   # 219 - 2*8 = 203 мм
    '273': {'wall_thickness': 8, 'inner_diameter': 257},   # 273 - 2*8 = 257 мм
    '325': {'wall_thickness': 8, 'inner_diameter': 309},   # 325 - 2*8 = 309 мм
    '426': {'wall_thickness': 8, 'inner_diameter': 410},   # 426 - 2*8 = 410 мм
    '530': {'wall_thickness': 8, 'inner_diameter': 514},   # 530 - 2*8 = 514 мм
}

def calculate_downhole_mixture_rate(Q_liq: float, water_cut_percent: float, volume_factor: float = 1.1) -> float:
    """
    Пересчет поверхностного дебита жидкости в дебит ГЖС в забойных условиях.
    
    Формула: Qгжс = Qж * [ (1 - B) * b + B ]
    где:
        Qж - дебит жидкости на поверхности, м³/сут
        B - обводненность в долях (water_cut_percent / 100)
        b - объемный коэффициент нефти (1.1)
    
    Returns:
        float: Дебит газожидкостной смеси (ГЖС) в забойных условиях, м³/сут
    """
    # Обводненность в долях
    wc = water_cut_percent / 100.0
    
    # Доля нефти
    oil_cut = 1.0 - wc
    
    # Расчет по формуле
    mixture_rate = Q_liq * (oil_cut * volume_factor + wc)
    
    return mixture_rate

def calculate_mixture_density(oil_density_relative, water_cut_percent):
    """
    Расчет плотности водонефтяной смеси в кг/м³.
    
    Parameters:
    -----------
    oil_density_relative : float
        Относительная плотность нефти (0.84, 0.86 и т.д.)
    water_cut_percent : float
        Обводненность в %
        
    Returns:
    --------
    float: Плотность смеси в кг/м³
    """
    if oil_density_relative is None:
        oil_density_relative = 0.85  # по умолчанию
    
    if water_cut_percent is None:
        water_cut_percent = 0
    
    # Абсолютная плотность нефти (кг/м³)
    oil_density_kg_m3 = oil_density_relative * 1000
    
    # Абсолютная плотность воды (кг/м³)
    water_density_kg_m3 = 1000  # при 20°C
    
    # Доля воды (от 0 до 1)
    water_fraction = water_cut_percent / 100
    
    # Плотность смеси (кг/м³)
    # ρ_смеси = ρ_нефти × (1 - W) + ρ_воды × W
    mixture_density_kg_m3 = (oil_density_kg_m3 * (1 - water_fraction) + 
                           water_density_kg_m3 * water_fraction)
    
    return mixture_density_kg_m3
    
def calculate_flow_velocity(total_flow_m3_per_hour, outer_diameter_mm, wall_thickness_mm=None):
    """
    Расчет средней скорости потока в трубопроводе.
    
    Parameters:
    -----------
    total_flow_m3_per_hour : float
        Суммарный дебит в м³/час
    outer_diameter_mm : int
        Наружный диаметр трубы в мм
    wall_thickness_mm : int, optional
        Толщина стенки в мм
        
    Returns:
    --------
    dict: Результаты расчета
    """
    # Если толщина не указана, берем из стандартов или по умолчанию 8 мм
    if wall_thickness_mm is None:
        if str(outer_diameter_mm) in STANDARD_PIPELINE_DIAMETERS:
            wall_thickness_mm = STANDARD_PIPELINE_DIAMETERS[str(outer_diameter_mm)]['wall_thickness']
        else:
            wall_thickness_mm = 8  # мм по умолчанию
    
    # Внутренний диаметр (мм)
    inner_diameter_mm = outer_diameter_mm - 2 * wall_thickness_mm
    
    # В метры для расчетов
    inner_diameter_m = inner_diameter_mm / 1000  # метры
    
    # Площадь сечения (м²)
    area_m2 = math.pi * (inner_diameter_m / 2) ** 2
    
    # Расход в м³/с
    flow_m3_per_sec = total_flow_m3_per_hour / 3600
    
    # Скорость (м/с)
    velocity_m_s = flow_m3_per_sec / area_m2 if area_m2 > 0 else 0
    
    return {
        'velocity_m_s': velocity_m_s,
        'inner_diameter_mm': inner_diameter_mm,
        'area_m2': area_m2,
        'flow_m3_per_sec': flow_m3_per_sec,
        'wall_thickness_mm': wall_thickness_mm
    }

def get_max_flow_for_velocity(velocity_m_s, outer_diameter_mm, wall_thickness_mm=None):
    """
    Расчет максимального расхода при заданной скорости.
    
    Returns:
    --------
    float: Максимальный расход в м³/час
    """
    result = calculate_flow_velocity(1, outer_diameter_mm, wall_thickness_mm)
    area_m2 = result['area_m2']
    
    # Q = v * A * 3600
    max_flow_m3_per_hour = velocity_m_s * area_m2 * 3600
    
    return max_flow_m3_per_hour
    
# ============================================================
# ФУНКЦИИ СОХРАНЕНИЯ И ВОССТАНОВЛЕНИЯ ДАННЫХ
# ============================================================

def save_data_to_file():
    """Сохраняет все данные в файл (только сериализуемые объекты)"""
    try:
        # ОСНОВНЫЕ ДАННЫЕ (всегда сериализуемы)
        data_to_save = {
            # Данные скважин и кустов
            'wells_data': st.session_state.get('wells_data', []),
            'clusters': st.session_state.get('clusters', {}),
            'selected_cits': st.session_state.get('selected_cits', 'ЦИТС VQ-BAD'),
            'selected_cdng': st.session_state.get('selected_cdng', 'ЦДНГ-1'),
            'selected_tpp': st.session_state.get('selected_tpp', 'VQ-BADнефтегаз'),
            'selected_cluster': st.session_state.get('selected_cluster', None),
        }
        
        # ИСТОРИЯ РАСЧЕТОВ (простые словари)
        if 'calculation_history' in st.session_state:
            data_to_save['calculation_history'] = st.session_state.calculation_history
            print(f"СОХРАНЕНИЕ: calculation_history = {st.session_state.calculation_history}")  # ОТЛАДКА
        
        # МОДУЛЬ 1: СТАБИЛИЗАЦИЯ ДАВЛЕНИЯ
        if 'last_optimization' in st.session_state and st.session_state.last_optimization:
            opt = st.session_state.last_optimization
            opt_copy = {
                'timestamp': opt.get('timestamp'),
                'cluster': opt.get('cluster'),
                'cits': opt.get('cits'),
                'cdng': opt.get('cdng'),
                'target_coefficient': opt.get('target_coefficient'),
                'phases_dict': opt.get('phases_dict', {}),
                'stats': opt.get('stats', {}),
                'current_time': opt.get('current_time'),
            }
            data_to_save['last_optimization'] = opt_copy
            print(f"СОХРАНЕНИЕ: last_optimization = {opt.get('cluster')}")  # ОТЛАДКА
        
        # МОДУЛЬ 2: ПАКЕТНЫЙ РАСЧЕТ КПР
        if 'batch_results_advanced' in st.session_state and st.session_state.batch_results_advanced is not None:
            batch_results = []
            for res in st.session_state.batch_results_advanced:
                if isinstance(res, dict):
                    res_copy = {k: v for k, v in res.items() if not k.startswith('_')}
                    batch_results.append(res_copy)
                else:
                    batch_results.append(res)
            data_to_save['batch_results_advanced'] = batch_results
            print(f"СОХРАНЕНИЕ: batch_results_advanced = {len(batch_results)} записей")  # ОТЛАДКА
        
        # МОДУЛЬ 2: АНАЛИЗ ПОТЕНЦИАЛА
        if 'potential_batch_results' in st.session_state and st.session_state.potential_batch_results is not None:
            potential_results = []
            for res in st.session_state.potential_batch_results:
                if isinstance(res, dict):
                    res_copy = {k: v for k, v in res.items() if not k.startswith('_')}
                    potential_results.append(res_copy)
                else:
                    potential_results.append(res)
            data_to_save['potential_batch_results'] = potential_results
            print(f"СОХРАНЕНИЕ: potential_batch_results = {len(potential_results)} записей")  # ОТЛАДКА
        
        # ДОБАВЛЕНО: Также сохраняем full_batch_results, если они есть
        if 'full_batch_results' in st.session_state and st.session_state.full_batch_results is not None:
            full_results = []
            for res in st.session_state.full_batch_results:
                if isinstance(res, dict):
                    res_copy = {k: v for k, v in res.items() if not k.startswith('_')}
                    full_results.append(res_copy)
                else:
                    full_results.append(res)
            data_to_save['full_batch_results'] = full_results
            print(f"СОХРАНЕНИЕ: full_batch_results = {len(full_results)} записей")  # ОТЛАДКА
        
        # МОДУЛЬ 2: ОДИНОЧНЫЙ РАСЧЕТ КПР
        if 'optimization_result' in st.session_state and st.session_state.optimization_result:
            opt_res = st.session_state.optimization_result
            if isinstance(opt_res, dict):
                opt_res_copy = {k: v for k, v in opt_res.items() if not k.startswith('_')}
                data_to_save['optimization_result'] = opt_res_copy
                print(f"СОХРАНЕНИЕ: optimization_result для {opt_res.get('well_name', 'unknown')}")  # ОТЛАДКА
        
        # МОДУЛЬ 3: ЗАМЕНА ЭЦН (оба режима)
        for mode in ['replace', 'optimize']:
            results_key = f'pump_calculation_results_{mode}'
            if results_key in st.session_state and st.session_state[results_key] is not None:
                data_to_save[results_key] = st.session_state[results_key]
                print(f"СОХРАНЕНИЕ: {results_key} = {len(st.session_state[results_key])} записей")  # ОТЛАДКА
            
            best_key = f'pump_best_variants_{mode}'
            if best_key in st.session_state and st.session_state[best_key] is not None:
                data_to_save[best_key] = st.session_state[best_key]
                print(f"СОХРАНЕНИЕ: {best_key} = {len(st.session_state[best_key])} записей")  # ОТЛАДКА
            
            params_key = f'pump_calculation_params_{mode}'
            if params_key in st.session_state and st.session_state[params_key] is not None:
                data_to_save[params_key] = st.session_state[params_key]
            
            indices_key = f'selected_wells_indices_{mode}'
            if indices_key in st.session_state:
                indices = st.session_state[indices_key]
                if isinstance(indices, set):
                    data_to_save[indices_key] = list(indices)
                else:
                    data_to_save[indices_key] = indices
        
        # МОДУЛЬ 5: ОПТИМИЗАЦИЯ НАГРУЗКИ
        if 'optimization_results' in st.session_state and st.session_state.optimization_results is not None:
            opt_results = st.session_state.optimization_results
            if isinstance(opt_results, dict):
                opt_copy = {k: v for k, v in opt_results.items() 
                          if not hasattr(v, '__class__') or v.__class__.__name__ in ['dict', 'list', 'str', 'int', 'float', 'bool', 'NoneType']}
                data_to_save['optimization_results'] = opt_copy
                print(f"СОХРАНЕНИЕ: optimization_results")  # ОТЛАДКА
        
        print(f"СОХРАНЕНИЕ: Всего ключей для сохранения: {len(data_to_save)}")  # ОТЛАДКА
        
        # Сохраняем в файл
        with open('povh_data.pkl', 'wb') as f:
            pickle.dump(data_to_save, f)
        
        return True
    except Exception as e:
        st.error(f"Ошибка сохранения: {e}")
        import traceback
        st.error(traceback.format_exc())
        return False

def load_data_from_file(force_reload=False):
    """Загружает все данные из файла (с восстановлением структур)"""
    try:
        if Path('povh_data.pkl').exists():
            with open('povh_data.pkl', 'rb') as f:
                data = pickle.load(f)
            
            # Загружаем все данные в session_state
            for key, value in data.items():
                # Конвертируем списки обратно в set для индексов
                if key.startswith('selected_wells_indices_') and isinstance(value, list):
                    st.session_state[key] = set(value)
                
                # Конвертируем словарь обратно в DataFrame для аналитики (НОВЫЙ ФОРМАТ)
                elif key in ['chess_raw_data', 'chess_enriched_data', 'filtered_analytics_data', 'current_analytics_data']:
                    if isinstance(value, dict) and 'data' in value and 'columns' in value:
                        # Это новый формат с метаданными
                        try:
                            import pandas as pd
                            df = pd.DataFrame(value['data'])
                            st.session_state[key] = df
                        except Exception as e:
                            st.warning(f"Ошибка восстановления DataFrame для {key}: {e}")
                            st.session_state[key] = value
                    elif isinstance(value, list) and value and isinstance(value[0], dict):
                        # Старый формат (список словарей)
                        try:
                            import pandas as pd
                            st.session_state[key] = pd.DataFrame(value)
                        except:
                            st.session_state[key] = value
                    else:
                        st.session_state[key] = value
                
                # Конвертируем список обратно в set для других индексов
                elif key == 'selected_wells_indices' and isinstance(value, list):
                    st.session_state[key] = set(value)
                
                # Все остальное загружаем как есть
                else:
                    st.session_state[key] = value
            
            # Специальная обработка для результатов расчетов, которые могли быть в старом формате
            # Модуль 1: стабилизация давления
            if 'last_optimization' in st.session_state and st.session_state.last_optimization:
                # Убеждаемся, что нет объектов классов
                if not isinstance(st.session_state.last_optimization, dict):
                    st.session_state.last_optimization = None
            
            # Модуль 2: пакетный расчет КПР
            if 'batch_results_advanced' in st.session_state and st.session_state.batch_results_advanced:
                # Очищаем от служебных полей, если они есть
                cleaned_results = []
                for res in st.session_state.batch_results_advanced:
                    if isinstance(res, dict):
                        cleaned_res = {k: v for k, v in res.items() if not k.startswith('_')}
                        cleaned_results.append(cleaned_res)
                    else:
                        cleaned_results.append(res)
                st.session_state.batch_results_advanced = cleaned_results
            
            # Модуль 3: замена ЭЦН - проверяем оба режима
            for mode in ['replace', 'optimize']:
                results_key = f'pump_calculation_results_{mode}'
                if results_key in st.session_state and st.session_state[results_key]:
                    # Убеждаемся, что это список
                    if not isinstance(st.session_state[results_key], list):
                        st.session_state[results_key] = []
            
            # Устанавливаем флаг инициализации
            st.session_state.app_initialized = True
            
            return True
    except Exception as e:
        st.warning(f"Ошибка загрузки данных: {e}")
        import traceback
        st.warning(traceback.format_exc())
    
    return False

def dataframe_to_serializable(df):
    """Конвертирует pandas DataFrame в сериализуемый формат"""
    if df is None:
        return None
    if hasattr(df, 'to_dict'):
        return {
            'data': df.to_dict('records'),
            'columns': df.columns.tolist(),
            'dtypes': {col: str(df[col].dtype) for col in df.columns}
        }
    return df

def serializable_to_dataframe(data):
    """Восстанавливает DataFrame из сериализованного формата"""
    if data is None:
        return None
    if isinstance(data, dict) and 'data' in data and 'columns' in data:
        import pandas as pd
        df = pd.DataFrame(data['data'])
        # Здесь можно восстановить типы данных, если нужно
        return df
    return data

def clear_all_data():
    """Очищает все данные (включая результаты расчетов)"""
    # Очищаем все данные
    st.session_state.wells_data = []
    st.session_state.clusters = {}
    st.session_state.calculation_history = []
    st.session_state.selected_cdng = "ЦДНГ-3"
    st.session_state.selected_cluster = None
    st.session_state.selected_cits = "ЦИТС VQ-BAD"
    st.session_state.selected_tpp = "VQ-BADнефтегаз"
    
    # Очищаем все результаты расчетов
    st.session_state.last_optimization = None
    st.session_state.show_results = False
    st.session_state.full_batch_results = None      
    st.session_state.full_batch_detailed = None     
    st.session_state.potential_batch_results = None
    st.session_state.pump_calculation_results = None
    st.session_state.optimization_result = None
    st.session_state.cycle_simulation = None
    st.session_state.inflow_curve = None
    st.session_state.batch_results_advanced = None
    
    # Очищаем результаты замены ЭЦН
    st.session_state.pump_calculation_results_replace = []
    st.session_state.pump_best_variants_replace = []
    st.session_state.pump_calculation_params_replace = {}
    st.session_state.pump_calculation_results_optimize = []
    st.session_state.pump_best_variants_optimize = []
    st.session_state.pump_calculation_params_optimize = {}
    st.session_state.current_conversion_tab = 'replace'
    st.session_state.selected_wells_indices_replace = set()
    st.session_state.selected_wells_indices_optimize = set()
    
    # Очищаем результаты потенциала
    st.session_state.potential_batch_results = []
    
    # Очищаем аналитику
    st.session_state.chess_raw_data = None
    st.session_state.chess_enriched_data = None
    st.session_state.filtered_analytics_data = None
    
    # Очищаем результаты оптимизации нагрузки
    st.session_state.load_optimizer_state = None
    st.session_state.current_load_analysis = None
    st.session_state.optimization_results = None
    st.session_state.pipeline_params = None 
    st.session_state.unsaved_changes = False
    
    # Удаляем файл с данными
    if Path('povh_data.pkl').exists():
        Path('povh_data.pkl').unlink()
    
    return True

def get_best_variant_for_well(variant_results):
    """Выбор лучшего варианта для скважины (с проверкой наличия колонок)"""
    if not variant_results:
        return None
    
    # Определяем название колонки с экономией
    savings_col = None
    for col in ['Экономия, руб/сут', 'Экономия, руб/месяц', 'Экономия, руб/год', 'total_effect_per_day']:
        if variant_results and col in variant_results[0]:
            savings_col = col
            break
    
    if not savings_col:
        # Если нет колонки с экономией, возвращаем первый вариант
        return variant_results[0] if variant_results else None
    
    # Ищем вариант с максимальной экономией
    best_variant = None
    max_savings = -float('inf')
    
    for variant in variant_results:
        savings = variant.get(savings_col, 0)
        if savings > max_savings:
            max_savings = savings
            best_variant = variant
    
    return best_variant

def save_current_calculations():
    """Ручное сохранение текущих результатов расчетов"""
    if save_data_to_file():
        st.success("✅ Результаты расчетов сохранены в файл")
    else:
        st.error("❌ Ошибка при сохранении")

def get_frequency_hz_safe(well_data):
    """Безопасное получение частоты в Гц"""
    freq = well_data.get('rotations_hz')
    if freq is None or freq == '':
        return 50
    try:
        freq_float = float(freq)
        return freq_float if freq_float > 0 else 50
    except:
        return 50

# ============================================================
# ФУНКЦИИ СОХРАНЕНИЯ РЕЗУЛЬТАТОВ ИЗ МОДУЛЕЙ
# ============================================================

def save_kpr_optimization_to_system(well_name, recommended_schedule, scenario_type):
    """Сохраняет результаты оптимизации КПР в основную систему"""
    if 'wells_data' not in st.session_state:
        return False
    
    wells_data = st.session_state.wells_data
    changes_made = False
    
    for i, well in enumerate(wells_data):
        if well['name'] == well_name:
            # Сохраняем новый график
            old_schedule = well.get('schedule')
            wells_data[i]['schedule'] = recommended_schedule
            
            # Устанавливаем режим
            wells_data[i]['mode'] = 'По времени'
            
            # Исключаем из сдвига только если скважина работает по давлению
            wells_data[i]['exclude_from_shift'] = False
            
            # Обновляем время модификации
            wells_data[i]['last_modified'] = datetime.now().strftime("%Y-%m-%d %H:%M")
            wells_data[i]['modification_source'] = f'kpr_optimization_{scenario_type}'
            
            changes_made = True
            break
    
    if changes_made:
        save_data_to_file()
        return True
    
    return False

def save_ecn_replacement_to_system(well_name, new_pump_type, new_pump_head, 
                                  recommended_schedule=None, kpr_mode='По давлению'):
    """Сохраняет результаты замены ЭЦН с переводом на КПР режим"""
    if 'wells_data' not in st.session_state:
        return False
    
    wells_data = st.session_state.wells_data
    changes_made = False
    
    for i, well in enumerate(wells_data):
        if well['name'] == well_name:
            # Меняем тип скважины с постоянной на КПР
            wells_data[i]['operation_mode'] = 'kpr'
            
            # Устанавливаем график КПР
            if recommended_schedule:
                wells_data[i]['schedule'] = recommended_schedule
            else:
                # Значения по умолчанию
                if new_pump_type == '125':
                    wells_data[i]['schedule'] = [40, 42]
                else:
                    wells_data[i]['schedule'] = [15*60, 45*60]
            
            # Устанавливаем режим КПР - СОХРАНЯЕМ В ОБА ПОЛЯ
            wells_data[i]['mode'] = kpr_mode
            wells_data[i]['kpr_mode'] = kpr_mode  # ← ДОБАВИТЬ ЭТУ СТРОКУ
            
            # Рассчитываем и сохраняем часы работы (ДОБАВИТЬ ЭТОТ БЛОК)
            if recommended_schedule and len(recommended_schedule) >= 2:
                work_min = recommended_schedule[0]
                pause_min = recommended_schedule[1]
                if work_min + pause_min > 0:
                    duty_cycle = work_min / (work_min + pause_min)
                    wells_data[i]['kpr_work_hours'] = duty_cycle * 24
            
            # Для режима по давлению исключаем из расчета сдвига
            wells_data[i]['exclude_from_shift'] = (kpr_mode == 'По давлению')
            
            # Обновляем марку насоса
            new_pump_mark = f"ЭЦН-{new_pump_type}-{int(new_pump_head)}"
            wells_data[i]['pump_mark'] = new_pump_mark
            
            # Обновляем время запуска по умолчанию
            wells_data[i]['base_launch_time'] = '08:00'
            
            changes_made = True
            break
    
    if changes_made:
        save_data_to_file()
        return True
    
    return False

def save_batch_kpr_changes(batch_results):
    """Сохраняет результаты пакетной оптимизации КПР в систему"""
    if 'wells_data' not in st.session_state or not batch_results:
        return {'total': 0, 'saved': 0, 'errors': 0}
    
    stats = {'total': len(batch_results), 'saved': 0, 'errors': 0}
    
    for result in batch_results:
        well_name = result['Скважина']
        
        # Парсим рекомендуемый режим из строки "480/1440 мин"
        try:
            if 'Рекомендуемый режим' in result and isinstance(result['Рекомендуемый режим'], str):
                # Удаляем " мин" и разделяем
                mode_str = result['Рекомендуемый режим'].replace(' мин', '')
                if '/' in mode_str:
                    work_str, pause_str = mode_str.split('/')
                    work_min = int(float(work_str.strip()))
                    pause_min = int(float(pause_str.strip()))
                    new_schedule = [work_min, pause_min]
                else:
                    new_schedule = [15*60, 45*60]
            else:
                new_schedule = [15*60, 45*60]
        except:
            new_schedule = [15*60, 45*60]
        
        # Определяем сценарий
        scenario_type = 'A' if 'Увеличение добычи' in result.get('Сценарий', '') else 'B'
        
        # Сохраняем изменения
        success = save_kpr_optimization_to_system(well_name, new_schedule, scenario_type)
        
        if success:
            stats['saved'] += 1
        else:
            stats['errors'] += 1
    
    return stats

@st.cache_data(ttl=3600)
def check_system_integrity():
    """Проверяет целостность данных перед оптимизацией давления"""
    issues = []
    
    if 'wells_data' not in st.session_state:
        return ["Нет данных о скважинах"]
    
    wells_data = st.session_state.wells_data
    
    # Проверяем КПР скважины
    kpr_wells = [w for w in wells_data if w.get('operation_mode') == 'kpr']
    
    for well in kpr_wells:
        # Проверка наличия графика
        if not well.get('schedule'):
            issues.append(f"Скважина {well['name']}: нет графика КПР")
        
        # Проверка режима
        if not well.get('mode'):
            issues.append(f"Скважина {well['name']}: не указан режим КПР")
        
        # Проверка времени запуска
        if not well.get('base_launch_time'):
            issues.append(f"Скважина {well['name']}: не указано время запуска")
    
    return issues

@st.cache_data(ttl=3600)
def find_clusters_with_multiple_kpr_wells(wells_data, min_kpr_count=2):
    """
    Находит кусты с 2 и более КПР скважинами для возможной оптимизации
    
    Args:
        wells_data: Список данных о скважинах
        min_kpr_count: Минимальное количество КПР скважин в кусте
    
    Returns:
        list: Список словарей с информацией о кустах
    """
    # Группируем скважины по ЦИТС -> ЦДНГ -> Куст
    cluster_data = {}
    
    for well in wells_data:
        if not well.get('is_active', True):
            continue
            
        cits = well.get('cits', 'ЦИТС VQ-BAD')
        cdng = well.get('cdng', 'ЦДНГ-1')
        cluster = well.get('cluster', 'Неизвестно')
        
        if cluster == 'Неизвестно':
            continue
        
        # Создаем структуру данных
        if cits not in cluster_data:
            cluster_data[cits] = {}
        
        if cdng not in cluster_data[cits]:
            cluster_data[cits][cdng] = {}
        
        if cluster not in cluster_data[cits][cdng]:
            cluster_data[cits][cdng][cluster] = {
                'total_wells': 0,
                'kpr_wells': 0,
                'constant_wells': 0,
                'kpr_well_names': [],
                'constant_well_names': [],
                'total_flow': 0,
                'kpr_flow': 0,
                'constant_flow': 0
            }
        
        # Обновляем статистику куста
        cluster_info = cluster_data[cits][cdng][cluster]
        cluster_info['total_wells'] += 1
        cluster_info['total_flow'] += well.get('flow_rate', 0)
        
        if well.get('operation_mode') == 'kpr':
            cluster_info['kpr_wells'] += 1
            cluster_info['kpr_flow'] += well.get('flow_rate', 0)
            cluster_info['kpr_well_names'].append({
                'name': well['name'],
                'flow_rate': well.get('flow_rate', 0),
                'schedule': well.get('schedule'),
                'mode': well.get('mode', 'По времени'),
                'base_launch_time': well.get('base_launch_time', '00:00'),
                'exclude_from_shift': well.get('exclude_from_shift', False)
            })
        else:
            cluster_info['constant_wells'] += 1
            cluster_info['constant_flow'] += well.get('flow_rate', 0)
            cluster_info['constant_well_names'].append({
                'name': well['name'],
                'flow_rate': well.get('flow_rate', 0)
            })
    
    # Фильтруем кусты с 2 и более КПР скважинами
    clusters_for_optimization = []
    
    for cits in cluster_data:
        for cdng in cluster_data[cits]:
            for cluster, info in cluster_data[cits][cdng].items():
                if info['kpr_wells'] >= min_kpr_count:
                    clusters_for_optimization.append({
                        'ЦИТС': cits,
                        'ЦДНГ': cdng,
                        'Куст': cluster,
                        'Всего скважин': info['total_wells'],
                        'КПР скважин': info['kpr_wells'],
                        'Постоянных скважин': info['constant_wells'],
                        'Общий дебит': round(info['total_flow'], 1),
                        'Дебит КПР': round(info['kpr_flow'], 1),
                        'Дебит постоянных': round(info['constant_flow'], 1),
                        'Доля КПР': f"{(info['kpr_wells'] / info['total_wells'] * 100):.0f}%" if info['total_wells'] > 0 else "0%",
                        'Статистика': info
                    })
    
    # Сортируем по количеству КПР скважин (по убыванию)
    clusters_for_optimization.sort(key=lambda x: x['КПР скважин'], reverse=True)
    
    return clusters_for_optimization

def show_automatic_cluster_search():
    """Интерфейс автоматического поиска кустов для оптимизации"""
    st.markdown("### 🎯 Автоматический поиск кустов для оптимизации")
    
    st.info("""
    **Поиск кустов для стабилизации давления:**
    - Находит кусты с 2 и более КПР скважинами
    - Оценивает потенциал оптимизации
    - Показывает статистику по каждому кусту
    - Позволяет сразу перейти к оптимизации выбранного куста
    """)
    
    # Проверка наличия данных
    if not st.session_state.get('wells_data'):
        st.warning("Сначала загрузите данные скважин через раздел 'Импорт техрежима'")
        return
    
    # Настройки поиска
    col_search1, col_search2, col_search3 = st.columns(3)
    
    with col_search1:
        min_kpr_count = st.number_input(
            "Минимальное количество КПР скважин",
            min_value=2,
            max_value=20,
            value=2,
            step=1,
            key="min_kpr_count_search"
        )
    
    with col_search2:
        min_total_wells = st.number_input(
            "Минимальное всего скважин в кусте",
            min_value=1,
            max_value=50,
            value=3,
            step=1,
            key="min_total_wells_search"
        )
    
    with col_search3:
        min_kpr_flow_share = st.slider(
            "Минимальная доля дебита КПР",
            min_value=0,
            max_value=100,
            value=30,
            step=5,
            key="min_kpr_flow_share",
            help="Минимальный процент дебита от КПР скважин от общего дебита куста"
        )
    
    # Кнопка поиска
    if st.button("🔍 Найти кусты для оптимизации", type="primary", key="search_clusters_btn"):
        with st.spinner("Поиск кустов..."):
            # Получаем все кусты с КПР скважинами
            all_clusters = find_clusters_with_multiple_kpr_wells(
                st.session_state.wells_data,
                min_kpr_count
            )
            
            # Фильтруем по дополнительным критериям
            filtered_clusters = []
            
            for cluster in all_clusters:
                # Проверяем минимальное количество скважин
                if cluster['Всего скважин'] < min_total_wells:
                    continue
                
                # Проверяем долю дебита КПР
                if cluster['Общий дебит'] > 0:
                    kpr_flow_share = (cluster['Дебит КПР'] / cluster['Общий дебит']) * 100
                    if kpr_flow_share < min_kpr_flow_share:
                        continue
                
                filtered_clusters.append(cluster)
            
            # Сохраняем результаты поиска
            st.session_state.found_clusters_for_optimization = filtered_clusters
            
            if filtered_clusters:
                st.success(f"✅ Найдено {len(filtered_clusters)} кустов для оптимизации")
            else:
                st.warning("⚠️ Не найдено кустов, соответствующих критериям")
    
    # Показ результатов
    if 'found_clusters_for_optimization' in st.session_state:
        clusters = st.session_state.found_clusters_for_optimization
        
        if clusters:
            st.markdown("---")
            st.markdown(f"#### 📊 Найдено {len(clusters)} кустов для оптимизации")
            
            # Статистика
            total_kpr_wells = sum(c['КПР скважин'] for c in clusters)
            total_wells = sum(c['Всего скважин'] for c in clusters)
            avg_kpr_per_cluster = total_kpr_wells / len(clusters) if clusters else 0
            
            col_stat1, col_stat2, col_stat3 = st.columns(3)
            
            with col_stat1:
                st.metric("Всего кустов", len(clusters))
            
            with col_stat2:
                st.metric("Всего КПР скважин", total_kpr_wells)
            
            with col_stat3:
                st.metric("Среднее КПР на куст", f"{avg_kpr_per_cluster:.1f}")
            
            # Таблица кустов
            st.markdown("##### 🗺️ Список кустов для оптимизации")
            
            # Подготавливаем данные для таблицы
            table_data = []
            for i, cluster in enumerate(clusters):
                table_data.append({
                    '№': i + 1,
                    'Куст': cluster['Куст'],
                    'ЦДНГ': cluster['ЦДНГ'],
                    'Всего скв.': cluster['Всего скважин'],
                    'КПР скв.': cluster['КПР скважин'],
                    'Дебит КПР': f"{cluster['Дебит КПР']:.0f} м³/сут",
                    'Общий дебит': f"{cluster['Общий дебит']:.0f} м³/сут",
                    'Доля КПР': cluster['Доля КПР'],
                    'Потенциал': 'Высокий' if cluster['КПР скважин'] >= 4 else 'Средний' if cluster['КПР скважин'] >= 2 else 'Низкий'
                })
            
            df_clusters = pd.DataFrame(table_data)
            st.dataframe(df_clusters, use_container_width=True, hide_index=True)
            
            # Выбор куста для оптимизации
            st.markdown("---")
            st.markdown("#### 🎯 Выбор куста для немедленной оптимизации")
            
            col_select1, col_select2 = st.columns(2)
            
            with col_select1:
                # Создаем список кустов для выбора
                cluster_options = [f"{c['Куст']} ({c['ЦДНГ']})" for c in clusters]
                selected_cluster_str = st.selectbox(
                    "Выберите куст для оптимизации",
                    cluster_options,
                    key="cluster_selection_auto"
                )
            
            with col_select2:
                # Получаем выбранный куст
                selected_index = cluster_options.index(selected_cluster_str)
                selected_cluster = clusters[selected_index]
                
                # Показываем информацию о выбранном кусте
                st.info(f"""
                **Информация о кусте:**
                - Куст: {selected_cluster['Куст']}
                - ЦДНГ: {selected_cluster['ЦДНГ']}
                - ЦИТС: {selected_cluster['ЦИТС']}
                - КПР скважин: {selected_cluster['КПР скважин']} из {selected_cluster['Всего скважин']}
                - Общий дебит: {selected_cluster['Общий дебит']} м³/сут
                """)
            
            # Детальная информация о скважинах в выбранном кусте
            st.markdown("##### 📋 Список КПР скважин в выбранном кусте")
            
            cluster_stats = selected_cluster['Статистика']
            
            # Таблица КПР скважин
            if cluster_stats['kpr_well_names']:
                kpr_table_data = []
                for well_info in cluster_stats['kpr_well_names']:
                    schedule = well_info.get('schedule', [0, 0])
                    if schedule and len(schedule) >= 2:
                        schedule_str = f"{schedule[0]}/{schedule[1]} мин"
                    else:
                        schedule_str = "Нет данных"
                    
                    kpr_table_data.append({
                        'Скважина': well_info['name'],
                        'Дебит': f"{well_info['flow_rate']:.1f} м³/сут",
                        'Режим': well_info.get('mode', 'По времени'),
                        'График': schedule_str,
                        'Запуск': well_info.get('base_launch_time', '00:00'),
                        'Искл. из сдвига': 'Да' if well_info.get('exclude_from_shift', False) else 'Нет'
                    })
                
                df_kpr_wells = pd.DataFrame(kpr_table_data)
                st.dataframe(df_kpr_wells, use_container_width=True, hide_index=True)
            else:
                st.info("В кусте нет КПР скважин")
            
            # Таблица постоянных скважин
            if cluster_stats['constant_well_names']:
                st.markdown("##### 🔄 Постоянные скважины в кусте")
                
                const_table_data = []
                for well_info in cluster_stats['constant_well_names']:
                    const_table_data.append({
                        'Скважина': well_info['name'],
                        'Дебит': f"{well_info['flow_rate']:.1f} м³/сут"
                    })
                
                df_const_wells = pd.DataFrame(const_table_data)
                st.dataframe(df_const_wells, use_container_width=True, hide_index=True)
            
            # Кнопка перехода к оптимизации
            st.markdown("---")
            st.markdown("#### 🚀 Запуск оптимизации")
            
            if st.button("⚙️ Оптимизировать этот куст", type="primary", key="optimize_selected_cluster"):
                # Сохраняем выбранный куст в session_state
                st.session_state.opt_selected_cluster = selected_cluster['Куст']
                st.session_state.selected_cdng = selected_cluster['ЦДНГ']
                st.session_state.selected_cits = selected_cluster['ЦИТС']
                
                # Устанавливаем целевую страницу
                st.session_state.current_page = "optimization"
                st.session_state.show_results = False
                st.session_state.last_optimization = None
                
                st.success(f"✅ Переход к оптимизации куста: {selected_cluster['Куст']}")
                st.rerun()
            
            # Экспорт результатов поиска
            st.markdown("---")
            st.markdown("#### 📥 Экспорт результатов поиска")
            
            def export_search_results():
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    # Основная таблица кустов
                    df_clusters.to_excel(writer, sheet_name='Кусты для оптимизации', index=False)
                    
                    # Детальная информация по выбранному кусту
                    selected_data = {
                        'Параметр': [
                            'Куст', 'ЦИТС', 'ЦДНГ', 'Всего скважин', 
                            'КПР скважин', 'Постоянных скважин',
                            'Общий дебит', 'Дебит КПР', 'Дебит постоянных',
                            'Доля КПР', 'Дата поиска'
                        ],
                        'Значение': [
                            selected_cluster['Куст'],
                            selected_cluster['ЦИТС'],
                            selected_cluster['ЦДНГ'],
                            selected_cluster['Всего скважин'],
                            selected_cluster['КПР скважин'],
                            selected_cluster['Постоянных скважин'],
                            f"{selected_cluster['Общий дебит']:.1f} м³/сут",
                            f"{selected_cluster['Дебит КПР']:.1f} м³/сут",
                            f"{selected_cluster['Дебит постоянных']:.1f} м³/сут",
                            selected_cluster['Доля КПР'],
                            datetime.now().strftime("%d.%m.%Y %H:%M")
                        ]
                    }
                    pd.DataFrame(selected_data).to_excel(writer, sheet_name='Выбранный куст', index=False)
                    
                    # КПР скважины
                    if cluster_stats['kpr_well_names']:
                        kpr_export_data = []
                        for well in cluster_stats['kpr_well_names']:
                            schedule = well.get('schedule', [0, 0])
                            if schedule and len(schedule) >= 2:
                                schedule_str = f"{schedule[0]}/{schedule[1]}"
                            else:
                                schedule_str = "Нет данных"
                            
                            kpr_export_data.append({
                                'Скважина': well['name'],
                                'Дебит, м³/сут': well['flow_rate'],
                                'Режим КПР': well.get('mode', 'По времени'),
                                'График, мин': schedule_str,
                                'Время запуска': well.get('base_launch_time', '00:00'),
                                'Исключена из сдвига': 'Да' if well.get('exclude_from_shift', False) else 'Нет'
                            })
                        
                        pd.DataFrame(kpr_export_data).to_excel(writer, sheet_name='КПР скважины', index=False)
                    
                    # Постоянные скважины
                    if cluster_stats['constant_well_names']:
                        const_export_data = []
                        for well in cluster_stats['constant_well_names']:
                            const_export_data.append({
                                'Скважина': well['name'],
                                'Дебит, м³/сут': well['flow_rate']
                            })
                        
                        pd.DataFrame(const_export_data).to_excel(writer, sheet_name='Постоянные скважины', index=False)
                
                return output.getvalue()
            
            excel_data = export_search_results()
            
            st.download_button(
                label="📊 Скачать отчет о поиске",
                data=excel_data,
                file_name=f"поиск_кустов_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        else:
            st.warning("Не найдено кустов, соответствующих критериям поиска")

def save_custom_selected_wells():
    """Интерфейс для сохранения выбранных вручную скважин из всех модулей"""
    st.markdown("### 💾 Выборочное сохранение скважин")
    st.info("""
    **Инструкция:**
    1. Вставьте список скважин в поле ниже (по одной на строку или через запятую)
    2. Выберите тип сохранения
    3. Настройте параметры фильтрации
    4. Сохраните только нужные скважины
    """)
    
    # Ввод списка скважин
    st.markdown("#### 📝 Вставьте список скважин")
    well_names_input = st.text_area(
        "Названия скважин",
        placeholder="""Скв_101
Скв_205
Скв_308, Скв_402
Скв_511 Скв_614""",
        height=150,
        help="Можно вводить по одной на строку, через запятые или пробелы"
    )
    
    if not well_names_input:
        st.warning("Введите список скважин")
        return
    
    # Парсим список скважин
    import re
    well_names = []
    for line in well_names_input.split('\n'):
        line = line.strip()
        if not line:
            continue
        
        # Разбиваем строку на отдельные имена
        names = re.split(r'[,\s]+', line)
        well_names.extend([name.strip() for name in names if name.strip()])
    
    # Удаляем дубликаты
    well_names = list(set(well_names))
    
    if not well_names:
        st.error("Не удалось распознать названия скважин")
        return
    
    st.success(f"Распознано {len(well_names)} уникальных скважин")
    
    # Выбор типа сохранения
    st.markdown("#### ⚙️ Тип сохранения")
    save_type = st.radio(
        "Выберите тип сохранения",
        ["Из пакетного расчета КПР", "Из замены ЭЦН", "Обновить режимы в системе"],
        horizontal=True
    )
    
    if save_type == "Из пакетного расчета КПР":
        if 'batch_results_advanced' not in st.session_state:
            st.warning("Сначала выполните пакетный расчет КПР")
            return
        
        batch_results = st.session_state.batch_results_advanced
        
        # Фильтруем только те скважины, которые есть в результатах
        available_results = []
        available_wells = []
        missing_wells = []
        
        for well_name in well_names:
            found = False
            for result in batch_results:
                if result['Скважина'] == well_name:
                    available_results.append(result)
                    available_wells.append(well_name)
                    found = True
                    break
            
            if not found:
                missing_wells.append(well_name)
        
        # Показываем статистику
        col_stat1, col_stat2 = st.columns(2)
        with col_stat1:
            st.metric("Найдено в результатах", len(available_wells))
        with col_stat2:
            if missing_wells:
                st.metric("Не найдено", len(missing_wells))
        
        if missing_wells:
            with st.expander("Скважины не найдены в результатах"):
                st.write(", ".join(missing_wells[:10]))
                if len(missing_wells) > 10:
                    st.caption(f"... и еще {len(missing_wells) - 10} скважин")
        
        if not available_results:
            st.error("Нет скважин для сохранения")
            return
        
        # Дополнительные фильтры
        st.markdown("#### 🔍 Дополнительные фильтры")
        
        col_filter1, col_filter2 = st.columns(2)
        
        with col_filter1:
            filter_scenario = st.multiselect(
                "Фильтр по сценарию",
                options=["Увеличение добычи", "Снижение затрат", "Нет оптимизации"],
                default=["Увеличение добычи", "Снижение затрат"]
            )
        
        with col_filter2:
            min_pnas = st.number_input(
                "Минимальное Pнас (атм)",
                min_value=0.0,
                value=0.0,
                step=5.0
            )
        
        # Применяем фильтры
        filtered_results = []
        for result in available_results:
            # Фильтр по сценарию
            if filter_scenario and result.get('Сценарий') not in filter_scenario:
                continue
            
            # Фильтр по Pнас
            if min_pnas > 0 and result.get('Pнас', 0) < min_pnas:
                continue
            
            filtered_results.append(result)
        
        st.info(f"После фильтрации осталось {len(filtered_results)} скважин")
        
        # Предпросмотр
        if st.checkbox("Показать предпросмотр", value=True):
            preview_data = []
            for result in filtered_results[:10]:  # Первые 10
                preview_data.append({
                    'Скважина': result['Скважина'],
                    'Сценарий': result['Сценарий'],
                    'Текущий режим': result['Текущий режим'],
                    'Рекомендуемый': result['Рекомендуемый режим'],
                    'Pнас': result.get('Pнас', '-')
                })
            
            if preview_data:
                st.dataframe(pd.DataFrame(preview_data), use_container_width=True)
        
        # Кнопка сохранения
        if st.button("💾 Сохранить выбранные скважины", type="primary"):
            stats = save_selected_kpr_changes(filtered_results, None)  # None = все из filtered_results
            
            if stats['saved'] > 0:
                st.success(f"✅ Сохранено {stats['saved']} скважин")
                
                # Предлагаем перейти к оптимизации
                if st.button("🎯 Перейти к стабилизации давления"):
                    st.session_state.current_page = "optimization"
                    st.rerun()
            else:
                st.error("❌ Не удалось сохранить скважины")
    
    elif save_type == "Из замены ЭЦН":
        if 'pump_calculation_results' not in st.session_state:
            st.warning("Сначала выполните расчет замены ЭЦН")
            return
        
        pump_results = st.session_state.pump_calculation_results
        
        # ФИЛЬТРАЦИЯ ПО ЭКОНОМИЧЕСКОЙ ЭФФЕКТИВНОСТИ
        st.markdown("#### 💰 Фильтр по экономической эффективности")
        
        col_econ1, col_econ2 = st.columns(2)
        
        with col_econ1:
            min_daily_savings = st.number_input(
                "Минимальная экономия (руб/сут)",
                min_value=0,
                value=100,
                step=50,
                help="Сохранять только скважины с экономией выше этого значения"
            )
        
        with col_econ2:
            min_yearly_savings = st.number_input(
                "Минимальная экономия (руб/год)",
                min_value=0,
                value=10000,
                step=1000,
                help="Сохранять только скважины с годовой экономией выше этого значения"
            )
        
        # Фильтруем по экономике
        filtered_results, selected_wells = filter_wells_by_economy(
            pump_results,
            min_daily_savings,
            min_yearly_savings
        )
        
        # Дополнительно фильтруем по введенным именам
        final_results = []
        final_wells = []
        
        for result in filtered_results:
            if result['Скважина'] in well_names:
                final_results.append(result)
                final_wells.append(result['Скважина'])
        
        # Статистика
        col_stat1, col_stat2, col_stat3 = st.columns(3)
        
        with col_stat1:
            st.metric("Всего введено", len(well_names))
        
        with col_stat2:
            st.metric("С положительной экономией", len(filtered_results))
        
        with col_stat3:
            st.metric("Будет сохранено", len(final_results))
        
        if not final_results:
            st.error("❌ Нет скважин, соответствующих критериям")
            return
        
        # Настройки КПР для замененных скважин
        st.markdown("#### ⚙️ Настройки КПР режима")
        
        col_kpr1, col_kpr2 = st.columns(2)
        
        with col_kpr1:
            kpr_mode = st.selectbox(
                "Режим КПР",
                ["По давлению", "По времени"]
            )
        
        with col_kpr2:
            if kpr_mode == "По давлению":
                default_schedule = "40/42"
                help_text = "работа/пауза (атм)"
            else:
                default_schedule = "900/2700"
                help_text = "работа/пауза (мин)"
            
            kpr_schedule = st.text_input(
                "График КПР",
                value=default_schedule,
                help=help_text
            )
        
        # Предпросмотр
        if final_results and st.checkbox("Показать предпросмотр экономии", value=True):
            preview_data = []
            for result in final_results[:10]:
                preview_data.append({
                    'Скважина': result['Скважина'],
                    'Экономия (руб/сут)': f"{result.get('Экономия, руб/сут', 0):.0f}",
                    'Экономия (руб/год)': f"{result.get('Экономия, руб/год', 0):.0f}",
                    'Новый насос': result.get('ЭЦН-после', '-')
                })
            
            st.dataframe(pd.DataFrame(preview_data), use_container_width=True)
        
        # Кнопка сохранения
        if final_results and st.button("💾 Сохранить экономически эффективные скважины", 
                                      type="primary"):
            
            # Парсим график
            try:
                if kpr_schedule and '/' in kpr_schedule:
                    work_str, pause_str = kpr_schedule.split('/')
                    if kpr_mode == "По давлению":
                        work = int(float(work_str))
                        pause = int(float(pause_str))
                    else:
                        work = int(float(work_str))
                        pause = int(float(pause_str))
                    schedule = [work, pause]
                else:
                    schedule = [40, 42] if kpr_mode == "По давлению" else [900, 2700]
            except:
                schedule = [40, 42] if kpr_mode == "По давлению" else [900, 2700]
            
            # Сохраняем каждую скважину
            saved_count = 0
            error_count = 0
            
            with st.spinner("Сохранение скважин..."):
                for result in final_results:
                    well_name = result['Скважина']
                    new_pump_type = st.session_state.get('new_pump_type', '125')
                    new_pump_head = st.session_state.get('new_pump_head', 1500)
                    
                    success = save_ecn_replacement_to_system(
                        well_name,
                        new_pump_type,
                        new_pump_head,
                        schedule,
                        kpr_mode
                    )
                    
                    if success:
                        saved_count += 1
                    else:
                        error_count += 1
            
            if saved_count > 0:
                st.success(f"✅ Сохранено {saved_count} экономически эффективных скважин!")
                st.info(f"""
                **Настройки сохранения:**
                - Режим: {kpr_mode}
                - График: {kpr_schedule}
                - Исключены из сдвига: {'Да' if kpr_mode == 'По давлению' else 'Нет'}
                - Минимальная экономия: {min_daily_savings} руб/сут
                """)
                
                if error_count > 0:
                    st.warning(f"Не удалось сохранить {error_count} скважин")
                
                # Предлагаем перейти к оптимизации
                if st.button("🎯 Оптимизировать время запуска КПР"):
                    st.session_state.current_page = "optimization"
                    st.rerun()
            else:
                st.error("❌ Не удалось сохранить ни одной скважины")
    
    else:  # Обновить режимы в системе
        st.markdown("#### 🔄 Обновление существующих скважин в системе")
        
        # Находим скважины в системе
        system_wells = st.session_state.get('wells_data', [])
        
        found_wells = []
        not_found_wells = []
        
        for well_name in well_names:
            found = False
            for well in system_wells:
                if well['name'] == well_name:
                    found_wells.append(well)
                    found = True
                    break
            
            if not found:
                not_found_wells.append(well_name)
        
        st.info(f"Найдено в системе: {len(found_wells)} скважин")
        
        if not_found_wells:
            st.warning(f"Не найдено в системе: {len(not_found_wells)} скважин")
            with st.expander("Список не найденных"):
                st.write(", ".join(not_found_wells[:20]))
        
        if found_wells:
            # Настройки обновления
            st.markdown("#### ⚙️ Настройки обновления")
            
            col_upd1, col_upd2 = st.columns(2)
            
            with col_upd1:
                new_operation_mode = st.selectbox(
                    "Тип скважины",
                    ["kpr", "constant"],
                    format_func=lambda x: "КПР" if x == "kpr" else "Постоянная"
                )
            
            with col_upd2:
                if new_operation_mode == "kpr":
                    exclude_from_shift = st.checkbox("Исключить из сдвига", value=False)
                else:
                    exclude_from_shift = False
            
            # Кнопка обновления
            if st.button("🔄 Обновить выбранные скважины", type="primary"):
                updated_count = 0
                
                for well_name in well_names:
                    for i, well in enumerate(system_wells):
                        if well['name'] == well_name:
                            # Сохраняем изменения
                            system_wells[i]['operation_mode'] = new_operation_mode
                            system_wells[i]['exclude_from_shift'] = exclude_from_shift
                            system_wells[i]['last_modified'] = datetime.now().strftime("%Y-%m-%d %H:%M")
                            system_wells[i]['modification_source'] = 'manual_update'
                            updated_count += 1
                            break
                
                if updated_count > 0:
                    save_data_to_file()
                    st.success(f"✅ Обновлено {updated_count} скважин")
                else:
                    st.error("❌ Не удалось обновить скважины")

@st.cache_data(ttl=3600)
def filter_wells_by_economy(pump_results, min_daily_savings=0, min_yearly_savings=0):
    """
    Фильтрует скважины по экономической эффективности
    
    Args:
        pump_results: Результаты расчета замены ЭЦН
        min_daily_savings: Минимальная экономия в руб/сут
        min_yearly_savings: Минимальная экономия в руб/год
    
    Returns:
        list: Отфильтрованные результаты
        list: Имена скважин с положительной экономией
    """
    if not pump_results:
        return [], []
    
    filtered_results = []
    selected_wells = []
    
    for result in pump_results:
        # Проверяем наличие данных об экономии
        daily_savings = result.get('Экономия, руб/сут', 0)
        yearly_savings = result.get('Экономия, руб/год', 0)
        
        # Проверяем условия
        meets_daily = daily_savings >= min_daily_savings
        meets_yearly = yearly_savings >= min_yearly_savings
        
        if meets_daily and meets_yearly:
            filtered_results.append(result)
            selected_wells.append(result['Скважина'])
    
    return filtered_results, selected_wells

def save_selected_kpr_changes(batch_results, selected_well_names=None):
    """
    Сохраняет выбранные результаты пакетной оптимизации КПР в систему
    
    Args:
        batch_results: Список результатов пакетного расчета
        selected_well_names: Список имен скважин для сохранения (None = все)
    """
    if 'wells_data' not in st.session_state or not batch_results:
        return {'total': 0, 'saved': 0, 'errors': 0, 'filtered': 0}
    
    stats = {'total': len(batch_results), 'saved': 0, 'errors': 0, 'filtered': 0}
    
    for result in batch_results:
        well_name = result['Скважина']
        
        # Проверяем, нужно ли сохранять эту скважину
        if selected_well_names is not None and well_name not in selected_well_names:
            stats['filtered'] += 1
            continue
        
        # Парсим рекомендуемый режим
        try:
            if 'Рекомендуемый режим' in result and isinstance(result['Рекомендуемый режим'], str):
                mode_str = result['Рекомендуемый режим']
                
                # Убираем текст в скобках если есть
                if '(' in mode_str:
                    mode_str = mode_str.split('(')[0].strip()
                
                # Удаляем " мин" и разделяем
                mode_str = mode_str.replace(' мин', '').strip()
                
                if '/' in mode_str:
                    work_str, pause_str = mode_str.split('/')
                    work_min = int(float(work_str.strip()))
                    pause_min = int(float(pause_str.strip()))
                    new_schedule = [work_min, pause_min]
                else:
                    new_schedule = [15*60, 45*60]  # 15/45 часов по умолчанию
            else:
                new_schedule = [15*60, 45*60]
        except Exception as e:
            new_schedule = [15*60, 45*60]
        
        # Определяем сценарий
        scenario = result.get('Сценарий', '')
        scenario_type = 'A' if 'Увеличение добычи' in scenario else 'B'
        
        # Сохраняем изменения
        success = save_kpr_optimization_to_system(well_name, new_schedule, scenario_type)
        
        if success:
            stats['saved'] += 1
        else:
            stats['errors'] += 1
    
    return stats
    
# ============================================================
# ОСНОВНОЙ КЛАСС С ПРАВИЛЬНОЙ МАТЕМАТИКОЙ
# ============================================================

class PressureStabilizationOptimizer:
    """
    Класс для оптимизации фаз КПР скважин для стабилизации давления.
    Использует математику: 60% - давление, 30% - равномерность, 10% - пики
    """
    
    def __init__(self, wells_data, target_coefficient=0.7):
        self.wells = wells_data
        self.active_wells = [w for w in wells_data if w.get('is_active', True)]
        self.kpr_wells = [w for w in self.active_wells if w.get('operation_mode') == 'kpr']
        self.constant_wells = [w for w in self.active_wells if w.get('operation_mode') == 'constant']
        self.target_coefficient = target_coefficient
        
        # Автоматический расчет целевого дебита
        self.target_flow = self._calculate_target_flow()
        
        # Предрасчет дебитов для скорости
        self._precalculate_flows()
    
    def _calculate_target_flow(self):
        """Рассчитывает целевой дебит (м³/час) для оптимального давления"""
        total_capacity = 0
        
        for well in self.active_wells:
            if well.get('operation_mode') == 'kpr':
                try:
                    work_time, pause_time = well['schedule']
                    duty_cycle = work_time / (work_time + pause_time)
                    # Мгновенный дебит во время работы
                    instant_flow = well['flow_rate'] / duty_cycle  # м³/сут
                except:
                    instant_flow = well['flow_rate']  # если ошибка, используем как есть
            else:
                instant_flow = well['flow_rate']  # постоянные скважины
            
            total_capacity += instant_flow  # м³/сут
        
        # Переводим в м³/час и берем заданный коэффициент от мощности
        target_flow_m3_per_hour = (total_capacity * self.target_coefficient) / 24
        return target_flow_m3_per_hour

    def calculate_working_wells_count(self, phases_dict):
        """
        Рассчитывает количество работающих скважин в каждый момент времени
        phases_dict: {имя_скважины: сдвиг_фазы_в_минутах}
        """
        # Временные точки (те же, что и для дебита)
        time_points = np.arange(0, 24 * 60, 5)  # 288 точек
        time_hours = time_points / 60
        
        working_count_before = []
        working_count_after = []
        
        # Нулевые сдвиги для "до"
        zero_phases = {name: 0 for name in phases_dict.keys()}
        
        for t in time_points:
            # Считаем работающие скважины ДО оптимизации
            count_before = 0
            for well_name, well_data in self.well_flows.items():
                if well_data['type'] == 'constant':
                    # Постоянные всегда работают
                    count_before += 1
                else:
                    # КПР скважины
                    work_time, pause_time = well_data['schedule']
                    cycle_time = work_time + pause_time
                    
                    base_time = well_data['base_time']
                    launch_time = (base_time + zero_phases.get(well_name, 0)) % (24 * 60)
                    
                    time_in_cycle = (t - launch_time) % cycle_time
                    if time_in_cycle < 0:
                        time_in_cycle += cycle_time
                    
                    if time_in_cycle < work_time:
                        count_before += 1
            
            # Считаем работающие скважины ПОСЛЕ оптимизации
            count_after = 0
            for well_name, well_data in self.well_flows.items():
                if well_data['type'] == 'constant':
                    count_after += 1
                else:
                    work_time, pause_time = well_data['schedule']
                    cycle_time = work_time + pause_time
                    
                    base_time = well_data['base_time']
                    launch_time = (base_time + phases_dict.get(well_name, 0)) % (24 * 60)
                    
                    time_in_cycle = (t - launch_time) % cycle_time
                    if time_in_cycle < 0:
                        time_in_cycle += cycle_time
                    
                    if time_in_cycle < work_time:
                        count_after += 1
            
            working_count_before.append(count_before)
            working_count_after.append(count_after)
        
        return {
            'time_hours': time_hours,
            'count_before': working_count_before,
            'count_after': working_count_after
        }
    
    def _precalculate_flows(self):
        """Предрасчет дебитов для каждой скважины (м³/час)"""
        self.well_flows = {}  # мгновенный дебит во время работы
        
        for well in self.active_wells:
            name = well['name']
            
            if well.get('operation_mode') == 'kpr':
                try:
                    work_time, pause_time = well['schedule']
                    duty_cycle = work_time / (work_time + pause_time)
                    # Мгновенный дебит во время работы в м³/час
                    instant_flow_hour = well['flow_rate'] / duty_cycle / 24
                    self.well_flows[name] = {
                        'flow': instant_flow_hour,
                        'type': 'kpr',
                        'schedule': well['schedule'],
                        'base_time': self._time_to_minutes(well.get('base_launch_time', '00:00')),
                        'mode': well.get('mode', 'По времени'),
                        'exclude_from_shift': well.get('exclude_from_shift', False)
                    }
                except:
                    # На случай ошибки в данных
                    self.well_flows[name] = {
                        'flow': well['flow_rate'] / 24,  # м³/час
                        'type': 'kpr',
                        'schedule': (15, 45),
                        'base_time': 0,
                        'mode': 'По времени',
                        'exclude_from_shift': well.get('exclude_from_shift', False)
                    }
            else:
                # Постоянная скважина
                self.well_flows[name] = {
                    'flow': well['flow_rate'] / 24,  # м³/час
                    'type': 'constant',
                    'schedule': None,
                    'base_time': None,
                    'mode': None,
                    'exclude_from_shift': False
                }
    
    def _time_to_minutes(self, time_str):
        """Конвертирует время 'ЧЧ:ММ' в минуты"""
        try:
            hours, minutes = map(int, time_str.split(':'))
            return hours * 60 + minutes
        except:
            return 0
    
    def calculate_total_flow_at_time(self, t_minutes, phases_dict):
        """
        Рассчитывает суммарный дебит в момент времени t (минуты)
        phases_dict: {имя_скважины: сдвиг_фазы_в_минутах}
        """
        total_flow = 0.0
        
        for well_name, well_data in self.well_flows.items():
            if well_data['type'] == 'constant':
                # Постоянная скважина всегда работает
                total_flow += well_data['flow']
            else:
                # КПР скважина - проверяем работает ли в момент t
                work_time, pause_time = well_data['schedule']
                cycle_time = work_time + pause_time
                
                # Получаем сдвиг фазы для этой скважины
                phase_shift = phases_dict.get(well_name, 0)
                
                # Абсолютное время запуска с учетом сдвига
                base_time = well_data['base_time']
                launch_time = (base_time + phase_shift) % (24 * 60)
                
                # Определяем, работает ли скважина в момент t
                time_in_cycle = (t_minutes - launch_time) % cycle_time
                if time_in_cycle < 0:
                    time_in_cycle += cycle_time
                
                if time_in_cycle < work_time:
                    total_flow += well_data['flow']
        
        return total_flow
    
    def objective_function(self, phases_array):
        """
        Целевая функция для оптимизации давления
        phases_array: массив сдвигов фаз в минутах для каждой КПР скважины
        """
        # Создаем словарь фаз {имя_скважины: сдвиг}
        phases_dict = {}
        kpr_names = []
        
        # Собираем только те КПР скважины, которые участвуют в оптимизации
        for well in self.kpr_wells:
            if not well.get('exclude_from_shift', False):  # Исключаем скважины по галочке
                kpr_names.append(well['name'])
        
        for i, well_name in enumerate(kpr_names):
            if i < len(phases_array):
                phases_dict[well_name] = phases_array[i]
            else:
                phases_dict[well_name] = 0
        
        # Для скважин с exclude_from_shift = True, используем нулевой сдвиг
        for well in self.kpr_wells:
            if well.get('exclude_from_shift', False):
                phases_dict[well['name']] = 0
        
        # Рассчитываем поток на 24 часа с шагом 5 минут
        time_points = np.arange(0, 24 * 60, 5)  # 288 точек
        total_flows = []
        
        for t in time_points:
            flow = self.calculate_total_flow_at_time(t, phases_dict)
            total_flows.append(flow)
        
        total_flows = np.array(total_flows)
        avg = np.mean(total_flows)
        std = np.std(total_flows)
        
        # 1. Штраф за отклонение от целевого дебита (60%)
        if self.target_flow > 0:
            pressure_penalty = abs(avg - self.target_flow) / self.target_flow
        else:
            pressure_penalty = 1.0
        
        # 2. Штраф за неравномерность (30%)
        if avg > 0:
            uniformity_penalty = std / avg
        else:
            uniformity_penalty = 1.0
        
        # 3. Штраф за пиковые нагрузки (10%) - robust метод
        peak_threshold = avg + 1.5 * std
        peaks = total_flows[total_flows > peak_threshold]
        
        if len(peaks) > 0:
            magnitude = (np.mean(peaks) - avg) / avg
            frequency = len(peaks) / len(total_flows)
            peak_penalty = 0.7 * magnitude + 0.3 * frequency
        else:
            peak_penalty = 0
        
        # Итоговая целевая функция с весами
        total_penalty = (0.6 * pressure_penalty + 
                        0.3 * uniformity_penalty + 
                        0.1 * peak_penalty)
        
        return total_penalty
    
    def optimize(self):
        """Основная функция оптимизации"""
        _load_scipy()
        # Собираем только КПР скважины, которые участвуют в оптимизации
        kpr_wells_for_optimization = [w for w in self.kpr_wells 
                                     if not w.get('exclude_from_shift', False)]
        
        if not kpr_wells_for_optimization:
            return {}, 0, "no_kpr_wells_for_optimization"
        
        # Подготовка границ для каждой КПР скважины
        bounds = []
        kpr_names = []
        
        for well in kpr_wells_for_optimization:
            kpr_names.append(well['name'])
            base_time = self._time_to_minutes(well.get('base_launch_time', '00:00'))
            
            # Границы: ±30 минут или 30% цикла (что меньше)
            try:
                work_time, pause_time = well['schedule']
                cycle = work_time + pause_time
                bound = min(30, cycle * 0.3)  # Не более 30% цикла
            except:
                bound = 30  # если ошибка в расписании
            
            bounds.append((-bound, bound))
        
        # Запуск оптимизации
        result = differential_evolution(
            self.objective_function,
            bounds,
            strategy='best1bin',
            maxiter=100,
            popsize=15,
            mutation=(0.5, 1.0),
            recombination=0.7,
            tol=0.001,
            disp=False,
            seed=42
        )
        
        # Формируем результат в виде словаря
        optimal_phases = {}
        
        # Сначала добавляем сдвиги для участвующих скважин
        for i, well_name in enumerate(kpr_names):
            if i < len(result.x):
                optimal_phases[well_name] = float(result.x[i])
            else:
                optimal_phases[well_name] = 0.0
        
        # Затем добавляем нулевые сдвиги для исключенных скважин
        for well in self.kpr_wells:
            if well.get('exclude_from_shift', False):
                optimal_phases[well['name']] = 0.0
        
        # Рассчитываем статистику
        stats = self.calculate_statistics(optimal_phases)
        
        return optimal_phases, result.fun, stats
    
    def calculate_statistics(self, phases_dict):
        """Рассчитывает статистику после оптимизации"""
        # Поток до оптимизации (нулевые сдвиги)
        zero_phases = {name: 0 for name in phases_dict.keys()}
        
        time_points = np.arange(0, 24 * 60, 5)
        
        flows_before = []
        flows_after = []
        
        for t in time_points:
            flow_before = self.calculate_total_flow_at_time(t, zero_phases)
            flow_after = self.calculate_total_flow_at_time(t, phases_dict)
            flows_before.append(flow_before)
            flows_after.append(flow_after)
        
        flows_before = np.array(flows_before)
        flows_after = np.array(flows_after)
        
        avg_before = np.mean(flows_before)
        avg_after = np.mean(flows_after)
        std_before = np.std(flows_before)
        std_after = np.std(flows_after)
        
        # Расчет пиков до/после
        peak_threshold_before = avg_before + 1.5 * std_before
        peak_threshold_after = avg_after + 1.5 * std_after
        
        peaks_before = flows_before[flows_before > peak_threshold_before]
        peaks_after = flows_after[flows_after > peak_threshold_after]
        
        # Расчет улучшений
        flow_improvement = ((avg_after - avg_before) / avg_before * 100) if avg_before > 0 else 0
        stability_improvement = ((std_before - std_after) / std_before * 100) if std_before > 0 else 0
        peaks_improvement = ((len(peaks_before) - len(peaks_after)) / len(peaks_before) * 100) if len(peaks_before) > 0 else 100
        
        # Эффективность по формуле давления
        target_deviation = abs(avg_after - self.target_flow) / self.target_flow if self.target_flow > 0 else 1
        target_achievement = max(0, 100 * (1 - target_deviation))
        
        efficiency = (
            0.6 * target_achievement +           # Достижение цели (60%)
            0.3 * min(100, max(0, stability_improvement)) +  # Стабильность (30%)
            0.1 * min(100, max(0, flow_improvement))         # Дебит (10%)
        )
        
        return {
            'target_flow': self.target_flow,
            'target_coefficient': self.target_coefficient,
            'avg_flow_before': avg_before,
            'avg_flow_after': avg_after,
            'std_before': std_before,
            'std_after': std_after,
            'peaks_before': len(peaks_before),
            'peaks_after': len(peaks_after),
            'peak_magnitude_before': np.mean(peaks_before) if len(peaks_before) > 0 else 0,
            'peak_magnitude_after': np.mean(peaks_after) if len(peaks_after) > 0 else 0,
            'flow_improvement': flow_improvement,
            'stability_improvement': stability_improvement,
            'peaks_improvement': peaks_improvement,
            'target_achievement': target_achievement,
            'efficiency': efficiency,
            'pressure_penalty': target_deviation * 100
        }

# ============================================================
# КЛАСС ОПТИМИЗАЦИИ НАГРУЗКИ НА СИСТЕМУ СБОРА
# ============================================================

@st.cache_data(ttl=3600)
def prepare_load_data_for_plotting(current_load, v_min_allowed, v_max_allowed):
    """Подготавливает данные для графика - это чистые данные, которые можно кэшировать"""
    return {
        'time_hours': current_load['time_hours'].tolist(),
        'total_load': current_load['total_load'].tolist(),
        'avg_load': float(np.mean(current_load['total_load'])),
        'v_min': v_min_allowed,
        'v_max': v_max_allowed
    }
        
class SystemLoadOptimizer:
    """
    Анализатор и оптимизатор нагрузки на систему сбора нефти.
    Теперь с учетом ограничений по скорости потока в трубопроводе.
    """
    
    def __init__(self, wells_data, selected_clusters=None, time_step_minutes=5):
        """
        Инициализация оптимизатора.
        """
        self.all_wells = wells_data
        self.selected_clusters = selected_clusters if selected_clusters else []
        self.time_step = time_step_minutes
        
        # Фильтруем скважины
        self.filtered_wells = self.filter_wells_by_clusters(selected_clusters)
        
        # Разделяем на типы
        self.kpr_wells = [w for w in self.filtered_wells 
                         if w.get('operation_mode') == 'kpr' and w.get('is_active', True)]
        self.constant_wells = [w for w in self.filtered_wells 
                              if w.get('operation_mode') == 'constant' and w.get('is_active', True)]
        
        # Временные точки для анализа
        self.time_points = np.arange(0, 24 * 60, time_step_minutes)  # минуты
        self.time_hours = self.time_points / 60  # часы
        
        # Исключенные интервалы
        self.excluded_intervals = [
            (13 * 60, 14 * 60),    # Обед 13:00-14:00
            (8 * 60, 8 * 60 + 30), # Утренняя пересменка 8:00-8:30
            (20 * 60, 20 * 60 + 30) # Вечерняя пересменка 20:00-20:30
        ]
        
        # Параметры трубопровода (будут установлены позже)
        self.pipeline_diameter_mm = None
        self.pipeline_wall_thickness_mm = None
        self.v_min_allowed = None
        self.v_max_allowed = None
        
        # Свойства смеси
        self.mixture_properties = self._calculate_mixture_properties()
        
        # Статистика
        self.results = {
            'current_load': None,
            'optimized_load': None,
            'optimal_phases': None,
            'comparison_stats': None,
            'tech_map': None,
            'sampling_schedule': None,
            'velocity_analysis': None
        }
    
    # ====================== ВСПОМОГАТЕЛЬНЫЕ МЕТОДЫ ======================
    
    def filter_wells_by_clusters(self, cluster_names):
        """Фильтрация скважин по названиям кустов."""
        if not cluster_names:
            return [w for w in self.all_wells if w.get('is_active', True)]
        
        filtered = []
        for well in self.all_wells:
            if not well.get('is_active', True):
                continue
                
            well_cluster = well.get('cluster', '')
            if well_cluster in cluster_names:
                filtered.append(well)
        
        return filtered
    
    def _time_to_minutes(self, time_str):
        """Конвертирует время 'ЧЧ:ММ' в минуты."""
        try:
            hours, minutes = map(int, time_str.split(':'))
            return hours * 60 + minutes
        except:
            return 0
    
    def _minutes_to_time(self, minutes):
        """Конвертирует минуты в строку времени 'ЧЧ:ММ'."""
        total_minutes = int(minutes) % (24 * 60)
        hours = total_minutes // 60
        mins = total_minutes % 60
        return f"{hours:02d}:{mins:02d}"
    
    def _calculate_mixture_properties(self):
        """Расчет усредненных свойств смеси для выбранных скважин."""
        if not self.filtered_wells:
            return None
        
        total_flow = 0
        weighted_oil_density = 0
        weighted_water_cut = 0
        
        for well in self.filtered_wells:
            flow = well.get('flow_rate', 0)
            oil_density_rel = well.get('oil_density')
            
            # Обработка None значений
            if oil_density_rel is None:
                oil_density_rel = 0.85  # значение по умолчанию
                
            water_cut = well.get('water_cut', 0)
            
            total_flow += flow
            weighted_oil_density += flow * oil_density_rel
            weighted_water_cut += flow * water_cut
        
        if total_flow > 0:
            avg_oil_density_rel = weighted_oil_density / total_flow
            avg_water_cut = weighted_water_cut / total_flow
        else:
            avg_oil_density_rel = 0.85
            avg_water_cut = 0
        
        # Плотность смеси
        mixture_density = calculate_mixture_density(avg_oil_density_rel, avg_water_cut)
        
        # Рекомендуемые скорости на основе обводненности
        if avg_water_cut > 70:
            recommended_min, recommended_max = 0.5, 3.0
            velocity_desc = "Обводненность >70% (вода)"
        elif avg_water_cut > 30:
            recommended_min, recommended_max = 0.6, 2.0
            velocity_desc = "Обводненность 30-70% (эмульсия)"
        else:
            recommended_min, recommended_max = 0.7, 2.5
            velocity_desc = "Обводненность <30% (нефть)"
        
        return {
            'avg_oil_density_relative': avg_oil_density_rel,
            'avg_water_cut_percent': avg_water_cut,
            'mixture_density_kg_m3': mixture_density,
            'recommended_min_velocity': recommended_min,
            'recommended_max_velocity': recommended_max,
            'velocity_description': velocity_desc,
            'total_flow_m3_per_day': total_flow
        }
    
    # ====================== НАСТРОЙКА ТРУБОПРОВОДА ======================
    
    def set_pipeline_parameters(self, diameter_mm, v_min=None, v_max=None):
        """
        Установка параметров трубопровода.
        """
        self.pipeline_diameter_mm = diameter_mm
        
        # Определяем толщину стенки
        if str(diameter_mm) in STANDARD_PIPELINE_DIAMETERS:
            self.pipeline_wall_thickness_mm = STANDARD_PIPELINE_DIAMETERS[str(diameter_mm)]['wall_thickness']
        else:
            self.pipeline_wall_thickness_mm = 8  # по умолчанию
        
        # Устанавливаем скорости
        if v_min is None and self.mixture_properties:
            self.v_min_allowed = self.mixture_properties['recommended_min_velocity']
        else:
            self.v_min_allowed = v_min or 0.7
        
        if v_max is None and self.mixture_properties:
            self.v_max_allowed = self.mixture_properties['recommended_max_velocity']
        else:
            self.v_max_allowed = v_max or 2.5
    
    # ====================== РАСЧЕТ СКОРОСТИ ======================
    
    def calculate_velocity_profile(self, load_array):
        """
        Расчет профиля скорости для массива нагрузки.
        
        Parameters:
        -----------
        load_array : np.array
            Массив нагрузки в м³/час
            
        Returns:
        --------
        dict: Профиль скорости
        """
        if self.pipeline_diameter_mm is None:
            raise ValueError("Сначала установите параметры трубопровода через set_pipeline_parameters()")
        
        velocities = []
        velocity_data = []
        
        for load in load_array:
            result = calculate_flow_velocity(
                load, 
                self.pipeline_diameter_mm,
                self.pipeline_wall_thickness_mm
            )
            
            velocity_m_s = result['velocity_m_s']
            velocities.append(velocity_m_s)
            
            # Определяем статус скорости
            if velocity_m_s > self.v_max_allowed * 1.2:
                status = "КРИТИЧЕСКОЕ ПРЕВЫШЕНИЕ"
                color = "red"
            elif velocity_m_s > self.v_max_allowed:
                status = "ПРЕВЫШЕНИЕ"
                color = "orange"
            elif velocity_m_s < self.v_min_allowed:
                status = "НИЗКАЯ"
                color = "yellow"
            else:
                status = "НОРМА"
                color = "green"
            
            velocity_data.append({
                'velocity_m_s': velocity_m_s,
                'status': status,
                'color': color,
                'load_m3_per_hour': load
            })
        
        # Статистика по скорости
        if velocities:
            velocity_stats = {
                'max_velocity': max(velocities),
                'min_velocity': min(velocities),
                'avg_velocity': np.mean(velocities),
                'std_velocity': np.std(velocities),
                'high_velocity_count': sum(1 for v in velocities if v > self.v_max_allowed),
                'low_velocity_count': sum(1 for v in velocities if v < self.v_min_allowed),
                'normal_velocity_count': sum(1 for v in velocities if self.v_min_allowed <= v <= self.v_max_allowed)
            }
        else:
            velocity_stats = {}
        
        return {
            'velocities': np.array(velocities),
            'velocity_data': velocity_data,
            'velocity_stats': velocity_stats,
            'v_min_allowed': self.v_min_allowed,
            'v_max_allowed': self.v_max_allowed
        }
    
    # ====================== РАСЧЕТ НАГРУЗКИ ======================
    
    def calculate_well_flow_at_time(self, well_data, current_time_minutes, phase_shift=0):
        """Расчет дебита скважины в конкретный момент времени."""
        # Для постоянных скважин - всегда работает
        if well_data.get('operation_mode') == 'constant':
            return well_data.get('flow_rate', 0) / 24  # м³/сут -> м³/час
        
        # Для КПР скважин - проверяем работает ли в данный момент
        schedule = well_data.get('schedule')
        if not schedule or len(schedule) < 2:
            return 0
        
        work_time, pause_time = schedule
        cycle_time = work_time + pause_time
        
        # Базовое время запуска
        base_launch = self._time_to_minutes(well_data.get('base_launch_time', '00:00'))
        
        # С учетом сдвига фазы
        actual_launch = (base_launch + phase_shift) % (24 * 60)
        
        # Определяем положение в цикле
        time_in_cycle = (current_time_minutes - actual_launch) % cycle_time
        if time_in_cycle < 0:
            time_in_cycle += cycle_time
        
        # Если в периоде работы - возвращаем дебит
        if time_in_cycle < work_time:
            duty_cycle = work_time / cycle_time if cycle_time > 0 else 1
            instant_flow = well_data.get('flow_rate', 0) / duty_cycle / 24  # м³/час
            return instant_flow
        else:
            return 0
    
    def calculate_current_hourly_load(self):
        """Расчет текущей нагрузки на систему по часам."""
        # Инициализация массивов
        total_load = np.zeros(len(self.time_points))
        constant_load = np.zeros(len(self.time_points))
        kpr_load = np.zeros(len(self.time_points))
        
        # Расчет нагрузки от постоянных скважин
        for well in self.constant_wells:
            well_flow = well.get('flow_rate', 0) / 24  # м³/час
            constant_load += well_flow
        
        # Расчет нагрузки от КПР скважин (без сдвигов)
        for well in self.kpr_wells:
            for i, t in enumerate(self.time_points):
                flow = self.calculate_well_flow_at_time(well, t, phase_shift=0)
                kpr_load[i] += flow
        
        # Суммарная нагрузка
        total_load = constant_load + kpr_load
        
        # Расчет скорости (если задан диаметр)
        velocity_profile = None
        if self.pipeline_diameter_mm is not None:
            try:
                velocity_profile = self.calculate_velocity_profile(total_load)
            except Exception as e:
                st.error(f"Ошибка расчета скорости: {str(e)}")
                velocity_profile = None
        
        # Статистика нагрузки
        stats = self._calculate_load_statistics(total_load)
        
        result = {
            'time_points': self.time_points.copy(),
            'time_hours': self.time_hours.copy(),
            'total_load': total_load,
            'constant_load': constant_load,
            'kpr_load': kpr_load,
            'stats': stats,
            'velocity_profile': velocity_profile
        }
        
        self.results['current_load'] = result
        return result
    
    def _calculate_load_statistics(self, load_array):
        """Расчет статистики нагрузки."""
        if len(load_array) == 0:
            return {}
        
        return {
            'max_load': np.max(load_array),
            'min_load': np.min(load_array),
            'avg_load': np.mean(load_array),
            'std_load': np.std(load_array),
            'load_range': np.max(load_array) - np.min(load_array),
            'cv_load': np.std(load_array) / np.mean(load_array) if np.mean(load_array) > 0 else 0
        }
    
    # ====================== АНАЛИЗ ПРОБЛЕМ ======================
    
    def find_problem_intervals(self, load_data=None):
        """
        Поиск проблемных интервалов (по дебиту и скорости).
        
        Returns:
        --------
        dict: Проблемные интервалы
        """
        if load_data is None:
            if self.results['current_load'] is None:
                self.calculate_current_hourly_load()
            load_data = self.results['current_load']
        
        total_load = load_data['total_load']
        time_points = load_data['time_points']
        
        # 1. Проблемы по дебиту (перегрузка/недогрузка)
        # Максимальная теоретическая мощность
        max_possible_load = sum(w.get('flow_rate', 0) / 24 for w in self.filtered_wells 
                               if w.get('is_active', True))
        
        threshold_high = max_possible_load * 0.8  # 80% от максимума
        threshold_low = max_possible_load * 0.4   # 40% от максимума
        
        # Находим интервалы
        flow_problems = self._find_interval_problems(
            total_load, time_points, 
            threshold_high, threshold_low,
            "ПЕРЕГРУЗКА", "НЕДОГРУЗКА"
        )
        
        # 2. Проблемы по скорости (если есть данные)
        velocity_problems = {}
        if load_data.get('velocity_profile'):
            velocities = load_data['velocity_profile']['velocities']
            
            # Проблемы по скорости
            velocity_problems = self._find_interval_problems(
                velocities, time_points,
                self.v_max_allowed, self.v_min_allowed,
                "ВЫСОКАЯ СКОРОСТЬ", "НИЗКАЯ СКОРОСТЬ"
            )
        
        # Объединяем результаты
        problems = {
            'flow_problems': flow_problems,
            'velocity_problems': velocity_problems,
            'max_possible_load': max_possible_load,
            'threshold_high': threshold_high,
            'threshold_low': threshold_low
        }
        
        # Добавляем статистику скорости
        if load_data.get('velocity_profile'):
            problems.update(load_data['velocity_profile']['velocity_stats'])
        
        return problems
    
    def _find_interval_problems(self, values, time_points, 
                               threshold_high, threshold_low,
                               problem_name_high, problem_name_low):
        """Вспомогательная функция для поиска проблемных интервалов."""
        high_intervals = []
        low_intervals = []
        
        # Поиск интервалов превышения
        in_high = False
        high_start = None
        
        for i, (t, value) in enumerate(zip(time_points, values)):
            if value > threshold_high and not in_high:
                in_high = True
                high_start = t
            elif value <= threshold_high and in_high:
                in_high = False
                if high_start is not None:
                    high_intervals.append({
                        'start': high_start,
                        'end': t,
                        'start_time': self._minutes_to_time(high_start),
                        'end_time': self._minutes_to_time(t),
                        'duration_min': t - high_start,
                        'max_value': max(values[max(0, i-10):i+1]),
                        'problem_type': problem_name_high
                    })
                    high_start = None
        
        # Если остались в превышении
        if in_high and high_start is not None:
            high_intervals.append({
                'start': high_start,
                'end': 24 * 60,
                'start_time': self._minutes_to_time(high_start),
                'end_time': "23:59",
                'duration_min': 24 * 60 - high_start,
                'max_value': max(values[-10:]),
                'problem_type': problem_name_high
            })
        
        # Поиск интервалов недогрузки
        in_low = False
        low_start = None
        
        for i, (t, value) in enumerate(zip(time_points, values)):
            if value < threshold_low and not in_low:
                in_low = True
                low_start = t
            elif value >= threshold_low and in_low:
                in_low = False
                if low_start is not None:
                    low_intervals.append({
                        'start': low_start,
                        'end': t,
                        'start_time': self._minutes_to_time(low_start),
                        'end_time': self._minutes_to_time(t),
                        'duration_min': t - low_start,
                        'min_value': min(values[max(0, i-10):i+1]),
                        'problem_type': problem_name_low
                    })
                    low_start = None
        
        if in_low and low_start is not None:
            low_intervals.append({
                'start': low_start,
                'end': 24 * 60,
                'start_time': self._minutes_to_time(low_start),
                'end_time': "23:59",
                'duration_min': 24 * 60 - low_start,
                'min_value': min(values[-10:]),
                'problem_type': problem_name_low
            })
        
        return {
            'high_intervals': high_intervals,
            'low_intervals': low_intervals,
            'threshold_high': threshold_high,
            'threshold_low': threshold_low
        }
    
    # ====================== ОПТИМИЗАЦИЯ ======================
    
    def _objective_function_with_velocity(self, phases_array):
        """
        Целевая функция с учетом ограничений по скорости.
        
        Веса:
        - 40%: Минимизация размаха нагрузки
        - 20%: Избегание пиков дебита
        - 25%: Соблюдение скоростных ограничений
        - 10%: Исключенные интервалы
        - 5%:  Стабильность смеси
        """
        # Создаем словарь фаз
        phases_dict = {}
        for i, well in enumerate(self.kpr_wells):
            if i < len(phases_array):
                # Исключаем скважины, работающие по давлению
                if not well.get('exclude_from_shift', False):
                    phases_dict[well['name']] = phases_array[i]
                else:
                    phases_dict[well['name']] = 0
        
        # Рассчитываем нагрузку с этими фазами
        total_load = np.zeros(len(self.time_points))
        
        # Постоянные скважины
        for well in self.constant_wells:
            well_flow = well.get('flow_rate', 0) / 24
            total_load += well_flow
        
        # КПР скважины со сдвигами
        for well in self.kpr_wells:
            well_name = well['name']
            phase_shift = phases_dict.get(well_name, 0)
            
            for i, t in enumerate(self.time_points):
                flow = self.calculate_well_flow_at_time(well, t, phase_shift)
                total_load[i] += flow
        
        # 1. Штраф за размах нагрузки (40%)
        load_range = np.max(total_load) - np.min(total_load)
        max_possible_range = sum(w.get('flow_rate', 0) / 24 for w in self.filtered_wells)
        
        if max_possible_range > 0:
            range_penalty = load_range / max_possible_range
        else:
            range_penalty = 1.0
        
        # 2. Штраф за пиковые нагрузки (20%)
        avg_load = np.mean(total_load)
        peak_threshold = avg_load * 1.3
        peaks = total_load[total_load > peak_threshold]
        
        if len(peaks) > 0:
            peak_magnitude = np.mean(peaks) / avg_load - 1
            peak_frequency = len(peaks) / len(total_load)
            peak_penalty = 0.7 * peak_magnitude + 0.3 * peak_frequency
        else:
            peak_penalty = 0
        
        # 3. Штраф за нарушение скоростных ограничений (25%)
        velocity_penalty = 0
        if self.pipeline_diameter_mm is not None:
            for load in total_load:
                result = calculate_flow_velocity(
                    load, 
                    self.pipeline_diameter_mm,
                    self.pipeline_wall_thickness_mm
                )
                velocity = result['velocity_m_s']
                
                # Штраф за превышение
                if velocity > self.v_max_allowed:
                    excess = (velocity - self.v_max_allowed) / self.v_max_allowed
                    velocity_penalty += excess ** 2  # квадратичный штраф
                
                # Штраф за слишком низкую скорость
                if velocity < self.v_min_allowed:
                    deficit = (self.v_min_allowed - velocity) / self.v_min_allowed
                    velocity_penalty += deficit * 0.5
        
        velocity_penalty = min(velocity_penalty / len(total_load), 1.0)
        
        # 4. Штраф за запуски в исключенные интервалы (10%)
        exclusion_penalty = 0
        for well in self.kpr_wells:
            well_name = well['name']
            phase_shift = phases_dict.get(well_name, 0)
            base_launch = self._time_to_minutes(well.get('base_launch_time', '00:00'))
            actual_launch = (base_launch + phase_shift) % (24 * 60)
            
            # Проверяем каждый исключенный интервал
            for excl_start, excl_end in self.excluded_intervals:
                if excl_start <= actual_launch <= excl_end:
                    exclusion_penalty += 0.1
        
        exclusion_penalty = min(exclusion_penalty / len(self.kpr_wells), 1.0)
        
        # 5. Штраф за нестабильность смеси (5%)
        # (если сильно меняется обводненность из-за запусков)
        stability_penalty = 0
        # Пока оставляем 0, можно добавить позже
        
        # Итоговая целевая функция
        total_penalty = (0.4 * range_penalty + 
                        0.2 * peak_penalty + 
                        0.25 * velocity_penalty + 
                        0.1 * exclusion_penalty + 
                        0.05 * stability_penalty)
        
        return total_penalty
    
    def optimize_launch_times(self, optimization_method="genetic", max_iterations=50):
        """
        Оптимизация времени запуска КПР скважин.
        
        Returns:
        --------
        dict: Результаты оптимизации
        """
        if not self.kpr_wells:
            return {"error": "Нет КПР скважин для оптимизации"}
        
        # Фильтруем только скважины, которые участвуют в оптимизации
        optimizable_wells = [w for w in self.kpr_wells 
                            if not w.get('exclude_from_shift', False)]
        
        if not optimizable_wells:
            return {"error": "Все КПР скважины исключены из оптимизации"}
        
        # Границы для сдвигов (в минутах)
        bounds = []
        for well in optimizable_wells:
            schedule = well.get('schedule', [15, 45])
            cycle_time = schedule[0] + schedule[1]
            max_shift = min(cycle_time / 2, 4 * 60)  # Не более 4 часов
            bounds.append((-max_shift, max_shift))
        
        # Оптимизация
        if optimization_method == "differential_evolution":
            result = differential_evolution(
                self._objective_function_with_velocity,
                bounds,
                strategy='best1bin',
                maxiter=max_iterations,
                popsize=15,
                mutation=(0.5, 1.0),
                recombination=0.7,
                tol=0.001,
                disp=False,
                seed=42
            )
            
            optimal_phases_array = result.x
        else:
            # Простой жадный алгоритм для начала
            optimal_phases_array = np.zeros(len(optimizable_wells))
        
        # Создаем полный словарь фаз
        optimal_phases = {}
        phase_idx = 0
        
        for well in self.kpr_wells:
            if well.get('exclude_from_shift', False):
                optimal_phases[well['name']] = 0
            else:
                if phase_idx < len(optimal_phases_array):
                    optimal_phases[well['name']] = float(optimal_phases_array[phase_idx])
                    phase_idx += 1
                else:
                    optimal_phases[well['name']] = 0
        
        # Рассчитываем оптимизированную нагрузку
        self.results['optimal_phases'] = optimal_phases
        self.results['optimized_load'] = self._calculate_optimized_load(optimal_phases)
        
        # Расчет скорости для оптимизированной нагрузки
        if self.pipeline_diameter_mm is not None:
            velocity_profile = self.calculate_velocity_profile(
                self.results['optimized_load']['total_load']
            )
            self.results['optimized_load']['velocity_profile'] = velocity_profile
        
        # Сравниваем с текущей нагрузкой
        if self.results['current_load'] is None:
            self.calculate_current_hourly_load()
        
        self.results['comparison_stats'] = self._compare_loads(
            self.results['current_load'],
            self.results['optimized_load']
        )
        
        # Генерируем документы
        self.results['tech_map'] = self.generate_technology_map(optimal_phases)
        self.results['sampling_schedule'] = self.generate_sampling_schedule(optimal_phases)
        
        return {
            'optimal_phases': optimal_phases,
            'optimized_load': self.results['optimized_load'],
            'comparison_stats': self.results['comparison_stats'],
            'tech_map': self.results['tech_map'],
            'sampling_schedule': self.results['sampling_schedule']
        }
    
    def _calculate_optimized_load(self, phases_dict):
        """Расчет нагрузки после оптимизации."""
        total_load = np.zeros(len(self.time_points))
        kpr_load = np.zeros(len(self.time_points))
        
        # Постоянные скважины
        for well in self.constant_wells:
            well_flow = well.get('flow_rate', 0) / 24
            total_load += well_flow
        
        # КПР скважины с оптимальными фазами
        for well in self.kpr_wells:
            well_name = well['name']
            phase_shift = phases_dict.get(well_name, 0)
            
            for i, t in enumerate(self.time_points):
                flow = self.calculate_well_flow_at_time(well, t, phase_shift)
                kpr_load[i] += flow
                total_load[i] += flow
        
        return {
            'time_points': self.time_points.copy(),
            'time_hours': self.time_hours.copy(),
            'total_load': total_load,
            'kpr_load': kpr_load,
            'stats': self._calculate_load_statistics(total_load)
        }
    
    def _compare_loads(self, current_load, optimized_load):
        """Сравнение нагрузок до и после оптимизации."""
        current_stats = current_load['stats']
        optimized_stats = optimized_load['stats']
        
        improvements = {
            'max_load_reduction': (current_stats['max_load'] - optimized_stats['max_load']) / current_stats['max_load'] * 100 if current_stats['max_load'] > 0 else 0,
            'min_load_increase': (optimized_stats['min_load'] - current_stats['min_load']) / current_stats['min_load'] * 100 if current_stats['min_load'] > 0 else 0,
            'range_reduction': (current_stats['load_range'] - optimized_stats['load_range']) / current_stats['load_range'] * 100 if current_stats['load_range'] > 0 else 0,
            'std_reduction': (current_stats['std_load'] - optimized_stats['std_load']) / current_stats['std_load'] * 100 if current_stats['std_load'] > 0 else 0
        }
        
        # Сравнение скорости (если есть)
        if current_load.get('velocity_profile') and optimized_load.get('velocity_profile'):
            current_vel = current_load['velocity_profile']['velocity_stats']
            optimized_vel = optimized_load['velocity_profile']['velocity_stats']
            
            improvements.update({
                'max_velocity_reduction': (current_vel['max_velocity'] - optimized_vel['max_velocity']) / current_vel['max_velocity'] * 100 if current_vel['max_velocity'] > 0 else 0,
                'high_velocity_count_reduction': (current_vel['high_velocity_count'] - optimized_vel['high_velocity_count']) / max(current_vel['high_velocity_count'], 1) * 100
            })
        
        # Оцениваем общее улучшение
        weights = [0.3, 0.2, 0.3, 0.2]  # веса для разных метрик
        weighted_sum = 0
        metric_values = [
            improvements.get('max_load_reduction', 0),
            improvements.get('min_load_increase', 0),
            improvements.get('range_reduction', 0),
            improvements.get('max_velocity_reduction', improvements.get('std_reduction', 0))
        ]
        
        for metric, weight in zip(metric_values, weights):
            weighted_sum += max(0, min(metric, 100)) * weight
        
        improvements['overall_improvement'] = weighted_sum
        
        return {
            'current_stats': current_stats,
            'optimized_stats': optimized_stats,
            'improvements': improvements,
            'comparison_data': {
                'time_hours': current_load['time_hours'],
                'current_load': current_load['total_load'],
                'optimized_load': optimized_load['total_load']
            }
        }
    
    # ====================== ДОКУМЕНТЫ ======================
    
    def generate_technology_map(self, optimal_phases):
        """Формирование технологической карты."""
        records = []
        
        for well in self.filtered_wells:
            if not well.get('is_active', True):
                continue
            
            record = {
                'Скважина': well['name'],
                'Куст': well.get('cluster', 'Неизвестно'),
                'Тип': 'Постоянная' if well.get('operation_mode') == 'constant' else 'КПР',
                'Дебит_м3_сут': well.get('flow_rate', 0),
                'Обводненность_%': well.get('water_cut', 0),
                'Плотность_нефти': well.get('oil_density', 0.85)
            }
            
            if well.get('operation_mode') == 'kpr':
                schedule = well.get('schedule', [15, 45])
                work_time, pause_time = schedule
                
                # Базовое время запуска
                base_launch = self._time_to_minutes(well.get('base_launch_time', '00:00'))
                
                # С учетом оптимального сдвига
                phase_shift = optimal_phases.get(well['name'], 0)
                actual_launch = (base_launch + phase_shift) % (24 * 60)
                
                # Время останова
                actual_stop = (actual_launch + work_time) % (24 * 60)
                
                record.update({
                    'Время_запуска': self._minutes_to_time(actual_launch),
                    'Время_останова': self._minutes_to_time(actual_stop),
                    'Продолжительность_работы_мин': work_time,
                    'Продолжительность_простоя_мин': pause_time,
                    'Режим': well.get('mode', 'По времени'),
                    'Сдвиг_фазы_мин': phase_shift
                })
            else:
                record.update({
                    'Время_запуска': '00:00',
                    'Время_останова': '23:59',
                    'Продолжительность_работы_мин': 24 * 60,
                    'Продолжительность_простоя_мин': 0,
                    'Режим': 'Постоянный',
                    'Сдвиг_фазы_мин': 0
                })
            
            records.append(record)
        
        df = pd.DataFrame(records)
        
        # Сортируем по времени запуска
        if 'Время_запуска' in df.columns:
            df['Время_запуска_мин'] = df['Время_запуска'].apply(self._time_to_minutes)
            df = df.sort_values('Время_запуска_мин')
            df = df.drop('Время_запуска_мин', axis=1)
        
        return df
    
    def generate_sampling_schedule(self, optimal_phases):
        """Формирование графика отбора проб для КПР скважин."""
        records = []
        
        for well in self.kpr_wells:
            if not well.get('is_active', True):
                continue
            
            schedule = well.get('schedule', [15, 45])
            work_time, _ = schedule
            
            # Базовое время запуска
            base_launch = self._time_to_minutes(well.get('base_launch_time', '00:00'))
            
            # С учетом оптимального сдвига
            phase_shift = optimal_phases.get(well['name'], 0)
            actual_launch = (base_launch + phase_shift) % (24 * 60)
            
            # Время для проб
            sample1 = actual_launch  # Начало работы
            sample2 = actual_launch + work_time / 2  # Середина
            sample3 = actual_launch + work_time  # Конец
            
            # Нормализуем в пределах суток
            sample1 = sample1 % (24 * 60)
            sample2 = sample2 % (24 * 60)
            sample3 = sample3 % (24 * 60)
            
            record = {
                'Скважина': well['name'],
                'Куст': well.get('cluster', 'Неизвестно'),
                'Проба_1_начало': self._minutes_to_time(sample1),
                'Проба_2_середина': self._minutes_to_time(sample2),
                'Проба_3_конец': self._minutes_to_time(sample3),
                'Режим_мин': f"{work_time}/{schedule[1]}",
                'Дебит_м3_сут': well.get('flow_rate', 0),
                'Обводненность_%': well.get('water_cut', 0),
                'Плотность_нефти': well.get('oil_density', 0.85),
                'Сдвиг_фазы_мин': phase_shift,
                'Примечание': 'КПР' if well.get('mode') == 'По времени' else 'КПР (по давлению)'
            }
            
            records.append(record)
        
        df = pd.DataFrame(records)
        
        # Сортируем по времени первой пробы
        if 'Проба_1_начало' in df.columns:
            df['Время_начала_мин'] = df['Проба_1_начало'].apply(self._time_to_minutes)
            df = df.sort_values('Время_начала_мин')
            df = df.drop('Время_начала_мин', axis=1)
        
        return df
    
    # ====================== ВИЗУАЛИЗАЦИЯ ======================
    
    def visualize_hourly_load(self, current_load, optimized_load=None):
        """Визуализация без декоратора, использует кэшированные данные"""
        
        # Подготавливаем данные для кэширования
        data_key = prepare_load_data_for_plotting(
            current_load, 
            self.v_min_allowed, 
            self.v_max_allowed
        )
        
        fig = go.Figure()
        
        # Текущая нагрузка
        fig.add_trace(go.Scatter(
            x=data_key['time_hours'],
            y=data_key['total_load'],
            mode='lines',
            name='📊 Текущая нагрузка',
            line=dict(color='red', width=2, dash='dash'),
            fill='tozeroy',
            fillcolor='rgba(255, 0, 0, 0.1)'
        ))
        
        # Оптимизированная нагрузка (не кэшируется, т.к. меняется)
        if optimized_load is not None:
            fig.add_trace(go.Scatter(
                x=optimized_load['time_hours'],
                y=optimized_load['total_load'],
                mode='lines',
                name='🎯 Оптимизированная нагрузка',
                line=dict(color='blue', width=2),
                fill='tozeroy',
                fillcolor='rgba(0, 0, 255, 0.1)'
            ))
        
        # Средняя нагрузка
        fig.add_hline(
            y=data_key['avg_load'],
            line_dash="dot",
            line_color="green",
            opacity=0.5,
            annotation_text=f"Средняя: {data_key['avg_load']:.1f} м³/час",
            annotation_position="bottom right"
        )
        
        # Максимальная допустимая нагрузка
        if self.pipeline_diameter_mm is not None and self.v_max_allowed is not None:
            max_flow = get_max_flow_for_velocity(
                self.v_max_allowed, 
                self.pipeline_diameter_mm,
                self.pipeline_wall_thickness_mm
            )
            
            fig.add_hline(
                y=max_flow,
                line_dash="dash",
                line_color="orange",
                opacity=0.7,
                annotation_text=f"Макс. при {self.v_max_allowed} м/с: {max_flow:.0f} м³/час",
                annotation_position="top right"
            )
        
        fig.update_layout(
            title='📈 Нагрузка на систему сбора (м³/час)',
            xaxis_title='Время суток, часы',
            yaxis_title='Нагрузка, м³/час',
            hovermode='x unified',
            height=500,
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
        )
        
        return fig

    def visualize_velocity_profile(self, load_data):
        """Визуализация профиля скорости."""
        if not load_data.get('velocity_profile'):
            # Пробуем рассчитать, если еще не рассчитано
            if 'total_load' in load_data and self.pipeline_diameter_mm:
                velocity_profile = self.calculate_velocity_profile(load_data['total_load'])
                load_data['velocity_profile'] = velocity_profile
            else:
                return None
        
        velocity_profile = load_data['velocity_profile']
        velocities = velocity_profile['velocities']
        v_min = velocity_profile['v_min_allowed']
        v_max = velocity_profile['v_max_allowed']
        
        fig = go.Figure()
        
        # Скорость потока
        fig.add_trace(go.Scatter(
            x=load_data['time_hours'],
            y=velocities,
            mode='lines',
            name='Скорость потока',
            line=dict(color='purple', width=2),
            fill='tozeroy',
            fillcolor='rgba(128, 0, 128, 0.2)'
        ))
        
        # Допустимый диапазон
        fig.add_hrect(
            y0=v_min, y1=v_max,
            fillcolor="green", opacity=0.2,
            annotation_text="Допустимый диапазон",
            annotation_position="top left"
        )
        
        # Линии ограничений
        fig.add_hline(
            y=v_max,
            line_dash="dash",
            line_color="red",
            opacity=0.7,
            annotation_text=f"Макс: {v_max} м/с",
            annotation_position="top right"
        )
        
        fig.add_hline(
            y=v_min,
            line_dash="dash", 
            line_color="orange",
            opacity=0.7,
            annotation_text=f"Мин: {v_min} м/с",
            annotation_position="bottom right"
        )
        
        # Статистика
        stats = velocity_profile.get('velocity_stats', {})
        if stats and 'avg_velocity' in stats:
            avg_velocity = stats['avg_velocity']
            fig.add_hline(
                y=avg_velocity,
                line_dash="dot",
                line_color="blue",
                opacity=0.5,
                annotation_text=f"Средняя: {avg_velocity:.2f} м/с",
                annotation_position="bottom left"
            )
        
        fig.update_layout(
            title='📊 Скорость потока в трубопроводе',
            xaxis_title='Время суток, часы',
            yaxis_title='Скорость, м/с',
            hovermode='x unified',
            height=400
        )
        
        return fig
    
    # ====================== EXCEL ОТЧЕТ ======================
    
    def create_excel_report(self):
        """Создание Excel отчета со всеми данными."""
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        output = BytesIO()
        wb = Workbook()
        
        # ================= ЛИСТ 1: СВОДКА =================
        ws1 = wb.active
        ws1.title = "Сводка_результатов"
        
        # Заголовок
        ws1.merge_cells('A1:H1')
        ws1['A1'] = "АНАЛИЗ И ОПТИМИЗАЦИЯ НАГРУЗКИ НА СИСТЕМУ СБОРА"
        ws1['A1'].font = Font(bold=True, size=14)
        ws1['A1'].alignment = Alignment(horizontal='center')
        
        ws1.merge_cells('A2:H2')
        ws1['A2'] = f"Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws1['A2'].font = Font(size=11)
        ws1['A2'].alignment = Alignment(horizontal='center')
        
        # Параметры трубопровода
        ws1['A4'] = "ПАРАМЕТРЫ ТРУБОПРОВОДА"
        ws1['A4'].font = Font(bold=True, size=12)
        
        # БЕЗОПАСНОЕ получение данных
        pipe_data = []
        if self.pipeline_diameter_mm:
            pipe_data.append(("Наружный диаметр, мм", self.pipeline_diameter_mm))
        if self.pipeline_wall_thickness_mm:
            pipe_data.append(("Толщина стенки, мм", self.pipeline_wall_thickness_mm))
        if self.v_min_allowed:
            pipe_data.append(("Минимальная скорость, м/с", self.v_min_allowed))
        if self.v_max_allowed:
            pipe_data.append(("Максимальная скорость, м/с", self.v_max_allowed))
        
        for i, (param, value) in enumerate(pipe_data, start=5):
            ws1.cell(row=i, column=1, value=param).font = Font(bold=True)
            ws1.cell(row=i, column=2, value=value)
        
        # Свойства смеси
        ws1['A10'] = "СВОЙСТВА СМЕСИ"
        ws1['A10'].font = Font(bold=True, size=12)
        
        if self.mixture_properties:
            mix_data = [
                ("Средняя обводненность, %", self.mixture_properties.get('avg_water_cut_percent', 0)),
                ("Средняя плотность нефти", self.mixture_properties.get('avg_oil_density_relative', 0.85)),
                ("Плотность смеси, кг/м³", self.mixture_properties.get('mixture_density_kg_m3', 0)),
                ("Рекомендуемые скорости", self.mixture_properties.get('velocity_description', 'Не определено'))
            ]
            
            for i, (param, value) in enumerate(mix_data, start=11):
                ws1.cell(row=i, column=1, value=param).font = Font(bold=True)
                ws1.cell(row=i, column=2, value=value)
        
        # Статистика улучшений
        ws1['A16'] = "СТАТИСТИКА УЛУЧШЕНИЙ"
        ws1['A16'].font = Font(bold=True, size=12)
        
        # БЕЗОПАСНЫЙ доступ к результатам
        if self.results and self.results.get('comparison_stats'):
            stats = self.results['comparison_stats']
            improvements = stats.get('improvements', {})
            
            headers = ['Показатель', 'До оптимизации', 'После оптимизации', 'Изменение, %']
            for col, header in enumerate(headers, start=1):
                cell = ws1.cell(row=17, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            # Безопасное получение данных
            current_stats = stats.get('current_stats', {})
            optimized_stats = stats.get('optimized_stats', {})
            
            data_rows = [
                ('Максимальная нагрузка, м³/час', 
                 current_stats.get('max_load', 0), 
                 optimized_stats.get('max_load', 0),
                 improvements.get('max_load_reduction', 0)),
                
                ('Минимальная нагрузка, м³/час',
                 current_stats.get('min_load', 0),
                 optimized_stats.get('min_load', 0),
                 improvements.get('min_load_increase', 0)),
                
                ('Размах нагрузки, м³/час',
                 current_stats.get('load_range', 0),
                 optimized_stats.get('load_range', 0),
                 improvements.get('range_reduction', 0))
            ]
            
            for i, row_data in enumerate(data_rows, start=18):
                for j, value in enumerate(row_data, start=1):
                    cell = ws1.cell(row=i, column=j, value=value)
                    if j == 4 and isinstance(value, (int, float)):
                        if value > 0:
                            cell.font = Font(color="00FF00")  # Зеленый
                        elif value < 0:
                            cell.font = Font(color="FF0000")  # Красный
            
            # Общее улучшение
            ws1['A23'] = "ОБЩЕЕ УЛУЧШЕНИЕ"
            ws1['A23'].font = Font(bold=True)
            ws1['B23'] = f"{improvements.get('overall_improvement', 0):.1f}%"
            if improvements.get('overall_improvement', 0) > 20:
                ws1['B23'].font = Font(color="00FF00", bold=True)
        
        # ================= ЛИСТ 2: ТЕХНОЛОГИЧЕСКАЯ КАРТА =================
        ws2 = wb.create_sheet("Технологическая_карта")
        
        if self.results and self.results.get('tech_map') is not None:
            df_tech = self.results['tech_map']
            
            # Заголовок
            ws2.merge_cells('A1:L1')
            ws2['A1'] = "ТЕХНОЛОГИЧЕСКАЯ КАРТА РЕЖИМОВ РАБОТЫ СКВАЖИН"
            ws2['A1'].font = Font(bold=True, size=14)
            ws2['A1'].alignment = Alignment(horizontal='center')
            
            # Заголовки столбцов
            headers = list(df_tech.columns)
            for col, header in enumerate(headers, start=1):
                cell = ws2.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            # Данные
            for r_idx, row in enumerate(df_tech.itertuples(index=False), start=4):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws2.cell(row=r_idx, column=c_idx, value=value)
        
        # ================= ЛИСТ 3: ГРАФИК ОТБОРА ПРОБ =================
        ws3 = wb.create_sheet("График_отбора_проб")
        
        if self.results and self.results.get('sampling_schedule') is not None:
            df_samp = self.results['sampling_schedule']
            
            # Заголовок
            ws3.merge_cells('A1:K1')
            ws3['A1'] = "ГРАФИК ОТБОРА ПРОБ ДЛЯ КПР СКВАЖИН"
            ws3['A1'].font = Font(bold=True, size=14)
            ws3['A1'].alignment = Alignment(horizontal='center')
            
            # Заголовки столбцов
            headers = list(df_samp.columns)
            for col, header in enumerate(headers, start=1):
                cell = ws3.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            # Данные
            for r_idx, row in enumerate(df_samp.itertuples(index=False), start=4):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws3.cell(row=r_idx, column=c_idx, value=value)
        
        # Удаляем дефолтный лист
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Настройка ширины столбцов
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем в BytesIO
        wb.save(output)
        output.seek(0)
        
        return output

# ============================================================
# ФУНКЦИИ ЗАГРУЗКИ ДАННЫХ ИЗ ТЕХРЕЖИМА
# ============================================================

def parse_cdng_cits(cdng_value, default_cits=None):
    """Парсим ЦДНГ и ЦИТС из значения типа 'ЦДНГ-1 (В)'"""
    if not isinstance(cdng_value, str):
        return "ЦДНГ-1", default_cits or "ЦИТС VQ-BAD"
    
    cdng_value = str(cdng_value).strip()
    
    # Определяем ЦИТС по букве в скобках
    cits_map = {
        'В': 'ЦИТС Аган',
        'П': 'ЦИТС VQ-BAD'
    }
    
    # По умолчанию
    cdng = cdng_value
    cits = default_cits or "ЦИТС VQ-BAD"
    
    # Ищем букву в скобках
    if '(' in cdng_value and ')' in cdng_value:
        try:
            # Выделяем часть в скобках
            start = cdng_value.find('(') + 1
            end = cdng_value.find(')')
            letter = cdng_value[start:end].strip()
            
            # Определяем ЦИТС
            if letter.upper() in cits_map:
                cits = cits_map[letter.upper()]
            
            # Убираем скобки из ЦДНГ
            cdng = cdng_value[:start-1].strip()
        except:
            pass
    
    return cdng, cits

def safe_float_convert(value, default=0):
    """Безопасное преобразование в float"""
    if pd.isna(value):
        return default
    
    try:
        # Пробуем преобразовать в float
        return float(value)
    except (ValueError, TypeError):
        # Если не получается (например, дата '11.09.2007'), возвращаем default
        return default
        
@st.cache_data(ttl=3600, show_spinner="Загрузка файла техрежима...")
def load_tech_regime_file(uploaded_file, selected_cits=None):
    """Загружает и парсит файл техрежима"""
    try:
        # Читаем Excel, начиная с 22 строки
        df = pd.read_excel(uploaded_file, skiprows=21, dtype=str, na_filter=False)
        
        st.info(f"Файл прочитан успешно: {len(df)} строк, {len(df.columns)} столбцов")
        
        wells_data = []
        
        # Преобразуем буквенные обозначения столбцов в индексы
        def col_to_index(col_letter):
            """Конвертирует букву столбца в индекс (A=0, B=1, ...)"""
            col_letter = col_letter.upper()
            index = 0
            for char in col_letter:
                index = index * 26 + (ord(char) - ord('A') + 1)
            return index - 1  # Индексы с 0
        
        # ОБНОВЛЕННЫЕ индексы столбцов (добавили R для диаметра ЭК)
        col_indices = {
            'D': col_to_index('D'),   # ЦДНГ - индекс 3
            'H': col_to_index('H'),   # Номер скважины - индекс 7
            'I': col_to_index('I'),   # Куст - индекс 8
            'U': col_to_index('U'),   # Тип установки (ЭЦН/ОРД/ШГН) - индекс 20
            'W': col_to_index('W'),   # Марка насоса с дебитом и напором - индекс 22
            'X': col_to_index('X'),   # Глубина насоса - индекс 23
            'AF': col_to_index('AF'), # Коэффициент подачи - индекс 31
            'AI': col_to_index('AI'), # Наработка на отказ - индекс 34
            'AL': col_to_index('AL'), # Буферное давление (Рб/ф) - индекс 37
            'AN': col_to_index('AN'), # Hдин - индекс 39
            'AQ': col_to_index('AQ'), # Pпр - индекс 42
            'AY': col_to_index('AY'), # Pпл - индекс 50
            'BC': col_to_index('BC'), # Коэффициент продуктивности - индекс 54
            'BD': col_to_index('BD'), # Qж - индекс 55
            'BE': col_to_index('BE'), # Обводненность - индекс 56
            'BG': col_to_index('BG'), # Дни работы+накопление - индекс 32
            'BH': col_to_index('BH'), # Период.режим, час раб - индекс 33
            'BI': col_to_index('BI'), # Период.режим, час накопл - индекс 34
            'BJ': col_to_index('BJ'), # Т раб.мес - индекс 35
            'CX': col_to_index('CX'), # Плотность нефти - индекс 100
            'DC': col_to_index('DC'), # Газовый фактор - индекс 105
            'DE': col_to_index('DE'), # Pнас - индекс 107
            'O': col_to_index('O'),   # Интервал ВДП - индекс 14
            'AO': col_to_index('AO'), # Pзат - индекс 40
            'AT': col_to_index('AT'), # Pзаб ВДП - индекс 44
            
            # ДОБАВЛЕННЫЕ СТОЛБЦЫ:
            'AP': col_to_index('AP'), # Удл. На Ндин - индекс 41
            'AD': col_to_index('AD'), # обороты - индекс 29
            'R': col_to_index('R'),   # Диаметр ЭК - индекс 17 (НОВОЕ!)
        }
        
        # Счетчики для статистики
        stats = {
            'total_rows': len(df),
            'skipped_no_well': 0,
            'skipped_no_flow': 0,
            'skipped_work_hours': 0,
            'processed': 0
        }
        
        # Вспомогательные функции
        def parse_casing_diameter(casing_str):
            """Парсит диаметр эксплуатационной колонны из строки (столбец R)"""
            if not casing_str or pd.isna(casing_str):
                return None
            
            casing_str = str(casing_str).strip()
            
            # Если несколько значений через точку с запятой, берём первое
            if ';' in casing_str:
                parts = casing_str.split(';')
                first_part = parts[0].strip()
                if first_part:
                    try:
                        return float(first_part.replace(',', '.'))
                    except:
                        pass
            
            # Если одно значение
            try:
                return float(casing_str.replace(',', '.'))
            except:
                return None
        
        def extract_pump_gabarit(pump_mark):
            """Извлекает габарит насоса из маркировки"""
            if not pump_mark or pd.isna(pump_mark):
                return '5'  # По умолчанию ЭЦН5
            
            pump_mark = str(pump_mark).upper().strip()
            
            patterns = [
                r'ЭЦН[А-Я]?(\d+[А-Я]?)',          # ЭЦН5, ЭЦН5А, ЭЦН5-...
                r'(\d+[А-Я]?)[Лл]?[Ээ][Цц][Нн]',  # 5ЭЦН, 5АЭЦН
                r'(\d+[А-Я]?)[-_]',               # 5-125, 5А-200
                r'(\d+[А-Я]?)$',                  # Просто 5, 5А
            ]
            
            import re
            for pattern in patterns:
                match = re.search(pattern, pump_mark)
                if match:
                    gabarit = match.group(1)
                    # Нормализация (русская А в английскую A)
                    if gabarit.endswith('А'):
                        gabarit = gabarit[:-1] + 'A'
                    elif 'А' in gabarit:
                        gabarit = gabarit.replace('А', 'A')
                    return gabarit
            
            return '5'  # По умолчанию
        
        def get_pump_dimensions(gabarit):
            """Возвращает размеры насоса по габариту"""
            PUMP_DIMENSIONS = {
                '2A': {'pump_OD': 69, 'min_casing_ID': 87, 'casing_OD': 114, 'typical_casing_ID': 102},
                '5': {'pump_OD': 92, 'min_casing_ID': 121.7, 'casing_OD': 146, 'typical_casing_ID': 130},
                '5A': {'pump_OD': 103, 'min_casing_ID': 126, 'casing_OD': 146, 'typical_casing_ID': 130},
                '6': {'pump_OD': 114, 'min_casing_ID': 143, 'casing_OD': 168, 'typical_casing_ID': 150},
                '7A': {'pump_OD': 136, 'min_casing_ID': 165, 'casing_OD': 194, 'typical_casing_ID': 178},
            }
            
            return PUMP_DIMENSIONS.get(gabarit, PUMP_DIMENSIONS['5'])
        
        for index, row in df.iterrows():
            try:
                # ПРАВИЛО 1: Проверяем столбец H (номер скважины)
                if col_indices['H'] >= len(row) or pd.isna(row.iloc[col_indices['H']]):
                    stats['skipped_no_well'] += 1
                    continue
                
                well_name = str(row.iloc[col_indices['H']]).strip()
                if not well_name or well_name.lower() in ['nan', 'none', '']:
                    stats['skipped_no_well'] += 1
                    continue
                
                # ПРАВИЛО 2: Проверяем столбец BJ (Т раб.мес)
                if col_indices['BJ'] >= len(row):
                    continue
                
                bj_value = row.iloc[col_indices['BJ']]
                try:
                    work_hours = float(str(bj_value).replace(',', '.'))
                    if work_hours <= 0:
                        stats['skipped_work_hours'] += 1
                        continue
                except:
                    stats['skipped_work_hours'] += 1
                    continue
                
                # Парсим ЦДНГ и ЦИТС (столбец D)
                cdng_value = str(row.iloc[col_indices['D']]) if col_indices['D'] < len(row) else "ЦДНГ-1"
                cdng, cits = parse_cdng_cits(cdng_value, selected_cits)
                
                # Определяем тип скважины по BH
                if col_indices['BH'] >= len(row):
                    continue
                
                bh_value = row.iloc[col_indices['BH']]
                
                # Проверяем, есть ли значение в BH
                has_bh_value = False
                if not pd.isna(bh_value):
                    bh_str = str(bh_value).strip()
                    if bh_str and bh_str.lower() not in ['nan', 'none', '']:
                        try:
                            bh_float = float(bh_str.replace(',', '.'))
                            if bh_float > 0:
                                has_bh_value = True
                        except:
                            pass
                
                if not has_bh_value:
                    # Постоянная скважина
                    well_type = 'constant'
                    work_time = None
                    pause_time = None
                    operation_mode = 'constant'
                    mode = None
                else:
                    # КПР скважина
                    well_type = 'kpr'
                    operation_mode = 'kpr'
                    mode = 'По времени'
                    
                    # Время работы (BH)
                    try:
                        work_time = round(float(str(bh_value).replace(',', '.')) * 60)   # в минуты
                    except:
                        work_time = 15 * 60  # по умолчанию 15 часов
                    
                    # Время накопления (BI)
                    if col_indices['BI'] < len(row):
                        bi_value = row.iloc[col_indices['BI']]
                        try:
                            pause_time = round(float(str(bi_value).replace(',', '.')) * 60)  # в минуты
                        except:
                            pause_time = 45 * 60  # по умолчанию 45 часов
                    else:
                        pause_time = 45 * 60
                
                # ПРАВИЛО 3: Дебит (BD - Qж)
                flow_rate = 0
                if col_indices['BD'] < len(row):
                    bd_value = row.iloc[col_indices['BD']]
                    try:
                        flow_rate = float(str(bd_value).replace(',', '.'))
                    except:
                        pass
                
                if flow_rate <= 0:
                    stats['skipped_no_flow'] += 1
                    continue
                
                # Обводненность (BE) - записываем но не показываем
                water_cut = 0
                if col_indices['BE'] < len(row):
                    be_value = row.iloc[col_indices['BE']]
                    try:
                        water_cut = float(str(be_value).replace(',', '.'))
                    except:
                        pass
                
                # Куст (I)
                cluster = "Неизвестно"
                if col_indices['I'] < len(row):
                    cluster_val = str(row.iloc[col_indices['I']]).strip()
                    if cluster_val and cluster_val.lower() not in ['nan', 'none', '']:
                        cluster = cluster_val
                
                # Tраб (BJ / BG)
                t_rab = 0
                if col_indices['BG'] < len(row):
                    bg_value = row.iloc[col_indices['BG']]
                    try:
                        bg_float = float(str(bg_value).replace(',', '.'))
                        if bg_float > 0:
                            t_rab = work_hours / bg_float
                    except:
                        pass
                
                # Технические параметры
                def parse_tech_param(col_key, default=None):
                    if col_indices[col_key] < len(row):
                        value = row.iloc[col_indices[col_key]]
                        if not pd.isna(value):
                            try:
                                return float(str(value).replace(',', '.'))
                            except:
                                pass
                    return default
                
                # ПАРСИМ МАРКУ НАСОСА ИЗ СТОЛБЦА W
                pump_mark = None
                pump_flow = None  # Дебит ГНО, м³/сут
                pump_head = None  # Напор, м
                
                if col_indices['W'] < len(row):
                    pump_mark_value = row.iloc[col_indices['W']]
                    if not pd.isna(pump_mark_value):
                        pump_mark = str(pump_mark_value).strip()
                        
                        # Парсим дебит и напор из марки насоса
                        # Формат: 115ЭЦН(НА)5-125-1500, где 125 - дебит, 1500 - напор
                        try:
                            # Ищем цифры через дефис
                            import re
                            # Ищем паттерн числа-числа в конце строки
                            matches = re.findall(r'(\d+)-(\d+)$', pump_mark)
                            if matches:
                                pump_flow = float(matches[0][0])  # Первое число после последнего дефиса
                                pump_head = float(matches[0][1])  # Второе число после последнего дефиса
                            
                            # Альтернативный вариант поиска
                            if pump_flow is None or pump_head is None:
                                # Ищем все числа в строке
                                numbers = re.findall(r'\d+', pump_mark)
                                if len(numbers) >= 3:
                                    # Обычно формат: код-дебит-напор
                                    pump_flow = float(numbers[-2])  # предпоследнее число
                                    pump_head = float(numbers[-1])  # последнее число
                        except:
                            pass
                
                # Получаем тип установки (столбец U)
                installation_type = "Неизвестно"
                if col_indices['U'] < len(row):
                    installation_value = row.iloc[col_indices['U']]
                    if not pd.isna(installation_value):
                        installation_type = str(installation_value).strip()
                
                # НОВОЕ: Парсим диаметр ЭК (столбец R)
                casing_diameter = None
                if col_indices['R'] < len(row):
                    casing_str = row.iloc[col_indices['R']]
                    casing_diameter = parse_casing_diameter(casing_str)
                
                # НОВОЕ: Определяем габарит насоса и размеры
                pump_gabarit = extract_pump_gabarit(pump_mark)
                pump_dimensions = get_pump_dimensions(pump_gabarit)
                
                # Если диаметр ЭК не указан, используем типовой по габариту
                if not casing_diameter:
                    casing_diameter = pump_dimensions['casing_OD']
                
                # Создаем словарь с данными скважины
                well_data = {
                    # Основные данные
                    'name': well_name,
                    'cluster': cluster,
                    'flow_rate': flow_rate,
                    'water_cut': water_cut,  # Храним но не показываем
                    'operation_mode': operation_mode,
                    'mode': mode,
                    'kpr_mode': mode,
                    'is_active': True,
                    
                    # Для КПР скважин
                    'schedule': [work_time, pause_time] if operation_mode == 'kpr' else None,
                    'base_launch_time': '08:00',  # Время запуска по умолчанию
                    'exclude_from_shift': False,
                    
                    # Иерархия
                    'tpp': 'VQ-BADнефтегаз',
                    'cits': cits,
                    'cdng': cdng,
                    
                    # Технические параметры (существующие)
                    'work_hours_month': work_hours,
                    't_rab': t_rab,
                    'pump_depth': parse_tech_param('X'),      # Глубина насоса
                    'pump_mark': pump_mark,                   # Марка насоса
                    'pump_flow': pump_flow,                   # Дебит ГНО, м³/сут
                    'pump_head': pump_head,                   # Напор, м
                    'delivery_coef': parse_tech_param('AF'),  # Коэффициент подачи
                    'mttf': parse_tech_param('AI'),           # Наработка на отказ
                    'h_din': parse_tech_param('AN'),          # Hдин
                    'p_pr': parse_tech_param('AQ'),           # Pпр
                    'p_pl': parse_tech_param('AY'),           # Pпл
                    'buffer_pressure': parse_tech_param('AL'), # Рб/ф (буферное давление)
                    'prod_coef': parse_tech_param('BC'),      # Коэффициент продуктивности
                    'work_days': parse_tech_param('BG'),      # Дни работы+накопление
                    'oil_density': parse_tech_param('CX'),    # Плотность нефти
                    'gas_factor': parse_tech_param('DC'),     # Газовый фактор
                    'p_nas': parse_tech_param('DE'),          # Pнас
                    'installation_type': installation_type,   # Тип установки (ЭЦН/ОРД/ШГН)
                    'vdp_interval': parse_tech_param('O'),    # Интервал ВДП
                    'p_zat': parse_tech_param('AO'),          # Pзат
                    'p_zab_vdp': parse_tech_param('AT'),      # Pзаб ВДП
                    'udl_na_hdin': parse_tech_param('AP'),    # Удл. На Ндин
                    
                    # НОВЫЕ ПОЛЯ ДЛЯ УЧЕТА ГАЗА И КОНСТРУКЦИИ:
                    'casing_diameter': casing_diameter,       # Диаметр ЭК, мм
                    'pump_gabarit': pump_gabarit,            # Габарит насоса (2A, 5, 5A, 6, 7A)
                    'pump_OD': pump_dimensions['pump_OD'],   # Диаметр корпуса насоса, мм
                    'min_casing_ID': pump_dimensions['min_casing_ID'],  # Мин. внутр. диаметр ЭК
                    'typical_casing_ID': pump_dimensions.get('typical_casing_ID', 130),  # Тип. внутр. диаметр
                    
                    # Константы для расчетов с газом (типовые значения)
                    'gas_density': 0.8,              # кг/м³ при нормальных условиях
                    'gas_molar_mass': 18.0,          # кг/кмоль (метан с примесями)
                    'oil_viscosity': 1.2,            # сПз (типовое для легкой нефти)
                    'water_viscosity': 0.8,          # сПз
                    'oil_volume_factor': 1.2,        # м³/м³ (типовое)
                    'formation_temp': 60,            # °C (типовое для Западной Сибири)
                    'surface_temp': 20,              # °C (стандартная температура на устье)
                    
                    # Метка импорта
                    'import_source': 'tech_regime',
                    'import_date': datetime.now().strftime("%Y-%m-%d %H:%M")
                }
                
                # ОБРАБОТКА СТОЛБЦА "ОБОРОТЫ" (AD) с особой логикой
                rotations_value = 0
                if col_indices['AD'] < len(row):
                    ad_value = row.iloc[col_indices['AD']]
                    if not pd.isna(ad_value):
                        try:
                            rotations_value = float(str(ad_value).replace(',', '.'))
                            
                            # Применяем логику преобразования оборотов в Гц
                            if rotations_value > 100:
                                # Если значение > 100, это обороты в об/мин
                                # Преобразуем в Гц: обороты * 50 / 2910
                                rotations_value = (rotations_value * 50) / 2910
                            # Если значение <= 100, оставляем как есть (предполагаем, что это уже Гц)
                        except:
                            rotations_value = 0
                
                well_data['rotations_hz'] = rotations_value  # Обороты в Гц
                
                wells_data.append(well_data)
                stats['processed'] += 1
                
            except Exception as e:
                continue
        
        # Показываем статистику обработки
        st.success(f"✅ Успешно обработано {len(wells_data)} скважин")
        
        # Дополнительная статистика по новым полям
        wells_with_casing_data = sum(1 for w in wells_data if w.get('casing_diameter') is not None)
        wells_with_gabarit = sum(1 for w in wells_data if w.get('pump_gabarit') != '5')
        
        st.info(f"""
        **Статистика обработки:**
        - Всего строк в файле: {stats['total_rows']}
        - Пропущено (нет номера скважины): {stats['skipped_no_well']}
        - Пропущено (дебит ≤ 0): {stats['skipped_no_flow']}
        - Пропущено (рабочие часы ≤ 0): {stats['skipped_work_hours']}
        - Обработано успешно: {stats['processed']}
        """)
        
        return wells_data
    
    except Exception as e:
        st.error(f"❌ Ошибка при чтении файла: {str(e)}")
        return []
 
@st.cache_data(ttl=3600)
def update_structure_from_wells(wells_data):
    """Обновляет структуру кустов на основе загруженных данных"""
    clusters = {}
    
    for well in wells_data:
        cits = well.get('cits', 'ЦИТС VQ-BAD')
        cdng = well.get('cdng', 'ЦДНГ-1')
        cluster = well.get('cluster', 'Неизвестно')
        
        # Создаем структуру: {ЦИТС: {ЦДНГ: [кусты]}}
        if cits not in clusters:
            clusters[cits] = {}
        
        if cdng not in clusters[cits]:
            clusters[cits][cdng] = []
        
        # Добавляем куст если его нет и не "Неизвестно"
        if cluster != 'Неизвестно' and cluster not in clusters[cits][cdng]:
            clusters[cits][cdng].append(cluster)
    
    return clusters

# ============================================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ============================================================
def calculate_working_stats(wells_data, phases_dict):
    """
    Рассчитывает статистику по работе скважин
    """
    optimizer = PressureStabilizationOptimizer(wells_data)
    count_data = optimizer.calculate_working_wells_count(phases_dict)
    
    # Статистика ДО
    before = np.array(count_data['count_before'])
    after = np.array(count_data['count_after'])
    
    stats = {
        'before': {
            'min': np.min(before),
            'max': np.max(before),
            'avg': np.mean(before),
            'std': np.std(before),
            'range': np.max(before) - np.min(before)
        },
        'after': {
            'min': np.min(after),
            'max': np.max(after),
            'avg': np.mean(after),
            'std': np.std(after),
            'range': np.max(after) - np.min(after)
        },
        'improvements': {
            'range_reduction': ((np.max(before) - np.min(before)) - (np.max(after) - np.min(after))) / (np.max(before) - np.min(before)) * 100 if (np.max(before) - np.min(before)) > 0 else 0,
            'std_reduction': (np.std(before) - np.std(after)) / np.std(before) * 100 if np.std(before) > 0 else 0
        }
    }
    
    return stats
        
@st.cache_data(ttl=3600)
def calculate_next_launch_times(wells_data, phases_dict, current_time_str):
    """
    КОРРЕКТНЫЙ расчет ближайших времен запуска с учетом текущего времени
    Возвращает список рекомендаций для оператора
    """
    try:
        current_hours, current_minutes = map(int, current_time_str.split(':'))
        current_total_minutes = current_hours * 60 + current_minutes
    except:
        # Если ошибка в формате времени, используем 00:00
        current_total_minutes = 0
    
    recommendations = []
    
    for well in wells_data:
        if not well.get('is_active', True):
            continue
            
        if well.get('operation_mode') != 'kpr':
            continue
        
        well_name = well['name']
        base_launch_str = well.get('base_launch_time', '00:00')
        
        try:
            base_hours, base_minutes = map(int, base_launch_str.split(':'))
            base_total_minutes = base_hours * 60 + base_minutes
        except:
            base_total_minutes = 0
        
        # Получаем оптимальный сдвиг для этой скважины
        phase_shift = phases_dict.get(well_name, 0)
        
        # Оптимальное абсолютное время запуска
        optimal_launch_minutes = (base_total_minutes + phase_shift) % (24 * 60)
        
        # Параметры цикла
        try:
            work_time, pause_time = well['schedule']
            cycle_time = work_time + pause_time
        except:
            D = 15, 45
            cycle_time = 60
        
        # РАСЧЕТ 1: Ближайший КПР запуск от текущего времени
        if current_total_minutes <= base_total_minutes:
            cycles_to_kpr = 0
        else:
            cycles_to_kpr = ((current_total_minutes - base_total_minutes) // cycle_time) + 1
        
        next_kpr_launch = base_total_minutes + cycles_to_kpr * cycle_time
        
        # Если вышли за пределы суток
        next_kpr_launch = next_kpr_launch % (24 * 60)
        
        # РАСЧЕТ 2: Ближайший ОПТИМАЛЬНЫЙ запуск от текущего времени
        if current_total_minutes <= optimal_launch_minutes:
            cycles_to_opt = 0
        else:
            cycles_to_opt = ((current_total_minutes - optimal_launch_minutes) // cycle_time) + 1
        
        next_opt_launch = optimal_launch_minutes + cycles_to_opt * cycle_time
        next_opt_launch = next_opt_launch % (24 * 60)
        
        # Форматирование времени
        kpr_hours = int(next_kpr_launch // 60)
        kpr_minutes = int(next_kpr_launch % 60)
        kpr_time_str = f"{kpr_hours:02d}:{kpr_minutes:02d}"
        
        opt_hours = int(next_opt_launch // 60)
        opt_minutes = int(next_opt_launch % 60)
        opt_time_str = f"{opt_hours:02d}:{opt_minutes:02d}"
        
        # Определяем действие для оператора
        # Проверяем, исключена ли скважина из расчета сдвига
        exclude_from_shift = well.get('exclude_from_shift', False)
        
        if exclude_from_shift:
            action = "НЕ МЕНЯТЬ (исключена)"
            change_text = "0 мин"
        else:
            time_diff = abs(next_kpr_launch - next_opt_launch)
            if time_diff < 5:  # Если разница менее 5 минут
                action = "ОСТАВИТЬ как есть"
                change_text = "≈0 мин"
            elif next_opt_launch < next_kpr_launch:
                action = "ЗАПУСТИТЬ РАНЬШЕ"
                change_minutes = next_opt_launch - next_kpr_launch
                if change_minutes < -12*60:  # Корректировка через полночь
                    change_minutes += 24*60
                change_text = f"{int(change_minutes):+d} мин"
            else:
                action = "ОТЛОЖИТЬ запуск"
                change_minutes = next_opt_launch - next_kpr_launch
                if change_minutes > 12*60:  # Корректировка через полночь
                    change_minutes -= 24*60
                change_text = f"{int(change_minutes):+d} мин"
        
        # Форматирование режима
        mode = well.get('mode', 'По времени')
        if mode == 'По времени':
            schedule_text = f"{work_time}/{pause_time} мин"
        else:
            schedule_text = f"{work_time}/{pause_time} атм"
        
        # Добавляем метку об исключении
        if exclude_from_shift:
            schedule_text += " ⚠️ (искл.)"
        
        recommendations.append({
            'Скважина': well_name,
            'КПР время': kpr_time_str,
            'Оптим. время': opt_time_str,
            'Действие': action,
            'Изменение': change_text,
            'Режим': schedule_text,
            'Тип режима': mode,
            'Дебит': f"{well['flow_rate']} м³/сут",
            'Участвует в сдвиге': "Нет" if exclude_from_shift else "Да"
        })
    
    return sorted(recommendations, key=lambda x: x['КПР время'])

@st.cache_data(ttl=3600)
def plot_wells_cluster(wells_data):
    """Визуализация куста скважин в ряд"""
    if not wells_data:
        return go.Figure()
    
    fig = go.Figure()
    
    # Располагаем скважины в ряд
    n_wells = len(wells_data)
    x_positions = np.linspace(0, n_wells * 2, n_wells)
    
    for i, well in enumerate(wells_data):
        x = x_positions[i]
        y = 0
        
        # Определяем цвет и символ
        if well['operation_mode'] == 'constant':
            color = '#1f77b4' if well['is_active'] else '#aec7e8'
            symbol = 'circle' if well['is_active'] else 'circle-open'
            size = 25 if well['is_active'] else 20
        else:
            # Для КПР скважин, исключенных из сдвига, используем специальный цвет
            if well.get('exclude_from_shift', False):
                color = '#9467bd' if well['is_active'] else '#c5b0d5'  # Фиолетовый для исключенных
            else:
                color = '#2ca02c' if well['is_active'] else '#ff7f0e'  # Зеленый для обычных КПР
            
            symbol = 'circle' if well['is_active'] else 'circle-open'
            size = 25 if well['is_active'] else 20
        
        # Основной маркер
        fig.add_trace(go.Scatter(
            x=[x], y=[y],
            mode='markers+text',
            marker=dict(
                size=size,
                color=color,
                symbol=symbol,
                line=dict(width=3, color='darkgray'),
                opacity=0.9
            ),
            text=well['name'],
            textposition="middle center",
            textfont=dict(
                size=14,
                color='white' if well['is_active'] else 'darkgray',
                weight='bold'
            ),
            name=well['name'],
            hovertemplate=(
                f"<b>🏭 {well['name']}</b><br>"
                f"<b>Тип:</b> {'🔵 Постоянная' if well['operation_mode'] == 'constant' else '🟢 КПР'}<br>"
                f"<b>Статус:</b> {'🟢 Активна' if well['is_active'] else '🔴 Остановлена'}<br>"
                f"<b>Дебит:</b> {well['flow_rate']} м³/сут<br>"
                f"{f'<b>Режим:</b> {well['schedule'][0]}/{well['schedule'][1]} ({well.get('mode', 'По времени')})' if well['operation_mode'] == 'kpr' and well.get('schedule') and well['is_active'] else ''}"
                f"{f'<br><b>КПР запуск:</b> {well.get('base_launch_time', '00:00')}' if well['operation_mode'] == 'kpr' and well['is_active'] else ''}"
                f"{f'<br><b>Участвует в сдвиге:</b> {'Нет ⚠️' if well.get('exclude_from_shift', False) else 'Да'}' if well['operation_mode'] == 'kpr' and well['is_active'] else ''}"
                f"<extra></extra>"
            ),
            showlegend=False
        ))
        
        # Подпись типа
        fig.add_trace(go.Scatter(
            x=[x], y=[-0.3],
            mode='text',
            text=['Пост.' if well['operation_mode'] == 'constant' else 'КПР'],
            textposition="middle center",
            textfont=dict(
                size=12,
                color=color,
                weight='bold'
            ),
            showlegend=False,
            hoverinfo='skip'
        ))
        
        # Подпись дебита
        fig.add_trace(go.Scatter(
            x=[x], y=[-0.6],
            mode='text',
            text=[f"{well['flow_rate']} м³/сут"],
            textposition="middle center",
            textfont=dict(
                size=11,
                color='gray'
            ),
            showlegend=False,
            hoverinfo='skip'
        ))
    
    # Настройка внешнего вида
    fig.update_layout(
        title=dict(
            text="🛢️ Визуализация куста скважин",
            x=0.5,
            font=dict(size=20, color='#0055A5')
        ),
        xaxis=dict(
            showgrid=False, 
            zeroline=False, 
            showticklabels=False,
            range=[-1, n_wells * 2 + 1]
        ),
        yaxis=dict(
            showgrid=False, 
            zeroline=False, 
            showticklabels=False,
            range=[-1, 1]
        ),
        showlegend=False,
        height=400,
        margin=dict(l=20, r=20, t=80, b=100),
        plot_bgcolor='rgba(240,240,240,0.1)'
    )
    
    # Легенда с учетом исключенных скважин
    fig.add_annotation(
        x=0.5, y=-0.9,
        text="🔵 Постоянные  🟢 КПР  🟣 КПР(искл.)  ○ Остановлена",
        showarrow=False,
        xref="paper", yref="paper",
        font=dict(size=12, color='gray'),
        bgcolor="white",
        bordercolor="lightgray",
        borderwidth=1,
        borderpad=4
    )
    
    return fig

@st.cache_data(ttl=3600)
def plot_working_wells_count(wells_data, phases_dict, current_time_str):
    """
    Визуализация количества работающих скважин до и после оптимизации
    """
    optimizer = PressureStabilizationOptimizer(wells_data)
    
    # Получаем данные о работающих скважинах
    count_data = optimizer.calculate_working_wells_count(phases_dict)
    
    fig = go.Figure()
    
    # График ДО оптимизации
    fig.add_trace(go.Scatter(
        x=count_data['time_hours'],
        y=count_data['count_before'],
        name='📊 До оптимизации',
        line=dict(color='red', width=2, dash='dash'),
        fill='tozeroy',
        fillcolor='rgba(255, 0, 0, 0.1)',
        hovertemplate='<b>Время: %{x:.2f} ч</b><br>Работает: %{y} скв.<extra></extra>'
    ))
    
    # График ПОСЛЕ оптимизации
    fig.add_trace(go.Scatter(
        x=count_data['time_hours'],
        y=count_data['count_after'],
        name='🎯 После оптимизации',
        line=dict(color='blue', width=2),
        fill='tozeroy',
        fillcolor='rgba(0, 0, 255, 0.1)',
        hovertemplate='<b>Время: %{x:.2f} ч</b><br>Работает: %{y} скв.<extra></extra>'
    ))
    
    # Горизонтальные линии для средних значений
    avg_before = np.mean(count_data['count_before'])
    avg_after = np.mean(count_data['count_after'])
    
    fig.add_hline(
        y=avg_before,
        line_dash="dot",
        line_color="red",
        opacity=0.5,
        annotation_text=f"Среднее до: {avg_before:.1f} скв",
        annotation_position="bottom left"
    )
    
    fig.add_hline(
        y=avg_after,
        line_dash="dot",
        line_color="blue",
        opacity=0.5,
        annotation_text=f"Среднее после: {avg_after:.1f} скв",
        annotation_position="bottom right"
    )
    
    # Текущее время
    try:
        current_hours, current_minutes = map(int, current_time_str.split(':'))
        current_total_hours = current_hours + current_minutes/60
        fig.add_vline(
            x=current_total_hours,
            line_dash="dash",
            line_color="orange",
            opacity=0.7,
            annotation_text=f"Сейчас: {current_time_str}",
            annotation_position="top"
        )
    except:
        pass
    
    fig.update_layout(
        title='🔄 Количество работающих скважин в течение суток',
        xaxis_title='Время, часы',
        yaxis_title='Количество работающих скважин',
        height=400,
        hovermode='x unified',
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        yaxis=dict(
            tickmode='linear',
            tick0=0,
            dtick=1
        )
    )
    
    fig.update_xaxes(showgrid=True, gridwidth=1, gridcolor='LightGray')
    fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor='LightGray')
    
    return fig

@st.cache_data(ttl=3600)
def plot_pressure_optimization_results(wells_data, phases_dict, current_time_str, target_flow):
    """Визуализация результатов оптимизации"""
    optimizer = PressureStabilizationOptimizer(wells_data)
    
    # Время для графиков
    time_points = np.arange(0, 24 * 60, 5)
    time_hours = time_points / 60
    
    # Поток до/после
    zero_phases = {name: 0 for name in phases_dict.keys()}
    flows_before = []
    flows_after = []
    
    for t in time_points:
        flow_before = optimizer.calculate_total_flow_at_time(t, zero_phases)
        flow_after = optimizer.calculate_total_flow_at_time(t, phases_dict)
        flows_before.append(flow_before)
        flows_after.append(flow_after)
    
    # График
    fig = go.Figure()
    
    fig.add_trace(go.Scatter(
        x=time_hours, y=flows_before,
        name='📊 До оптимизации',
        line=dict(color='red', width=2, dash='dash'),
        hovertemplate='<b>Время: %{x:.2f} ч</b><br>Дебит: %{y:.1f} м³/час<extra></extra>'
    ))
    
    fig.add_trace(go.Scatter(
        x=time_hours, y=flows_after,
        name='🎯 После оптимизации',
        line=dict(color='blue', width=2),
        hovertemplate='<b>Время: %{x:.2f} ч</b><br>Дебит: %{y:.1f} м³/час<extra></extra>'
    ))
    
    # Линия целевого дебита
    fig.add_hline(
        y=target_flow,
        line_dash="dot",
        line_color="green",
        opacity=0.7,
        annotation_text=f"Цель: {target_flow:.1f} м³/час",
        annotation_position="bottom right"
    )
    
    # Текущее время
    try:
        current_hours, current_minutes = map(int, current_time_str.split(':'))
        current_total_hours = current_hours + current_minutes/60
        fig.add_vline(
            x=current_total_hours,
            line_dash="dash",
            line_color="orange",
            opacity=0.7,
            annotation_text=f"Сейчас: {current_time_str}",
            annotation_position="top"
        )
    except:
        pass
    
    fig.update_layout(
        title='📊 Сравнение дебитов до и после оптимизации',
        xaxis_title='Время, часы',
        yaxis_title='Суммарный дебит, м³/час',
        height=500,
        hovermode='x unified',
        legend=dict(yanchor="top", y=0.99, xanchor="left", x=0.01)
    )
    
    fig.update_xaxes(showgrid=True, gridwidth=1, gridcolor='LightGray')
    fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor='LightGray')
    
    return fig

@st.cache_data(ttl=3600)
def create_comparison_table(stats):
    """Создает таблицу сравнения до/после"""
    comparison_data = {
        'Показатель': [
            'Средний дебит, м³/час',
            'Стандартное отклонение, м³/час',
            'Количество пиков',
            'Средняя величина пиков, м³/час',
            'Достижение цели, %',
            'Общая эффективность, %'
        ],
        'До оптимизации': [
            f"{stats['avg_flow_before']:.1f}",
            f"{stats['std_before']:.2f}",
            f"{stats['peaks_before']}",
            f"{stats['peak_magnitude_before']:.1f}" if stats['peaks_before'] > 0 else "0",
            "-",
            "-"
        ],
        'После оптимизации': [
            f"{stats['avg_flow_after']:.1f}",
            f"{stats['std_after']:.2f}",
            f"{stats['peaks_after']}",
            f"{stats['peak_magnitude_after']:.1f}" if stats['peaks_after'] > 0 else "0",
            f"{stats['target_achievement']:.1f}%",
            f"{stats['efficiency']:.1f}%"
        ],
        'Изменение': [
            f"{stats['flow_improvement']:+.1f}%",
            f"{stats['stability_improvement']:+.1f}%",
            f"{stats['peaks_improvement']:+.1f}%",
            f"{(stats['peak_magnitude_after'] - stats['peak_magnitude_before']):+.1f}" if stats['peaks_before'] > 0 and stats['peaks_after'] > 0 else "-",
            "-",
            "-"
        ]
    }
    
    return pd.DataFrame(comparison_data)

def export_to_excel(recommendations, stats, wells_data, phases_dict, current_time):
    """Экспорт всех данных в Excel файл"""
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        # 1. Рекомендации
        if recommendations:
            df_rec = pd.DataFrame(recommendations)
            df_rec.to_excel(writer, sheet_name='Рекомендации', index=False)
        
        # 2. Сравнение результатов
        comp_df = create_comparison_table(stats)
        comp_df.to_excel(writer, sheet_name='Сравнение', index=False)
        
        # 3. Исходные данные скважин
        wells_df_data = []
        for well in wells_data:
            wells_df_data.append({
                'Скважина': well['name'],
                'Тип': 'Постоянная' if well['operation_mode'] == 'constant' else 'КПР',
                'Дебит, м³/сут': well['flow_rate'],
                'Режим': f"{well['schedule'][0]}/{well['schedule'][1]} ({well.get('mode', 'По времени')})" if well['operation_mode'] == 'kpr' else '-',
                'Запуск КПР': well.get('base_launch_time', '-') if well['operation_mode'] == 'kpr' else '-',
                'Статус': 'Активна' if well['is_active'] else 'Остановлена',
                'Сдвиг фазы, мин': phases_dict.get(well['name'], 0) if well['operation_mode'] == 'kpr' else '-',
                'Участвует в сдвиге': 'Нет' if well.get('exclude_from_shift', False) else 'Да' if well['operation_mode'] == 'kpr' else '-'
            })
        
        wells_df = pd.DataFrame(wells_df_data)
        wells_df.to_excel(writer, sheet_name='Скважины', index=False)
        
        # 4. Статистика
        stats_df = pd.DataFrame([{
            'Параметр': 'Целевой дебит',
            'Значение': f"{stats['target_flow']:.1f} м³/час",
            'Примечание': f"Коэффициент загрузки: {stats['target_coefficient']*100:.0f}%"
        }, {
            'Параметр': 'Общая эффективность',
            'Значение': f"{stats['efficiency']:.1f}%",
            'Примечание': '60% цель + 30% стабильность + 10% дебит'
        }, {
            'Параметр': 'Время оптимизации',
            'Значение': datetime.now().strftime("%Y-%m-%d %H:%M"),
            'Примечание': f"Текущее время: {current_time}"
        }])
        
        stats_df.to_excel(writer, sheet_name='Статистика', index=False)
    
    output.seek(0)
    return output

# ============================================================
# РАЗДЕЛЫ ИНТЕРФЕЙСА
# ============================================================

def show_dashboard():
    """Главная страница приложения"""
    # Загружаем plotly только если нужен
    _load_plotly()
    
    st.markdown("""
    <div style='text-align: center; margin-bottom: 30px;'>
        <h1 style='color: #0055A5;'>🛢️ PovhEquilibrium</h1>
        <h3 style='color: #666;'>Система оптимизации скважин КПР</h3>
    </div>
    """, unsafe_allow_html=True)
    
    # Карточки с фиксированной высотой
    col1, col2, col3, col4 = st.columns(4)
    
    # CSS для фиксированной высоты
    card_style = """
    <style>
    .dashboard-card {
        padding: 20px;
        border-radius: 10px;
        min-height: 160px;
        height: 160px;
        display: flex;
        flex-direction: column;
        justify-content: space-between;
        margin-bottom: 15px;
    }
    </style>
    """
    st.markdown(card_style, unsafe_allow_html=True)
    
    with col1:
        st.markdown("""
        <div style='background: #f0f8ff; padding: 20px; border-radius: 10px; border-left: 5px solid #0055A5; min-height: 160px; height: 160px; display: flex; flex-direction: column; justify-content: space-between;'>
            <h4 style='color: #0055A5; margin: 0;'>🛢️ Управление кустами</h4>
            <p style='color: #0055A5; margin: 0;'>Добавление, редактирование скважин, импорт данных</p>
        </div>
        """, unsafe_allow_html=True)
        if st.button("Перейти к кустам →", use_container_width=True, key="btn_wells"):
            st.session_state.current_page = "wells"
            st.rerun()
    
    with col2:
        st.markdown("""
        <div style='background: #f0fff0; padding: 20px; border-radius: 10px; border-left: 5px solid #2ca02c; min-height: 160px; height: 160px; display: flex; flex-direction: column; justify-content: space-between;'>
            <h4 style='color: #2ca02c; margin: 0;'>⚙️ Оптимизация</h4>
            <p style='color: #2ca02c; margin: 0;'>Расчет оптимального времени запуска КПР скважин</p>
        </div>
        """, unsafe_allow_html=True)
        if st.button("Запустить оптимизацию →", use_container_width=True, key="btn_optimization"):
            st.session_state.current_page = "optimization"
            st.rerun()
    
    with col3:
        st.markdown("""
        <div style='background: #fff8f0; padding: 20px; border-radius: 10px; border-left: 5px solid #ff7f0e; min-height: 160px; height: 160px; display: flex; flex-direction: column; justify-content: space-between;'>
            <h4 style='color: #ff7f0e; margin: 0;'>📊 Отчеты</h4>
            <p style='color: #ff7f0e; margin: 0;'>История расчетов, экспорт в Excel</p>
        </div>
        """, unsafe_allow_html=True)
        if st.button("Смотреть отчеты →", use_container_width=True, key="btn_reports"):
            st.session_state.current_page = "reports"
            st.rerun()
    
    with col4:
        st.markdown("""
        <div style='background: #f8f0ff; padding: 20px; border-radius: 10px; border-left: 5px solid #9467bd; min-height: 160px; height: 160px; display: flex; flex-direction: column; justify-content: space-between;'>
            <h4 style='color: #9467bd; margin: 0;'>💾 Выборочное сохранение</h4>
            <p style='color: #9467bd; margin: 0;'>Сохранение выбранных скважин из расчетов</p>
        </div>
        """, unsafe_allow_html=True)
        if st.button("Выбрать скважины →", use_container_width=True, key="btn_custom_save"):
            st.session_state.current_page = "custom_save"
            st.rerun()
    
    st.markdown("---")
    
    # Последние расчеты по всем модулям (ИСПРАВЛЕНО)
    st.subheader("📈 Последние расчеты по всем модулям")
    
    # Собираем все последние расчеты из разных модулей
    recent_calculations = []
    
    # 1. Расчеты стабилизации давления (Модуль 1)
    if 'calculation_history' in st.session_state and st.session_state.calculation_history:
        for calc in st.session_state.calculation_history[-3:]:  # Последние 3
            recent_calculations.append({
                'Дата': calc.get('Дата', 'Не указано'),
                'Тип': '📊 Стабилизация давления',
                'Объект': calc.get('Куст', 'Не указан'),
                'Результат': calc.get('Эффективность', 'Нет данных'),
                'ЦИТС': calc.get('ЦИТС', 'Не указан')
            })
    
    # 2. Пакетные расчеты КПР (Модуль 2) - ИСПРАВЛЕНО
    if 'batch_results_advanced' in st.session_state and st.session_state.batch_results_advanced:
        batch_data = st.session_state.batch_results_advanced
        if batch_data:
            # Подсчитываем статистику
            df_batch = pd.DataFrame(batch_data)
            scenario_a = len(df_batch[df_batch['Сценарий'] == 'A']) if 'Сценарий' in df_batch.columns else 0
            scenario_b = len(df_batch[df_batch['Сценарий'] == 'B']) if 'Сценарий' in df_batch.columns else 0
            total_effect = df_batch['Эффект (₽/сут)'].sum() if 'Эффект (₽/сут)' in df_batch.columns else 0
            
            recent_calculations.append({
                'Дата': datetime.now().strftime("%Y-%m-%d %H:%M"),
                'Тип': '📦 Пакетный КПР',
                'Объект': f"{len(batch_data)} скважин",
                'Результат': f"A:{scenario_a} B:{scenario_b} | {total_effect:,.0f} ₽/сут",
                'ЦИТС': 'Пакетный анализ'
            })
    
    # 3. Анализ потенциала (Модуль 2) - НОВОЕ
    if 'potential_batch_results' in st.session_state and st.session_state.potential_batch_results:
        potential_data = st.session_state.potential_batch_results
        if potential_data:
            df_potential = pd.DataFrame(potential_data)
            eligible = len(df_potential[df_potential['Проходит фильтры'] == 'Да']) if 'Проходит фильтры' in df_potential.columns else 0
            total_effect = df_potential['Эффект, ₽/сут'].sum() if 'Эффект, ₽/сут' in df_potential.columns else 0
            
            recent_calculations.append({
                'Дата': datetime.now().strftime("%Y-%m-%d %H:%M"),
                'Тип': '📈 Потенциал КПР',
                'Объект': f"{len(potential_data)} скважин",
                'Результат': f"Прошли: {eligible} | {total_effect:,.0f} ₽/сут",
                'ЦИТС': 'Анализ потенциала'
            })
    
    # 4. Расчеты замены ЭЦН (Модуль 3) - ИСПРАВЛЕНО
    ecn_calculations = []
    
    # Проверяем replace режим
    if 'pump_calculation_results_replace' in st.session_state and st.session_state.pump_calculation_results_replace:
        replace_data = st.session_state.pump_calculation_results_replace
        if replace_data:
            df_replace = pd.DataFrame(replace_data)
            total_savings = df_replace['Экономия, руб/сут'].sum() if 'Экономия, руб/сут' in df_replace.columns else 0
            unique_wells = df_replace['Скважина'].nunique() if 'Скважина' in df_replace.columns else len(replace_data)
            
            ecn_calculations.append({
                'Дата': datetime.now().strftime("%Y-%m-%d %H:%M"),
                'Тип': '🔄 Замена (пост→КПР)',
                'Объект': f"{unique_wells} скважин",
                'Результат': f"{total_savings:,.0f} ₽/сут",
                'ЦИТС': 'Замена насосов'
            })
    
    # Проверяем optimize режим
    if 'pump_calculation_results_optimize' in st.session_state and st.session_state.pump_calculation_results_optimize:
        optimize_data = st.session_state.pump_calculation_results_optimize
        if optimize_data:
            df_optimize = pd.DataFrame(optimize_data)
            total_savings = df_optimize['Экономия, руб/сут'].sum() if 'Экономия, руб/сут' in df_optimize.columns else 0
            unique_wells = df_optimize['Скважина'].nunique() if 'Скважина' in df_optimize.columns else len(optimize_data)
            
            ecn_calculations.append({
                'Дата': datetime.now().strftime("%Y-%m-%d %H:%M"),
                'Тип': '🔄 Оптимизация (КПР→КПР)',
                'Объект': f"{unique_wells} скважин",
                'Результат': f"{total_savings:,.0f} ₽/сут",
                'ЦИТС': 'Оптимизация КПР'
            })
    
    # Добавляем все расчеты ЭЦН
    recent_calculations.extend(ecn_calculations)
    
    # 5. Одиночный расчет КПР (если есть)
    if 'optimization_result' in st.session_state and st.session_state.optimization_result:
        result = st.session_state.optimization_result
        if 'well_name' in result:
            recent_calculations.append({
                'Дата': datetime.now().strftime("%Y-%m-%d %H:%M"),
                'Тип': '⚡ Одиночный КПР',
                'Объект': result.get('well_name', 'Не указан'),
                'Результат': f"Сценарий {result.get('best_scenario', {}).get('scenario', 'A')}" if result.get('best_scenario') else "Нет данных",
                'ЦИТС': 'Анализ скважины'
            })
    
    # Показываем таблицу если есть расчеты
    if recent_calculations:
        # Сортируем по дате (новые сверху)
        recent_calculations.sort(key=lambda x: x['Дата'], reverse=True)
        history_df = pd.DataFrame(recent_calculations[:5])  # Показываем 5 последних
        
        # Стилизация таблицы
        st.dataframe(
            history_df,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Дата": st.column_config.TextColumn(width="medium"),
                "Тип": st.column_config.TextColumn(width="medium"),
                "Объект": st.column_config.TextColumn(width="medium"),
                "Результат": st.column_config.TextColumn(width="large"),
                "ЦИТС": st.column_config.TextColumn(width="small")
            }
        )
        
        # Показываем сколько всего расчетов
        st.caption(f"Показано 5 из {len(recent_calculations)} последних расчетов")
    else:
        st.info("""
        Пока нет выполненных расчетов. Запустите первый расчет в любом модуле:
        
        1. **🎯 Стабилизация давления** - Оптимизация времени запуска КПР
        2. **⚡ Пакетный расчет КПР** - Подбор оптимального режима работы
        3. **📈 Анализ потенциала** - Поиск скважин для увеличения дебита
        4. **🔄 Замена ЭЦН** - Расчет экономической эффективности замены насосов
        """)
    
    # Быстрая статистика (ИСПРАВЛЕНО)
    st.markdown("---")
    st.subheader("📊 Общая статистика системы")
    
    col_stats1, col_stats2, col_stats3 = st.columns(3)
    
    with col_stats1:
        total_wells = len(st.session_state.get('wells_data', []))
        st.metric("Всего скважин", total_wells)
    
    with col_stats2:
        # Считаем количество кустов с учетом новой структуры
        total_clusters = 0
        if 'clusters' in st.session_state:
            for cits in st.session_state.clusters:
                for cdng in st.session_state.clusters[cits]:
                    total_clusters += len(st.session_state.clusters[cits][cdng])
        st.metric("Кустов в системе", total_clusters)
    
    with col_stats3:
        # Подсчитываем ВСЕ расчеты по всем модулям (ИСПРАВЛЕНО)
        total_all_calculations = 0
        
        # Модуль 1: расчеты стабилизации давления
        if 'calculation_history' in st.session_state:
            total_all_calculations += len(st.session_state.calculation_history)
        
        # Модуль 2: пакетный расчет КПР
        if 'batch_results_advanced' in st.session_state and st.session_state.batch_results_advanced:
            total_all_calculations += 1
        
        # Модуль 2: анализ потенциала
        if 'potential_batch_results' in st.session_state and st.session_state.potential_batch_results:
            total_all_calculations += 1
        
        # Модуль 2: одиночный расчет КПР
        if 'optimization_result' in st.session_state and st.session_state.optimization_result:
            total_all_calculations += 1
        
        # Модуль 3: расчет замены ЭЦН (оба режима)
        if 'pump_calculation_results_replace' in st.session_state and st.session_state.pump_calculation_results_replace:
            total_all_calculations += 1
        
        if 'pump_calculation_results_optimize' in st.session_state and st.session_state.pump_calculation_results_optimize:
            total_all_calculations += 1
        
        st.metric("Всего расчетов", total_all_calculations)
    
    # Дополнительная информация
    st.markdown("---")
    st.subheader("ℹ️ Информация о системе")
    
    col_info1, col_info2 = st.columns(2)
    
    with col_info1:
        st.info("""
        **🛢️ Что можно делать в системе:**
        
        1. **Загружать данные** скважин из техрежима
        2. **Оптимизировать время** запуска КПР скважин
        3. **Рассчитывать оптимальные** режимы работы КПР
        4. **Анализировать замену** насосов ЭЦН
        5. **Генерировать отчеты** в формате Excel
        """)
    
    with col_info2:
        st.info("""
        **📊 Доступные модули:**
        
        1. **Модуль 1** - Стабилизация давления в коллекторе
        2. **Модуль 2** - Потенциал и оптимизация КПР
        3. **Модуль 3** - Замена ЭЦН с переводом на КПР
        4. **Модуль 4** - Аналитика скважин (Шахматки)  
        **💾 Данные сохраняются** автоматически при выходе
        """)

def show_wells_management():
    """Страница управления кустами и скважинами"""
    st.title("🛢️ Управление кустами и скважинами")
    
    # ИЗМЕНЕНО: теперь 3 вкладки вместо 2
    tab1, tab2, tab3 = st.tabs(["📋 Управление скважинами", "📥 Импорт техрежима", "📡 Импорт данных с ТМ"])
    
    with tab1:
        show_wells_management_tab()
    
    with tab2:
        show_import_tab()
    
    # ДОБАВЛЕНО: третья вкладка
    with tab3:
        show_tm_import_tab()

def normalize_cdng_with_cits(cdng_number, cits):
    """
    Преобразует номер цеха из ТМ в формат системы (БЕЗ буквы в скобках!)
    """
    # В системе ЦДНГ хранится без скобок, просто "ЦДНГ-1"
    return f"ЦДНГ-{cdng_number}"

def show_tm_import_tab():
    """
    Вкладка импорта данных с телемеханики (ТМ)
    Два отдельных этапа с контролем пользователя
    """
    st.markdown("### 📡 Импорт данных с телемеханики (ТМ)")
    
    # Функции очистки (как в работающей версии)
    def clean_cell(cell):
        """Очищает ячейку от артефактов"""
        if pd.isna(cell) or cell == '':
            return ''
        cell_str = str(cell)
        cell_str = cell_str.replace('_x000D_', '')
        cell_str = cell_str.strip()
        return cell_str
    
    def clean_dataframe(df):
        """Очищает весь DataFrame"""
        for col in df.columns:
            df[col] = df[col].apply(clean_cell)
        return df

    def is_date_in_range(date_str, current_date=None):
        """
        Проверяет, попадает ли дата в диапазон: с 1-го числа предыдущего месяца по текущую дату
        Поддерживает форматы: ДД.ММ.ГГГГ и ДД.ММ.ГГ
        """
        if not date_str or pd.isna(date_str):
            return False
        
        try:
            # Парсим дату из строки
            date_str = str(date_str).strip()
            
            # Извлекаем дату (берем первую часть до пробела если есть)
            if ' ' in date_str:
                date_part = date_str.split(' ')[0]
            else:
                date_part = date_str
            
            # Разбираем на составляющие
            parts = date_part.split('.')
            if len(parts) != 3:
                return False
            
            day = int(parts[0])
            month = int(parts[1])
            year_str = parts[2]
            
            # Обработка года (2 или 4 цифры)
            if len(year_str) == 2:
                # Предполагаем, что 00-49 → 2000-2049, 50-99 → 1950-1999
                year = 2000 + int(year_str) if int(year_str) < 50 else 1900 + int(year_str)
            else:
                year = int(year_str)
            
            # Создаем объект datetime
            item_date = datetime(year, month, day)
            
            # Текущая дата для сравнения
            if current_date is None:
                current_date = datetime.now()
            else:
                current_date = datetime.strptime(current_date, '%Y-%m-%d') if isinstance(current_date, str) else current_date
            
            # Вычисляем первый день предыдущего месяца
            if current_date.month == 1:
                # Если январь, то предыдущий месяц - декабрь прошлого года
                first_day_prev_month = datetime(current_date.year - 1, 12, 1)
            else:
                first_day_prev_month = datetime(current_date.year, current_date.month - 1, 1)
            
            # Проверяем диапазон
            is_in_range = first_day_prev_month <= item_date <= current_date
            
            # Для отладки
            if not is_in_range:
                print(f"Дата {item_date} вне диапазона {first_day_prev_month} - {current_date}")
            
            return is_in_range
            
        except Exception as e:
            print(f"Ошибка парсинга даты '{date_str}': {e}")
            return False
    
    # ========== УНИВЕРСАЛЬНЫЕ ФУНКЦИИ ПАРСИНГА (из VBA) ==========
    
    def extract_number_after_hyphen_or_space(input_string, keep_letters=False):
        """
        Универсальная функция для извлечения значения после дефиса или пробела
        keep_letters = True - сохранять буквы (для Куста), False - только цифры (для ЦДНГ)
        """
        if not input_string:
            return ""
        
        input_str = str(input_string)
        
        # Находим нужный префикс в зависимости от того, что ищем
        if keep_letters:
            # Для Куста - ищем "Куст-"
            prefix = "Куст-"
            prefix_pos = input_str.find(prefix)
            if prefix_pos >= 0:
                # Берем всё после "Куст-"
                after_prefix = input_str[prefix_pos + len(prefix):]
                # Ищем точку или конец строки
                dot_pos = after_prefix.find('.')
                if dot_pos >= 0:
                    return after_prefix[:dot_pos].strip()
                else:
                    return after_prefix.strip()
        else:
            # Для ЦДНГ - ищем "ЦДНГ-"
            prefix = "ЦДНГ-"
            prefix_pos = input_str.find(prefix)
            if prefix_pos >= 0:
                # Берем всё после "ЦДНГ-"
                after_prefix = input_str[prefix_pos + len(prefix):]
                # Ищем точку или конец строки
                dot_pos = after_prefix.find('.')
                if dot_pos >= 0:
                    result = after_prefix[:dot_pos]
                else:
                    result = after_prefix
                
                # Если есть дефис (как в "1-2"), берем только первое число
                if '-' in result:
                    result = result.split('-')[0]
                
                # Оставляем только цифры
                cleaned = ''
                for char in result:
                    if char.isdigit():
                        cleaned += char
                    else:
                        break
                return cleaned.strip()
        
        # Если не нашли по префиксу, используем старую логику
        dash_pos = input_str.find('-')
        space_pos = input_str.find(' ')
        
        if dash_pos >= 0 and space_pos >= 0:
            separator_pos = min(dash_pos, space_pos)
        elif dash_pos >= 0:
            separator_pos = dash_pos
        elif space_pos >= 0:
            separator_pos = space_pos
        else:
            separator_pos = -1
        
        if separator_pos >= 0:
            temp_string = input_str[separator_pos + 1:]
            dot_pos = temp_string.find('.')
            
            if dot_pos >= 0:
                result = temp_string[:dot_pos]
            else:
                result = temp_string
            
            if '-' in result and not keep_letters:
                result = result.split('-')[0]
            
            if not keep_letters:
                cleaned_result = ''
                for char in result:
                    if char.isdigit():
                        cleaned_result += char
                    else:
                        break
                result = cleaned_result
            
            return result.strip()
        else:
            return ""
    
    def extract_number_after_skv(input_string):
        """
        Функция для извлечения номера скважины (с буквой, если есть)
        """
        if not input_string:
            return ""
        
        input_str = str(input_string)
        
        # Находим позицию текста "Скв" (регистронезависимо)
        skv_pos = input_str.lower().find('скв')
        
        if skv_pos < 0:
            return ""
        
        # Берем все, что после "Скв"
        temp_string = input_str[skv_pos + 3:]
        
        # Пропускаем пробелы и дефисы
        num_start = 0
        for i, char in enumerate(temp_string):
            if char not in [' ', '-']:
                num_start = i
                break
        
        # Если не нашли начало номера
        if num_start == 0 and (temp_string and temp_string[0] in [' ', '-']):
            return ""
        
        # Ищем конец номера (до точки или до конца строки)
        num_end = len(temp_string)
        for i in range(num_start, len(temp_string)):
            if temp_string[i] == '.':
                num_end = i
                break
        
        # Извлекаем номер (сохраняем ВСЕ символы, включая буквы)
        if num_end > num_start:
            result = temp_string[num_start:num_end]
        else:
            result = ""
        
        return result.strip()
    
    def extract_first_line(input_string):
        """
        Извлекает первую строку из многострочного текста (для уставок)
        """
        if not input_string:
            return ""
        
        input_str = str(input_string)
        
        # Проверяем наличие переноса строки
        if '\n' in input_str:
            # Берем текст до первого переноса строки
            return input_str.split('\n')[0].strip()
        else:
            # Если переноса нет, берем всю строку
            return input_str.strip()

    def extract_date_from_cell(input_string):
        """
        Извлекает дату из многострочной ячейки
        Пример: "10\n26.02.2026 13:06:13" -> "26.02.2026"
        """
        if not input_string:
            return ""
        
        input_str = str(input_string)
        
        # Разбиваем текст по переносам строк
        lines = input_str.split('\n')
        
        # Ищем строку с датой
        for line in lines:
            line = line.strip()
            # Ищем паттерн даты: две цифры.две цифры.четыре цифры
            if '.' in line:
                # Разбиваем по пробелу и берем первую часть
                parts = line.split(' ')
                for part in parts:
                    part = part.strip()
                    # Проверяем, похоже ли на дату (содержит две точки)
                    if part.count('.') == 2:
                        # Проверяем формат ДД.ММ.ГГГГ
                        date_parts = part.split('.')
                        if len(date_parts) == 3 and len(date_parts[2]) == 4:
                            return part
        
        # Если не нашли, пробуем вторую строку
        if len(lines) >= 2:
            second_line = lines[1].strip()
            if ' ' in second_line:
                return second_line.split(' ')[0].strip()
        
        return ""
    
    def extract_time_from_cell(input_string):
        """
        Извлекает время из многострочной ячейки (для времени запуска)
        Пример: "10\n26.02.2026 13:06:13" -> "13:06"
        """
        if not input_string:
            return ""
        
        input_str = str(input_string)
        
        # Разбиваем текст по переносам строк
        lines = input_str.split('\n')
        
        # Ищем строку, содержащую дату и время (формат "ДД.ММ.ГГГГ ЧЧ:ММ:СС")
        for line in lines:
            line = line.strip()
            # Проверяем, есть ли точка (признак даты) и двоеточие (признак времени)
            if '.' in line and ':' in line:
                # Разбиваем по пробелу
                parts = line.split(' ')
                if len(parts) >= 2:
                    # Последняя часть должна быть временем
                    time_part = parts[-1].strip()
                    if ':' in time_part:
                        # Берем только часы и минуты
                        if time_part.count(':') >= 2:
                            return ':'.join(time_part.split(':')[:2])
                        return time_part
        
        # Если не нашли, пробуем вторую строку
        if len(lines) >= 2:
            second_line = lines[1].strip()
            if ' ' in second_line:
                time_part = second_line.split(' ')[-1].strip()
                if ':' in time_part:
                    if time_part.count(':') >= 2:
                        return ':'.join(time_part.split(':')[:2])
                    return time_part
        
        return ""
    
    # ========== ШАГ 1: Выбор ЦИТС ==========
    cits_options = ["ЦИТС VQ-BAD", "ЦИТС Аган"]
    selected_cits = st.radio(
        "Выберите ЦИТС для импорта:",
        cits_options,
        horizontal=True,
        key="tm_import_cits"
    )
    
    st.markdown(f"**Выбран:** {selected_cits}")
    st.markdown("---")
    
    # ========== ЭТАП 1: ИМПОРТ УСТАВОК ==========
    st.markdown("### 📥 Этап 1: Импорт уставок (время работы/накопления)")
    st.info("""
    **Формат файла для уставок:**
    - Столбец A: составная строка (Аганское.ЦДНГ-1-2.Куст-24.АГЗУ-1.Скв-2053Л.СУ-2053Л.)
    - Столбец B: время работы (первая строка - число)
    - Столбец C: время накопления (первая строка - число)
    
    **Результат:** Столбцы D(ЦДНГ), E(Куст), F(Скважина), G(работа), H(накопление)
    """)
    
    schedule_file = st.file_uploader(
        "Загрузите файл с уставками",
        type=['xlsx', 'xls', 'xlsm'],
        key="tm_schedule_file"
    )
    
    if schedule_file is not None:
        # Показываем предпросмотр
        try:
            df_preview = pd.read_excel(schedule_file, dtype=str, header=None, nrows=5)
            df_preview = clean_dataframe(df_preview)
            
            preview_data = []
            for i in range(min(5, len(df_preview))):
                col_a = df_preview.iloc[i, 0] if len(df_preview.columns) > 0 else ''
                col_b = df_preview.iloc[i, 1] if len(df_preview.columns) > 1 else ''
                col_c = df_preview.iloc[i, 2] if len(df_preview.columns) > 2 else ''
                
                col_a_short = str(col_a)[:50] + "..." if len(str(col_a)) > 50 else str(col_a)
                col_b_short = str(col_b).replace('\n', '\\n')[:30] if str(col_b) else ''
                col_c_short = str(col_c).replace('\n', '\\n')[:30] if str(col_c) else ''
                
                preview_data.append({
                    'Столбец A (составная)': col_a_short,
                    'Столбец B (работа)': col_b_short,
                    'Столбец C (накопление)': col_c_short
                })
            
            st.dataframe(pd.DataFrame(preview_data), use_container_width=True, hide_index=True)
        except Exception as e:
            st.error(f"Ошибка предпросмотра: {str(e)}")
        
        if st.button("🚀 Импортировать уставки", key="import_schedule_btn"):
            with st.spinner("Импорт уставок..."):
                try:
                    # Читаем и очищаем файл
                    df_schedule = pd.read_excel(schedule_file, dtype=str, header=None)
                    df_schedule = clean_dataframe(df_schedule)
                    
                    wells_data = st.session_state.wells_data
                    
                    schedule_stats = {
                        'updated': 0,
                        'converted': 0,
                        'not_found': 0,
                        'parsing_errors': 0,
                        'skipped_by_date': 0  # ← добавить
                    }
                    
                    # Обрабатываем каждую строку
                    for idx, row in df_schedule.iterrows():
                        try:
                            if len(row) < 3:
                                schedule_stats['parsing_errors'] += 1
                                continue
                            
                            # Парсим столбец A
                            col_a = str(row[0]) if pd.notna(row[0]) else ''
                            col_b = str(row[1]) if pd.notna(row[1]) else ''
                            col_c = str(row[2]) if pd.notna(row[2]) else ''
                            
                            if not col_a:
                                schedule_stats['parsing_errors'] += 1
                                continue
                            
                            # Извлекаем компоненты
                            tm_cdng = extract_number_after_hyphen_or_space(col_a, keep_letters=False)
                            tm_cluster = extract_number_after_hyphen_or_space(col_a, keep_letters=True)
                            tm_well = extract_number_after_skv(col_a)
                            
                            # Извлекаем время (первая строка)
                            tm_work = extract_first_line(col_b)
                            tm_pause = extract_first_line(col_c)
                            
                            # ========== ПРОВЕРКА ДАТЫ ==========
                            # Извлекаем даты из столбцов B и C
                            work_date_str = col_b.replace(tm_work, '').strip() if tm_work else ''
                            pause_date_str = col_c.replace(tm_pause, '').strip() if tm_pause else ''
                            
                            # Проверяем, попадают ли даты в нужный диапазон
                            work_date_valid = is_date_in_range(work_date_str)
                            pause_date_valid = is_date_in_range(pause_date_str)
                            
                            # Для DEBUG: показываем информацию о датах
                            with st.sidebar.expander(f"📅 Проверка даты (строка {idx+1})", expanded=False):
                                st.write(f"**Дата работы:** {work_date_str} → {'✅ в диапазоне' if work_date_valid else '❌ вне диапазона'}")
                                st.write(f"**Дата накопления:** {pause_date_str} → {'✅ в диапазоне' if pause_date_valid else '❌ вне диапазона'}")
                                st.write(f"**Текущая дата:** {datetime.now().strftime('%d.%m.%Y')}")
                                st.write(f"**Диапазон:** с 01.{datetime.now().month-1 if datetime.now().month > 1 else 12}.{datetime.now().year if datetime.now().month > 1 else datetime.now().year-1} по {datetime.now().strftime('%d.%m.%Y')}")
                            
                            # Если ни одна дата не в диапазоне - пропускаем строку
                            if not work_date_valid and not pause_date_valid:
                                schedule_stats['skipped_by_date'] = schedule_stats.get('skipped_by_date', 0) + 1
                                continue
                            # ========== /ПРОВЕРКА ДАТЫ ==========
                            
                            if not tm_well or not tm_work or not tm_pause:
                                schedule_stats['parsing_errors'] += 1
                                continue
                            
                            if not tm_well or not tm_work or not tm_pause:
                                schedule_stats['parsing_errors'] += 1
                                continue
                            
                            try:
                                work_min = float(tm_work)
                                pause_min = float(tm_pause)
                            except ValueError:
                                schedule_stats['parsing_errors'] += 1
                                continue
                            
                            target_cdng = f"ЦДНГ-{tm_cdng}"

                            # ========== DEBUG ==========
                            st.sidebar.markdown("### 🔍 DEBUG Информация")
                            with st.sidebar.expander(f"Строка {idx+1}", expanded=False):
                                st.write("**Исходные данные:**")
                                st.write(f"- col_a: `{col_a}`")
                                st.write(f"- col_b: `{col_b}`")
                                st.write(f"- col_c: `{col_c}`")
                                
                                st.write("**Результат парсинга:**")
                                st.write(f"- tm_cdng: `{tm_cdng}`")
                                st.write(f"- tm_cluster: `{tm_cluster}`")
                                st.write(f"- tm_well: `{tm_well}`")
                                st.write(f"- tm_work: `{tm_work}`")
                                st.write(f"- tm_pause: `{tm_pause}`")
                                
                                st.write("**Поиск в системе:**")
                                st.write(f"- Ищем: ЦДНГ-{tm_cdng}, Куст {tm_cluster}, Скв {tm_well}")
                                
                                # Показываем все скважины из этого куста для сравнения
                                similar_wells = []
                                for w in wells_data:
                                    if w.get('cdng') == f"ЦДНГ-{tm_cdng}" and w.get('cluster') == tm_cluster:
                                        similar_wells.append({
                                            'name': w.get('name'),
                                            'cluster': w.get('cluster'),
                                            'cdng': w.get('cdng')
                                        })
                                
                                if similar_wells:
                                    st.write("**Скважины в этом кусте:**")
                                    for sw in similar_wells[:5]:  # первые 5
                                        st.write(f"  - `{sw['name']}`")
                                else:
                                    st.write("**Нет скважин в этом кусте!**")
                            # ========== /DEBUG ==========

                            # Поиск скважины
                            found = False
                            for i, well in enumerate(wells_data):
                                if (well.get('cdng') == target_cdng and 
                                    well.get('cluster') == tm_cluster and 
                                    well.get('name') == tm_well and
                                    well.get('cits') == selected_cits):
                                    
                                    # Обновляем уставки
                                    wells_data[i]['schedule'] = [work_min, pause_min]
                                    
                                    # Если была постоянной - конвертируем
                                    if well.get('operation_mode') == 'constant':
                                        wells_data[i]['operation_mode'] = 'kpr'
                                        wells_data[i]['mode'] = 'По времени'
                                        wells_data[i]['converted_from_constant'] = True
                                        schedule_stats['converted'] += 1
                                    
                                    wells_data[i]['last_modified'] = datetime.now().strftime("%Y-%m-%d %H:%M")
                                    wells_data[i]['modification_source'] = 'tm_import_schedule'
                                    
                                    schedule_stats['updated'] += 1
                                    found = True
                                    break
                            
                            if not found:
                                schedule_stats['not_found'] += 1
                                
                        except Exception:
                            schedule_stats['parsing_errors'] += 1
                            continue
                    
                    # Сохраняем
                    if 'save_data_to_file' in globals():
                        save_data_to_file()
                    
                    # Показываем результат
                    st.success("✅ Уставки импортированы!")
                    
                    col1, col2, col3, col4, col5 = st.columns(5)
                    with col1:
                        st.metric("Обновлено уставок", schedule_stats['updated'])
                    with col2:
                        st.metric("Конвертировано в КПР", schedule_stats['converted'])
                    with col3:
                        st.metric("Не найдено", schedule_stats['not_found'])
                    with col4:
                        st.metric("Ошибок парсинга", schedule_stats['parsing_errors'])
                    with col5:
                        st.metric("Пропущено по дате", schedule_stats['skipped_by_date'])
                    
                    # Запоминаем, что уставки загружены
                    st.session_state.schedule_imported = True
                    
                except Exception as e:
                    st.error(f"Ошибка при импорте: {str(e)}")
    
    st.markdown("---")
    
    # ========== ЭТАП 2: ИМПОРТ ВРЕМЕНИ ЗАПУСКА ==========
    st.markdown("### 🕒 Этап 2: Импорт времени запуска")
    st.warning("Время запуска добавляется ТОЛЬКО для скважин, которые уже являются КПР!")
    
    # Проверяем, загружены ли уставки
    if not st.session_state.get('schedule_imported', False):
        st.info("👆 Сначала импортируйте уставки")
    else:
        st.info("""
        **Формат файла для времени запуска:**
        - Столбец A: составная строка (Аганское.ЦДНГ-1-2.Куст-24.АГЗУ-1.Скв-2053Л.СУ-2053Л.)
        - Столбец B: ячейка с временем запуска (например: "10\\n26.02.2026 13:06:13")
        
        **Результат:** Столбцы C(ЦДНГ), D(Куст), E(Скважина), F(Время запуска)
        **Важно:** Обновляется ТОЛЬКО для КПР скважин!
        """)
        
        launch_file = st.file_uploader(
            "Загрузите файл с временем запуска",
            type=['xlsx', 'xls', 'xlsm'],
            key="tm_launch_file"
        )
        
        if launch_file is not None:
            
            # Показываем предпросмотр
            try:
                df_preview = pd.read_excel(launch_file, dtype=str, header=None, nrows=5)
                df_preview = clean_dataframe(df_preview)
                
                preview_data = []
                for i in range(min(5, len(df_preview))):
                    col_a = df_preview.iloc[i, 0] if len(df_preview.columns) > 0 else ''
                    col_b = df_preview.iloc[i, 1] if len(df_preview.columns) > 1 else ''
                    
                    col_a_short = str(col_a)[:50] + "..." if len(str(col_a)) > 50 else str(col_a)
                    col_b_short = str(col_b).replace('\n', '\\n')[:40] if str(col_b) else ''
                    
                    preview_data.append({
                        'Столбец A (составная)': col_a_short,
                        'Столбец B (с временем)': col_b_short
                    })
                
                st.dataframe(pd.DataFrame(preview_data), use_container_width=True, hide_index=True)
                
                # Кнопка Preview парсинга
                if st.button("🔍 Предпросмотр парсинга времени", key="preview_launch"):
                    preview_parsed = []
                    for i in range(min(5, len(df_preview))):
                        col_a = str(df_preview.iloc[i, 0]) if len(df_preview.columns) > 0 else ''
                        col_b = str(df_preview.iloc[i, 1]) if len(df_preview.columns) > 1 else ''
                        
                        cdng_part = extract_number_after_hyphen_or_space(col_a, keep_letters=False)
                        kust_part = extract_number_after_hyphen_or_space(col_a, keep_letters=True)
                        skv_part = extract_number_after_skv(col_a)
                        time_part = extract_time_from_cell(col_b)
                        
                        preview_parsed.append({
                            'ЦДНГ (C)': cdng_part,
                            'Куст (D)': kust_part,
                            'Скважина (E)': skv_part,
                            'Время запуска (F)': time_part
                        })
                    
                    st.success("✅ Пример парсинга:")
                    st.dataframe(pd.DataFrame(preview_parsed), use_container_width=True, hide_index=True)
                
            except Exception as e:
                st.error(f"Ошибка предпросмотра: {str(e)}")
            
            if st.button("🚀 Импортировать время запуска", key="import_launch_btn"):
                with st.spinner("Импорт времени запуска..."):
                    try:
                        # Читаем и очищаем файл
                        df_launch = pd.read_excel(launch_file, dtype=str, header=None)
                        df_launch = clean_dataframe(df_launch)
                        
                        wells_data = st.session_state.wells_data
                        
                        launch_stats = {
                            'updated': 0,
                            'skipped_constant': 0,
                            'not_found': 0,
                            'parsing_errors': 0,
                            'skipped_by_date': 0
                        }
                        
                        # Обрабатываем каждую строку
                        for idx, row in df_launch.iterrows():
                            try:
                                if len(row) < 2:
                                    launch_stats['parsing_errors'] += 1
                                    continue
                                
                                # Парсим столбец A
                                col_a = str(row[0]) if pd.notna(row[0]) else ''
                                col_b = str(row[1]) if pd.notna(row[1]) else ''
                                
                                if not col_a or not col_b:
                                    launch_stats['parsing_errors'] += 1
                                    continue
                                
                                # Извлекаем компоненты
                                tm_cdng = extract_number_after_hyphen_or_space(col_a, keep_letters=False)
                                tm_cluster = extract_number_after_hyphen_or_space(col_a, keep_letters=True)
                                tm_well = extract_number_after_skv(col_a)
                                
                                # Извлекаем время и дату
                                launch_time = extract_time_from_cell(col_b)
                                launch_date = extract_date_from_cell(col_b)
                                
                                # ========== ПРОВЕРКА ПАРСИНГА ==========
                                parsing_ok = True
                                if not tm_cdng:
                                    st.sidebar.error(f"❌ Нет ЦДНГ в строке {idx+1}")
                                    parsing_ok = False
                                if not tm_cluster:
                                    st.sidebar.error(f"❌ Нет куста в строке {idx+1}")
                                    parsing_ok = False
                                if not tm_well:
                                    st.sidebar.error(f"❌ Нет скважины в строке {idx+1}")
                                    parsing_ok = False
                                if not launch_time:
                                    st.sidebar.error(f"❌ Нет времени в строке {idx+1}")
                                    parsing_ok = False
                                if not launch_date:
                                    st.sidebar.error(f"❌ Нет даты в строке {idx+1}")
                                    parsing_ok = False
                                
                                if not parsing_ok:
                                    launch_stats['parsing_errors'] += 1
                                    continue
                                # ========== /ПРОВЕРКА ПАРСИНГА ==========
                                
                                # ========== ПРОВЕРКА ДАТЫ ==========
                                launch_date_valid = is_date_in_range(launch_date)
                                
                                # DEBUG даты
                                with st.sidebar.expander(f"📅 Дата запуска (строка {idx+1})", expanded=False):
                                    st.write(f"**Исходная ячейка:** {repr(col_b)}")
                                    st.write(f"**Извлеченная дата:** {launch_date}")
                                    st.write(f"**Извлеченное время:** {launch_time}")
                                    st.write(f"**В диапазоне:** {'✅' if launch_date_valid else '❌'}")
                                
                                # Если дата не в диапазоне - пропускаем строку
                                if not launch_date_valid:
                                    launch_stats['skipped_by_date'] += 1
                                    continue
                                # ========== /ПРОВЕРКА ДАТЫ ==========
                                
                                # ========== ПРОВЕРКА ФОРМАТА ВРЕМЕНИ ==========
                                try:
                                    if ':' in launch_time:
                                        parts = launch_time.split(':')
                                        if len(parts) >= 2:
                                            hour = int(parts[0])
                                            minute = int(parts[1])
                                            if 0 <= hour <= 23 and 0 <= minute <= 59:
                                                launch_time_formatted = f"{hour:02d}:{minute:02d}"
                                            else:
                                                launch_stats['parsing_errors'] += 1
                                                continue
                                        else:
                                            launch_stats['parsing_errors'] += 1
                                            continue
                                    else:
                                        launch_stats['parsing_errors'] += 1
                                        continue
                                except (ValueError, IndexError):
                                    launch_stats['parsing_errors'] += 1
                                    continue
                                # ========== /ПРОВЕРКА ФОРМАТА ВРЕМЕНИ ==========
                                
                                target_cdng = f"ЦДНГ-{tm_cdng}"
                                
                                # ========== DEBUG ПЕРЕД ПОИСКОМ ==========
                                st.sidebar.markdown("### 🔍 DEBUG Информация (Время запуска)")
                                with st.sidebar.expander(f"Строка {idx+1}", expanded=False):
                                    st.write("**Исходные данные:**")
                                    st.write(f"- col_a: `{col_a}`")
                                    st.write(f"- col_b: `{col_b}`")
                                    
                                    st.write("**Результат парсинга:**")
                                    st.write(f"- tm_cdng: `{tm_cdng}`")
                                    st.write(f"- tm_cluster: `{tm_cluster}`")
                                    st.write(f"- tm_well: `{tm_well}`")
                                    st.write(f"- launch_time: `{launch_time_formatted}`")
                                    st.write(f"- launch_date: `{launch_date}`")
                                    st.write(f"- launch_date_valid: `{launch_date_valid}`")
                                    
                                    st.write("**Поиск в системе:**")
                                    st.write(f"- Ищем: {target_cdng}, Куст {tm_cluster}, Скв {tm_well}")
                                    
                                    # Показываем все скважины из этого куста
                                    similar_wells = []
                                    for w in wells_data:
                                        if w.get('cdng') == target_cdng and w.get('cluster') == tm_cluster:
                                            similar_wells.append({
                                                'name': w.get('name'),
                                                'cluster': w.get('cluster'),
                                                'cdng': w.get('cdng'),
                                                'mode': w.get('operation_mode')
                                            })
                                    
                                    if similar_wells:
                                        st.write("**Скважины в этом кусте:**")
                                        for sw in similar_wells:
                                            st.write(f"  - `{sw['name']}` (режим: {sw['mode']})")
                                    else:
                                        st.write("**❌ Нет скважин в этом кусте!**")
                                        
                                        # Проверяем, есть ли вообще такие ЦДНГ и куст
                                        all_cdng = set(w.get('cdng') for w in wells_data)
                                        all_clusters = set(w.get('cluster') for w in wells_data)
                                        st.write(f"**Доступные ЦДНГ:** {sorted(all_cdng)[:10]}...")
                                        st.write(f"**Доступные кусты:** {sorted(all_clusters)[:10]}...")
                                # ========== /DEBUG ПЕРЕД ПОИСКОМ ==========
                                
                                # ========== ПОИСК СКВАЖИНЫ ==========
                                found = False
                                for i, well in enumerate(wells_data):
                                    if (well.get('cdng') == target_cdng and 
                                        well.get('cluster') == tm_cluster and 
                                        well.get('name') == tm_well and
                                        well.get('cits') == selected_cits):
                                        
                                        # ========== НАШЛИ СКВАЖИНУ ==========
                                        st.sidebar.markdown(f"### 🔥 НАШЛИ СКВАЖИНУ! Строка {idx+1}")
                                        st.sidebar.write(f"**Скважина в системе:** {well.get('name')}")
                                        st.sidebar.write(f"**Режим:** {well.get('operation_mode')}")
                                        st.sidebar.write(f"**Текущее время запуска:** {well.get('base_launch_time', 'Не указано')}")
                                        st.sidebar.write(f"**Новое время:** {launch_time_formatted}")
                                        st.sidebar.write(f"**Дата в диапазоне:** {launch_date_valid}")
                                        # ========== /НАШЛИ СКВАЖИНУ ==========
                                        
                                        # Проверяем режим
                                        if well.get('operation_mode') == 'kpr':
                                            # КПР - обновляем время
                                            old_time = well.get('base_launch_time', 'Не указано')
                                            wells_data[i]['base_launch_time'] = launch_time_formatted
                                            wells_data[i]['last_modified'] = datetime.now().strftime("%Y-%m-%d %H:%M")
                                            wells_data[i]['modification_source'] = 'tm_import_launch'
                                            
                                            # ========== ПОДТВЕРЖДЕНИЕ ==========
                                            st.sidebar.success(f"✅ Время обновлено: {old_time} → {launch_time_formatted}")
                                            # ========== /ПОДТВЕРЖДЕНИЕ ==========
                                            
                                            launch_stats['updated'] += 1
                                        else:
                                            # ========== ПРОПУСК ==========
                                            st.sidebar.warning(f"⏭️ Скважина {well.get('name')} не КПР (режим: {well.get('operation_mode')})")
                                            # ========== /ПРОПУСК ==========
                                            launch_stats['skipped_constant'] += 1
                                        
                                        found = True
                                        break
                                
                                if not found:
                                    # ========== НЕ НАШЛИ ==========
                                    st.sidebar.error(f"❌ НЕ НАШЛИ: {target_cdng}, Куст {tm_cluster}, Скв {tm_well}")
                                    # ========== /НЕ НАШЛИ ==========
                                    launch_stats['not_found'] += 1
                                # ========== /ПОИСК СКВАЖИНЫ ==========
                                    
                            except Exception as e:
                                st.sidebar.error(f"❌ Ошибка в строке {idx+1}: {str(e)}")
                                launch_stats['parsing_errors'] += 1
                                continue
                        
                        # Сохраняем
                        if 'save_data_to_file' in globals():
                            save_data_to_file()
                        
                        # Показываем результат
                        st.success("✅ Время запуска импортировано!")
                        
                        col1, col2, col3, col4, col5 = st.columns(5)
                        with col1:
                            st.metric("Обновлено время запуска", launch_stats['updated'])
                        with col2:
                            st.metric("Пропущено (постоянные)", launch_stats['skipped_constant'])
                        with col3:
                            st.metric("Не найдено", launch_stats['not_found'])
                        with col4:
                            st.metric("Ошибок парсинга", launch_stats['parsing_errors'])
                        with col5:
                            st.metric("Пропущено по дате", launch_stats['skipped_by_date'])
                        
                        # Показываем примеры
                        with st.expander("🔍 Примеры обновленных скважин"):
                            examples = 0
                            for well in wells_data:
                                if well.get('modification_source', '').startswith('tm_import') and examples < 10:
                                    st.write(f"**{well['name']}** (ЦДНГ: {well.get('cdng')}, Куст: {well.get('cluster')})")
                                    st.write(f"- Время запуска: {well.get('base_launch_time', 'Не указано')}")
                                    st.write("---")
                                    examples += 1
                            
                            if examples == 0:
                                st.info("Нет примеров для отображения")
                        
                    except Exception as e:
                        st.error(f"Ошибка при импорте: {str(e)}")

def show_wells_management_tab():
    """Вкладка управления скважинами"""

    _load_plotly()
    
    st.markdown("### Иерархия предприятия")
    
    col_h1, col_h2, col_h3 = st.columns(3)
    with col_h1:
        st.info(f"**ТПП:** {st.session_state.get('selected_tpp', 'VQ-BADнефтегаз')}")
    with col_h2:
        # ВЫБОР ЦИТС
        cits_options = ["ЦИТС VQ-BAD", "ЦИТС Аган"]
        selected_cits = st.selectbox(
            "ЦИТС",
            cits_options,
            key="cits_select_wells"
        )
        st.session_state.selected_cits = selected_cits
    with col_h3:
        # ВЫБОР ЦДНГ
        selected_cits = st.session_state.selected_cits
        wells_data = st.session_state.get('wells_data', [])
        
        # Получаем ЦДНГ для выбранного ЦИТС из загруженных скважин
        cdng_list = []
        for well in wells_data:
            if well.get('cits') == selected_cits and well.get('cdng'):
                if well['cdng'] not in cdng_list:
                    cdng_list.append(well['cdng'])
        
        # Добавляем ЦДНГ из структуры кустов
        if selected_cits in st.session_state.clusters:
            for cdng in st.session_state.clusters[selected_cits].keys():
                if cdng not in cdng_list:
                    cdng_list.append(cdng)
        
        # Если все еще пусто, добавляем по умолчанию
        if not cdng_list:
            cdng_list = [f"ЦДНГ-{i}" for i in range(1, 7)]
        
        selected_cdng = st.selectbox(
            "ЦДНГ",
            cdng_list,
            key="cdng_select_wells"
        )
        st.session_state.selected_cdng = selected_cdng
    
    st.markdown("---")
    
    col_wells1, col_wells2 = st.columns([1, 2])
    
    with col_wells1:
        st.subheader("📁 Управление кустами")
        
        with st.expander("➕ Создать новый куст", expanded=False):
            new_cluster_name = st.text_input("Название куста", placeholder="Куст-1")
            
            col_btn1, col_btn2 = st.columns(2)
            with col_btn1:
                if st.button("Создать", use_container_width=True):
                    if new_cluster_name and new_cluster_name.strip():
                        selected_cits = st.session_state.selected_cits
                        selected_cdng = st.session_state.selected_cdng
                        
                        # Инициализируем структуру, если нет
                        if selected_cits not in st.session_state.clusters:
                            st.session_state.clusters[selected_cits] = {}
                        
                        if selected_cdng not in st.session_state.clusters[selected_cits]:
                            st.session_state.clusters[selected_cits][selected_cdng] = []
                        
                        # Добавляем куст
                        if new_cluster_name not in st.session_state.clusters[selected_cits][selected_cdng]:
                            st.session_state.clusters[selected_cits][selected_cdng].append(new_cluster_name)
                            st.session_state.selected_cluster = new_cluster_name
                            st.success(f"Куст '{new_cluster_name}' создан!")
                            st.rerun()
                        else:
                            st.warning(f"Куст '{new_cluster_name}' уже существует!")
                    else:
                        st.warning("Введите название куста!")
            
            with col_btn2:
                if st.button("Отмена", use_container_width=True, type="secondary"):
                    pass
        
        st.markdown("### Выбор куста")
        
        # Получаем кусты для выбранного ЦИТС и ЦДНГ
        selected_cits = st.session_state.selected_cits
        
        if (selected_cits in st.session_state.clusters and 
            st.session_state.selected_cdng in st.session_state.clusters[selected_cits]):
            clusters = st.session_state.clusters[selected_cits][st.session_state.selected_cdng]
        else:
            clusters = []
        
        if clusters:
            # Используем selectbox для удобного выбора
            clusters_with_default = ["-- Выберите куст --"] + clusters
            
            selected_cluster_dropdown = st.selectbox(
                "Выберите куст для редактирования:",
                clusters_with_default,
                key="cluster_select_dropdown"
            )
            
            if selected_cluster_dropdown and selected_cluster_dropdown != "-- Выберите куст --":
                st.session_state.selected_cluster = selected_cluster_dropdown
                
                # Показываем информацию о выбранном кусте
                well_count = len([w for w in st.session_state.get('wells_data', []) 
                                if w.get('cits') == selected_cits and
                                w.get('cdng') == st.session_state.selected_cdng and 
                                w.get('cluster') == selected_cluster_dropdown])
                
                st.info(f"📊 Выбран куст **{selected_cluster_dropdown}** ({well_count} скважин)")
                
                # Кнопка для перехода к редактированию
                if st.button("📝 Редактировать куст", use_container_width=True):
                    st.rerun()
            else:
                st.session_state.selected_cluster = None
                st.info("👈 Выберите куст из списка")
        else:
            st.info("Создайте первый куст")
    
    with col_wells2:
        if st.session_state.get('selected_cluster'):
            st.subheader(f"📋 Управление скважинами куста: {st.session_state.selected_cluster}")
            
            current_wells = [w for w in st.session_state.get('wells_data', []) 
                           if w.get('cits') == st.session_state.selected_cits and
                           w.get('cdng') == st.session_state.selected_cdng and 
                           w.get('cluster') == st.session_state.selected_cluster]
            
            if current_wells:
                # Создаем редактируемую таблицу
                edited_data = []
                
                for i, well in enumerate(current_wells):
                    # Формируем текст режима КПР
                    if well['operation_mode'] == 'kpr' and well.get('schedule'):
                        schedule_text = f"{well['schedule'][0]}/{well['schedule'][1]}"
                    else:
                        schedule_text = '-'
                    
                    well_info = {
                        '№': i + 1,
                        'Скважина': well['name'],
                        'Тип': 'Постоянная' if well['operation_mode'] == 'constant' else 'КПР',
                        'Дебит, м³/сут': well['flow_rate'],
                        'Время запуска': well.get('base_launch_time', '08:00'),
                        'Режим КПР': schedule_text,
                        'Статус': 'Активна' if well['is_active'] else 'Остановлена',
                        'Искл. из сдвига': 'Да' if well.get('exclude_from_shift', False) else 'Нет',
                        'idx': i
                    }
                    edited_data.append(well_info)
                
                # Создаем DataFrame для редактирования
                df_editable = pd.DataFrame(edited_data)
                
                # Редактируемая таблица
                edited_df = st.data_editor(
                    df_editable,
                    column_config={
                        "№": st.column_config.NumberColumn(disabled=True),
                        "Скважина": st.column_config.TextColumn(disabled=True),
                        "Тип": st.column_config.SelectboxColumn(
                            options=["Постоянная", "КПР"]
                        ),
                        "Дебит, м³/сут": st.column_config.NumberColumn(
                            min_value=0,
                            max_value=1000,
                            step=1
                        ),
                        "Время запуска": st.column_config.TextColumn(),
                        "Режим КПР": st.column_config.TextColumn(),
                        "Статус": st.column_config.SelectboxColumn(
                            options=["Активна", "Остановлена"]
                        ),
                        "Искл. из сдвига": st.column_config.SelectboxColumn(
                            options=["Да", "Нет"]
                        ),
                        "idx": None
                    },
                    hide_index=True,
                    use_container_width=True,
                    num_rows="dynamic",
                    key=f"well_editor_{st.session_state.selected_cluster}"
                )
                
                # Кнопка сохранения изменений
                if st.button("💾 Сохранить изменения", use_container_width=True, type="primary", 
                           key=f"save_btn_{st.session_state.selected_cluster}"):
                    # Находим индексы скважин в основном списке
                    all_wells = st.session_state.wells_data
                    
                    for idx, row in edited_df.iterrows():
                        original_idx = int(row['idx'])
                        if original_idx < len(current_wells):
                            well_name = current_wells[original_idx]['name']
                            
                            # Находим скважину в основном списке
                            for i, well in enumerate(all_wells):
                                if well['name'] == well_name and \
                                   well.get('cits') == st.session_state.selected_cits and \
                                   well.get('cdng') == st.session_state.selected_cdng and \
                                   well.get('cluster') == st.session_state.selected_cluster:
                                    
                                    old_schedule = well.get('schedule')
                                    old_mode = well.get('mode')
                                    
                                    # Обновляем основные данные
                                    all_wells[i]['flow_rate'] = float(row['Дебит, м³/сут'])
                                    all_wells[i]['is_active'] = (row['Статус'] == 'Активна')
                                    all_wells[i]['base_launch_time'] = row['Время запуска']
                                    all_wells[i]['exclude_from_shift'] = (row['Искл. из сдвига'] == 'Да')
                                    
                                    if row['Тип'] == 'Постоянная':
                                        all_wells[i]['operation_mode'] = 'constant'
                                        all_wells[i]['schedule'] = None
                                        all_wells[i]['mode'] = None
                                    else:
                                        all_wells[i]['operation_mode'] = 'kpr'
                                        all_wells[i]['mode'] = old_mode if old_mode else 'По времени'
                                        
                                        try:
                                            if row['Режим КПР'] != '-' and row['Статус'] == 'Активна':
                                                work, pause = map(float, row['Режим КПР'].split('/'))
                                                all_wells[i]['schedule'] = [int(round(work)), int(round(pause))]
                                            else:
                                                all_wells[i]['schedule'] = old_schedule if old_schedule else [15, 45]
                                        except:
                                            all_wells[i]['schedule'] = old_schedule if old_schedule else [15, 45]
                                    
                                    break
                    
                    st.success("✅ Изменения сохранены!")
                    st.rerun()
                
                st.markdown("### 🎨 Визуализация куста")
                cluster_fig = plot_wells_cluster(current_wells)
                st.plotly_chart(cluster_fig, use_container_width=True)
                
                col_actions1, col_actions2 = st.columns(2)
                
                with col_actions1:
                    if st.button("🗑️ Очистить куст", use_container_width=True, type="secondary",
                               key=f"clear_btn_{st.session_state.selected_cluster}"):
                        indices_to_remove = []
                        for i, well in enumerate(st.session_state.wells_data):
                            if well.get('cits') == st.session_state.selected_cits and \
                               well.get('cdng') == st.session_state.selected_cdng and \
                               well.get('cluster') == st.session_state.selected_cluster:
                                indices_to_remove.append(i)
                        
                        for idx in sorted(indices_to_remove, reverse=True):
                            st.session_state.wells_data.pop(idx)
                        
                        st.success(f"Куст '{st.session_state.selected_cluster}' очищен!")
                        st.rerun()
            
            else:
                st.info("В кусте нет скважин")
            
            # ФОРМА РУЧНОГО ДОБАВЛЕНИЯ СКВАЖИНЫ - НОВАЯ ВЕРСИЯ
            st.markdown("---")
            st.subheader("➕ Добавить скважину вручную")
            
            # Создаем контейнер для формы, чтобы она могла обновляться
            form_container = st.container()
            
            with form_container:
                # Инициализируем состояние для формы
                if 'well_form_operation_mode' not in st.session_state:
                    st.session_state.well_form_operation_mode = "constant"
                
                if 'well_form_kpr_mode' not in st.session_state:
                    st.session_state.well_form_kpr_mode = "По времени"
                
                # Поля формы - используем уникальные ключи
                well_name = st.text_input("Название скважины", placeholder="Скв_1", 
                                         key="add_well_name")
                flow_rate = st.number_input("Дебит (м³/сутки)", 1, 1000, 100, 
                                          key="add_well_flow")
                
                # Радиокнопка для типа скважины - реагирует сразу
                operation_mode = st.radio(
                    "Тип скважины",
                    ["constant", "kpr"],
                    format_func=lambda x: "Постоянная" if x == "constant" else "КПР",
                    key="add_well_mode",
                    horizontal=True
                )
                
                # Сохраняем выбор в session_state
                st.session_state.well_form_operation_mode = operation_mode
                
                is_active = st.checkbox("Скважина активна", True, key="add_well_active")
                
                # ДИНАМИЧЕСКИЕ ПОЛЯ ДЛЯ КПР СКВАЖИН
                if operation_mode == "kpr":
                    st.markdown("---")
                    st.subheader("⚙️ Параметры КПР")
                    
                    # Режим работы КПР
                    kpr_mode = st.radio(
                        "Режим работы",
                        ["По времени", "По давлению"],
                        key="add_well_kpr_mode",
                        horizontal=True
                    )
                    
                    # Сохраняем выбор режима КПР
                    st.session_state.well_form_kpr_mode = kpr_mode
                    
                    col3, col4 = st.columns(2)
                    with col3:
                        work_val = st.number_input(f"Работа (мин)", 1, 720, 10,  # 15 часов по умолчанию
                                                  key="add_well_work")
                    with col4:
                        pause_val = st.number_input(f"Пауза (мин)", 1, 720, 20,  # 45 часов по умолчанию
                                                   key="add_well_pause")
                    
                    # Время запуска
                    base_launch_time = st.text_input("Время запуска КПР (ЧЧ:ММ)", "08:00", 
                                                    key="add_well_launch")
                    
                    # Галочка для скважин, работающих по давлению
                    if kpr_mode == "По давлению":
                        exclude_from_shift = st.checkbox(
                            "Не рассчитывать сдвиг времени для этой скважины",
                            value=False,
                            key="add_well_exclude_shift",
                            help="Если включено, скважина будет учитываться в расчете дебита, но для нее не будет рассчитываться оптимальный сдвиг времени"
                        )
                    else:
                        exclude_from_shift = False
                    
                    # Валидация времени
                    time_valid = True
                    try:
                        hours, minutes = map(int, base_launch_time.split(':'))
                        if not (0 <= hours <= 23 and 0 <= minutes <= 59):
                            st.error("⚠️ Некорректное время! Используйте ЧЧ:ММ (00:00-23:59)")
                            time_valid = False
                    except:
                        st.error("⚠️ Некорректный формат! Используйте ЧЧ:ММ")
                        time_valid = False
                else:
                    # Для постоянной скважины - нет параметров КПР
                    kpr_mode = None
                    exclude_from_shift = False
                    work_val = None
                    pause_val = None
                    base_launch_time = '00:00'
                    time_valid = True
                
                # Кнопки добавления
                col_btn1, col_btn2 = st.columns(2)
                with col_btn1:
                    add_button = st.button("✅ Добавить скважину", use_container_width=True,
                                          key="add_well_submit")
                
                with col_btn2:
                    clear_button = st.button("❌ Очистить форму", use_container_width=True,
                                            type="secondary", key="add_well_clear")
                
                # Обработка добавления скважины
                if add_button:
                    if operation_mode == "kpr" and not time_valid:
                        st.error("❌ Исправьте ошибки в форме перед сохранением!")
                    elif not well_name or not well_name.strip():
                        st.error("❌ Введите название скважины!")
                    else:
                        # Проверяем, нет ли уже скважины с таким именем в этом кусте
                        existing_names = [w['name'] for w in current_wells]
                        if well_name in existing_names:
                            st.error(f"❌ Скважина с именем '{well_name}' уже существует в этом кусте!")
                        else:
                            new_well = {
                                'name': well_name,
                                'flow_rate': flow_rate,
                                'operation_mode': operation_mode,
                                'schedule': [work_val, pause_val] if operation_mode == "kpr" else None,
                                'base_launch_time': base_launch_time,
                                'mode': kpr_mode if operation_mode == "kpr" else None,
                                'exclude_from_shift': exclude_from_shift if operation_mode == "kpr" and kpr_mode == "По давлению" else False,
                                'is_active': is_active,
                                'tpp': st.session_state.selected_tpp,
                                'cits': st.session_state.selected_cits,
                                'cdng': st.session_state.selected_cdng,
                                'cluster': st.session_state.selected_cluster,
                                'import_source': 'manual_add',
                                'import_date': datetime.now().strftime("%Y-%m-%d %H:%M")
                            }
                            
                            st.session_state.wells_data.append(new_well)
                            st.success(f"✅ Скважина '{well_name}' добавлена!")
                            
                            # Очищаем поля формы
                            st.session_state.well_form_operation_mode = "constant"
                            st.session_state.well_form_kpr_mode = "По времени"
                            st.rerun()
                
                # Обработка очистки формы
                if clear_button:
                    st.session_state.well_form_operation_mode = "constant"
                    st.session_state.well_form_kpr_mode = "По времени"
                    st.rerun()
        
        else:
            st.info("👈 Выберите куст для управления скважинами")

def show_import_tab():
    """Вкладка импорта данных из техрежима"""
    st.markdown("### 📥 Импорт данных из техрежима")
    st.info("""
    **Инструкция:**
    1. Выберите ЦИТС
    2. Загрузите файл техрежима (Excel)
    3. Система автоматически определит структуру
    4. Данные будут добавлены в систему
    """)
    
    # Выбор ЦИТС для импорта
    cits_options = ["ЦИТС VQ-BAD", "ЦИТС Аган"]
    import_cits = st.selectbox(
        "Выберите ЦИТС для импорта",
        cits_options,
        key="import_cits_select"
    )
    
    uploaded_file = st.file_uploader("Выберите файл техрежима (Excel)", type=['xlsx', 'xls', 'xlsm'])
    
    if uploaded_file is not None:
        with st.spinner("Чтение файла..."):
            wells_data = load_tech_regime_file(uploaded_file, import_cits)
            
            if wells_data:
                st.success(f"✅ Прочитано {len(wells_data)} скважин")
                
                # Показываем предпросмотр
                st.subheader("Предпросмотр данных")
                
                preview_data = []
                for i, well in enumerate(wells_data[:10]):  # Первые 10 для предпросмотра
                    preview_data.append({
                        'Скважина': well['name'],
                        'Куст': well['cluster'],
                        'ЦДНГ': well['cdng'],
                        'ЦИТС': well['cits'],
                        'Тип': 'КПР' if well['operation_mode'] == 'kpr' else 'Постоянная',
                        'Дебит': f"{well['flow_rate']} м³/сут",
                        'Обводненность': f"{well['water_cut']}%"
                    })
                
                preview_df = pd.DataFrame(preview_data)
                st.dataframe(preview_df, use_container_width=True, hide_index=True)
                
                if len(wells_data) > 10:
                    st.caption(f"... и еще {len(wells_data) - 10} скважин")
                
                st.subheader("Структура из файла:")
                
                # Группируем по ЦИТС и ЦДНГ
                cits_data = {}
                for well in wells_data:
                    cits = well.get('cits', 'ЦИТС VQ-BAD')
                    cdng = well.get('cdng', 'ЦДНГ-1')
                    
                    if cits not in cits_data:
                        cits_data[cits] = {}
                    
                    if cdng not in cits_data[cits]:
                        cits_data[cits][cdng] = {
                            'wells': [],
                            'clusters': set()
                        }
                    
                    cits_data[cits][cdng]['wells'].append(well['name'])
                    if well.get('cluster') and well['cluster'] != 'Неизвестно':
                        cits_data[cits][cdng]['clusters'].add(well['cluster'])
                
                # Показываем структуру
                for cits, cdng_dict in cits_data.items():
                    st.write(f"**{cits}:**")
                    for cdng, data in cdng_dict.items():
                        st.write(f"  - **{cdng}**: {len(data['wells'])} скважин, {len(data['clusters'])} кустов")
                        if data['clusters']:
                            clusters_list = sorted(data['clusters'])
                            if len(clusters_list) <= 5:
                                st.write(f"    Кусты: {', '.join(clusters_list)}")
                            else:
                                st.write(f"    Кусты: {', '.join(clusters_list[:5])}... (ещё {len(clusters_list)-5})")
                
                # Кнопка импорта
                col_imp1, col_imp2 = st.columns(2)
                with col_imp1:
                    if st.button("📥 Импортировать данные", type="primary", use_container_width=True):
                        # Сохраняем данные
                        for well in wells_data:
                            # Проверяем, нет ли уже такой скважина
                            existing_idx = -1
                            for i, existing_well in enumerate(st.session_state.wells_data):
                                if existing_well['name'] == well['name'] and \
                                   existing_well.get('cits') == well['cits']:
                                    existing_idx = i
                                    break
                            
                            if existing_idx >= 0:
                                # Обновляем существующую
                                st.session_state.wells_data[existing_idx].update(well)
                            else:
                                # Добавляем новую
                                st.session_state.wells_data.append(well)
                        
                        # Обновляем структуру кустов
                        new_clusters = update_structure_from_wells(wells_data)
                        
                        # Объединяем с существующей структурой
                        for cits, cdng_dict in new_clusters.items():
                            if cits not in st.session_state.clusters:
                                st.session_state.clusters[cits] = {}
                            
                            for cdng, clusters_list in cdng_dict.items():
                                if cdng not in st.session_state.clusters[cits]:
                                    st.session_state.clusters[cits][cdng] = []
                                
                                for cluster in clusters_list:
                                    if cluster not in st.session_state.clusters[cits][cdng]:
                                        st.session_state.clusters[cits][cdng].append(cluster)
                        
                        # Автоматически сохраняем
                        save_data_to_file()
                        
                        st.success(f"✅ Данные успешно импортированы! Добавлено/обновлено {len(wells_data)} скважин")
                        st.rerun()
            else:
                st.error("❌ Не удалось прочитать данные из файла")

def show_optimization():
    """Страница оптимизации"""
    st.title("⚙️ Оптимизация")
    
    tab1, tab2, tab3, tab4 = st.tabs([
        "🎯 Оптимизация времени КПР", 
        "📈 Расчет потенциала КПР",
        "🔄 Перевод постоянных в КПР",
        "📊 Анализ нагрузки"  # ← НОВАЯ ВКЛАДКА
    ])
    
    with tab1:
        show_optimization_tab()
    
    with tab2:
        show_kpr_potential_tab_corrected()
    
    with tab3:
        show_pump_conversion_system()
    
    with tab4:
        show_load_analysis_tab()  # ← НОВАЯ ФУНКЦИЯ

def show_optimization_tab():
    _load_scipy()
    _load_plotly()
    """Вкладка оптимизации времени КПР"""
    st.markdown("### Шаг 1: Выбор куста для оптимизации")
    
    # ДОБАВИЛИ ТРЕТИЙ СТОЛБЕЦ ДЛЯ ЦИТС
    col_opt1, col_opt2, col_opt3 = st.columns(3)
    
    # БЛОК 1: ВЫБОР ЦИТС (НОВЫЙ)
    with col_opt1:
        # Получаем список всех ЦИТС из загруженных скважин
        wells_data = st.session_state.get('wells_data', [])
        
        cits_list = []
        for well in wells_data:
            cits = well.get('cits', 'ЦИТС VQ-BAD')
            if cits and cits not in cits_list:
                cits_list.append(cits)
        
        # Если нет данных, используем значения по умолчанию
        if not cits_list:
            cits_list = ["ЦИТС VQ-BAD", "ЦИТС Аган"]
        
        # Используем уникальный ключ для этого выбора
        selected_cits = st.selectbox(
            "ЦИТС",
            cits_list,
            key="optimization_cits_select",
            index=0
        )
    
    # БЛОК 2: ВЫБОР ЦДНГ ДЛЯ ВЫБРАННОГО ЦИТС
    with col_opt2:
        # Получаем ЦДНГ для выбранного ЦИТС
        cdng_list = []
        for well in wells_data:
            if well.get('cits', 'ЦИТС VQ-BAD') == selected_cits and well.get('cdng'):
                if well['cdng'] not in cdng_list:
                    cdng_list.append(well['cdng'])
        
        # Добавляем ЦДНГ из структуры кустов
        if selected_cits in st.session_state.clusters:
            for cdng in st.session_state.clusters[selected_cits].keys():
                if cdng not in cdng_list:
                    cdng_list.append(cdng)
        
        # Если все еще пусто, добавляем по умолчанию
        if not cdng_list:
            cdng_list = [f"ЦДНГ-{i}" for i in range(1, 7)]
        
        selected_cdng = st.selectbox(
            "ЦДНГ",
            cdng_list,
            key="cdng_opt_select"
        )
    
    # БЛОК 3: ВЫБОР КУСТА ДЛЯ ВЫБРАННЫХ ЦИТС и ЦДНГ
    with col_opt3:
        # Получаем кусты для выбранного ЦИТС и ЦДНГ
        if (selected_cits in st.session_state.clusters and 
            selected_cdng in st.session_state.clusters[selected_cits]):
            clusters = st.session_state.clusters[selected_cits][selected_cdng]
        else:
            clusters = []
        
        if clusters:
            # Используем selectbox для удобного выбора
            clusters_with_default = ["-- Выберите куст --"] + clusters
            
            selected_cluster = st.selectbox(
                "Куст",
                clusters_with_default,
                key="cluster_opt_select"
            )
            
            if selected_cluster and selected_cluster != "-- Выберите куст --":
                st.session_state.opt_selected_cluster = selected_cluster
            else:
                selected_cluster = None
                st.session_state.opt_selected_cluster = None
        else:
            st.warning("В выбранном ЦДНГ нет кустов")
            selected_cluster = None
    
    # ============================================================
    # АВТОМАТИЧЕСКИЙ ПОИСК КУСТОВ С КПР СКВАЖИНАМИ
    # (ЭТОТ БЛОК ДОЛЖЕН БЫТЬ ПОСЛЕ ВЫБОРА КУСТА, НЕ ВНУТРИ col_opt3)
    # ============================================================
    st.markdown("---")
    st.markdown("### 🔍 Автоматический поиск кустов для оптимизации")
    
    # Кнопка автоматического поиска
    col_auto1, col_auto2 = st.columns([3, 1])
    with col_auto1:
        if st.button("🔍 Найти все кусты с КПР скважинами", 
                    key="auto_search_btn", 
                    use_container_width=True,
                    help="Автоматический поиск кустов с 2 и более КПР скважинами для оптимизации"):
            st.session_state.show_auto_search = True
            st.rerun()
    
    with col_auto2:
        if st.session_state.get('show_auto_search', False):
            st.success("✅ Активен")
    
    # Показ автоматического поиска
    if st.session_state.get('show_auto_search', False):
        show_automatic_cluster_search()
        
        # Кнопка возврата к ручному выбору
        if st.button("← Вернуться к выбору куста", key="back_to_manual"):
            st.session_state.show_auto_search = False
            st.rerun()
        
        st.markdown("---")
    
    # ============================================================
    # РУЧНОЙ ВЫБОР КУСТА (ОРИГИНАЛЬНЫЙ КОД)
    # ============================================================
    
    # ПРОДОЛЖАЕМ ОСТАЛЬНОЙ КОД
    if selected_cluster and selected_cluster != "-- Выберите куст --":
        # Фильтруем скважины по выбранным ЦИТС, ЦДНГ и кусту
        current_wells = [w for w in st.session_state.get('wells_data', []) 
                       if w.get('cits', 'ЦИТС VQ-BAD') == selected_cits and
                       w.get('cdng') == selected_cdng and 
                       w.get('cluster') == selected_cluster]
        
        if current_wells:
            st.markdown("---")
            st.markdown(f"### Шаг 2: Предпросмотр куста '{selected_cluster}'")
            
            col_preview1, col_preview2 = st.columns(2)
            
            with col_preview1:
                total_flow = sum(w['flow_rate'] for w in current_wells if w.get('is_active', True))
                kpr_count = sum(1 for w in current_wells if w.get('operation_mode') == 'kpr')
                const_count = sum(1 for w in current_wells if w.get('operation_mode') == 'constant')
                
                st.metric("Суммарный дебит", f"{total_flow} м³/сут")
                st.metric("КПР / Постоянные", f"{kpr_count} / {const_count}")
            
            with col_preview2:
                cluster_fig = plot_wells_cluster(current_wells)
                st.plotly_chart(cluster_fig, use_container_width=True, height=300)
            
            st.markdown("---")
            st.markdown("### Шаг 3: Настройка параметров оптимизации")

            # Целевой коэффициент
            target_coeff = st.slider(
                "Коэффициент загрузки для целевого дебита",
                min_value=0.5,
                max_value=0.9,
                value=0.7,
                step=0.05,
                help="Процент от максимальной мощности системы (рекомендуется 70%)",
                key="target_coeff_slider"
            )
            
            # Текущее время
            current_time = st.text_input(
                "Текущее время (ЧЧ:ММ)",
                value=datetime.now().strftime("%H:%M"),
                key="current_time_opt"
            )
            
            st.markdown("---")
            st.markdown("### Шаг 4: Запуск расчета")
            
            if st.button("🚀 Запустить оптимизацию", use_container_width=True, type="primary"):
                with st.spinner("Выполняется оптимизация..."):
                    try:
                        optimizer = PressureStabilizationOptimizer(current_wells, target_coeff)
                        phases_dict, objective_value, stats = optimizer.optimize()
                        
                        optimization_result = {
                            'timestamp': datetime.now(),
                            'cluster': selected_cluster,
                            'cits': selected_cits,
                            'cdng': selected_cdng,
                            'target_coefficient': target_coeff,
                            'phases_dict': phases_dict,
                            'stats': stats,
                            'current_time': current_time,
                            'wells_data': current_wells,
                            'optimizer': optimizer
                        }
                        
                        if 'calculation_history' not in st.session_state:
                            st.session_state.calculation_history = []
                        
                        st.session_state.calculation_history.append({
                            'Дата': datetime.now().strftime("%Y-%m-%d %H:%M"),
                            'Куст': selected_cluster,
                            'ЦИТС': selected_cits,
                            'ЦДНГ': selected_cdng,
                            'Эффективность': f"{stats['efficiency']:.1f}%",
                            'Целевой дебит': f"{stats['target_flow']:.1f} м³/час"
                        })
                        
                        st.session_state.last_optimization = optimization_result
                        st.session_state.show_results = True

                        save_data_to_file()
                        
                        st.success("✅ Оптимизация завершена!")
                        st.rerun()
                        
                    except Exception as e:
                        st.error(f"❌ Ошибка оптимизации: {str(e)}")
            
            # Показ результатов
            if st.session_state.get('show_results', False) and st.session_state.get('last_optimization'):
                results = st.session_state.last_optimization
                
                st.markdown("---")
                st.markdown("### 📊 Результаты оптимизации")
                
                # Сравнительная таблица
                st.subheader("📈 Сравнение до и после")
                comparison_df = create_comparison_table(results['stats'])
                st.dataframe(comparison_df, use_container_width=True, hide_index=True)
                
                # Основные метрики
                col_res1, col_res2, col_res3, col_res4 = st.columns(4)
                
                with col_res1:
                    delta_eff = results['stats']['efficiency'] - 50
                    st.metric(
                        "🎯 Общая эффективность",
                        f"{results['stats']['efficiency']:.1f}%",
                        delta=f"{delta_eff:+.1f}%" if delta_eff > 0 else None,
                        delta_color="normal" if delta_eff > 0 else "off"
                    )
                
                with col_res2:
                    st.metric(
                        "📈 Достижение цели",
                        f"{results['stats']['target_achievement']:.1f}%",
                        delta=f"Цель: {results['stats']['target_flow']:.1f} м³/час"
                    )
                
                with col_res3:
                    st.metric(
                        "🔄 Улучшение стабильности",
                        f"+{results['stats']['stability_improvement']:.1f}%",
                        delta_color="normal"
                    )
                
                with col_res4:
                    color = "normal" if results['stats']['peaks_improvement'] > 0 else "off"
                    st.metric(
                        "⚡ Снижение пиков",
                        f"+{results['stats']['peaks_improvement']:.1f}%",
                        delta=f"{results['stats']['peaks_before']} → {results['stats']['peaks_after']}",
                        delta_color=color
                    )
                
                # Рекомендации
                st.subheader("⏰ Рекомендации для оператора")
                st.info(f"Текущее время: **{current_time}**")
                
                recommendations = calculate_next_launch_times(
                    current_wells,
                    results['phases_dict'],
                    current_time
                )
                
                if recommendations:
                    df_recommendations = pd.DataFrame(recommendations)
                    st.dataframe(df_recommendations, use_container_width=True, hide_index=True)
                    
                    # Инструкция
                    st.info("""
                    **📋 ИНСТРУКЦИЯ ДЛЯ ОПЕРАТОРА:**
                    1. **ЗАПУСТИТЬ РАНЬШЕ** - установите время запуска как в колонке "Оптим. время"
                    2. **ОТЛОЖИТЬ запуск** - пропустите ближайший КПР запуск, дождитесь оптимального времени
                    3. **ОСТАВИТЬ как есть** - скважина уже работает в оптимальном режиме
                    4. **НЕ МЕНЯТЬ (исключена)** - скважина работает по давлению, оставьте текущий график
                    
                    **Контроль:** После внедрения отслеживайте стабильность давления в коллекторе.
                    """)
                
                # Графики
                st.subheader("📊 Визуализация")
                
                # График дебитов (сверху)
                fig_debits = plot_pressure_optimization_results(
                    current_wells,
                    results['phases_dict'],
                    current_time,
                    results['stats']['target_flow']
                )
                st.plotly_chart(fig_debits, use_container_width=True)
                
                # График количества работающих скважин (снизу)
                fig_working = plot_working_wells_count(
                    current_wells,
                    results['phases_dict'],
                    current_time
                )
                st.plotly_chart(fig_working, use_container_width=True)
                
                # Статистика по работе скважин
                working_stats = calculate_working_stats(current_wells, results['phases_dict'])
                
                col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
                
                with col_stat1:
                    st.metric(
                        "Макс. одновременно",
                        f"{working_stats['after']['max']} скв",
                        delta=f"{working_stats['after']['max'] - working_stats['before']['max']:+.0f}"
                    )
                
                with col_stat2:
                    st.metric(
                        "Мин. одновременно",
                        f"{working_stats['after']['min']} скв",
                        delta=f"{working_stats['after']['min'] - working_stats['before']['min']:+.0f}"
                    )
                
                with col_stat3:
                    st.metric(
                        "Разброс",
                        f"{working_stats['after']['range']:.0f} скв",
                        delta=f"{working_stats['improvements']['range_reduction']:.1f}%",
                        delta_color="inverse" if working_stats['improvements']['range_reduction'] > 0 else "off"
                    )
                
                with col_stat4:
                    st.metric(
                        "Равномерность",
                        f"{working_stats['after']['std']:.2f}",
                        delta=f"{working_stats['improvements']['std_reduction']:.1f}%",
                        delta_color="inverse" if working_stats['improvements']['std_reduction'] > 0 else "off"
                    )
                
                # Экспорт в Excel
                st.subheader("📥 Экспорт результатов")
                
                if st.button("📊 Создать детальный отчет в Excel", use_container_width=True):
                    # Добавляем дополнительную информацию для отчета
                    results['tpp'] = st.session_state.get('selected_tpp', 'VQ-BADнефтегаз')
                    results['timestamp'] = datetime.now()
                    results['current_time'] = current_time
                    
                    excel_file = create_pressure_stabilization_report(results)
                    
                    st.download_button(
                        label="📥 Скачать отчет в Excel",
                        data=excel_file.getvalue(),
                        file_name=f"стабилизация_давления_{selected_cluster}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

        
        else:
            st.warning(f"В кусте '{selected_cluster}' нет скважин")
    else:
        st.info("👈 Выберите ЦИТС, ЦДНГ и куст для начала оптимизации")
        
import streamlit as st
import pandas as pd
import numpy as np
import math
from typing import Dict, List, Tuple, Optional, Union, Any
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from datetime import datetime
from io import BytesIO
import json
import os
import traceback
import sys

# ============================================================
# НАСТРОЙКА ЛОГИРОВАНИЯ ДЛЯ ОТЛАДКИ
# ============================================================

class DebugLogger:
    """Логгер для отладки расчетов (вывод в консоль)"""
    
    def __init__(self, enabled=True):
        self.enabled = enabled
        self.indent_level = 0
    
    def log(self, message: str, level: str = "INFO"):
        """Запись в лог"""
        if self.enabled:
            indent = "  " * self.indent_level
            timestamp = datetime.now().strftime("%H:%M:%S")
            print(f"[{timestamp}] {indent}{level}: {message}", file=sys.stderr)
    
    def section(self, title: str):
        """Начало секции"""
        if self.enabled:
            print(f"\n{'='*60}", file=sys.stderr)
            print(f"{' ' * self.indent_level}{title}", file=sys.stderr)
            print(f"{'='*60}", file=sys.stderr)
    
    def data(self, name: str, value: Any):
        """Вывод данных"""
        if self.enabled:
            indent = "  " * self.indent_level
            print(f"{indent}📊 {name}: {value}", file=sys.stderr)
    
    def enter(self):
        """Увеличить отступ"""
        self.indent_level += 1
    
    def exit(self):
        """Уменьшить отступ"""
        self.indent_level = max(0, self.indent_level - 1)

# Глобальный логгер
DEBUG = DebugLogger(enabled=True)

# ============================================================
# ИСПРАВЛЕННЫЙ КЛАСС ФИЗИЧЕСКИХ РАСЧЕТОВ
# ============================================================

class CorrectedKPRPhysics:
    """Исправленный класс физических расчетов с правильной физикой КПР"""
    
    # Константы
    G = 9.81  # м/с²
    ATM_TO_PA = 101325  # 1 атм = 101325 Па
    WATER_DENSITY = 1000  # кг/м³
    GAS_DENSITY_ZABOI = 1.3  # кг/м³ при забойных условиях
    K_DISPERSION = 0.7  # Коэффициент дисперсии газ в НКТ/газ в затрубе
    ANNULAR_CAPACITY = 9.5  # л/м (емкость кольцевого пространства на 1 м)
    
    # Геометрические параметры
    D_COLUMN = 0.146  # м, диаметр колонны
    D_TUBING = 0.073  # м, диаметр НКТ

    WATER_CUT_TABLE = [0, 20, 50, 80, 100]  # обводненность, %
    DENSITY_TABLE = [840, 880, 930, 970, 1010]  # плотность, кг/м³
    VISCOSITY_TABLE = [20, 35, 120, 15, 1]  # вязкость, сП
    
    def __init__(self, well_data: Dict):
        """Инициализация с данными скважины"""
        self.well = well_data
        
        # Площадь поперечного сечения затрубного пространства
        self.area_annulus = math.pi * (self.D_COLUMN**2 - self.D_TUBING**2) / 4  # м²
        
        DEBUG.log(f"Инициализация физики для скважины {well_data.get('name', 'Unknown')}")
        DEBUG.data("area_annulus", f"{self.area_annulus:.6f} м²")
    
    @staticmethod
    def safe_float(value: Optional[Union[str, float, int]], default: float = 0.0) -> float:
        """Безопасное преобразование в float"""
        if value is None:
            return default
        try:
            return float(value)
        except (ValueError, TypeError):
            return default

    def interpolate_table(self, x: float, x_vals: list, y_vals: list) -> float:
        """
        Линейная интерполяция по таблице (как в VBA)
        """
        # Если x меньше минимального
        if x <= x_vals[0]:
            return y_vals[0]
        
        # Если x больше максимального
        if x >= x_vals[-1]:
            return y_vals[-1]
        
        # Поиск интервала
        for i in range(len(x_vals) - 1):
            if x >= x_vals[i] and x <= x_vals[i + 1]:
                # Линейная интерполяция
                t = (x - x_vals[i]) / (x_vals[i + 1] - x_vals[i])
                return y_vals[i] + t * (y_vals[i + 1] - y_vals[i])
        
        return y_vals[-1]
    
    def convert_rotations_to_hz(self, rotations_value: Optional[Union[str, float, int]]) -> float:
        """Конвертация оборотов в Гц с валидацией"""
        # Значение по умолчанию
        DEFAULT_FREQ = 50.0
        
        if rotations_value is None or rotations_value == '':
            DEBUG.data("Частота отсутствует", f"используется {DEFAULT_FREQ} Гц")
            return DEFAULT_FREQ
        
        try:
            val = float(rotations_value)
            
            # Проверка на 0 или отрицательное
            if val <= 0:
                DEBUG.data(f"Частота {val} <= 0", f"используется {DEFAULT_FREQ} Гц")
                return DEFAULT_FREQ
            
            if val > 100:
                # об/мин → Гц: (об/мин * 50) / 2910
                hz = (val * 50.0) / 2910.0
                # Проверка результата
                if hz <= 0 or hz > 100:
                    DEBUG.data(f"Результат конвертации {hz:.1f} вне диапазона", f"используется {DEFAULT_FREQ} Гц")
                    return DEFAULT_FREQ
                DEBUG.data("Конвертация оборотов", f"{val:.0f} об/мин → {hz:.1f} Гц")
                return hz
            else:
                # уже в Гц, проверяем диапазон
                if val < 30 or val > 70:
                    DEBUG.data(f"Частота {val:.1f} Гц вне рабочего диапазона", f"используется {DEFAULT_FREQ} Гц")
                    return DEFAULT_FREQ
                DEBUG.data("Частота", f"{val:.1f} Гц")
                return val
        except:
            DEBUG.data("Ошибка конвертации", f"используется {DEFAULT_FREQ} Гц")
            return DEFAULT_FREQ
    
    def safe_frequency(self, freq_value: Optional[Union[str, float, int]]) -> float:
        """Безопасное преобразование частоты (по умолчанию 50 Гц)"""
        if freq_value is None:
            return 50.0
        
        try:
            freq = float(freq_value)
            if freq <= 0 or freq > 100:
                DEBUG.data("Частота невалидна", f"{freq} → 50 Гц")
                return 50.0
            DEBUG.data("Частота", f"{freq:.1f} Гц")
            return freq
        except (ValueError, TypeError):
            DEBUG.data("Ошибка частоты", "используется 50 Гц")
            return 50.0

    def calculate_gas_free(self, p_intake: float, p_sat: float, 
                           gas_factor_m3_per_t: float, q_oil_t_per_day: float,
                           oil_density_ton: float, q_nom: float, freq: float) -> dict:
        """
        Расчет свободного газа по формуле из VBA
        
        Returns:
            dict: {
                'q_gas_free': объем свободного газа, м³/сут
                'gas_fraction': доля газа в насосе
                'k_degr': коэффициент деградации (1 - газ_fraction)
            }
        """
        DEBUG.enter()
        DEBUG.log(f"Расчет свободного газа: Pприем={p_intake:.1f}, Pнас={p_sat:.1f}, ГФ={gas_factor_m3_per_t:.0f}")
        
        result = {
            'q_gas_free': 0.0,
            'gas_fraction': 0.0,
            'k_degr': 1.0
        }
        
        # Если нет давления насыщения или Pприем >= Pнас - газа нет
        if p_sat <= 0 or p_intake >= p_sat:
            DEBUG.log("Нет свободного газа (Pприем ≥ Pнас)")
            DEBUG.exit()
            return result
        
        # Переводим газовый фактор из м³/т в м³/м³
        # ГФ_м3_м3 = ГФ_м3_т * плотность_нефти_т_м3
        gor_m3_m3 = gas_factor_m3_per_t * oil_density_ton
        DEBUG.data("ГФ м³/м³", f"{gor_m3_m3:.1f}")
        
        # Формула из VBA:
        # Qг своб = Qн * (ГФ - (ГФ/Pнас)*Pприем) * 0.12114 * (Pприем/10)^(-1.02207)
        if p_sat > 0:
            term1 = gor_m3_m3 - (gor_m3_m3 / p_sat) * p_intake
            if term1 < 0:
                term1 = 0
            
            term2 = 0.12114 * (p_intake / 10) ** (-1.02207)
            
            q_gas_free = q_oil_t_per_day * term1 * term2
            
            if q_gas_free < 0:
                q_gas_free = 0
                
            DEBUG.data("term1", f"{term1:.3f}")
            DEBUG.data("term2", f"{term2:.3f}")
            DEBUG.data("Qг своб", f"{q_gas_free:.1f} м³/сут")
        else:
            q_gas_free = 0
        
        # Доля газа в насосе
        # gas_fraction = Qг_своб / (Qном * (f/50))
        denominator = q_nom * (freq / 50.0)
        if denominator > 0:
            gas_fraction = q_gas_free / denominator
            if gas_fraction > 1:
                gas_fraction = 1
        else:
            gas_fraction = 0
        
        DEBUG.data("Доля газа", f"{gas_fraction*100:.1f}%")
        
        # Коэффициент деградации
        k_degr = 1.0 - gas_fraction
        if k_degr < 0.2:
            k_degr = 0.2  # минимум 20% эффективности
            DEBUG.data("K дегр ограничен", f"{k_degr:.3f} (минимум)")
        
        DEBUG.data("K деградации", f"{k_degr:.3f}")
        
        result['q_gas_free'] = q_gas_free
        result['gas_fraction'] = gas_fraction
        result['k_degr'] = k_degr
        
        DEBUG.exit()
        return result

    @st.cache_data(ttl=3600)
    def calculate_gas_fraction_in_annulus(_self, p_zab: float, p_nas: float, gas_factor: float) -> float:
        """
        Доля свободного газа в затрубе (в долях единицы)
        """
        DEBUG.enter()
        DEBUG.log(f"Расчет доли газа: Pзаб={p_zab:.1f}, Pнас={p_nas:.1f}, ГФ={gas_factor:.0f}")
        
        if p_nas <= 0 or p_zab >= p_nas:
            DEBUG.log("Нет свободного газа (Pзаб ≥ Pнас или Pнас=0)")
            DEBUG.exit()
            return 0.0
        
        # Относительное снижение давления
        relative_drop = (p_nas - p_zab) / p_nas
        DEBUG.data("Относительное падение", f"{relative_drop:.3f}")
        
        # Максимальная доля газа при полном выделении (эмпирически 30%)
        max_gas_fraction = 0.3
        
        # Экспоненциальная зависимость
        gas_fraction = max_gas_fraction * (1.0 - math.exp(-3.0 * relative_drop))
        
        # Ограничение сверху
        result = min(gas_fraction, 0.4)
        DEBUG.data("Доля газа в затрубе", f"{result:.3f}")
        DEBUG.exit()
        
        return result
    
    def get_mixture_densities(self, p_zab_vdp: float, temperature_c: float = 84.0) -> Tuple[float, float]:
        """
        Возвращает (ρ_annulus, ρ_tubing) в кг/м³
        """
        DEBUG.enter()
        DEBUG.log(f"Расчет плотности смеси: Pзаб(ВДП)={p_zab_vdp:.1f} атм")
        
        # Получаем данные скважины
        water_cut_percent = self.safe_float(self.well.get('water_cut', 0.0))
        oil_density_ton_per_m3 = self.safe_float(self.well.get('oil_density', 0.84))  # т/м³!
        p_nas = self.safe_float(self.well.get('p_nas', 0.0))
        gas_factor = self.safe_float(self.well.get('gas_factor', 50.0))
        
        DEBUG.data("Входные данные", f"Обв={water_cut_percent:.1f}%, ρнефти={oil_density_ton_per_m3:.3f} т/м³, Pнас={p_nas:.1f}, ГФ={gas_factor:.0f}")
        
        # Преобразование единиц
        wc = water_cut_percent / 100.0  # доля
        oil_density_kg_per_m3 = oil_density_ton_per_m3 * 1000.0  # т/м³ → кг/м³
        
        # ===== НОВЫЙ КОД: интерполяция плотности по таблице =====
        # Плотность чистой воды всегда 1000, чистой нефти - из данных
        # Но для смеси используем таблицу из VBA для большей точности
        if water_cut_percent <= 100:
            # Интерполируем по таблице
            rho_liquid = self.interpolate_table(
                water_cut_percent, 
                self.WATER_CUT_TABLE, 
                self.DENSITY_TABLE
            )
        else:
            # Если обводненность >100% (бывает?), используем старую формулу
            rho_liquid = wc * self.WATER_DENSITY + (1.0 - wc) * oil_density_kg_per_m3
        
        # Также получаем вязкость для дальнейших расчетов
        self.current_viscosity_cP = self.interpolate_table(
            water_cut_percent,
            self.WATER_CUT_TABLE,
            self.VISCOSITY_TABLE
        )
        self.current_viscosity_Pa_s = self.current_viscosity_cP * 0.001  # сП → Па·с
        DEBUG.data("Плотность жидкости без газа", f"{rho_liquid:.1f} кг/м³")
        
        # 2. Проверка на свободный газ
        if p_nas <= 0 or p_zab_vdp >= p_nas:
            # Нет свободного газа
            DEBUG.log("Нет свободного газа → ρзатр = ρнкт = ρжидкости")
            DEBUG.data("Результат", f"ρ={rho_liquid:.1f} кг/м³")
            DEBUG.exit()
            return rho_liquid, rho_liquid
        
        # 3. Доля газа в затрубе
        gas_fraction_annulus = self.calculate_gas_fraction_in_annulus(p_zab_vdp, p_nas, gas_factor)
        
        # 4. Плотность газа при забойных условиях
        rho_gas = self.GAS_DENSITY_ZABOI
        DEBUG.data("Плотность газа", f"{rho_gas:.1f} кг/м³")
        
        # 5. Плотность в затрубе (с газом)
        rho_annulus = ((1.0 - gas_fraction_annulus) * rho_liquid + 
                       gas_fraction_annulus * rho_gas)
        DEBUG.data("ρ_затрубное", f"{rho_annulus:.1f} кг/м³ (газа {gas_fraction_annulus*100:.1f}%)")
        
        # 6. Доля газа в НКТ (меньше из-за дисперсии)
        gas_fraction_tubing = gas_fraction_annulus * self.K_DISPERSION
        DEBUG.data("Доля газа в НКТ", f"{gas_fraction_tubing*100:.1f}% (Kдисп={self.K_DISPERSION})")
        
        # 7. Плотность в НКТ
        rho_tubing = ((1.0 - gas_fraction_tubing) * rho_liquid + 
                      gas_fraction_tubing * rho_gas)
        DEBUG.data("ρ_нкт", f"{rho_tubing:.1f} кг/м³")
        
        DEBUG.log(f"Итог: ρзатр={rho_annulus:.1f}, ρнкт={rho_tubing:.1f}")
        DEBUG.exit()
        
        return rho_annulus, rho_tubing
    
    def has_critical_gas_problem(self, p_zab_calculated: float, p_nas: float) -> bool:
        """Критическая газовая проблема: РАСЧИТАННОЕ Pзаб < 0.75*Pнас"""
        if p_nas <= 0:
            return False
        critical = p_zab_calculated < (0.75 * p_nas)
        if critical:
            DEBUG.log(f"КРИТИЧЕСКАЯ газовая проблема: Pзаб={p_zab_calculated:.1f} < 0.75*Pнас={0.75*p_nas:.1f}")
        return critical
    
    def has_gas_problem(self, p_zab_calculated: float, p_nas: float) -> bool:
        """Есть ли газовая проблема: РАСЧИТАННОЕ Pзаб < Pнас"""
        if p_nas <= 0:
            DEBUG.log("Нет Pнас → нет анализа газа")
            return False
        problem = p_zab_calculated < p_nas
        if problem:
            DEBUG.log(f"Газовая проблема: Pзаб={p_zab_calculated:.1f} < Pнас={p_nas:.1f}")
        return problem

    def calculate_pressure_at_pump_intake(self, well_data: Dict) -> float:
        """
        Расчет давления на приеме насоса
        (Перенесено из класса WellPotentialAnalyzer)
        """
        DEBUG.enter()
        DEBUG.log("Расчет давления на приеме насоса")
        
        # Пробуем взять из данных
        p_pr_data = self.safe_float(well_data.get('p_pr', 0))
        
        if p_pr_data > 0:
            DEBUG.data("Pпр из данных", f"{p_pr_data:.1f} атм")
            DEBUG.exit()
            return p_pr_data
        
        # Если нет в данных, рассчитываем
        p_zat = self.safe_float(well_data.get('p_zat', 0))
        pump_depth = self.safe_float(well_data.get('pump_depth', 0))
        h_din = self.safe_float(well_data.get('h_din', 0))
        
        DEBUG.data("Исходные для расчета", f"Pзатр={p_zat:.1f}, Lнас={pump_depth:.1f}, Hдин={h_din:.1f}")
        
        if pump_depth <= 0 or h_din <= 0:
            DEBUG.log("Ошибка: нет данных для расчета")
            DEBUG.exit()
            return 0.0
        
        # Расчет плотности жидкости
        water_cut_percent = self.safe_float(well_data.get('water_cut', 0.0))
        oil_density_ton_per_m3 = self.safe_float(well_data.get('oil_density', 0.84))
        
        wc = water_cut_percent / 100.0
        oil_density_kg_per_m3 = oil_density_ton_per_m3 * 1000.0
        
        rho_liquid = wc * self.WATER_DENSITY + (1.0 - wc) * oil_density_kg_per_m3
        
        # P_прием = P_затрубное + ρgh
        height_diff = pump_depth - h_din
        p_hydrostatic = rho_liquid * self.G * height_diff / self.ATM_TO_PA
        
        p_pr = p_zat + p_hydrostatic
        
        DEBUG.data("Параметры расчета", f"ρ={rho_liquid:.1f} кг/м³, Δh={height_diff:.1f} м")
        DEBUG.data("P_гидростатическое", f"{p_hydrostatic:.1f} атм")
        DEBUG.data("P_прием расчетное", f"{p_pr:.1f} атм = {p_zat:.1f} + {p_hydrostatic:.1f}")
        
        DEBUG.exit()
        return p_pr
        
    def calculate_pwf_comprehensive_corrected(self, h_din: float, q_liquid_instant: float,
                                            pump_head_pasport: float, pump_depth: Optional[float] = None,
                                            is_pump_working: bool = True) -> float:
        """
        Расчет забойного давления с правильной физикой и формулами из VBA
        """
        DEBUG.enter()
        DEBUG.log(f"Расчет Pзаб: Hдин={h_din:.1f} м, Q={q_liquid_instant/24:.1f} м³/сут, Напор(паспорт)={pump_head_pasport:.0f} м, Работа={is_pump_working}")
        
        try:
            if pump_depth is None:
                pump_depth = self.safe_float(self.well.get('pump_depth', 0.0))
            
            # Получаем p_zab_vdp из данных скважины для расчета плотности
            p_zab_vdp = self.safe_float(self.well.get('p_zab_vdp', 0.0))
            DEBUG.data("Исходные", f"Lнас={pump_depth:.1f} м, Pзаб(ВДП)={p_zab_vdp:.1f} атм")
            
            # Получаем плотности для этого Pзаб (используем p_zab_vdp из данных)
            rho_annulus, rho_tubing = self.get_mixture_densities(p_zab_vdp)
            
            # Сохраняем вязкость для расчетов трения
            if hasattr(self, 'current_viscosity_Pa_s'):
                viscosity = self.current_viscosity_Pa_s
            else:
                viscosity = 0.001  # запасной вариант
            
            # 1. Столб в затрубе (от забоя до динамического уровня)
            liquid_column_annulus = pump_depth - h_din
            p_hydrostatic_annulus = rho_annulus * self.G * liquid_column_annulus / self.ATM_TO_PA
            DEBUG.data("Столб затруба", f"высота={liquid_column_annulus:.1f} м, давление={p_hydrostatic_annulus:.1f} атм")
            
            # 2. Столб в НКТ (от забоя до устья)
            p_hydrostatic_tubing = rho_tubing * self.G * pump_depth / self.ATM_TO_PA
            DEBUG.data("Столб НКТ", f"давление={p_hydrostatic_tubing:.1f} атм")
            
            if is_pump_working:
                # 3. Трение в НКТ с формулой Альтшуля (как в VBA)
                d_nkt = self.D_TUBING
                l_nkt = pump_depth - h_din if pump_depth > h_din else pump_depth
                a_cross = math.pi * d_nkt**2 / 4.0
                
                # Скорость жидкости в НКТ
                v_liquid = ((q_liquid_instant/24) / (24.0 * 3600.0)) / a_cross if a_cross > 0 else 0.0
                DEBUG.data("Скорость в НКТ", f"{v_liquid:.3f} м/с, площадь={a_cross:.6f} м²")
                
                if v_liquid > 0:
                    # Число Рейнольдса с РЕАЛЬНОЙ вязкостью
                    re = (rho_tubing * v_liquid * d_nkt) / viscosity
                    DEBUG.data("Число Рейнольдса", f"{re:.0f}")
                    
                    # Абсолютная шероховатость (как в VBA)
                    k_abs = 0.00015  # м
                    
                    if re < 2320.0:
                        f = 64.0 / re
                        DEBUG.data("Режим", "ламинарный")
                    else:
                        # Формула Альтшуля для турбулентного режима
                        f = 0.11 * (k_abs / d_nkt + 68.0 / re) ** 0.25
                        DEBUG.data("Режим", "турбулентный (Альтшуль)")
                    
                    # Потери в метрах (как в VBA)
                    h_loss = f * (l_nkt / d_nkt) * (v_liquid**2 / (2 * self.G))
                    
                    # Переводим в атмосферы
                    p_friction_atm = h_loss * rho_tubing * self.G / self.ATM_TO_PA
                    DEBUG.data("Потери на трение", f"{p_friction_atm:.2f} атм (H_loss={h_loss:.2f} м)")
                else:
                    p_friction_atm = 0.0
                    h_loss = 0.0
                
                # ===== НОВЫЙ КОД: расчет напора, который реально развивает насос =====
                # Получаем данные скважины
                p_buf = self.safe_float(self.well.get('buffer_pressure', 0))
                oil_density_ton = self.safe_float(self.well.get('oil_density', 0.86))
                raw_freq = self.well.get('rotations_hz')
                freq_hz = self.convert_rotations_to_hz(raw_freq)
                
                # Средняя плотность для расчетов
                rho_mix = (rho_annulus + rho_tubing) / 2.0
                
                # Перевод буферного давления в метры СТОЛБА ЖИДКОСТИ (как в VBA)
                # В VBA: (data.PUst) * 10.33 * (1000 / data.Density)
                h_from_pbuf = p_buf * 10.33 * (1000.0 / rho_mix)
                
                # Требуемый напор = Hдин + напор от буфера + потери на трение
                required_head = h_din + h_from_pbuf + h_loss
                
                DEBUG.data("H от буфера", f"{h_from_pbuf:.1f} м")
                DEBUG.data("H потерь", f"{h_loss:.1f} м")
                DEBUG.data("Требуемый напор (с запасом)", f"{required_head:.1f} м")
                
                # Максимально возможный напор насоса при текущей частоте
                max_possible_head = pump_head_pasport * (freq_hz / 50.0) ** 2
                DEBUG.data("Макс. возможный напор", f"{max_possible_head:.1f} м (при {freq_hz:.1f} Гц)")
                
                # Фактический напор, который развивает насос
                actual_head = required_head
                if actual_head > max_possible_head:
                    DEBUG.data("Напор ограничен!", f"{actual_head:.1f} > {max_possible_head:.1f}, используем макс.")
                    actual_head = max_possible_head
                
                DEBUG.data("Фактический напор", f"{actual_head:.1f} м")
                # ===== КОНЕЦ НОВОГО КОДА =====
                
                # 4. Давление, создаваемое насосом
                p_pump_head = rho_mix * self.G * actual_head / self.ATM_TO_PA
                DEBUG.data("Давление насоса", f"{p_pump_head:.1f} атм (ρср={rho_mix:.1f} кг/м³)")
                
                # ИТОГО при работе
                p_zab = (p_hydrostatic_tubing + 
                        p_friction_atm - 
                        p_pump_head + 
                        p_hydrostatic_annulus)
                DEBUG.data("Pзаб при работе", f"{p_zab:.1f} = {p_hydrostatic_tubing:.1f} + {p_friction_atm:.1f} - {p_pump_head:.1f} + {p_hydrostatic_annulus:.1f}")
                
            else:
                # При накоплении - только гидростатика
                p_zab = p_hydrostatic_tubing + p_hydrostatic_annulus
                DEBUG.data("Pзаб при накоплении", f"{p_zab:.1f} = {p_hydrostatic_tubing:.1f} + {p_hydrostatic_annulus:.1f}")
            
            result = max(0.0, p_zab)
            DEBUG.log(f"Расчетное Pзаб = {result:.1f} атм")
            DEBUG.exit()
            
            return result
            
        except Exception as e:
            DEBUG.log(f"Ошибка расчета Pзаб: {str(e)}", "ERROR")
            # Упрощенный расчет при ошибке
            water_cut = self.safe_float(self.well.get('water_cut', 0.0))
            oil_density = self.safe_float(self.well.get('oil_density', 0.84))
            oil_density_kg = oil_density * 1000.0
            
            wc = water_cut / 100.0
            rho_mix = wc * self.WATER_DENSITY + (1.0 - wc) * oil_density_kg
            
            liquid_column = pump_depth - h_din
            p_hydrostatic = rho_mix * self.G * liquid_column / self.ATM_TO_PA
            
            DEBUG.log(f"Упрощенный расчет: Pзаб={p_hydrostatic:.1f} атм")
            DEBUG.exit()
            
            return p_hydrostatic
    
    def simulate_pwf_during_work(self, t_work_minutes: float, t_pause_minutes: float, 
                               freq_hz: float = 50.0) -> Dict:
        """
        Моделирование изменения Pзаб во время работы насоса
        """
        DEBUG.enter()
        DEBUG.section("Моделирование цикла работы")
        DEBUG.data("Входные", f"tраб={t_work_minutes:.0f} мин, tнак={t_pause_minutes:.0f} мин, f={freq_hz:.1f} Гц")
        
        # Данные скважины
        q_tech = self.safe_float(self.well.get('flow_rate', 0.0))  # м³/сут
        pump_depth = self.safe_float(self.well.get('pump_depth', 0.0))
        h_din_start = self.safe_float(self.well.get('h_din', pump_depth))  # если нет данных, на забое
        pump_head = self.safe_float(self.well.get('pump_head', 1000.0))
        
        DEBUG.data("Параметры скважины", f"Q={q_tech:.1f} м³/сут, Lнас={pump_depth:.1f} м, Hдин={h_din_start:.1f} м, Напор={pump_head:.0f} м")
        
        # Расчет мгновенного дебита
        work_hours_current = (t_work_minutes / (t_work_minutes + t_pause_minutes)) * 24.0
        q_instant_current = q_tech * (24.0 / work_hours_current) if work_hours_current > 0 else q_tech
        q_instant_with_freq = q_instant_current * (freq_hz / 50.0)
        
        DEBUG.data("Мгновенный дебит", f"{q_instant_with_freq:.1f} м³/сут (рабочих часов={work_hours_current:.1f})")
        DEBUG.data("Площадь затруба", f"{self.area_annulus:.6f} м²")
        
        # Создаем точки времени для моделирования
        time_points = np.linspace(0, t_work_minutes, 100)
        pwf_points = []
        h_din_points = []
        
        DEBUG.log("Начинаем моделирование...")
        
        for idx, t in enumerate(time_points):
            # Прогресс в работе
            progress = t / t_work_minutes
            
            # Объем откачки к моменту времени t
            volume_pumped = q_instant_with_freq * (t / 60.0) / 24.0  # м³
            DEBUG.data(f"Точка {idx}", f"t={t:.1f} мин, прогресс={progress:.2f}, объем={volume_pumped:.3f} м³")
            
            # Падение динамического уровня
            delta_h_din = volume_pumped / self.area_annulus if self.area_annulus > 0 else 0.0
            h_din_current = h_din_start + delta_h_din
            DEBUG.data("Hдин текущий", f"{h_din_current:.1f} м (Δ={delta_h_din:.1f} м)")
            
            # Расчет Pзаб при работе (используем рассчитанное p_zab)
            p_zab = self.calculate_pwf_comprehensive_corrected(
                h_din_current, q_instant_with_freq, pump_head, pump_depth, is_pump_working=True
            )
            
            pwf_points.append(p_zab)
            h_din_points.append(h_din_current)
            
            if idx % 20 == 0:
                DEBUG.data(f"Результат t={t:.1f} мин", f"Pзаб={p_zab:.1f} атм, Hдин={h_din_current:.1f} м")
        
        # Pзаб в начале работы (насос выключен)
        p_zab_start = self.calculate_pwf_comprehensive_corrected(
            h_din_start, 0.0, pump_head, pump_depth, is_pump_working=False
        )
        
        DEBUG.data("Pзаб начало", f"{p_zab_start:.1f} атм (насос выключен)")
        DEBUG.data("Pзаб конец", f"{pwf_points[-1] if pwf_points else p_zab_start:.1f} атм")
        DEBUG.data("Hдин конец", f"{h_din_points[-1] if h_din_points else h_din_start:.1f} м")
        
        result = {
            'time_points': time_points.tolist(),
            'pwf_points': pwf_points,
            'h_din_points': h_din_points,
            'p_zab_end': pwf_points[-1] if pwf_points else p_zab_start,  # РАСЧИТАННОЕ!
            'p_zab_start': p_zab_start,
            'q_instant': q_instant_with_freq,
            'work_hours': work_hours_current,
            'h_din_end': h_din_points[-1] if h_din_points else h_din_start
        }
        
        DEBUG.log("Моделирование завершено")
        DEBUG.exit()
        
        return result
    
    def analyze_gas_problem_comprehensive(self, t_work_minutes: float, t_pause_minutes: float,
                                        freq_hz: float = 50.0) -> Dict:
        """
        Комплексный анализ газовой проблемы с использованием РАСЧИТАННОГО Pзаб
        """
        DEBUG.enter()
        DEBUG.section("АНАЛИЗ ГАЗОВОЙ ПРОБЛЕМЫ")
        
        p_nas = self.safe_float(self.well.get('p_nas', 0.0))
        DEBUG.data("Pнас", f"{p_nas:.1f} атм")
        
        if p_nas <= 0:
            DEBUG.log("Нет Pнас → анализ невозможен")
            DEBUG.exit()
            return {
                'has_gas_problem': False,
                'severity': 'none',
                'p_zab_end': 0.0,
                'p_nas': p_nas,
                'p_zab_relative': 1.0,
                'time_to_critical': None,
                'recommendation': 'Нет данных по Pнас'
            }
        
        # Моделирование работы - получаем РАСЧИТАННОЕ p_zab
        simulation = self.simulate_pwf_during_work(t_work_minutes, t_pause_minutes, freq_hz)
        p_zab_end = simulation['p_zab_end']  # ← РАСЧИТАННОЕ!
        
        DEBUG.data("Pзаб_конец (расчетное)", f"{p_zab_end:.1f} атм")
        DEBUG.data("Pзаб/Pнас", f"{p_zab_end:.1f}/{p_nas:.1f} = {p_zab_end/p_nas:.3f}")
        
        # Анализ по РАСЧИТАННОМУ p_zab
        result = {
            'has_gas_problem': False,
            'severity': 'none',
            'p_zab_end': p_zab_end,  # РАСЧИТАННОЕ!
            'p_nas': p_nas,
            'p_zab_relative': p_zab_end / p_nas if p_nas > 0 else 1.0,
            'time_to_critical': None,
            'recommendation': ''
        }
        
        # Проверяем газовую проблему по РАСЧИТАННОМУ p_zab
        if p_zab_end < p_nas:
            result['has_gas_problem'] = True
            DEBUG.log("ГАЗОВАЯ ПРОБЛЕМА ОБНАРУЖЕНА!")
            
            # Определяем время до критического уровня (0.75*Pнас)
            time_points = simulation['time_points']
            pwf_points = simulation['pwf_points']
            p_critical = 0.75 * p_nas
            
            DEBUG.data("Критический уровень", f"0.75*Pнас = {p_critical:.1f} атм")
            
            # Находим пересечение с критическим уровнем
            critical_time = None
            for i in range(1, len(pwf_points)):
                if pwf_points[i] <= p_critical and pwf_points[i-1] > p_critical:
                    # Линейная интерполяция
                    t1, t2 = time_points[i-1], time_points[i]
                    p1, p2 = pwf_points[i-1], pwf_points[i]
                    critical_time = t1 + (t2 - t1) * (p_critical - p1) / (p2 - p1)
                    result['time_to_critical'] = critical_time
                    DEBUG.data("Время до критического", f"{critical_time:.1f} мин")
                    break
            
            # Определяем степень серьезности по РАСЧИТАННОМУ p_zab
            if p_zab_end < 0.75 * p_nas:
                result['severity'] = 'critical'
                result['recommendation'] = 'Срочно сократить время работы!'
                DEBUG.log("КРИТИЧЕСКАЯ степень: Pзаб < 0.75*Pнас")
            elif p_zab_end < 0.85 * p_nas:
                result['severity'] = 'warning'
                result['recommendation'] = 'Рекомендуется сократить время работы'
                DEBUG.log("СРЕДНЯЯ степень: Pзаб < 0.85*Pнас")
            else:
                result['severity'] = 'mild'
                result['recommendation'] = 'Возможна оптимизация режима'
                DEBUG.log("ЛЕГКАЯ степень: Pзаб < Pнас но > 0.85*Pнас")
        else:
            DEBUG.log("Газовой проблемы НЕТ: Pзаб ≥ Pнас")
            result['recommendation'] = 'Режим безопасен по газу'
        
        DEBUG.data("Рекомендация", result['recommendation'])
        DEBUG.exit()
        
        return result

# ============================================================
# ИСПРАВЛЕННЫЙ КЛАСС ЭКОНОМИЧЕСКОГО РАСЧЕТА
# ============================================================

class EconomicCalculatorCorrected:
    """Исправленный экономический расчет с правильными формулами"""
    
    def __init__(self, oil_price_rub_per_ton: float = 50000.0, energy_price_rub_per_kwh: float = 6.0):
        self.oil_price = oil_price_rub_per_ton
        self.energy_price = energy_price_rub_per_kwh
        
        DEBUG.log(f"Инициализация экономики: нефть={oil_price_rub_per_ton:.0f} ₽/т, энергия={energy_price_rub_per_kwh:.2f} ₽/кВт·ч")
        
        # Каталог насосов (из вашего VBA)
        self.PUMP_CATALOG = {
            '25': self.create_stage_data([0, 6.4, 41, 0], [10, 6.3, 42, 17.4], [18, 6, 42.9, 28.6]),
            '40': self.create_stage_data([0, 6.5, 55, 0], [15, 6.3, 55, 19.5], [30, 5.65, 54.5, 35.3]),
            '60': self.create_stage_data([0, 6.2, 53, 0], [15, 6.2, 55, 19.2], [30, 6.1, 58, 35.8],
                                         [40, 6, 61.5, 44.3], [50, 5.9, 66, 50.8], [60, 5.6, 69.2, 55.1]),
            '80': self.create_stage_data([0, 6.3, 64, 0], [30, 6.15, 69, 30.4], [60, 5.8, 79, 50]),
            '125': self.create_stage_data([0, 6, 76, 0], [30, 5.9, 86, 23], [60, 5.8, 96, 41]),
            '160': self.create_stage_data([0, 6.4, 41, 0], [50, 6.3, 42, 17.4], [100, 6.0, 42.9, 28.6]),
            '200': self.create_stage_data([0, 7.4, 110, 0], [50, 7.3, 150, 27.6], [100, 7.1, 178, 45.3])
        }

    def calculate_utilization_factor(self, q_tech: float, work_min: float, pause_min: float,
                                   q_pump: float, freq_hz: Optional[float] = None) -> float:
        """Расчет коэффициента загрузки насоса (алиас для calculate_k_pod_corrected)"""
        return self.calculate_k_pod_corrected(q_tech, work_min, pause_min, q_pump, freq_hz)

    def get_pump_efficiency(self, pump_type: str) -> float:
        """
        Возвращает КПД насоса по типу (из VBA)
        """
        # Нормализуем тип
        pump_type = pump_type.upper().strip()
        
        # Таблица КПД из VBA
        efficiency_map = {
            # Габарит 2А
            'ЭЦН2А-30': 0.36, 'ЭЦН2А-35': 0.36, 'ЭЦН2А-45': 0.36,
            'ЭЦН2А-50': 0.36, 'ЭЦН2А-60': 0.36, 'ЭЦН2А-80': 0.41,
            'ЭЦН2А-100': 0.41, 'ЭЦН2А-125': 0.37,
            
            # Габарит 5
            'ЭЦН5-25': 0.26, 'ЭЦН5-30': 0.26, 'ЭЦН5-40': 0.29, 
            'ЭЦН5-40Э': 0.34, 'ЭЦН5-50': 0.35, 'ЭЦН5-60': 0.40, 
            'ЭЦН5-80': 0.41, 'ЭЦН5-80Э': 0.43, 'ЭЦН5-125': 0.42, 
            'ЭЦН5-125Э': 0.46, 'ЭЦН5-200': 0.46,
            
            # Габарит 5А
            'ЭЦН5А-60': 0.40, 'ЭЦН5А-160': 0.48, 'ЭЦН5А-200': 0.47, 
            'ЭЦН5А-250': 0.46, 'ЭЦН5А-400': 0.46, 'ЭЦН5А-500': 0.46, 
            'ЭЦН5А-700': 0.50,
            
            # Габарит 6
            'ЭЦН6-1000': 0.45,
            
            # Габарит 7А
            'ЭЦН7А-800': 0.52, 'ЭЦН7А-1000': 0.52, 'ЭЦН7А-1550': 0.52, 
            'ЭЦН7А-2000': 0.52,
        }
        
        # Ищем точное совпадение
        for key, value in efficiency_map.items():
            if key in pump_type:
                DEBUG.data(f"КПД для {pump_type}", f"{value:.2f}")
                return value
        
        # Если не нашли, пытаемся извлечь базовый тип
        import re
        # Ищем паттерн типа ЭЦН5-60 или ЭЦН5А-160
        match = re.search(r'(ЭЦН[0-9А]*-[0-9]+)', pump_type)
        if match:
            base_type = match.group(1)
            if base_type in efficiency_map:
                DEBUG.data(f"КПД для {base_type}", f"{efficiency_map[base_type]:.2f}")
                return efficiency_map[base_type]
        
        # По умолчанию для ЭЦН5
        if 'ЭЦН5' in pump_type:
            DEBUG.data("КПД по умолчанию для ЭЦН5", "0.40")
            return 0.40
        # По умолчанию для ЭЦН5А
        elif 'ЭЦН5А' in pump_type:
            DEBUG.data("КПД по умолчанию для ЭЦН5А", "0.45")
            return 0.45
        # По умолчанию для остальных
        else:
            DEBUG.data("КПД по умолчанию", "0.40")
            return 0.40
        
    @staticmethod
    def safe_float(value, default=0.0):
        """Безопасное преобразование в float"""
        if value is None:
            return default
        try:
            return float(value)
        except (ValueError, TypeError):
            return default
            
    def create_stage_data(self, *rows: List) -> List:
        """Создание массива данных ступени"""
        return list(rows)
    
    def extract_pump_type(self, pump_mark: str) -> str:
        """Извлечение типа насоса из марки"""
        if not pump_mark:
            DEBUG.data("Тип насоса", "не указан → 60 по умолчанию")
            return '60'  # по умолчанию
        
        import re
        pump_mark = str(pump_mark).upper()
        DEBUG.data("Марка насоса", pump_mark)
        
        # Ищем числа в марке
        numbers = re.findall(r'\d+', pump_mark)
        if numbers:
            # Первое или второе число может быть типом
            for num in numbers:
                if num in ['25', '40', '60', '80', '125', '160', '200']:
                    DEBUG.data("Найден тип", num)
                    return num
        
        DEBUG.data("Тип не определен", "60 по умолчанию")
        return '60'  # по умолчанию
    
    def safe_frequency(self, freq_value: Optional[Union[str, float, int]]) -> float:
        """Безопасное преобразование частоты (по умолчанию 50 Гц)"""
        if freq_value is None:
            return 50.0
        
        try:
            freq = float(freq_value)
            if freq <= 0 or freq > 100:
                DEBUG.data("Частота невалидна", f"{freq} → 50 Гц")
                return 50.0
            return freq
        except (ValueError, TypeError):
            DEBUG.data("Ошибка частоты", "используется 50 Гц")
            return 50.0

    def calculate_wear_factor(self, well_data: Dict, q_nom: float) -> float:
        """
        Расчет коэффициента износа по двухфакторной модели из VBA
        
        wear_time = 1 + (наработка/1000) * 0.1
        wear_mode = Qном / (Qфакт * (50/f))
        k_iznos = wear_time * wear_mode
        """
        DEBUG.enter()
        DEBUG.log("Расчет коэффициента износа")
        
        # Наработка в сутках
        narabotka = self.safe_float(well_data.get('mttf', 0))
        if narabotka <= 0:
            narabotka = 0
        DEBUG.data("Наработка", f"{narabotka:.0f} сут")
        
        # 1. Износ от времени работы
        wear_time = 1 + (narabotka / 1000.0) * 0.1
        if wear_time > 2:
            wear_time = 2
        DEBUG.data("Износ от времени", f"{wear_time:.3f}")
        
        # 2. Износ от режима работы
        q_tech = self.safe_float(well_data.get('flow_rate', 0))
        freq = self.safe_frequency(well_data.get('rotations_hz'))
        
        # Qфакт, приведенный к 50 Гц
        if freq > 0:
            q_fact_50 = q_tech * (50.0 / freq)
        else:
            q_fact_50 = q_tech
        DEBUG.data("Qфакт при 50Гц", f"{q_fact_50:.1f} м³/сут")
        
        if q_fact_50 > 0 and q_nom > 0:
            wear_mode = q_nom / q_fact_50
        else:
            wear_mode = 1.0
        
        # Ограничения
        if wear_mode < 0.5:
            wear_mode = 0.5
            DEBUG.data("wear_mode ограничен снизу", "0.5")
        if wear_mode > 3:
            wear_mode = 3.0
            DEBUG.data("wear_mode ограничен сверху", "3.0")
        
        DEBUG.data("Износ от режима", f"{wear_mode:.3f}")
        
        # 3. Итоговый коэффициент износа
        k_iznos = wear_time * wear_mode
        if k_iznos > 5:
            k_iznos = 5.0
            DEBUG.data("k_iznos ограничен", "5.0")
        
        DEBUG.data("Итоговый K износа", f"{k_iznos:.3f}")
        DEBUG.exit()
        
        return k_iznos

    def recalculate_flow_rate_for_new_regime(self, q_current: float, 
                                            old_work_min: float, old_pause_min: float, old_freq: float,
                                            new_work_min: float, new_pause_min: float, new_freq: float) -> float:
        """
        ПРАВИЛЬНЫЙ пересчет дебита при изменении параметров
        """
        DEBUG.enter()
        DEBUG.log(f"Пересчет дебита: Qтек={q_current:.1f}, режим {old_work_min:.0f}/{old_pause_min:.0f}→{new_work_min:.0f}/{new_pause_min:.0f}")
        DEBUG.data("Частота", f"{old_freq:.1f}→{new_freq:.1f} Гц")
        
        # Старые рабочие часы
        old_cycle = old_work_min + old_pause_min
        if old_cycle <= 0:
            old_hours = 0
        else:
            old_hours = (old_work_min / old_cycle) * 24.0
        
        # Новые рабочие часы
        new_cycle = new_work_min + new_pause_min
        if new_cycle <= 0:
            new_hours = 0
        else:
            new_hours = (new_work_min / new_cycle) * 24.0
        
        DEBUG.data("Рабочие часы", f"старые={old_hours:.1f} ч, новые={new_hours:.1f} ч")
        
        # Расчет нового дебита
        if old_hours > 0:
            q_new = q_current * (new_hours / old_hours) * (new_freq / old_freq)
            DEBUG.data("Коэффициент", f"(часы={new_hours/old_hours:.3f})×(частота={new_freq/old_freq:.3f})={new_hours/old_hours*new_freq/old_freq:.3f}")
        else:
            q_new = q_current * (new_freq / old_freq)
            DEBUG.data("Коэффициент", f"(частота={new_freq/old_freq:.3f})")
        
        DEBUG.data("Новый дебит", f"{q_new:.1f} м³/сут")
        DEBUG.exit()
        
        return q_new
    
    def calculate_k_pod_corrected(self, q_tech: float, work_min: float, pause_min: float,
                                q_nom: float, freq: float) -> float:
        """
        ПРАВИЛЬНЫЙ расчет коэффициента подачи
        K_под = Qфакт / (duty_cycle * Qном * (f/50))
        где Qфакт - дебит жидкости на поверхности
        """
        DEBUG.enter()
        DEBUG.log(f"Расчет K_под: Q={q_tech:.1f}, режим={work_min:.0f}/{pause_min:.0f}, Qном={q_nom:.0f}, f={freq:.1f} Гц")
        
        if q_nom <= 0:
            DEBUG.log("Ошибка: Qном ≤ 0")
            DEBUG.exit()
            return 0.0
        
        cycle = work_min + pause_min
        if cycle <= 0:
            DEBUG.log("Ошибка: цикл ≤ 0")
            DEBUG.exit()
            return 0.0
        
        duty_cycle = work_min / cycle
        freq_factor = freq / 50.0
        
        denominator = duty_cycle * q_nom * freq_factor
        
        DEBUG.data("Параметры", f"duty_cycle={duty_cycle:.3f}, freq_factor={freq_factor:.3f}")
        DEBUG.data("Знаменатель", f"{denominator:.1f} = {duty_cycle:.3f}×{q_nom:.0f}×{freq_factor:.3f}")
        
        if denominator <= 0:
            DEBUG.log("Ошибка: знаменатель ≤ 0")
            DEBUG.exit()
            return 0.0
        
        k_pod = q_tech / denominator
        DEBUG.data("K_под", f"{k_pod:.3f} = {q_tech:.1f}/{denominator:.1f}")
        
        DEBUG.exit()
        return k_pod
    
    def calculate_instant_flow_rate(self, q_tech: float, work_min: float, pause_min: float,
                                  freq_hz: Optional[float] = None) -> Tuple[float, float]:
        """
        Расчет мгновенного дебита и рабочих часов в сутках
        """
        DEBUG.enter()
        DEBUG.log(f"Расчет мгновенного дебита: Q={q_tech:.1f}, режим={work_min:.0f}/{pause_min:.0f}")
        
        cycle_min = work_min + pause_min
        if cycle_min <= 0:
            DEBUG.log("Ошибка: цикл ≤ 0")
            DEBUG.exit()
            return 0.0, 0.0
        
        work_hours_per_day = (work_min / cycle_min) * 24.0
        DEBUG.data("Рабочие часы/сут", f"{work_hours_per_day:.1f}")
        
        if work_hours_per_day > 0:
            q_instant_without_freq = q_tech * (24.0 / work_hours_per_day)
            DEBUG.data("Q_мгновенный без частоты", f"{q_instant_without_freq:.1f} м³/сут")
        else:
            q_instant_without_freq = q_tech
            DEBUG.data("Q_мгновенный без частоты", f"{q_instant_without_freq:.1f} м³/сут (work_hours=0)")
        
        # Учет частоты
        freq_hz_safe = self.safe_frequency(freq_hz)
        q_instant = q_instant_without_freq * (freq_hz_safe / 50.0)
        
        DEBUG.data("Частота", f"{freq_hz_safe:.1f} Гц")
        DEBUG.data("Коэффициент частоты", f"{freq_hz_safe/50.0:.3f}")
        DEBUG.data("Q_мгновенный итог", f"{q_instant:.1f} м³/сут")
        
        DEBUG.exit()
        
        return q_instant, work_hours_per_day
    
    def interpolate_pump_parameters(self, pump_type: str, q_m3_per_day_50hz: float) -> Tuple[Optional[float], Optional[float], Optional[float]]:
        """Интерполяция параметров ступени по каталогу"""
        DEBUG.enter()
        DEBUG.log(f"Интерполяция параметров насоса: тип={pump_type}, Q={q_m3_per_day_50hz:.1f} м³/сут")
        
        if pump_type not in self.PUMP_CATALOG:
            DEBUG.log(f"Ошибка: тип насоса '{pump_type}' не найден в каталоге")
            DEBUG.exit()
            return None, None, None  # H_stage, N_stage, efficiency
        
        data = self.PUMP_CATALOG[pump_type]
        
        # Если дебит меньше минимального
        if q_m3_per_day_50hz <= data[0][0]:
            DEBUG.data("Дебит меньше минимального", f"Q={q_m3_per_day_50hz:.1f} ≤ {data[0][0]:.1f}")
            DEBUG.data("Параметры", f"H={data[0][1]:.1f} м, N={data[0][2]:.1f} Вт, КПД={data[0][3]:.1f}%")
            DEBUG.exit()
            return data[0][1], data[0][2], data[0][3]
        
        # Если дебит больше максимального
        if q_m3_per_day_50hz >= data[-1][0]:
            DEBUG.data("Дебит больше максимального", f"Q={q_m3_per_day_50hz:.1f} ≥ {data[-1][0]:.1f}")
            DEBUG.data("Параметры", f"H={data[-1][1]:.1f} м, N={data[-1][2]:.1f} Вт, КПД={data[-1][3]:.1f}%")
            DEBUG.exit()
            return data[-1][1], data[-1][2], data[-1][3]
        
        # Поиск интервала для интерполяции
        DEBUG.data("Ищем интервал", "для интерполяции...")
        for i in range(1, len(data)):
            if q_m3_per_day_50hz <= data[i][0]:
                q1, q2 = data[i-1][0], data[i][0]
                h1, h2 = data[i-1][1], data[i][1]
                n1, n2 = data[i-1][2], data[i][2]
                eff1, eff2 = data[i-1][3], data[i][3]
                
                DEBUG.data("Найден интервал", f"Q={q1:.1f}-{q2:.1f} м³/сут")
                DEBUG.data("Значения в точках", f"H={h1:.1f}-{h2:.1f} м, N={n1:.1f}-{n2:.1f} Вт, КПД={eff1:.1f}-{eff2:.1f}%")
                
                # Линейная интерполяция
                t = (q_m3_per_day_50hz - q1) / (q2 - q1)
                h_stage = h1 + (h2 - h1) * t
                n_stage = n1 + (n2 - n1) * t
                efficiency = eff1 + (eff2 - eff1) * t
                
                DEBUG.data("Коэффициент t", f"{t:.3f}")
                DEBUG.data("Итоговые параметры", f"H={h_stage:.1f} м, N={n_stage:.1f} Вт, КПД={efficiency:.1f}%")
                
                DEBUG.exit()
                return h_stage, n_stage, efficiency
        
        DEBUG.log("Интервал не найден, возвращаем последние значения")
        DEBUG.exit()
        return data[-1][1], data[-1][2], data[-1][3]
    
    def calculate_pump_power(self, pump_type: str, q_instant: float, pump_head: float, 
                             freq_hz: float, well_data: Dict = None,
                             k_degr: float = 1.0, k_iznos: float = 1.0) -> float:
        """
        Расчет мощности насоса с учетом газа и износа (как в VBA)
        
        Параметры:
            pump_type: тип насоса (ключ для каталога)
            q_instant: мгновенный дебит, м³/сут
            pump_head: паспортный напор, м
            freq_hz: частота, Гц
            well_data: данные скважины (для пересчета в ГЖС)
            k_degr: коэффициент деградации от газа (0.2-1.0)
            k_iznos: коэффициент износа (1.0-5.0)
        """
        DEBUG.enter()
        DEBUG.section("РАСЧЕТ МОЩНОСТИ НАСОСА (С УЧЕТОМ ГАЗА И ИЗНОСА)")
        DEBUG.data("Входные", f"тип={pump_type}, Q={q_instant:.1f} м³/сут, H={pump_head:.0f} м, f={freq_hz:.1f} Гц")
        DEBUG.data("Коэф. деградации", f"{k_degr:.3f}")
        DEBUG.data("Коэф. износа", f"{k_iznos:.3f}")
        
        # ===== ШАГ 1: Пересчет дебита в ГЖС (если есть данные) =====
        q_for_power = q_instant
        if well_data is not None:
            water_cut = well_data.get('water_cut', 0)
            q_for_power = calculate_downhole_mixture_rate(q_instant, water_cut, 1.1)
            DEBUG.data("Пересчет в ГЖС", f"{q_instant:.1f} → {q_for_power:.1f} м³/сут")
        
        # ===== ШАГ 2: Коррекция подачи с учетом газа (как в VBA) =====
        # В VBA: Qcorr = Qфакт * (f/50) * KDegr
        q_corrected = q_for_power * (freq_hz / 50.0) * k_degr
        DEBUG.data("Q скорректированная", f"{q_corrected:.1f} м³/сут")
        
        # Дебит для каталога (при 50 Гц)
        q_50hz = q_corrected * (50.0 / freq_hz)
        DEBUG.data("Q для каталога (50 Гц)", f"{q_50hz:.1f} м³/сут")
        
        # ===== ШАГ 3: Интерполяция параметров ступени =====
        h_stage_50hz, n_stage_50hz, efficiency = self.interpolate_pump_parameters(pump_type, q_50hz)
        
        if h_stage_50hz is None or n_stage_50hz is None:
            DEBUG.log("Параметры ступени не найдены, используем упрощенный расчет")
            # Упрощенный расчет (запасной вариант)
            density = 850.0
            efficiency_simple = 0.6
            power_watts = q_for_power * pump_head * 9.81 * density / (3600.0 * efficiency_simple)
            power_kw = power_watts / 1000.0
            power_kw = power_kw * k_iznos  # применяем износ
            DEBUG.data("Упрощенный расчет", f"ρ={density:.0f}, η={efficiency_simple:.2f}")
            DEBUG.data("Мощность с износом", f"{power_kw:.1f} кВт")
            DEBUG.exit()
            return power_kw
        
        DEBUG.data("Параметры ступени (50 Гц)", 
                   f"H={h_stage_50hz:.1f} м, N={n_stage_50hz:.1f} Вт, η={efficiency:.1f}%")
        
        # ===== ШАГ 4: Коррекция напора ступени по частоте и газу (как в VBA) =====
        # В VBA: data.HStage = data.HStage * (data.Freq / 50) ^ 2 * data.KDegr
        h_stage_work = h_stage_50hz * (freq_hz / 50.0) ** 2 * k_degr
        DEBUG.data("H ступени раб.", f"{h_stage_work:.1f} м")
        
        # ===== ШАГ 5: Число ступеней =====
        if h_stage_work > 0:
            num_stages = math.ceil(pump_head / h_stage_work)
        else:
            num_stages = 0
        DEBUG.data("Число ступеней", f"{num_stages:.0f}")
        
        # ===== ШАГ 6: Мощность ступени при рабочей частоте =====
        n_stage_work = n_stage_50hz * (freq_hz / 50.0) ** 3
        DEBUG.data("N ступени раб.", f"{n_stage_work:.1f} Вт")
        
        # ===== ШАГ 7: Базовая мощность =====
        power_base_kw = (num_stages * n_stage_work) / 1000.0
        DEBUG.data("Мощность базовая", f"{power_base_kw:.1f} кВт")
        
        # ===== ШАГ 8: Мощность с учетом износа (САМОЕ ВАЖНОЕ!) =====
        power_kw = power_base_kw * k_iznos
        DEBUG.data("Мощность с износом", f"{power_kw:.1f} кВт")
        
        DEBUG.exit()
        return power_kw
    
    def calculate_daily_flow_rate(self, q_tech_current: float, work_min_current: float, pause_min_current: float,
                                freq_current: float, work_min_new: float, pause_min_new: float,
                                freq_new: float) -> float:
        """
        Расчет нового суточного дебита жидкости при изменении параметров
        """
        DEBUG.enter()
        DEBUG.log("Расчет нового суточного дебита")
        DEBUG.data("Текущий", f"Q={q_tech_current:.1f}, режим={work_min_current:.0f}/{pause_min_current:.0f}, f={freq_current:.1f}")
        DEBUG.data("Новый", f"режим={work_min_new:.0f}/{pause_min_new:.0f}, f={freq_new:.1f}")
        
        # Текущие рабочие часы
        cycle_current = work_min_current + pause_min_current
        if cycle_current <= 0:
            work_hours_current = 0.0
        else:
            work_hours_current = (work_min_current / cycle_current) * 24.0
        
        # Новые рабочие часы
        cycle_new = work_min_new + pause_min_new
        if cycle_new <= 0:
            work_hours_new = 0.0
        else:
            work_hours_new = (work_min_new / cycle_new) * 24.0
        
        DEBUG.data("Рабочие часы", f"старые={work_hours_current:.1f}, новые={work_hours_new:.1f}")
        
        # Расчет нового дебита
        if work_hours_current > 0:
            q_tech_new = q_tech_current * (freq_new / freq_current) * (work_hours_new / work_hours_current)
            DEBUG.data("Коэффициенты", f"частота={freq_new/freq_current:.3f}, часы={work_hours_new/work_hours_current:.3f}")
        else:
            q_tech_new = q_tech_current * (freq_new / freq_current)
            DEBUG.data("Коэффициент", f"частота={freq_new/freq_current:.3f} (work_hours_current=0)")
        
        DEBUG.data("Новый дебит", f"{q_tech_new:.1f} м³/сут")
        DEBUG.exit()
        
        return q_tech_new
    
    def calculate_economic_effect_comprehensive(self, well_data: Dict, 
                                              old_schedule: List[float], new_schedule: List[float],
                                              old_freq: Optional[float] = None, 
                                              new_freq: Optional[float] = None,
                                              has_gas_problem_old: bool = False) -> Dict:
        """
        Полный экономический расчет с формулами из VBA
        """
        DEBUG.enter()
        DEBUG.section("ПОЛНЫЙ ЭКОНОМИЧЕСКИЙ РАСЧЕТ (VBA)")
        
        # ===== БАЗОВЫЕ ДАННЫЕ =====
        q_tech_current = self.safe_float(well_data.get('flow_rate', 0.0))  # м³/сут
        wct_percent = self.safe_float(well_data.get('water_cut', 0.0))
        oil_density_ton_per_m3 = self.safe_float(well_data.get('oil_density', 0.84))  # т/м³
        pump_mark = well_data.get('pump_mark', '')
        pump_head = self.safe_float(well_data.get('pump_head', 1000.0))
        q_pump = self.safe_float(well_data.get('pump_flow', 60.0))
        
        # Данные для газа и износа
        gas_factor = self.safe_float(well_data.get('gas_factor', 0))
        p_nas = self.safe_float(well_data.get('p_nas', 0))
        p_intake = self.safe_float(well_data.get('p_intake', 0))
        narabotka = self.safe_float(well_data.get('mttf', 0))
        
        DEBUG.data("Базовые данные", f"Q={q_tech_current:.1f} м³/сут, Обв={wct_percent:.1f}%, ρ={oil_density_ton_per_m3:.3f} т/м³")
        DEBUG.data("Параметры насоса", f"марка={pump_mark}, H={pump_head:.0f} м, Qном={q_pump:.0f} м³/сут")
        DEBUG.data("Газ", f"ГФ={gas_factor:.0f} м³/т, Pнас={p_nas:.1f}, Pприем={p_intake:.1f}")
        DEBUG.data("Наработка", f"{narabotka:.0f} сут")
        
        # Частоты
        old_freq_safe = self.safe_frequency(old_freq)
        new_freq_safe = self.safe_frequency(new_freq) if new_freq is not None else old_freq_safe
        
        DEBUG.data("Частоты", f"старая={old_freq_safe:.1f} Гц, новая={new_freq_safe:.1f} Гц")
        
        # Доля воды
        wct = wct_percent / 100.0
        
        # Режимы
        old_work, old_pause = old_schedule
        new_work, new_pause = new_schedule
        
        DEBUG.data("Режимы", f"старый={old_work:.0f}/{old_pause:.0f}, новый={new_work:.0f}/{new_pause:.0f}")
        
        # ===== ШАГ 1: Дебит нефти =====
        q_oil_t_per_day = q_tech_current * (1.0 - wct) * oil_density_ton_per_m3  # т/сут (приблизительно)
        q_oil_new_t_per_day = q_tech_new * (1.0 - wct) if 'q_tech_new' in locals() else 0
        
        # Данные для газа и износа
        gas_factor = self.safe_float(well_data.get('gas_factor', 0))
        p_nas = self.safe_float(well_data.get('p_nas', 0))
        
        # ===== ИСПРАВЛЕНО: получаем давление на приеме с расчетом =====
        # Сначала пробуем взять из данных
        p_intake_raw = well_data.get('p_intake')
        if p_intake_raw is not None and p_intake_raw > 0:
            p_intake = self.safe_float(p_intake_raw)
            DEBUG.data("Pприем из данных", f"{p_intake:.1f} атм")
        else:
            # Если нет в данных, создаем физику и рассчитываем
            if not hasattr(self, 'physics'):
                self.physics = CorrectedKPRPhysics(well_data)
            p_intake = self.physics.calculate_pressure_at_pump_intake(well_data)
            DEBUG.data("Pприем рассчитано", f"{p_intake:.1f} атм")
            
            # Дополнительная защита от нуля
            if p_intake <= 0:
                DEBUG.log("⚠️ Рассчитанное Pприем <= 0, устанавливаем 1.0 атм")
                p_intake = 1.0
        
        narabotka = self.safe_float(well_data.get('mttf', 0))
        
        # ===== ШАГ 2: Свободный газ (только если есть физика) =====
        # Создаем объект физики, если его нет (уже создан выше, если было нужно)
        if not hasattr(self, 'physics'):
            self.physics = CorrectedKPRPhysics(well_data)
        
        # Расчет свободного газа для старого режима
        gas_data = self.physics.calculate_gas_free(
            p_intake, p_nas, gas_factor, q_oil_t_per_day, 
            oil_density_ton_per_m3, q_pump, old_freq_safe
        )
        k_degr_old = gas_data['k_degr']
        q_gas_free = gas_data['q_gas_free']
        gas_fraction = gas_data['gas_fraction']
        
        DEBUG.data("Свободный газ", f"{q_gas_free:.1f} м³/сут")
        DEBUG.data("Доля газа", f"{gas_fraction*100:.1f}%")
        DEBUG.data("K деградации", f"{k_degr_old:.3f}")
        
        # ===== ШАГ 3: Коэффициент износа =====
        k_iznos = self.calculate_wear_factor(well_data, q_pump)
        DEBUG.data("K износа", f"{k_iznos:.3f}")
        
        # ===== ШАГ 4: Рабочие часы =====
        old_cycle = old_work + old_pause
        new_cycle = new_work + new_pause
        
        old_hours = (old_work / old_cycle) * 24.0 if old_cycle > 0 else 0.0
        new_hours = (new_work / new_cycle) * 24.0 if new_cycle > 0 else 0.0
        
        DEBUG.data("Рабочие часы", f"старые={old_hours:.1f} ч/сут, новые={new_hours:.1f} ч/сут")
        
        # ===== ШАГ 5: Мгновенные дебиты =====
        q_instant_old, _ = self.calculate_instant_flow_rate(q_tech_current, old_work, old_pause, old_freq_safe)
        q_instant_new = q_instant_old * (new_freq_safe / old_freq_safe) * (new_hours / old_hours) if old_hours > 0 else q_instant_old
        
        DEBUG.data("Мгновенные дебиты", f"старый={q_instant_old:.1f} м³/сут, новый={q_instant_new:.1f} м³/сут")
        
        # ===== ШАГ 6: Новые суточные дебиты =====
        q_tech_new = self.calculate_daily_flow_rate(
            q_tech_current, old_work, old_pause, old_freq_safe,
            new_work, new_pause, new_freq_safe
        )
        
        DEBUG.data("Суточные дебиты", f"старый={q_tech_current:.1f} м³/сут, новый={q_tech_new:.1f} м³/сут")
        
        # ===== ШАГ 7: Дебит нефти (новый) =====
        q_oil_new = q_tech_new * (1.0 - wct) * oil_density_ton_per_m3  # т/сут
        
        DEBUG.data("Дебиты нефти", f"старый={q_oil_t_per_day:.2f} т/сут, новый={q_oil_new:.2f} т/сут")
        DEBUG.data("Прирост нефти", f"{q_oil_new - q_oil_t_per_day:+.2f} т/сут")
        
        # ===== ШАГ 8: Мощность насоса (с учетом газа и износа!) =====
        pump_type = self.extract_pump_type(pump_mark)
        
        # Для старого режима - с газом и износом
        power_old = self.calculate_pump_power(
            pump_type, q_instant_old, pump_head, old_freq_safe, 
            well_data, k_degr_old, k_iznos
        )
        
        # Для нового режима - пока не знаем газ, используем тот же коэффициент
        # (в реальности газ может уменьшиться, но для оценки оставим так)
        power_new = self.calculate_pump_power(
            pump_type, q_instant_new, pump_head, new_freq_safe,
            well_data, k_degr_old, k_iznos
        )
        
        DEBUG.data("Мощность насоса (с учетом всего)", f"старая={power_old:.1f} кВт, новая={power_new:.1f} кВт")
        
        # ===== ШАГ 9: Энергопотребление =====
        energy_old = power_old * old_hours  # кВт·ч/сут
        energy_new = power_new * new_hours  # кВт·ч/сут
        
        DEBUG.data("Энергопотребление", f"старое={energy_old:.0f} кВт·ч/сут, новое={energy_new:.0f} кВт·ч/сут")
        
        # ===== ШАГ 10: Экономические расчеты =====
        oil_revenue_old = q_oil_t_per_day * self.oil_price  # ₽/сут
        oil_revenue_new = q_oil_new * self.oil_price  # ₽/сут
        
        energy_cost_old = energy_old * self.energy_price  # ₽/сут
        energy_cost_new = energy_new * self.energy_price  # ₽/сут
        
        delta_revenue = oil_revenue_new - oil_revenue_old
        delta_cost = energy_cost_new - energy_cost_old
        total_effect = delta_revenue - delta_cost
        
        DEBUG.data("Доход от нефти", f"старый={oil_revenue_old:.0f} ₽/сут, новый={oil_revenue_new:.0f} ₽/сут")
        DEBUG.data("Затраты на энергию", f"старые={energy_cost_old:.0f} ₽/сут, новые={energy_cost_new:.0f} ₽/сут")
        DEBUG.data("Эффект", f"{total_effect:.0f} ₽/сут")
        
        # ===== ШАГ 11: Расчет загрузки =====
        k_util_old = self.calculate_utilization_factor(
            q_tech_current, old_work, old_pause, q_pump, old_freq_safe
        )
        
        k_util_new = self.calculate_utilization_factor(
            q_tech_new, new_work, new_pause, q_pump, new_freq_safe
        )
        
        DEBUG.data("K_под", f"старый={k_util_old:.2f}, новый={k_util_new:.2f}")
        
        # ===== ШАГ 12: Расчет УРЭ и диагностика (как в VBA) =====
        pump_efficiency = self.get_pump_efficiency(pump_type)
        base_ure = 0.2725 / (pump_efficiency * 0.71)  # 0.71 - КПД двигателя
        ure_et = base_ure * (pump_head / 100.0)
        
        if q_tech_current > 0:
            ure_fact_old = energy_old / q_tech_current
            ure_fact_new = energy_new / q_tech_new if q_tech_new > 0 else 0
        else:
            ure_fact_old = 0
            ure_fact_new = 0
        
        # Приведенный УРЭ (с учетом реального напора)
        # Для этого нужен required_head - его нужно передавать или рассчитывать
        # Пока используем упрощение
        ure_priv_old = ure_fact_old
        ure_priv_new = ure_fact_new
        
        if ure_et > 0:
            deviation_old = (ure_priv_old / ure_et - 1) * 100
            deviation_new = (ure_priv_new / ure_et - 1) * 100
        else:
            deviation_old = 0
            deviation_new = 0
        
        # Диагноз для старого режима
        if deviation_old < -20:
            diagnosis_old = "НИЖЕ НОРМЫ"
        elif deviation_old <= 20:
            diagnosis_old = "НОРМА"
        elif deviation_old <= 100:
            diagnosis_old = "ПОВЫШЕННЫЙ РАСХОД"
        else:
            diagnosis_old = "КРИТИЧЕСКИЙ ПЕРЕРАСХОД"
        
        DEBUG.data("УРЭ эт", f"{ure_et:.3f}")
        DEBUG.data("УРЭ факт старый", f"{ure_fact_old:.3f}")
        DEBUG.data("Отклонение", f"{deviation_old:.1f}%")
        DEBUG.data("Диагноз", diagnosis_old)
        
        # ===== ШАГ 13: Сбор результата =====
        result = {
            # Режимы
            'old_regime': f"{old_work:.0f}/{old_pause:.0f} мин",
            'new_regime': f"{new_work:.0f}/{new_pause:.0f} мин",
            'work_hours_old': old_hours,
            'work_hours_new': new_hours,
            
            # Дебиты
            'q_tech_current': q_tech_current,
            'q_tech_new': q_tech_new,
            'q_instant_old': q_instant_old,
            'q_instant_new': q_instant_new,
            'q_oil_old': q_oil_t_per_day,
            'q_oil_new': q_oil_new,
            'delta_q_oil': q_oil_new - q_oil_t_per_day,
            
            # Газ
            'q_gas_free': q_gas_free,
            'gas_fraction': gas_fraction,
            'k_degr': k_degr_old,
            
            # Износ
            'k_iznos': k_iznos,
            
            # Энергетика
            'power_old': power_old,
            'power_new': power_new,
            'energy_old': energy_old,
            'energy_new': energy_new,
            'delta_energy': energy_new - energy_old,
            
            # Частоты
            'freq_old': old_freq_safe,
            'freq_new': new_freq_safe,
            
            # Экономика
            'oil_revenue_old': oil_revenue_old,
            'oil_revenue_new': oil_revenue_new,
            'energy_cost_old': energy_cost_old,
            'energy_cost_new': energy_cost_new,
            'delta_revenue': delta_revenue,
            'delta_cost': delta_cost,
            'total_effect_per_day': total_effect,
            'total_effect_per_month': total_effect * 30.0,
            'total_effect_per_year': total_effect * 365.0,
            
            # Показатели
            'k_util_old': k_util_old,
            'k_util_new': k_util_new,
            'specific_energy_old': energy_old / q_oil_t_per_day if q_oil_t_per_day > 0 else 0.0,
            'specific_energy_new': energy_new / q_oil_new if q_oil_new > 0 else 0.0,
            
            # УРЭ и диагностика
            'ure_et': ure_et,
            'ure_fact_old': ure_fact_old,
            'ure_fact_new': ure_fact_new,
            'ure_priv_old': ure_priv_old,
            'ure_priv_new': ure_priv_new,
            'deviation_old': deviation_old,
            'deviation_new': deviation_new,
            'diagnosis_old': diagnosis_old,
            
            # Статус
            'is_profitable': total_effect > 0,
            'has_gas_problem_old': has_gas_problem_old,
            'oil_price': self.oil_price,
            'energy_price': self.energy_price
        }
        
        DEBUG.data("Итоговый эффект", f"{total_effect:.0f} ₽/сут")
        DEBUG.data("Прибыльно", "ДА" if total_effect > 0 else "НЕТ")
        DEBUG.data("Диагноз (старый)", diagnosis_old)
        DEBUG.exit()
        
        return result

# ============================================================
# ИСПРАВЛЕННЫЙ ОПТИМИЗАТОР СЦЕНАРИЯ A И B
# ============================================================

class KPROptimizerCorrected:
    """Исправленный оптимизатор с правильной логикой"""
    
    def __init__(self, physics_calculator: CorrectedKPRPhysics, economic_calculator: EconomicCalculatorCorrected):
        self.physics = physics_calculator
        self.economic = economic_calculator
        DEBUG.log("Инициализация KPROptimizerCorrected")
    
    def can_increase_work_time(self, well_data: Optional[Dict] = None) -> bool:
        """
        Проверка можно ли увеличивать время работы
        
        Правило: если Kпр < 0.8 И (L_pump - H_din) < 300 м → НЕЛЬЗЯ увеличивать
        """
        DEBUG.enter()
        DEBUG.log("Проверка возможности увеличения времени работы")
        
        if well_data is None:
            well_data = self.physics.well
        
        kpr = self.physics.safe_float(well_data.get('prod_coef', 10.0))
        pump_depth = self.physics.safe_float(well_data.get('pump_depth', 0.0))
        h_din = self.physics.safe_float(well_data.get('h_din', 0.0))
        
        DEBUG.data("Kпр", f"{kpr:.2f}")
        DEBUG.data("Разница глубин", f"{pump_depth:.1f} - {h_din:.1f} = {pump_depth - h_din:.1f} м")
        
        # Правило: если Kпр < 0.8 И (L_pump - H_din) < 300 м → НЕЛЬЗЯ увеличивать
        if kpr < 0.8 and (pump_depth - h_din) < 300.0:
            DEBUG.log("❌ НЕЛЬЗЯ увеличивать: Kпр<0.8 и разница<300 м")
            DEBUG.exit()
            return False
        
        DEBUG.log("✅ МОЖНО увеличивать время работы")
        DEBUG.exit()
        return True
    
    def optimize_scenario_a_gas_problem(self, t_work_current: float, t_pause_current: float,
                                      freq_current: float = 50.0) -> Dict:
        """
        ИСПРАВЛЕННАЯ логика Сценария A с расчетом через динамический уровень
        
        1. Определяем цель: Pзаб ≥ 0.75 × Pнас
        2. Рассчитываем ΔP/мин = (Pзаб_start - Pзаб_end) / t_work_current
        3. Находим сколько нужно уменьшить время работы
        4. Рассчитываем новые уровни через ΔH/мин
        5. Рассчитываем новое время накопления через ΔH_per_min_pause
        6. Округляем в целые числа
        """
        DEBUG.enter()
        DEBUG.section("СЦЕНАРИЙ A: Оптимизация при газовой проблеме")
        
        well_data = self.physics.well
        p_nas = self.physics.safe_float(well_data.get('p_nas', 0.0))
        q_current = self.physics.safe_float(well_data.get('flow_rate', 0.0))
        pump_depth = self.physics.safe_float(well_data.get('pump_depth', 0.0))
        h_din_current = self.physics.safe_float(well_data.get('h_din', pump_depth))
        
        DEBUG.data("Входные", f"Pнас={p_nas:.1f}, Q={q_current:.1f}, режим={t_work_current:.0f}/{t_pause_current:.0f}")
        DEBUG.data("Глубины", f"Lнас={pump_depth:.1f}, Hдин={h_din_current:.1f}")
        
        if p_nas <= 0:
            DEBUG.log("❌ Нет Pнас → сценарий A неприменим")
            DEBUG.exit()
            return {'has_gas_problem': False, 'reason': 'Нет Pнас'}
        
        # 1. Анализ газовой проблемы
        gas_analysis = self.physics.analyze_gas_problem_comprehensive(
            t_work_current, t_pause_current, freq_current
        )
        
        result = {
            'scenario': 'A',
            'has_gas_problem': gas_analysis['has_gas_problem'],
            'p_zab_end_current': gas_analysis['p_zab_end'],
            'p_nas': p_nas,
            'current_work_time': t_work_current,
            'current_pause_time': t_pause_current,
            'current_freq': freq_current
        }
        
        if not gas_analysis['has_gas_problem']:
            DEBUG.log("✅ Нет газовой проблемы → оставляем текущий режим")
            result['reason'] = 'Нет газовой проблемы'
            result['recommended_work_time'] = t_work_current
            result['recommended_pause_time'] = t_pause_current
            result['recommended_freq'] = freq_current
            DEBUG.exit()
            return result
        
        DEBUG.log("⚠️ Обнаружена газовая проблема, начинаем оптимизацию...")
        
        # 2. Получаем Pзаб_start и Pзаб_end из моделирования
        simulation = self.physics.simulate_pwf_during_work(
            t_work_current, t_pause_current, freq_current
        )
        
        p_zab_start = simulation.get('p_zab_start', 0)
        p_zab_end = simulation.get('p_zab_end', 0)
        
        if p_zab_start <= 0:
            # Если нет из моделирования, оцениваем через Pзаб(ВДП)
            p_zab_vdp = self.physics.safe_float(well_data.get('p_zab_vdp', 0))
            p_zab_start = p_zab_vdp if p_zab_vdp > 0 else p_nas * 1.5
        
        DEBUG.data("Pзаб параметры", f"start={p_zab_start:.1f}, end={p_zab_end:.1f}")
        
        # 3. Целевое Pзаб
        target_p_zab = 0.75 * p_nas
        DEBUG.data("Целевое Pзаб", f"{target_p_zab:.1f} атм (0.75*{p_nas:.1f})")
        
        # 4. Расчет ΔP/мин
        if t_work_current > 0:
            delta_p_per_min = (p_zab_start - p_zab_end) / t_work_current
            DEBUG.data("ΔP/мин", f"{delta_p_per_min:.2f} атм/мин")
        else:
            DEBUG.log("Ошибка: t_work_current ≤ 0")
            result['reason'] = 'Ошибка: время работы ≤ 0'
            result['recommended_work_time'] = t_work_current
            result['recommended_pause_time'] = t_pause_current
            DEBUG.exit()
            return result
        
        # 5. Сколько нужно уменьшить время работы?
        delta_p_needed = target_p_zab - p_zab_end
        if delta_p_needed <= 0:
            DEBUG.log("Уже достигнута цель Pзаб")
            result['reason'] = 'Уже достигнута цель Pзаб'
            result['recommended_work_time'] = t_work_current
            result['recommended_pause_time'] = t_pause_current
            DEBUG.exit()
            return result
        
        delta_t_work_needed = delta_p_needed / delta_p_per_min
        new_t_work_float = t_work_current - delta_t_work_needed
        
        DEBUG.data("Необходимое уменьшение", f"ΔP={delta_p_needed:.1f} атм, Δt={delta_t_work_needed:.1f} мин")
        DEBUG.data("Новое t_work (float)", f"{new_t_work_float:.1f} мин")
        
        # 6. Округление и ограничения
        new_t_work = max(1.0, round(new_t_work_float))  # Минимум 1 минута, округляем
        if new_t_work >= t_work_current:
            new_t_work = max(1.0, t_work_current - 1.0)  # Хотя бы на 1 минуту уменьшаем
        
        DEBUG.data("Новое t_work (округл)", f"{new_t_work:.0f} мин")
        
        # 7. Расчет через динамический уровень
        
        # 7.1. ANNULAR_CAPACITY из физики
        annular_capacity = self.physics.ANNULAR_CAPACITY  # л/м
        
        # 7.2. Расчет мгновенного дебита q_moment
        cycle_current = t_work_current + t_pause_current
        work_hours_current = (t_work_current / cycle_current) * 24.0 if cycle_current > 0 else 0
        q_moment_current = q_current * (24.0 / work_hours_current) if work_hours_current > 0 else q_current
        
        DEBUG.data("Q мгновенный", f"{q_moment_current:.1f} м³/сут")
        
        # 7.3. Расчет h_din_pause (Hдин после накопления) по формуле 1.6
        # h_din_pause = h_din_pump + ((q_daily - q_moment) × 1000) / (ANNULAR_CAPACITY × 24)
        h_din_pump = h_din_current
        numerator = (q_current - q_moment_current) * 1000.0
        denominator = annular_capacity * 24.0
        
        if denominator > 0:
            delta_h_pause_to_pump = numerator / denominator  # Отрицательное значение
            h_din_pause = h_din_pump + delta_h_pause_to_pump
        else:
            h_din_pause = h_din_pump
        
        DEBUG.data("Hдин уровни", f"насос={h_din_pump:.1f} м, после накопления={h_din_pause:.1f} м")
        
        # 7.4. ΔH за цикл
        delta_h_cycle = h_din_pump - h_din_pause
        DEBUG.data("ΔH за цикл", f"{delta_h_cycle:.1f} м")
        
        # 7.5. ΔH за минуту работы и накопления
        if t_work_current > 0:
            delta_h_per_min_work = delta_h_cycle / t_work_current
            DEBUG.data("ΔH/мин работа", f"{delta_h_per_min_work:.2f} м/мин")
        else:
            delta_h_per_min_work = 0
        
        if t_pause_current > 0:
            delta_h_per_min_pause = delta_h_cycle / t_pause_current
            DEBUG.data("ΔH/мин накопление", f"{delta_h_per_min_pause:.2f} м/мин")
        else:
            delta_h_per_min_pause = 0
        
        # 8. Новое падение уровня за работу
        delta_h_new_work = delta_h_per_min_work * new_t_work
        DEBUG.data("Новое ΔH за работу", f"{delta_h_new_work:.1f} м")
        
        # 9. Новое время накопления
        if delta_h_per_min_pause > 0:
            new_t_pause_float = delta_h_new_work / delta_h_per_min_pause
        else:
            new_t_pause_float = t_pause_current * (new_t_work / t_work_current)  # пропорция
        
        DEBUG.data("Новое t_pause (float)", f"{new_t_pause_float:.1f} мин")
        
        # 10. Округление времени накопления
        new_t_pause = max(1.0, round(new_t_pause_float))
        DEBUG.data("Новое t_pause (округл)", f"{new_t_pause:.0f} мин")
        
        # 11. Проверка ограничений
        # Минимальный цикл: хотя бы 5 минут
        min_cycle = 5.0
        if (new_t_work + new_t_pause) < min_cycle:
            scale = min_cycle / (new_t_work + new_t_pause)
            new_t_work = max(1.0, round(new_t_work * scale))
            new_t_pause = max(1.0, round(new_t_pause * scale))
        
        # 12. Новый дебит
        new_cycle = new_t_work + new_t_pause
        new_work_hours = (new_t_work / new_cycle) * 24.0 if new_cycle > 0 else 0
        
        if work_hours_current > 0:
            q_tech_new = q_current * (new_work_hours / work_hours_current)
        else:
            q_tech_new = q_current
        
        DEBUG.data("Новый дебит", f"{q_tech_new:.1f} м³/сут")
        
        # 13. Проверка частоты (при необходимости)
        new_freq = freq_current
        
        # Если все еще проблема, пробуем уменьшить частоту
        if new_t_work <= 1.0 and freq_current > 40.0:
            DEBUG.log("Не удалось решить через t_work, пробуем частоту...")
            # Уменьшаем частоту на 10%
            new_freq = max(40.0, freq_current * 0.9)
            DEBUG.data("Новая частота", f"{new_freq:.1f} Гц")
        
        # 14. Формирование результата
        result.update({
            'recommended_work_time': new_t_work,
            'recommended_pause_time': new_t_pause,
            'recommended_freq': new_freq,
            'q_tech_new': q_tech_new,
            'q_tech_current': q_current,
            'reason': f'Газовая проблема: Pзаб={p_zab_end:.1f} < Pнас={p_nas:.1f}',
            'delta_h_per_min_work': delta_h_per_min_work,
            'delta_h_per_min_pause': delta_h_per_min_pause,
            'h_din_pump': h_din_pump,
            'h_din_pause': h_din_pause
        })

        # ========== НОВЫЙ КОД: Сохраняем детальные данные симуляции ==========
        # Получаем симуляцию еще раз (или можно было сохранить ее ранее в переменную)
        # Чтобы не делать двойную работу, лучше сохранить simulation в начале и использовать его.
        # Но для простоты и гарантии, что данные есть, пересчитаем.
        # В идеале, в начале метода нужно сохранить: simulation = self.physics.simulate_pwf_during_work(...)
        # и потом использовать его и здесь, и в расчетах выше.
        # Так как мы меняем только визуализацию, для надежности сделаем перерасчет.
        final_simulation = self.physics.simulate_pwf_during_work(
            new_t_work, new_t_pause, new_freq
        )
        result['simulation_data'] = {
            'time_points': final_simulation['time_points'],
            'pwf_points': final_simulation['pwf_points'],
            'h_din_points': final_simulation['h_din_points'],
            'p_zab_start': final_simulation['p_zab_start'],
            'p_zab_end': final_simulation['p_zab_end'],
        }
        # ========== КОНЕЦ НОВОГО КОДА ==========

        DEBUG.data("Итоговый режим", f"{new_t_work:.0f}/{new_t_pause:.0f}, f={new_freq:.1f}")
        DEBUG.data("Прирост/снижение", f"{((q_tech_new/q_current)-1)*100:+.1f}%")

        DEBUG.exit()
        return result
        
    def optimize_scenario_b_pump_load(self, t_work_current: float, t_pause_current: float,
                                      freq_current: float = 50.0, include_freq_optimization: bool = True) -> Dict:
        """
        СЦЕНАРИЙ B: Оптимизация загрузки насоса путем замены типоразмера
        
        Логика:
        1. Рассчитываем текущий K_под = Qж / ( (t_раб/(t_раб+t_пауз)) * Qном * (f/50) )
        2. Если K_под < 0.75 (НЕДОГРУЗКА) → ставим насос МЕНЬШЕГО типоразмера
        3. Если K_под > 1.25 (ПЕРЕГРУЗКА) → ставим насос БОЛЬШЕГО типоразмера (+5% к дебиту)
        4. Время работы оставляем таким же
        5. Пересчитываем время накопления: T_пауз = (Qном_нов * T_раб / Qгжс_нов) - T_раб
        6. Дебит нефти для экономики: Qн = Qж * (1 - обв) * плотность_нефти
        """
        DEBUG.enter()
        DEBUG.section("СЦЕНАРИЙ B: ОПТИМИЗАЦИЯ ЗАМЕНОЙ НАСОСА")
        
        try:
            well_data = self.physics.well
            physics = self.physics
            economic = self.economic
            
            # ========== 1. Базовые данные ==========
            q_liq_surface = physics.safe_float(well_data.get('flow_rate', 0.0))  # м³/сут (поверхностный)
            water_cut = physics.safe_float(well_data.get('water_cut', 0.0))  # %
            oil_density = physics.safe_float(well_data.get('oil_density', 0.85))  # т/м³
            q_nom_current = physics.safe_float(well_data.get('pump_flow', 60.0))  # м³/сут
            pump_mark = well_data.get('pump_mark', '')
            pump_depth = physics.safe_float(well_data.get('pump_depth', 0.0))
            h_din_current = physics.safe_float(well_data.get('h_din', pump_depth))
            k_prod = physics.safe_float(well_data.get('prod_coef', 0.0))
            
            DEBUG.data("Исходные данные", 
                      f"Qж={q_liq_surface:.1f} м³/сут, Обв={water_cut:.1f}%, "
                      f"ρн={oil_density:.3f} т/м³, Qном={q_nom_current:.0f} м³/сут, "
                      f"режим={t_work_current:.0f}/{t_pause_current:.0f} мин, f={freq_current:.1f} Гц")
            
            # ========== 2. Пересчет в ГЖС (забойные условия) ==========
            OIL_VOLUME_FACTOR = 1.1  # объемный коэффициент нефти
            wc = water_cut / 100.0  # доля воды
            oil_cut = 1.0 - wc  # доля нефти
            
            # Qгжс = Qж * [ (1 - B) * bн + B ]
            q_downhole_current = q_liq_surface * (oil_cut * OIL_VOLUME_FACTOR + wc)
            DEBUG.data("Пересчет в ГЖС", f"{q_liq_surface:.1f} → {q_downhole_current:.1f} м³/сут")
            
            # ========== 3. Текущий K_под ==========
            cycle_current = t_work_current + t_pause_current
            duty_cycle = t_work_current / cycle_current if cycle_current > 0 else 0
            freq_correction = freq_current / 50.0
            
            # Расчетная производительность установки (по ГЖС!)
            q_install_current = duty_cycle * q_nom_current * freq_correction
            
            # Коэффициент подачи (по ГЖС!)
            k_pod_current = q_downhole_current / q_install_current if q_install_current > 0 else 0
            
            DEBUG.data("K_под текущий", 
                      f"{k_pod_current:.3f} = {q_downhole_current:.1f} / ({duty_cycle:.3f}×{q_nom_current:.0f}×{freq_correction:.3f})")
            
            # ========== 4. Если уже в норме - выходим ==========
            if 0.75 <= k_pod_current <= 1.25:
                DEBUG.log("✅ K_под уже в оптимальном диапазоне 0.75-1.25")
                
                # Рассчитываем дебит нефти для экономики
                q_oil_current = q_liq_surface * oil_cut * oil_density  # т/сут
                
                result = {
                    'scenario': 'B',
                    'problem': 'optimal',
                    'k_pod_current': k_pod_current,
                    'k_pod_new': k_pod_current,
                    'current_regime': f"{t_work_current:.0f}/{t_pause_current:.0f}",
                    'recommended_regime': f"{t_work_current:.0f}/{t_pause_current:.0f}",
                    'current_freq': freq_current,
                    'recommended_freq': freq_current,
                    'current_work_time': t_work_current,
                    'current_pause_time': t_pause_current,
                    'recommended_work_time': t_work_current,
                    'recommended_pause_time': t_pause_current,
                    'q_liq_surface': q_liq_surface,
                    'q_liq_new': q_liq_surface,
                    'q_downhole_current': q_downhole_current,
                    'q_downhole_new': q_downhole_current,
                    'q_oil_current': q_oil_current,
                    'q_oil_new': q_oil_current,
                    'reason': 'K_под уже в оптимальном диапазоне 0.75-1.25',
                    'is_physical_limit': False
                }
                
                DEBUG.exit()
                return result
            
            # ========== 5. Определяем новый типоразмер насоса ==========
            # Иерархия насосов по возрастанию
            PUMP_HIERARCHY = ['25', '40', '60', '80', '125', '160', '200', '250']
            PUMP_NOMINAL_FLOWS = {
                '25': 25, '40': 40, '60': 60, '80': 80,
                '125': 125, '160': 160, '200': 200, '250': 250
            }
            
            # Определяем текущий тип насоса из марки или по номинальной подаче
            current_pump_type = None
            
            # Пытаемся извлечь из марки
            if pump_mark:
                import re
                numbers = re.findall(r'\d+', str(pump_mark))
                for num in numbers:
                    if num in PUMP_HIERARCHY:
                        current_pump_type = num
                        break
            
            # Если не удалось, определяем по номинальной подаче
            if not current_pump_type:
                for pump_type, nom_flow in PUMP_NOMINAL_FLOWS.items():
                    if abs(nom_flow - q_nom_current) < 10:  # погрешность 10 м³/сут
                        current_pump_type = pump_type
                        break
            
            # Если все еще не определили, определяем по диапазону
            if not current_pump_type:
                if q_nom_current <= 30:
                    current_pump_type = '25'
                elif q_nom_current <= 50:
                    current_pump_type = '40'
                elif q_nom_current <= 70:
                    current_pump_type = '60'
                elif q_nom_current <= 100:
                    current_pump_type = '80'
                elif q_nom_current <= 140:
                    current_pump_type = '125'
                elif q_nom_current <= 180:
                    current_pump_type = '160'
                elif q_nom_current <= 220:
                    current_pump_type = '200'
                else:
                    current_pump_type = '250'
            
            DEBUG.data("Текущий тип насоса", current_pump_type)
            
            # Находим индекс текущего насоса в иерархии
            try:
                current_idx = PUMP_HIERARCHY.index(current_pump_type)
            except ValueError:
                # Если не нашли, используем ближайший
                current_flow = PUMP_NOMINAL_FLOWS.get(current_pump_type, q_nom_current)
                closest_type = min(PUMP_HIERARCHY, 
                                  key=lambda x: abs(PUMP_NOMINAL_FLOWS[x] - current_flow))
                current_idx = PUMP_HIERARCHY.index(closest_type)
                current_pump_type = closest_type
                DEBUG.data("Скорректированный тип", current_pump_type)
            
            # ========== 6. Выбираем новый насос по правилам ==========
            new_pump_type = current_pump_type
            new_q_liq_surface = q_liq_surface  # по умолчанию дебит не меняем
            
            if k_pod_current < 0.75:  # НЕДОГРУЗКА
                DEBUG.log("⚠️ НЕДОГРУЗКА: нужен насос МЕНЬШЕГО типоразмера")
                
                # Берем насос на один шаг меньше, если это возможно
                if current_idx > 0:
                    new_pump_type = PUMP_HIERARCHY[current_idx - 1]
                    new_q_liq_surface = q_liq_surface  # дебит не меняем
                    reason = f"Недогрузка: K_под={k_pod_current:.2f} < 0.75, замена {current_pump_type} → {new_pump_type}"
                else:
                    # Уже минимальный насос, оставляем текущий
                    new_pump_type = current_pump_type
                    new_q_liq_surface = q_liq_surface
                    reason = f"Недогрузка, но это минимальный типоразмер {current_pump_type}"
                    DEBUG.log("⚠️ Это минимальный типоразмер, замена невозможна")
            
            elif k_pod_current > 1.25:  # ПЕРЕГРУЗКА
                DEBUG.log("⚠️ ПЕРЕГРУЗКА: нужен насос БОЛЬШЕГО типоразмера (+5% к дебиту)")
                
                # Берем насос на один шаг больше, если это возможно
                if current_idx < len(PUMP_HIERARCHY) - 1:
                    new_pump_type = PUMP_HIERARCHY[current_idx + 1]
                    new_q_liq_surface = q_liq_surface * 1.05  # +5% к дебиту
                    reason = f"Перегрузка: K_под={k_pod_current:.2f} > 1.25, замена {current_pump_type} → {new_pump_type}, дебит +5%"
                else:
                    # Уже максимальный насос, оставляем текущий
                    new_pump_type = current_pump_type
                    new_q_liq_surface = q_liq_surface * 1.05  # все равно +5%
                    reason = f"Перегрузка, это максимальный типоразмер {current_pump_type}, дебит +5%"
                    DEBUG.log("⚠️ Это максимальный типоразмер, только +5% к дебиту")
            
            DEBUG.data("Новый тип насоса", new_pump_type)
            DEBUG.data("Новый дебит Qж", f"{new_q_liq_surface:.1f} м³/сут")
            
            # ========== 7. Номинальная подача нового насоса ==========
            q_nom_new = PUMP_NOMINAL_FLOWS.get(new_pump_type, 60)
            DEBUG.data("Номинальная подача нового насоса", f"{q_nom_new:.0f} м³/сут")
            
            # ========== 8. Пересчет в ГЖС для нового дебита ==========
            q_downhole_new = new_q_liq_surface * (oil_cut * OIL_VOLUME_FACTOR + wc)
            DEBUG.data("Новый ГЖС", f"{new_q_liq_surface:.1f} → {q_downhole_new:.1f} м³/сут")
            
            # ========== 9. Расчет нового времени накопления ==========
            # T_пауз = (Qном_нов * T_раб * (f/50) / Qгжс_нов) - T_раб
            new_pause_time = (q_nom_new * t_work_current * freq_correction / q_downhole_new) - t_work_current
            new_pause_time = max(1.0, round(new_pause_time))  # минимум 1 минута, округляем
            
            DEBUG.data("Новое время накопления", 
                      f"{new_pause_time:.0f} мин = ({q_nom_new:.0f}×{t_work_current:.0f}×{freq_correction:.3f}/{q_downhole_new:.1f}) - {t_work_current:.0f}")
            
            # ========== 10. Проверка минимального цикла ==========
            min_cycle = 5.0
            if (t_work_current + new_pause_time) < min_cycle:
                scale = min_cycle / (t_work_current + new_pause_time)
                new_t_work = max(1.0, round(t_work_current * scale))
                new_pause_time = max(1.0, round(new_pause_time * scale))
                DEBUG.log(f"⚠️ Корректировка минимального цикла: {t_work_current:.0f}/{new_pause_time:.0f} → {new_t_work:.0f}/{new_pause_time:.0f}")
            else:
                new_t_work = t_work_current
            
            # ========== 11. Новый коэффициент подачи ==========
            new_cycle = new_t_work + new_pause_time
            new_duty_cycle = new_t_work / new_cycle if new_cycle > 0 else 0
            q_install_new = new_duty_cycle * q_nom_new * freq_correction
            k_pod_new = q_downhole_new / q_install_new if q_install_new > 0 else 0
            
            DEBUG.data("Новый K_под", 
                      f"{k_pod_new:.3f} = {q_downhole_new:.1f} / ({new_duty_cycle:.3f}×{q_nom_new:.0f}×{freq_correction:.3f})")
            
            # ========== 12. Дебит нефти для экономики ==========
            q_oil_current = q_liq_surface * oil_cut * oil_density  # т/сут
            q_oil_new = new_q_liq_surface * oil_cut * oil_density  # т/сут
            
            DEBUG.data("Дебит нефти", f"текущий={q_oil_current:.2f} т/сут, новый={q_oil_new:.2f} т/сут")
            
            # ========== 13. Экономический расчет ==========
            economic_result = economic.calculate_economic_effect_comprehensive(
                well_data=well_data,
                old_schedule=[t_work_current, t_pause_current],
                new_schedule=[new_t_work, new_pause_time],
                old_freq=freq_current,
                new_freq=freq_current,  # частоту не меняем
                has_gas_problem_old=False
            )
            
            # ========== 14. Формирование результата ==========
            result = {
                'scenario': 'B',
                'problem': 'underload' if k_pod_current < 0.75 else 'overload',
                'k_pod_current': k_pod_current,
                'k_pod_new': k_pod_new,
                'current_regime': f"{t_work_current:.0f}/{t_pause_current:.0f}",
                'recommended_regime': f"{new_t_work:.0f}/{new_pause_time:.0f}",
                'current_freq': freq_current,
                'recommended_freq': freq_current,
                'current_work_time': t_work_current,
                'current_pause_time': t_pause_current,
                'recommended_work_time': new_t_work,
                'recommended_pause_time': new_pause_time,
                'current_pump_type': current_pump_type,
                'recommended_pump_type': new_pump_type,
                'q_liq_surface': q_liq_surface,
                'q_liq_new': new_q_liq_surface,
                'q_downhole_current': q_downhole_current,
                'q_downhole_new': q_downhole_new,
                'q_oil_current': q_oil_current,
                'q_oil_new': q_oil_new,
                'reason': reason,
                'economic_analysis': economic_result,
                'economic_effect': economic_result.get('total_effect_per_day', 0),
                'is_physical_limit': (current_idx == 0 and k_pod_current < 0.75) or (current_idx == len(PUMP_HIERARCHY)-1 and k_pod_current > 1.25),
                'k_prod': k_prod,
                'duty_cycle': duty_cycle,
                'new_duty_cycle': new_duty_cycle
            }
            
            DEBUG.data("Итог", f"K_под {k_pod_current:.2f} → {k_pod_new:.2f}, насос {current_pump_type} → {new_pump_type}")
            DEBUG.data("Эффект", f"{economic_result.get('total_effect_per_day', 0):.0f} ₽/сут")
            
            DEBUG.exit()
            return result
            
        except Exception as e:
            DEBUG.log(f"Ошибка: {str(e)}", "ERROR")
            import traceback
            DEBUG.log(traceback.format_exc(), "ERROR")
            
            return {
                'scenario': 'B',
                'problem': 'error',
                'k_pod_current': 0,
                'k_pod_new': 0,
                'current_regime': f"{t_work_current:.0f}/{t_pause_current:.0f}",
                'recommended_regime': f"{t_work_current:.0f}/{t_pause_current:.0f}",
                'current_freq': freq_current,
                'recommended_freq': freq_current,
                'current_work_time': t_work_current,
                'current_pause_time': t_pause_current,
                'recommended_work_time': t_work_current,
                'recommended_pause_time': t_pause_current,
                'q_liq_surface': q_liq_surface if 'q_liq_surface' in locals() else 0,
                'q_liq_new': q_liq_surface if 'q_liq_surface' in locals() else 0,
                'reason': f"Ошибка расчета: {str(e)[:100]}",
                'economic_effect': 0,
                'is_physical_limit': False
            }
    
    def optimize_integrated(self, t_work_current: float, t_pause_current: float,
                            freq_current: float = 50.0, include_freq_optimization: bool = True) -> Dict:
        """
        ИСПРАВЛЕННАЯ интегрированная оптимизация
        
        1. Сначала Сценарий A (газ) - используем исправленную версию
        2. Если нет газа → Сценарий B (загрузка) - исправленная версия
        """
        DEBUG.enter()
        DEBUG.section("ИНТЕГРИРОВАННАЯ ОПТИМИЗАЦИЯ (A+B)")
        
        # 1. Проверяем газовую проблему
        scenario_a_result = self.optimize_scenario_a_gas_problem(
            t_work_current, t_pause_current, freq_current
        )
        
        # Если есть газовая проблема, возвращаем рекомендации Сценария A
        if scenario_a_result['has_gas_problem']:
            DEBUG.log("✅ Применен СЦЕНАРИЙ A (газовая проблема)")
            
            well_data = self.physics.well
            q_nom = self.physics.safe_float(well_data.get('pump_flow', 60.0))
            
            k_pod_before = self.economic.calculate_k_pod_corrected(
                scenario_a_result['q_tech_current'], t_work_current, t_pause_current, 
                q_nom, freq_current
            )
            
            k_pod_after = self.economic.calculate_k_pod_corrected(
                scenario_a_result['q_tech_new'], scenario_a_result['recommended_work_time'], 
                scenario_a_result['recommended_pause_time'], q_nom, scenario_a_result['recommended_freq']
            )
            
            result = {
                'scenario': 'A',
                'has_gas_problem': True,
                'reason': scenario_a_result.get('reason', 'Газовая проблема'),
                'current_regime': f"{t_work_current:.0f}/{t_pause_current:.0f}",
                'recommended_regime': f"{scenario_a_result['recommended_work_time']:.0f}/{scenario_a_result['recommended_pause_time']:.0f}",
                'recommended_work_time': scenario_a_result['recommended_work_time'],
                'recommended_pause_time': scenario_a_result['recommended_pause_time'],
                'recommended_freq': scenario_a_result.get('recommended_freq', freq_current),
                'q_tech_current': scenario_a_result.get('q_tech_current', 0),
                'q_tech_new': scenario_a_result.get('q_tech_new', 0),
                'p_zab_before': scenario_a_result.get('p_zab_end_current', 0),
                'p_zab_after': scenario_a_result.get('p_zab_end_new', 0),
                'current_work_time': t_work_current, 
                'current_pause_time': t_pause_current, 
                'current_freq': freq_current,
                'k_pod_before': k_pod_before,
                'k_pod_after': k_pod_after,
                'k_pod_problem_after': 'underload' if k_pod_after < 0.75 else 
                                      'overload' if k_pod_after > 1.25 else 'optimal'
            }
            
            DEBUG.data("Итог Сценария A", f"режим {result['current_regime']} → {result['recommended_regime']}")
            DEBUG.exit()
            return result
        
        # 2. Если нет газовой проблемы → Сценарий B
        DEBUG.log("✅ Нет газовой проблемы → применяем СЦЕНАРИЙ B")
        scenario_b_result = self.optimize_scenario_b_pump_load(
            t_work_current, t_pause_current, freq_current, include_freq_optimization
        )
        
        result = {
            'scenario': 'B',
            'has_gas_problem': False,
            'reason': scenario_b_result.get('reason', 'Оптимизация загрузки'),
            'current_regime': f"{t_work_current:.0f}/{t_pause_current:.0f}",
            'recommended_regime': f"{scenario_b_result['recommended_work_time']:.0f}/{scenario_b_result['recommended_pause_time']:.0f}",
            'recommended_work_time': scenario_b_result['recommended_work_time'],
            'recommended_pause_time': scenario_b_result['recommended_pause_time'],
            'recommended_freq': scenario_b_result.get('recommended_freq', freq_current),
            'q_tech_current': scenario_b_result.get('q_tech_current', 0),
            'q_tech_new': scenario_b_result.get('q_tech_new', 0),
            'k_pod_before': scenario_b_result['k_pod_current'],
            'k_pod_after': scenario_b_result['k_pod_new'],
            'bad_well': scenario_b_result.get('bad_well', False),
            'optimization_method': scenario_b_result.get('optimization_method', ''),
            'current_work_time': t_work_current,  # ← ДОБАВИТЬ
            'current_pause_time': t_pause_current,  # ← ДОБАВИТЬ
            'current_freq': freq_current,  # ← ДОБАВИТЬ
            'problem': scenario_b_result.get('problem', 'optimal')
        }
        
        DEBUG.data("Итог Сценария B", f"K_под {result['k_pod_before']:.2f} → {result['k_pod_after']:.2f}")
        DEBUG.exit()
        return result
        
    def calculate_optimal_pause_time(self, new_t_work: float, t_pause_current: float,
                                   t_work_current: float, q_tech: float,
                                   freq_current: float = 50.0) -> float:
        """
        Расчет оптимального времени накопления после изменения времени работы
        """
        DEBUG.enter()
        DEBUG.log("Расчет оптимального t_pause")
        
        # Расчет коэффициента изменения объема
        volume_change_factor = (new_t_work / t_work_current)
        
        # Безопасный коэффициент (запас на восстановление)
        safety_factor = 1.2
        
        # Новое время накопления
        new_t_pause = t_pause_current * volume_change_factor * safety_factor
        
        DEBUG.data("Коэффициенты", f"объем={volume_change_factor:.3f}, безопасность={safety_factor}")
        DEBUG.data("Исходное t_pause", f"{t_pause_current:.0f} мин")
        
        # Ограничения
        min_pause = 5.0  # Минимум 5 минут
        max_pause = 72.0 * 60.0  # Максимум 3 суток
        
        if new_t_pause < min_pause:
            new_t_pause = min_pause
            DEBUG.data("Ограничение снизу", f"{min_pause:.0f} мин")
        elif new_t_pause > max_pause:
            new_t_pause = max_pause
            DEBUG.data("Ограничение сверху", f"{max_pause:.0f} мин")
        
        DEBUG.data("Новое t_pause", f"{new_t_pause:.0f} мин")
        DEBUG.exit()
        
        return new_t_pause

# ============================================================
# КЛАСС ДЛЯ СЦЕНАРИЯ "ПОТЕНЦИАЛ СКВАЖИН" (ИСПРАВЛЕННЫЙ)
# ============================================================

class WellPotentialAnalyzer:
    """Анализ потенциала скважин для увеличения дебита"""
    
    def __init__(self, physics_calculator: CorrectedKPRPhysics, 
                 economic_calculator: EconomicCalculatorCorrected,
                 min_depth_diff: float = 400.0,
                 min_p_pr: float = 40.0,
                 min_k_prod: float = 0.7):
        """Инициализация с параметрами фильтров"""
        self.physics = physics_calculator
        self.economic = economic_calculator
        
        # Константы из вашей формулы (теперь передаются как параметры)
        self.ANNULAR_CAPACITY = 9.5  # л/м (емкость кольцевого пространства на 1 м)
        self.MIN_DEPTH_DIFF = min_depth_diff  # минимальная разница глубины, м
        self.MIN_P_PR = min_p_pr         # минимальное давление на приеме, атм
        self.MIN_K_PROD = min_k_prod        # минимальный коэффициент продуктивности
        
        DEBUG.log(f"Инициализация WellPotentialAnalyzer с фильтрами:")
        DEBUG.data("Фильтры", f"ΔL-Hдин>{self.MIN_DEPTH_DIFF} м, Pпр>{self.MIN_P_PR} атм, Kпр>{self.MIN_K_PROD}")
    
    def calculate_momentaneous_flow(self, q_daily: float, t_work: float, t_pause: float) -> float:
        """
        Расчет мгновенного дебита (формула 1.4-1.5)
        """
        DEBUG.enter()
        DEBUG.log(f"Расчет мгновенного дебита: Q={q_daily:.1f}, режим={t_work:.0f}/{t_pause:.0f}")
        
        if t_work + t_pause <= 0:
            DEBUG.log("Ошибка: сумма времени ≤ 0")
            DEBUG.exit()
            return q_daily
        
        work_hours = (t_work / (t_work + t_pause)) * 24.0
        
        DEBUG.data("Рабочие часы", f"{work_hours:.1f} ч/сут")
        
        if work_hours > 0:
            q_moment = q_daily * 24.0 / work_hours
            DEBUG.data("Q_мгновенный", f"{q_moment:.1f} м³/сут = {q_daily:.1f}×24/{work_hours:.1f}")
        else:
            q_moment = q_daily
            DEBUG.data("Q_мгновенный", f"{q_moment:.1f} м³/сут (work_hours=0)")
        
        DEBUG.exit()
        return q_moment
    
    def calculate_h_din_after_pause(self, q_daily: float, q_moment: float, 
                                   h_din_pump: float) -> float:
        """
        Расчет H_дин после накопления (формула 1.6)
        """
        DEBUG.enter()
        DEBUG.log(f"Расчет H_дин после накопления: Qсут={q_daily:.1f}, Qмгн={q_moment:.1f}, Hдин_насос={h_din_pump:.1f}")
        
        # Решаем уравнение относительно H_дин_накопления
        numerator = (q_daily - q_moment) * 1000.0
        denominator = self.ANNULAR_CAPACITY * 24.0
        
        DEBUG.data("Числитель", f"{numerator:.1f} = ({q_daily:.1f}-{q_moment:.1f})×1000")
        DEBUG.data("Знаменатель", f"{denominator:.1f} = {self.ANNULAR_CAPACITY}×24")
        
        if denominator == 0:
            DEBUG.log("Ошибка: знаменатель = 0")
            DEBUG.exit()
            return h_din_pump
            
        h_din_pause = h_din_pump + (numerator / denominator)
        
        DEBUG.data("H_дин после накопления", f"{h_din_pause:.1f} м = {h_din_pump:.1f} - ({numerator:.1f}/{denominator:.1f})")
        DEBUG.exit()
        
        return h_din_pause
    
    def calculate_pressure_at_pump_intake(self, well_data: Dict) -> float:
        """
        Расчет давления на приеме насоса
        """
        DEBUG.enter()
        DEBUG.log("Расчет давления на приеме насоса")
        
        # Пробуем взять из данных
        p_pr_data = self.physics.safe_float(well_data.get('p_pr', 0))
        
        if p_pr_data > 0:
            DEBUG.data("Pпр из данных", f"{p_pr_data:.1f} атм")
            DEBUG.exit()
            return p_pr_data
        
        # Если нет в данных, рассчитываем
        p_zat = self.physics.safe_float(well_data.get('p_zat', 0))
        pump_depth = self.physics.safe_float(well_data.get('pump_depth', 0))
        h_din = self.physics.safe_float(well_data.get('h_din', 0))
        
        DEBUG.data("Исходные для расчета", f"Pзатр={p_zat:.1f}, Lнас={pump_depth:.1f}, Hдин={h_din:.1f}")
        
        if pump_depth <= 0 or h_din <= 0:
            DEBUG.log("Ошибка: нет данных для расчета")
            DEBUG.exit()
            return 0.0
        
        # Расчет плотности жидкости
        water_cut_percent = self.physics.safe_float(well_data.get('water_cut', 0.0))
        oil_density_ton_per_m3 = self.physics.safe_float(well_data.get('oil_density', 0.84))
        
        wc = water_cut_percent / 100.0
        oil_density_kg_per_m3 = oil_density_ton_per_m3 * 1000.0
        
        rho_liquid = wc * self.physics.WATER_DENSITY + (1.0 - wc) * oil_density_kg_per_m3
        
        # P_прием = P_затрубное + ρgh
        height_diff = pump_depth - h_din
        p_hydrostatic = rho_liquid * self.physics.G * height_diff / self.physics.ATM_TO_PA
        
        p_pr = p_zat + p_hydrostatic
        
        DEBUG.data("Параметры расчета", f"ρ={rho_liquid:.1f} кг/м³, Δh={height_diff:.1f} м")
        DEBUG.data("P_гидростатическое", f"{p_hydrostatic:.1f} атм")
        DEBUG.data("P_прием расчетное", f"{p_pr:.1f} атм = {p_zat:.1f} + {p_hydrostatic:.1f}")
        
        DEBUG.exit()
        return p_pr
    
    def check_potential_filters(self, well_data: Dict, 
                               t_work: float, t_pause: float,
                               min_k_pod: float = 1.25) -> Dict:
        """
        Проверка всех фильтров для сценария "Потенциал" с настраиваемым min_k_pod
        """
        DEBUG.enter()
        DEBUG.section("ПРОВЕРКА ФИЛЬТРОВ ДЛЯ СЦЕНАРИЯ 'ПОТЕНЦИАЛ'")
        DEBUG.data("Скважина", well_data.get('name', 'Unknown'))
        DEBUG.data("Параметры фильтров", f"Kпод>{min_k_pod}, ΔL-Hдин>{self.MIN_DEPTH_DIFF}м, Kпр>{self.MIN_K_PROD}, Pпр>{self.MIN_P_PR}атм")
        
        result = {
            'eligible': True,
            'failed_filters': [],
            'values': {}
        }
        
        # 1. Данные скважины
        q_daily = self.physics.safe_float(well_data.get('flow_rate', 0))
        pump_depth = self.physics.safe_float(well_data.get('pump_depth', 0))
        h_din_pump = self.physics.safe_float(well_data.get('h_din', 0))
        k_prod = self.physics.safe_float(well_data.get('prod_coef', 0))
        q_pump = self.physics.safe_float(well_data.get('pump_flow', 60.0))
        freq = self.economic.safe_frequency(well_data.get('rotations_hz'))
        
        DEBUG.data("Базовые данные", f"Q={q_daily:.1f}, Lнас={pump_depth:.1f}, Hдин={h_din_pump:.1f}")
        DEBUG.data("Kпр", f"{k_prod:.2f}, Qнасос={q_pump:.0f}, f={freq:.1f}")
        
        # 2. Расчет K_подачи (используем ваш метод)
        k_pod = self.economic.calculate_k_pod_corrected(
            q_daily, t_work, t_pause, q_pump, freq
        )
        
        result['values']['k_pod'] = k_pod
        DEBUG.data("K_под", f"{k_pod:.2f}")
        
        # Фильтр 1: K_подачи > min_k_pod (перегрузка)
        if k_pod <= min_k_pod:
            result['eligible'] = False
            result['failed_filters'].append(f'K_подачи={k_pod:.2f} ≤ {min_k_pod:.2f}')
            DEBUG.log(f"❌ Фильтр 1: K_под={k_pod:.2f} ≤ {min_k_pod:.2f} (не перегружен)")
        else:
            DEBUG.log(f"✅ Фильтр 1: K_под={k_pod:.2f} > {min_k_pod:.2f} (перегрузка)")
        
        # 3. Разница глубин
        depth_diff = pump_depth - h_din_pump
        result['values']['depth_diff'] = depth_diff
        
        DEBUG.data("Разница глубин", f"{depth_diff:.0f} м = {pump_depth:.0f} - {h_din_pump:.0f}")
        
        # Фильтр 2: Разница > MIN_DEPTH_DIFF м
        if depth_diff <= self.MIN_DEPTH_DIFF:
            result['eligible'] = False
            result['failed_filters'].append(f'Разница глубин={depth_diff:.0f} м ≤ {self.MIN_DEPTH_DIFF} м')
            DEBUG.log(f"❌ Фильтр 2: ΔL-Hдин={depth_diff:.0f} м ≤ {self.MIN_DEPTH_DIFF} м")
        else:
            DEBUG.log(f"✅ Фильтр 2: ΔL-Hдин={depth_diff:.0f} м > {self.MIN_DEPTH_DIFF} м")
        
        # 4. Коэффициент продуктивности
        result['values']['k_prod'] = k_prod
        
        DEBUG.data("K_продуктивности", f"{k_prod:.2f}")
        
        # Фильтр 3: K_продуктивности > MIN_K_PROD
        if k_prod <= self.MIN_K_PROD:
            result['eligible'] = False
            result['failed_filters'].append(f'K_продуктивности={k_prod:.2f} ≤ {self.MIN_K_PROD}')
            DEBUG.log(f"❌ Фильтр 3: Kпр={k_prod:.2f} ≤ {self.MIN_K_PROD}")
        else:
            DEBUG.log(f"✅ Фильтр 3: Kпр={k_prod:.2f} > {self.MIN_K_PROD}")
        
        # 5. Давление на приеме
        p_pr = self.calculate_pressure_at_pump_intake(well_data)
        result['values']['p_pr'] = p_pr
        
        DEBUG.data("P_прием", f"{p_pr:.1f} атм")
        
        # Фильтр 4: P_прием > MIN_P_PR атм
        if p_pr <= self.MIN_P_PR:
            result['eligible'] = False
            result['failed_filters'].append(f'P_прием={p_pr:.1f} атм ≤ {self.MIN_P_PR} атм')
            DEBUG.log(f"❌ Фильтр 4: Pпр={p_pr:.1f} атм ≤ {self.MIN_P_PR} атм")
        else:
            DEBUG.log(f"✅ Фильтр 4: Pпр={p_pr:.1f} атм > {self.MIN_P_PR} атм")
        
        DEBUG.data("Итог проверки", "ПРОШЕЛ" if result['eligible'] else "НЕ ПРОШЕЛ")
        DEBUG.data("Не пройдено фильтров", f"{len(result['failed_filters'])}")
        
        DEBUG.exit()
        return result
    
    def analyze_potential_increase(self, well_data: Dict, 
                                  min_k_pod: float = 1.25) -> Dict:
        """
        Полный анализ потенциала увеличения дебита с настраиваемым min_k_pod
        """
        DEBUG.enter()
        DEBUG.section(f"АНАЛИЗ ПОТЕНЦИАЛА: {well_data.get('name', 'Unknown')}")
        DEBUG.data("Параметры фильтров", f"Kпод>{min_k_pod}, ΔL-Hдин>{self.MIN_DEPTH_DIFF}м, Kпр>{self.MIN_K_PROD}, Pпр>{self.MIN_P_PR}атм")
        
        try:
            # 1. Исходные данные
            q_daily = self.physics.safe_float(well_data.get('flow_rate', 0))
            schedule = well_data.get('schedule', [15.0 * 60.0, 45.0 * 60.0])
            
            DEBUG.data("Исходные", f"Q={q_daily:.1f} м³/сут")
            
            if len(schedule) < 2:
                DEBUG.log("Ошибка: нет данных о графике работы")
                DEBUG.exit()
                return {'error': 'Нет данных о графике работы'}
            
            t_work = float(schedule[0])
            t_pause = float(schedule[1])
            h_din_pump = self.physics.safe_float(well_data.get('h_din', 0))
            pump_depth = self.physics.safe_float(well_data.get('pump_depth', 0))
            
            DEBUG.data("Режим", f"{t_work:.0f}/{t_pause:.0f} мин")
            DEBUG.data("Глубины", f"Lнас={pump_depth:.1f} м, Hдин={h_din_pump:.1f} м")
            
            # 2. Проверка фильтров с настраиваемым min_k_pod
            filter_check = self.check_potential_filters(well_data, t_work, t_pause, min_k_pod)
            
            if not filter_check['eligible']:
                DEBUG.log("❌ Не проходит фильтры")
                result = {
                    'eligible': False,
                    'reason': f"Не проходит фильтры: {', '.join(filter_check['failed_filters'])}",
                    'filters': filter_check
                }
                DEBUG.exit()
                return result
            
            DEBUG.log("✅ Проходит все фильтры, продолжаем анализ...")
            
            # 3. Расчет Q_мгновенного
            q_moment = self.calculate_momentaneous_flow(q_daily, t_work, t_pause)
            DEBUG.data("Q_мгновенный", f"{q_moment:.1f} м³/сут")
            
            # 4. Расчет H_дин после накопления
            h_din_pause = self.calculate_h_din_after_pause(q_daily, q_moment, h_din_pump)
            DEBUG.data("H_дин после накопления", f"{h_din_pause:.1f} м")
            
            # 5. ΔH за цикл и за минуту
            delta_h_cycle = h_din_pump - h_din_pause
            delta_h_per_min = delta_h_cycle / t_work if t_work > 0 else 0
            
            DEBUG.data("ΔH за цикл", f"{delta_h_cycle:.1f} м")
            DEBUG.data("ΔH за минуту", f"{delta_h_per_min:.3f} м/мин")
            
            # 6. Запас для увеличения
            current_diff = pump_depth - h_din_pump
            available_reserve = current_diff - self.MIN_DEPTH_DIFF
            
            DEBUG.data("Текущая разница", f"{current_diff:.0f} м")
            DEBUG.data("Доступный запас", f"{available_reserve:.0f} м = {current_diff:.0f} - {self.MIN_DEPTH_DIFF}")
            
            if available_reserve <= 0:
                DEBUG.log("❌ Нет запаса по уровню")
                result = {
                    'eligible': False,
                    'reason': f'Нет запаса по уровню. Текущая разница: {current_diff:.0f} м',
                    'filters': filter_check
                }
                DEBUG.exit()
                return result
            
            DEBUG.log(f"✅ Есть запас: {available_reserve:.0f} м")
            
            # 7. Новое время работы
            additional_minutes = available_reserve / delta_h_per_min if delta_h_per_min > 0 else 0
            new_t_work = t_work + additional_minutes
            
            DEBUG.data("Дополнительное время", f"{additional_minutes:.0f} мин = {available_reserve:.0f}/{delta_h_per_min:.3f}")
            DEBUG.data("Новое t_work", f"{new_t_work:.0f} мин")
            
            # 8. Новое время накопления (оставляем прежним, как в вашем примере)
            new_t_pause = t_pause
            DEBUG.data("Новое t_pause", f"{new_t_pause:.0f} мин (без изменений)")
            
            # 9. Новый дебит жидкости
            old_work_hours = (t_work / (t_work + t_pause)) * 24.0 if (t_work + t_pause) > 0 else 0
            new_work_hours = (new_t_work / (new_t_work + new_t_pause)) * 24.0 if (new_t_work + new_t_pause) > 0 else 0
            
            DEBUG.data("Рабочие часы", f"старые={old_work_hours:.1f}, новые={new_work_hours:.1f}")
            
            if old_work_hours > 0:
                new_q_daily = q_daily * (new_work_hours / old_work_hours)
                DEBUG.data("Коэффициент дебита", f"{new_work_hours/old_work_hours:.3f}")
            else:
                new_q_daily = q_daily
            
            DEBUG.data("Новый дебит", f"{new_q_daily:.1f} м³/сут")
            
            # 10. Экономический расчет
            freq = self.economic.safe_frequency(well_data.get('rotations_hz'))
            
            economic_result = self.economic.calculate_economic_effect_comprehensive(
                well_data=well_data,
                old_schedule=[t_work, t_pause],
                new_schedule=[new_t_work, new_t_pause],
                old_freq=freq,
                new_freq=freq,  # частоту не меняем
                has_gas_problem_old=False
            )
            
            # 11. Формируем результат
            result = {
                'eligible': True,
                'current_regime': {
                    't_work': t_work,
                    't_pause': t_pause,
                    'q_daily': q_daily,
                    'work_hours': old_work_hours,
                    'h_din_pump': h_din_pump,
                    'h_din_pause': h_din_pause,
                    'delta_h_per_min': delta_h_per_min
                },
                'new_regime': {
                    't_work': new_t_work,
                    't_pause': new_t_pause,
                    'q_daily': new_q_daily,
                    'work_hours': new_work_hours,
                    'additional_minutes': additional_minutes,
                    'available_reserve': available_reserve
                },
                'filters': filter_check,
                'economic': economic_result,
                'summary': {
                    'increase_q_percent': ((new_q_daily / q_daily) - 1) * 100 if q_daily > 0 else 0,
                    'daily_effect': economic_result.get('total_effect_per_day', 0),
                    'monthly_effect': economic_result.get('total_effect_per_month', 0),
                    'yearly_effect': economic_result.get('total_effect_per_year', 0)
                }
            }
            
            DEBUG.data("Прирост дебита", f"{result['summary']['increase_q_percent']:.1f}%")
            DEBUG.data("Эффект в сутки", f"{result['summary']['daily_effect']:.0f} ₽")
            DEBUG.log("✅ Анализ потенциала завершен успешно")
            DEBUG.exit()
            
            return result
            
        except Exception as e:
            DEBUG.log(f"Ошибка анализа потенциала: {str(e)}", "ERROR")
            import traceback
            DEBUG.log(traceback.format_exc(), "ERROR")
            DEBUG.exit()
            
            return {
                'eligible': False,
                'error': str(e),
                'traceback': traceback.format_exc()
            }

# ============================================================
# ФУНКЦИИ ДЛЯ ПАКЕТНОГО РАСЧЕТА
# ============================================================

def analyze_gas_problem_for_batch(physics, well, work_min, pause_min):
    """
    Анализ газовой проблемы для пакетного расчета
    Использует РАСЧИТАННОЕ p_zab для анализа!
    """
    DEBUG.enter()
    DEBUG.log(f"Пакетный анализ газа: {well.get('name', 'Unknown')}")
    
    result = {
        'has_gas_problem': False,
        'severity': 'none',
        'p_zab_min': 0,
        'p_zab_end': 0,
        'p_zab_start': 0,
        'p_zab_mid': 0
    }
    
    p_nas = physics.safe_float(well.get('p_nas', 0))
    if p_nas <= 0:
        DEBUG.log("Нет Pнас → анализ невозможен")
        DEBUG.exit()
        return result
    
    # Параметры для расчета Pзаб - используем РАСЧИТАННОЕ p_zab
    p_buffer = physics.safe_float(well.get('buffer_pressure', 0))
    pump_depth = physics.safe_float(well.get('pump_depth', 0))
    h_din = physics.safe_float(well.get('h_din', 0))
    q_current = physics.safe_float(well.get('flow_rate', 0))
    pump_head = well.get('pump_head', 1000)
    
    # Расчет мгновенного дебита
    work_hours = work_min / 60
    q_instant = q_current * (24 / work_hours) if work_hours > 0 else q_current
    
    DEBUG.data("Параметры", f"Pнас={p_nas:.1f}, L={pump_depth:.1f}, Hдин={h_din:.1f}, Q={q_current:.1f}")
    
    # 1. Pзаб в начале работы (насос ВЫКЛЮЧЕН) - РАСЧИТАННОЕ
    p_zab_start = physics.calculate_pwf_comprehensive_corrected(
        h_din, 0, pump_head, pump_depth, is_pump_working=False
    )
    
    # 2. Pзаб в конце работы (насос РАБОТАЕТ) - РАСЧИТАННОЕ
    p_zab_end = physics.calculate_pwf_comprehensive_corrected(
        h_din, q_instant, pump_head, pump_depth, is_pump_working=True
    )
    
    # 3. Pзаб в середине работы (для анализа) - РАСЧИТАННОЕ
    p_zab_mid = physics.calculate_pwf_comprehensive_corrected(
        h_din, q_instant * 0.5, pump_head, pump_depth, is_pump_working=True
    )
    
    result['p_zab_start'] = p_zab_start
    result['p_zab_end'] = p_zab_end
    result['p_zab_mid'] = p_zab_mid
    result['p_zab_min'] = min(p_zab_start, p_zab_end, p_zab_mid)
    
    DEBUG.data("Pзаб расчетные", f"начало={p_zab_start:.1f}, середина={p_zab_mid:.1f}, конец={p_zab_end:.1f}")
    
    # Анализ проблемы по РАСЧИТАННОМУ p_zab_end
    if p_zab_end < p_nas:
        result['has_gas_problem'] = True
        
        if physics.has_critical_gas_problem(p_zab_end, p_nas):  # ← используем РАСЧИТАННОЕ
            result['severity'] = 'critical'
            DEBUG.log(f"КРИТИЧЕСКАЯ газовая проблема: {p_zab_end:.1f} < 0.75*{p_nas:.1f}")
        elif p_zab_end < 0.9 * p_nas:
            result['severity'] = 'warning'
            DEBUG.log(f"СРЕДНЯЯ газовая проблема: {p_zab_end:.1f} < 0.9*{p_nas:.1f}")
        else:
            result['severity'] = 'mild'
            DEBUG.log(f"ЛЕГКАЯ газовая проблема: {p_zab_end:.1f} < {p_nas:.1f}")
    else:
        DEBUG.log("Газовой проблемы НЕТ")
    
    DEBUG.exit()
    return result

def analyze_pump_load_for_batch(economic_calc, well, work_min, pause_min):
    """
    Анализ загрузки насоса для пакетного расчета
    """
    DEBUG.enter()
    DEBUG.log(f"Пакетный анализ загрузки: {well.get('name', 'Unknown')}")
    
    result = {
        'k_util_current': 0,
        'problem': 'none'
    }
    
    # Данные
    q_tech = well.get('flow_rate', 0)
    q_pump = well.get('pump_flow', 60)
    freq_raw = well.get('rotations_hz')
    
    freq = economic_calc.safe_frequency(freq_raw)
    
    if q_pump <= 0:
        DEBUG.log("Ошибка: Qном ≤ 0")
        DEBUG.exit()
        return result
    
    # Расчет загрузки
    k_util = economic_calc.calculate_utilization_factor(
        q_tech, work_min, pause_min, q_pump, freq
    )
    
    result['k_util_current'] = k_util
    
    DEBUG.data("K_под", f"{k_util:.2f}")
    
    # Критерии
    if k_util < 0.75:
        result['problem'] = 'underload'
        DEBUG.log("НЕДОГРУЗКА")
    elif k_util > 1.25:
        result['problem'] = 'overload'
        DEBUG.log("ПЕРЕГРУЗКА")
    else:
        result['problem'] = 'optimal'
        DEBUG.log("ОПТИМАЛЬНО")
    
    DEBUG.exit()
    return result

@st.cache_data(ttl=3600, show_spinner="Анализ скважин...")
def analyze_well_for_batch(well_data: Dict, economic_params: Dict) -> Dict:
    """
    Анализ одной скважины для пакетного расчета
    Использует РАСЧИТАННОЕ p_zab для анализа газовой проблемы!
    """
    DEBUG.enter()
    DEBUG.section(f"ПАКЕТНЫЙ АНАЛИЗ СКВАЖИНЫ: {well_data.get('name', 'Unknown')}")
    
    try:
        # Инициализация классов
        physics = CorrectedKPRPhysics(well_data)
        economic_calc = EconomicCalculatorCorrected(
            oil_price_rub_per_ton=economic_params.get('oil_price_rub_per_ton', 50000.0),
            energy_price_rub_per_kwh=economic_params.get('energy_price_rub_per_kwh', 6.0)
        )
        optimizer = KPROptimizerCorrected(physics, economic_calc)
        
        # Получаем текущий график
        schedule = well_data.get('schedule', [15.0 * 60.0, 45.0 * 60.0])
        if schedule and len(schedule) >= 2:
            work_min_current = float(schedule[0])
            pause_min_current = float(schedule[1])
        else:
            work_min_current = 15.0 * 60.0
            pause_min_current = 45.0 * 60.0
        
        DEBUG.data("Текущий режим", f"{work_min_current:.0f}/{pause_min_current:.0f}")
        
        # Частота
        freq_raw = well_data.get('rotations_hz')
        freq_current = economic_calc.safe_frequency(freq_raw)
        
        DEBUG.data("Частота", f"{freq_current:.1f} Гц")
        
        # 1. Анализ газовой проблемы с использованием РАСЧИТАННОГО p_zab
        gas_analysis = physics.analyze_gas_problem_comprehensive(
            work_min_current, pause_min_current, freq_current
        )
        
        # 2. Расчет загрузки насоса
        q_tech = physics.safe_float(well_data.get('flow_rate', 0.0))
        q_pump = physics.safe_float(well_data.get('pump_flow', 60.0))
        k_util_current = economic_calc.calculate_utilization_factor(
            q_tech, work_min_current, pause_min_current, q_pump, freq_current
        )
        
        DEBUG.data("K_под текущий", f"{k_util_current:.2f}")
        
        # 3. Интегрированная оптимизация
        optimization_result = optimizer.optimize_integrated(
            work_min_current, pause_min_current, freq_current, include_freq_optimization=True
        )
        
        DEBUG.data("Сценарий оптимизации", optimization_result['scenario'])
        
        # 4. Экономический расчет
        economic_result = None
        if 'recommended_work_time' in optimization_result:
            has_gas_problem = gas_analysis['has_gas_problem']
            
            economic_result = economic_calc.calculate_economic_effect_comprehensive(
                well_data=well_data,
                old_schedule=[work_min_current, pause_min_current],
                new_schedule=[
                    optimization_result['recommended_work_time'],
                    optimization_result['recommended_pause_time']
                ],
                old_freq=freq_current,
                new_freq=optimization_result.get('recommended_freq', freq_current),
                has_gas_problem_old=has_gas_problem
            )
        
        # 5. Формирование результата
        result = {
            # Идентификационные данные
            'Скважина': well_data.get('name', 'Unknown'),
            'Куст': well_data.get('cluster', '-'),
            'ЦИТС': well_data.get('cits', '-'),
            'ЦДНГ': well_data.get('cdng', '-'),
            
            # Текущие параметры
            'Текущий режим': f"{int(work_min_current)}/{int(pause_min_current)}",
            'Дебит жидкости': q_tech,
            'Обводненность': well_data.get('water_cut', 0.0),
            'Pнас': physics.safe_float(well_data.get('p_nas', 0.0)),
            'Pпл': physics.safe_float(well_data.get('p_pl', 0.0)),
            'Pбуф': physics.safe_float(well_data.get('buffer_pressure', 0.0)),
            'Частота': freq_current,
            
            # Диагностика с использованием РАСЧИТАННОГО p_zab
            'Pзаб_конец': gas_analysis['p_zab_end'],  # ← РАСЧИТАННОЕ!
            'Газовая проблема': 'Да' if gas_analysis['has_gas_problem'] else 'Нет',
            'Степень газа': gas_analysis['severity'],
            'K_util_текущий': k_util_current,
            'Проблема загрузки': (
                'underload' if k_util_current < 0.75 else
                'overload' if k_util_current > 1.25 else
                'optimal'
            ),
            
            # Результаты оптимизации
            'Сценарий': optimization_result['scenario'],
            'Рекомендуемый режим': (
                f"{int(optimization_result['recommended_work_time'])}/{int(optimization_result['recommended_pause_time'])}"
                if 'recommended_work_time' in optimization_result else 'Нет'
            ),
            'Новая частота': optimization_result.get('recommended_freq', freq_current),
            'Причина': optimization_result.get('reason', ''),
            
            # Экономика
            'Эффект (₽/сут)': economic_result['total_effect_per_day'] if economic_result else 0.0,
            'Дебит нефти новый': economic_result['q_oil_new'] if economic_result else 0.0,
            'Энергия новая': economic_result['energy_new'] if economic_result else 0.0,
            'Мощность новая': economic_result['power_new'] if economic_result else 0.0,
            'Загрузка новая': economic_result['k_util_new'] if economic_result else 0.0,
            'Прибыльно': 'Да' if economic_result and economic_result['is_profitable'] else 'Нет',
            
            # Дополнительные данные
            'Тип_сценария': optimization_result['scenario'],
            
            # Детальные данные для отладки
            '_detailed': {
                'gas_analysis': gas_analysis,
                'optimization_result': optimization_result,
                'economic_result': economic_result
            }
        }
        
        DEBUG.data("Итог", f"Сценарий: {optimization_result['scenario']}, Эффект: {result['Эффект (₽/сут)']:.0f} ₽/сут")
        DEBUG.log("✅ Пакетный анализ завершен успешно")
        DEBUG.exit()
        
        return result
        
    except Exception as e:
        DEBUG.log(f"Ошибка пакетного анализа: {str(e)}", "ERROR")
        import traceback
        error_msg = str(e)
        
        DEBUG.exit()
        
        return {
            'Скважина': well_data.get('name', 'Error'),
            'Куст': well_data.get('cluster', '-'),
            'Текущий режим': '-',
            'Сценарий': 'Ошибка',
            'Причина': f"Ошибка расчета: {error_msg[:100]}",
            '_error': True,
            '_error_details': traceback.format_exc()
        }

def run_comprehensive_batch_optimization(filtered_wells, economic_params):
    """
    ПОЛНОЦЕННЫЙ пакетный расчет с использованием РАСЧИТАННОГО p_zab
    """

    _load_scipy()  
    _load_plotly() 

    DEBUG.enter()
    DEBUG.section(f"ПОЛНЫЙ ПАКЕТНЫЙ РАСЧЕТ: {len(filtered_wells)} скважин")
    
    batch_results = []
    detailed_results = []
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    for idx, well in enumerate(filtered_wells):
        try:
            status_text.text(f"Обработка скважины {idx+1}/{len(filtered_wells)}: {well['name']}")
            DEBUG.log(f"Скважина {idx+1}/{len(filtered_wells)}: {well['name']}")
            
            # Анализ скважины (использует РАСЧИТАННОЕ p_zab!)
            result = analyze_well_for_batch(well, economic_params)
            
            batch_results.append(result)
            detailed_results.append({
                'well': well['name'],
                'analysis_result': result,
                'detailed': result.get('_detailed', {})
            })
            
            progress_bar.progress((idx + 1) / len(filtered_wells))
            
        except Exception as e:
            DEBUG.log(f"Ошибка при обработке {well.get('name', 'Unknown')}: {str(e)}", "ERROR")
            import traceback
            error_details = traceback.format_exc()
            
            error_result = {
                'Скважина': well.get('name', 'Ошибка'),
                'Куст': well.get('cluster', '-'),
                'Текущий режим': '-',
                'Сценарий': 'Ошибка',
                'Причина': f"Ошибка расчета: {str(e)[:100]}",
                '_error': True,
                '_error_details': error_details
            }
            batch_results.append(error_result)
    
    progress_bar.empty()
    status_text.empty()
    
    DEBUG.log(f"Пакетный расчет завершен: {len(batch_results)} результатов")
    DEBUG.exit()
    
    # Убедитесь, что здесь данные сохраняются
    st.session_state.batch_results_advanced = batch_results
    st.session_state.batch_results_detailed = detailed_results
    st.session_state.full_batch_results = batch_results  
    
    save_data_to_file()
    
    return batch_results, detailed_results

@st.cache_data(ttl=3600, show_spinner="Пакетный анализ...")
def analyze_potential_batch(filtered_wells, economic_params, 
                           min_k_pod=1.25, min_depth_diff=400, 
                           min_k_prod=0.7, min_p_pr=40):
    """
    Пакетный анализ потенциала скважин с настраиваемыми фильтрами
    """
    DEBUG.enter()
    DEBUG.section(f"ПАКЕТНЫЙ АНАЛИЗ ПОТЕНЦИАЛА: {len(filtered_wells)} скважин")
    DEBUG.data("Параметры фильтров", f"Kпод>{min_k_pod}, ΔL-Hдин>{min_depth_diff}м, Kпр>{min_k_prod}, Pпр>{min_p_pr}атм")
    
    batch_results = []
    
    for well in filtered_wells:
        try:
            DEBUG.log(f"Анализ потенциала: {well.get('name', 'Unknown')}")
            
            # Инициализация классов с правильными параметрами фильтров
            physics = CorrectedKPRPhysics(well)
            economics = EconomicCalculatorCorrected(
                oil_price_rub_per_ton=economic_params.get('oil_price_rub_per_ton', 50000.0),
                energy_price_rub_per_kwh=economic_params.get('energy_price_rub_per_kwh', 6.0)
            )
            
            # Инициализация анализатора с настраиваемыми параметрами фильтров
            analyzer = WellPotentialAnalyzer(
                physics, economics,
                min_depth_diff=min_depth_diff,
                min_p_pr=min_p_pr,
                min_k_prod=min_k_prod
            )
            
            # Анализ потенциала с настраиваемым min_k_pod
            result = analyzer.analyze_potential_increase(well, min_k_pod=min_k_pod)
            
            # Формируем результат для таблицы
            table_row = {
                'Скважина': well.get('name', 'Unknown'),
                'Куст': well.get('cluster', '-'),
                'ЦИТС': well.get('cits', '-'),
                'ЦДНГ': well.get('cdng', '-'),
                'Проходит фильтры': 'Да' if result.get('eligible', False) else 'Нет',
                'Текущий режим': f"{result.get('current_regime', {}).get('t_work', 0):.0f}/{result.get('current_regime', {}).get('t_pause', 0):.0f}",
                'Текущий Q, м³/сут': result.get('current_regime', {}).get('q_daily', 0),
                'Новый режим': f"{result.get('new_regime', {}).get('t_work', 0):.0f}/{result.get('new_regime', {}).get('t_pause', 0):.0f}",
                'Новый Q, м³/сут': result.get('new_regime', {}).get('q_daily', 0),
                'Прирост Q, %': result.get('summary', {}).get('increase_q_percent', 0),
                'Эффект, ₽/сут': result.get('summary', {}).get('daily_effect', 0),
                'Эффект, ₽/мес': result.get('summary', {}).get('monthly_effect', 0),
                'Причина': result.get('reason', '') if not result.get('eligible', False) else 'Рекомендуется увеличение',
                '_detailed': result
            }
            
            batch_results.append(table_row)
            
            DEBUG.data(f"Результат {well.get('name')}", f"Проходит: {table_row['Проходит фильтры']}, Эффект: {table_row['Эффект, ₽/сут']:.0f} ₽")
            
        except Exception as e:
            DEBUG.log(f"Ошибка анализа потенциала {well.get('name', 'Unknown')}: {str(e)}", "ERROR")
            batch_results.append({
                'Скважина': well.get('name', 'Error'),
                'Куст': well.get('cluster', '-'),
                'Проходит фильтры': 'Ошибка',
                'Причина': f"Ошибка расчета: {str(e)[:100]}",
                '_error': True
            })
    
    DEBUG.log(f"Анализ потенциала завершен: {len(batch_results)} результатов")
    DEBUG.exit()
    
    return batch_results

# ============================================================
# ИНТЕГРАЦИОННЫЕ ФУНКЦИИ
# ============================================================

@st.cache_data(ttl=3600)
def get_well_diagnostics(well_data: Dict) -> Dict:
    """
    Полная диагностика скважины с использованием РАСЧИТАННОГО p_zab
    """
    DEBUG.enter()
    DEBUG.log(f"Диагностика скважины: {well_data.get('name', 'Unknown')}")
    
    physics = CorrectedKPRPhysics(well_data)
    economic_calc = EconomicCalculatorCorrected()
    
    # Базовые данные
    q_tech = physics.safe_float(well_data.get('flow_rate', 0.0))
    wct = physics.safe_float(well_data.get('water_cut', 0.0))
    p_pl = physics.safe_float(well_data.get('p_pl', 0.0))
    p_nas = physics.safe_float(well_data.get('p_nas', 0.0))
    p_buffer = physics.safe_float(well_data.get('buffer_pressure', 0.0))
    kpr = physics.safe_float(well_data.get('prod_coef', 0.0))
    pump_depth = physics.safe_float(well_data.get('pump_depth', 0.0))
    h_din = physics.safe_float(well_data.get('h_din', 0.0))
    pump_head = physics.safe_float(well_data.get('pump_head', 1000.0))
    q_pump = physics.safe_float(well_data.get('pump_flow', 60.0))
    
    DEBUG.data("Базовые параметры", f"Q={q_tech:.1f}, Обв={wct:.1f}%, Pпл={p_pl:.1f}, Pнас={p_nas:.1f}")
    DEBUG.data("Параметры скважины", f"Kпр={kpr:.2f}, L={pump_depth:.1f}, Hдин={h_din:.1f}, Hнасос={pump_head:.0f}")
    
    # Текущий режим
    schedule = well_data.get('schedule', [15.0 * 60.0, 45.0 * 60.0])
    if schedule and len(schedule) >= 2:
        work_min = float(schedule[0])
        pause_min = float(schedule[1])
    else:
        work_min = 15.0 * 60.0
        pause_min = 45.0 * 60.0
    
    freq_raw = well_data.get('rotations_hz')
    freq = economic_calc.safe_frequency(freq_raw)
    
    DEBUG.data("Режим", f"{work_min:.0f}/{pause_min:.0f}, f={freq:.1f} Гц")
    
    # Расчеты
    work_hours = (work_min / (work_min + pause_min)) * 24.0 if (work_min + pause_min) > 0 else 0.0
    q_instant = q_tech * (24.0 / work_hours) if work_hours > 0 else q_tech
    
    DEBUG.data("Рабочие часы", f"{work_hours:.1f} ч/сут, Qмгн={q_instant:.1f} м³/сут")
    
    # Оценка Pзаб в конце работы - РАСЧИТАННОЕ!
    p_zab_end_est = physics.calculate_pwf_comprehensive_corrected(
        h_din, q_instant, pump_head, pump_depth, is_pump_working=True
    )
    
    DEBUG.data("Pзаб расчетное", f"{p_zab_end_est:.1f} атм")
    
    # Загрузка насоса
    k_util = economic_calc.calculate_utilization_factor(q_tech, work_min, pause_min, q_pump, freq)
    
    DEBUG.data("K_под", f"{k_util:.2f}")
    
    # Анализ газа по РАСЧИТАННОМУ p_zab
    has_gas_problem = physics.has_gas_problem(p_zab_end_est, p_nas)
    has_critical_gas = physics.has_critical_gas_problem(p_zab_end_est, p_nas)
    
    DEBUG.data("Газовая проблема", "ЕСТЬ" if has_gas_problem else "НЕТ")
    DEBUG.data("Критическая", "ДА" if has_critical_gas else "НЕТ")
    
    # Разница глубины
    depth_diff = pump_depth - h_din
    
    result = {
        'q_tech': q_tech,
        'wct': wct,
        'p_pl': p_pl,
        'p_nas': p_nas,
        'p_buffer': p_buffer,
        'kpr': kpr,
        'pump_depth': pump_depth,
        'h_din': h_din,
        'depth_diff': depth_diff,
        'work_min': work_min,
        'pause_min': pause_min,
        'work_hours': work_hours,
        'freq': freq,
        'q_instant': q_instant,
        'p_zab_end_est': p_zab_end_est,  # ← РАСЧИТАННОЕ!
        'k_util': k_util,
        'has_gas_problem': has_gas_problem,  # ← на основе РАСЧИТАННОГО
        'has_critical_gas': has_critical_gas,  # ← на основе РАСЧИТАННОГО
        'can_increase_work_time': (
            kpr >= 0.8 or depth_diff >= 300.0
        ) if pump_depth > 0 else True
    }
    
    DEBUG.data("Можно увеличивать t_work", "ДА" if result['can_increase_work_time'] else "НЕТ")
    DEBUG.log("✅ Диагностика завершена")
    DEBUG.exit()
    
    return result

# ============================================================
# ФУНКЦИИ ДЛЯ Streamlit ИНТЕРФЕЙСА
# ============================================================

def plot_multiple_cycles(well_data, simulation_data, work_time, pause_time, p_nas, p_pl,
                         num_cycles=3, debit_units="м³/час"):
    """
    Строит два графика друг под другом для нескольких циклов
    
    Параметры:
    - well_data: данные скважины
    - simulation_data: данные симуляции одного цикла
    - work_time: время работы в минутах
    - pause_time: время накопления в минутах
    - p_nas: давление насыщения
    - p_pl: пластовое давление (используется для ограничения, но не отображается)
    - num_cycles: количество отображаемых циклов
    - debit_units: единицы измерения дебита
    """
    _load_plotly()
    
    # Данные для одного цикла из симуляции
    p_start = simulation_data['p_zab_start']  # Pзаб в начале работы
    p_end = simulation_data['p_zab_end']      # Pзаб в конце работы
    
    # ЕСЛИ НАЧАЛЬНОЕ ДАВЛЕНИЕ БОЛЬШЕ ПЛАСТОВОГО - ПРИРАВНИВАЕМ
    if p_pl and p_pl > 0 and p_start > p_pl:
        p_start = p_pl  # ← ТИХО ПОДМЕНЯЕМ
    
    # Общее время для нескольких циклов
    cycle_total = work_time + pause_time
    total_minutes = num_cycles * cycle_total
    time_points_min = np.arange(0, total_minutes, 1)  # Шаг 1 минута
    
    # Массивы для результатов
    pwf_points = []
    q_inst_points = []
    
    # Текущие параметры скважины
    q_liq_day = well_data.get('flow_rate', 0)  # м³/сут
    
    # Мгновенный дебит во время работы
    duty_cycle = work_time / cycle_total if cycle_total > 0 else 1
    q_instant_day = q_liq_day / duty_cycle  # м³/сут (мгновенный)
    q_instant_hour = q_instant_day / 24  # м³/час
    
    # Конвертируем в выбранные единицы
    if debit_units == "м³/час":
        q_instant_display = q_instant_hour
        q_avg_display = q_liq_day / 24
        y_axis_title = "Дебит, м³/час"
    else:  # м³/сут
        q_instant_display = q_instant_day
        q_avg_display = q_liq_day
        y_axis_title = "Дебит, м³/сут"
    
    # Генерируем данные для каждого момента времени
    for t in time_points_min:
        # Определяем номер текущего цикла и позицию в нем
        cycle_num = t // cycle_total
        pos_in_cycle = t % cycle_total
        
        if pos_in_cycle < work_time:
            # Фаза работы - равномерное снижение давления
            progress = pos_in_cycle / work_time  # от 0 до 1
            p_current = p_start + (p_end - p_start) * progress
            pwf_points.append(p_current)
            q_inst_points.append(q_instant_display)
        else:
            # Фаза накопления - равномерное восстановление
            progress = (pos_in_cycle - work_time) / pause_time  # от 0 до 1
            # Восстанавливаемся от p_end до p_start
            p_current = p_end + (p_start - p_end) * progress
            pwf_points.append(p_current)
            q_inst_points.append(0)
    
    # Создаем фигуру с двумя подграфиками
    from plotly.subplots import make_subplots
    
    fig = make_subplots(
        rows=2, cols=1,
        shared_xaxes=True,
        vertical_spacing=0.1,
        subplot_titles=('Забойное давление (Pзаб)', f'Дебит жидкости ({debit_units})'),
        row_heights=[0.6, 0.4]
    )
    
    # ========== ВЕРХНИЙ ГРАФИК: Забойное давление ==========
    # Основная линия Pзаб
    fig.add_trace(
        go.Scatter(
            x=time_points_min,
            y=pwf_points,
            mode='lines',
            name='Pзаб',
            line=dict(color='red', width=2),
            hovertemplate='Время: %{x:.0f} мин<br>Pзаб: %{y:.1f} атм<extra></extra>'
        ),
        row=1, col=1
    )
    
    # Линия 0.75 * Pнас (критический уровень)
    if p_nas and p_nas > 0:
        p_critical = 0.75 * p_nas
        fig.add_trace(
            go.Scatter(
                x=[0, total_minutes],
                y=[p_critical, p_critical],
                mode='lines',
                name=f'0.75·Pнас ({p_critical:.1f} атм)',
                line=dict(color='orange', width=2, dash='dash'),
                hovertemplate='Критический уровень: %{y:.1f} атм<extra></extra>'
            ),
            row=1, col=1
        )
    
    # Добавляем точки начала и конца циклов для наглядности
    for cycle in range(num_cycles):
        start_time = cycle * cycle_total
        end_work_time = start_time + work_time
        end_cycle_time = start_time + cycle_total
        
        # Точка начала цикла
        fig.add_trace(
            go.Scatter(
                x=[start_time],
                y=[p_start],
                mode='markers',
                marker=dict(size=8, color='darkred', symbol='circle'),
                name='Начало цикла' if cycle == 0 else '',
                showlegend=(cycle == 0),
                hovertemplate='Начало цикла: %{y:.1f} атм<extra></extra>'
            ),
            row=1, col=1
        )
        
        # Точка конца работы
        fig.add_trace(
            go.Scatter(
                x=[end_work_time],
                y=[p_end],
                mode='markers',
                marker=dict(size=8, color='darkorange', symbol='circle'),
                name='Конец работы' if cycle == 0 else '',
                showlegend=(cycle == 0),
                hovertemplate='Конец работы: %{y:.1f} атм<extra></extra>'
            ),
            row=1, col=1
        )
    
    # Вертикальные линии разделения циклов
    for cycle in range(1, num_cycles):
        x_pos = cycle * cycle_total
        fig.add_vline(
            x=x_pos, 
            line_width=1, 
            line_dash="dot", 
            line_color="gray", 
            opacity=0.5,
            row=1, col=1
        )
    
    fig.update_yaxes(title_text="Давление, атм", row=1, col=1)
    
    # ========== НИЖНИЙ ГРАФИК: Дебит ==========
    # Мгновенный дебит с заливкой
    fig.add_trace(
        go.Scatter(
            x=time_points_min,
            y=q_inst_points,
            mode='lines',
            name=f'Qж мгновенный ({debit_units})',
            line=dict(color='blue', width=2),
            fill='tozeroy',
            fillcolor='rgba(0, 0, 255, 0.1)',
            hovertemplate='Время: %{x:.0f} мин<br>Qж: %{y:.1f} ' + debit_units + '<extra></extra>'
        ),
        row=2, col=1
    )
    
    # Среднесуточный дебит (пунктирная линия)
    fig.add_trace(
        go.Scatter(
            x=[0, total_minutes],
            y=[q_avg_display, q_avg_display],
            mode='lines',
            name=f'Qж ср. ({q_avg_display:.1f} {debit_units})',
            line=dict(color='green', width=2, dash='dash'),
            hovertemplate='Среднесуточный: %{y:.1f} ' + debit_units + '<extra></extra>'
        ),
        row=2, col=1
    )
    
    # Вертикальные линии разделения циклов на нижнем графике
    for cycle in range(1, num_cycles):
        x_pos = cycle * cycle_total
        fig.add_vline(
            x=x_pos, 
            line_width=1, 
            line_dash="dot", 
            line_color="gray", 
            opacity=0.5,
            row=2, col=1
        )
    
    # Подписи осей
    fig.update_xaxes(title_text="Время, минуты", row=2, col=1)
    fig.update_yaxes(title_text=y_axis_title, row=2, col=1)
    
    # Общие настройки
    fig.update_layout(
        title=f'Динамика работы скважины (показано {num_cycles} цикла{"ов" if num_cycles>4 else ""})',
        hovermode='x unified',
        height=650,
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
    )
    
    # Добавляем аннотации с временем работы/паузы
    for cycle in range(min(num_cycles, 3)):  # Показываем только первые 3 цикла, чтобы не загромождать
        start = cycle * cycle_total
        mid_work = start + work_time/2
        mid_pause = start + work_time + pause_time/2
        
        # Аннотация для работы
        fig.add_annotation(
            x=mid_work,
            y=max(pwf_points) * 0.9,
            text=f"Работа<br>{work_time} мин",
            showarrow=False,
            font=dict(size=9, color="blue"),
            row=1, col=1
        )
        
        # Аннотация для паузы
        fig.add_annotation(
            x=mid_pause,
            y=max(pwf_points) * 0.9,
            text=f"Пауза<br>{pause_time} мин",
            showarrow=False,
            font=dict(size=9, color="gray"),
            row=1, col=1
        )
    
    return fig

def show_optimization_results_corrected(result, economic_params=None):
    """Отображение результатов оптимизации с расширенной экономикой"""
    DEBUG.enter()
    DEBUG.section("ОТОБРАЖЕНИЕ РЕЗУЛЬТАТОВ ОПТИМИЗАЦИИ")
    
    import streamlit as st
    
    if economic_params is None:
        economic_params = {
            'oil_price_rub_per_ton': 50000,
            'energy_price_rub_per_kwh': 6.0,
            'days_per_month': 30
        }

    # Основные метрики - сначала проверяем, какие ключи есть в result
    col1, col2, col3 = st.columns(3)
    
    with col1:
        # Проверяем разные возможные ключи для времени работы
        work_time_key = None
        current_work_key = None
        
        # Ищем ключи для рекомендованного времени работы
        possible_work_keys = ['recommended_work_time', 'new_work_time', 'work_time_new']
        for key in possible_work_keys:
            if key in result:
                work_time_key = key
                break
        
        # Ищем ключи для текущего времени работы
        possible_current_keys = ['current_work_time', 'work_time_current', 'current_work_min']
        for key in possible_current_keys:
            if key in result:
                current_work_key = key
                break
        
        if work_time_key and current_work_key:
            delta_work = result[work_time_key] - result[current_work_key]
            st.metric(
                "Время работы",
                f"{int(result[work_time_key])} мин",
                delta=f"{int(delta_work)} мин"
            )
            DEBUG.data("Время работы", f"{result[work_time_key]:.0f} мин (Δ={delta_work:.0f})")
        elif work_time_key:
            st.metric(
                "Время работы",
                f"{int(result[work_time_key])} мин"
            )
            DEBUG.data("Время работы", f"{result[work_time_key]:.0f} мин")
        else:
            st.metric("Время работы", "Нет данных")
            DEBUG.data("Время работы", "Нет данных")
    
    with col2:
        # Аналогично для времени накопления
        pause_time_key = None
        current_pause_key = None
        
        # Ищем ключи для рекомендованного времени накопления
        possible_pause_keys = ['recommended_pause_time', 'new_pause_time', 'pause_time_new']
        for key in possible_pause_keys:
            if key in result:
                pause_time_key = key
                break
        
        # Ищем ключи для текущего времени накопления
        possible_current_pause_keys = ['current_pause_time', 'pause_time_current', 'current_pause_min']
        for key in possible_current_pause_keys:
            if key in result:
                current_pause_key = key
                break
        
        if pause_time_key and current_pause_key:
            delta_pause = result[pause_time_key] - result[current_pause_key]
            st.metric(
                "Время накопления",
                f"{int(result[pause_time_key])} мин",
                delta=f"{int(delta_pause)} мин"
            )
            DEBUG.data("Время накопления", f"{result[pause_time_key]:.0f} мин (Δ={delta_pause:.0f})")
        elif pause_time_key:
            st.metric(
                "Время накопления",
                f"{int(result[pause_time_key])} мин"
            )
            DEBUG.data("Время накопления", f"{result[pause_time_key]:.0f} мин")
        else:
            st.metric("Время накопления", "Нет данных")
            DEBUG.data("Время накопления", "Нет данных")
    
    with col3:
        # Для частоты
        freq_key = None
        current_freq_key = None
        
        possible_freq_keys = ['recommended_freq', 'new_freq', 'freq_new']
        for key in possible_freq_keys:
            if key in result:
                freq_key = key
                break
        
        possible_current_freq_keys = ['current_freq', 'freq_current', 'freq']
        for key in possible_current_freq_keys:
            if key in result:
                current_freq_key = key
                break
        
        if freq_key and current_freq_key:
            delta_freq = result[freq_key] - result[current_freq_key]
            st.metric(
                "Частота",
                f"{result[freq_key]:.1f} Гц",
                delta=f"{delta_freq:.1f} Гц"
            )
            DEBUG.data("Частота", f"{result[freq_key]:.1f} Гц (Δ={delta_freq:.1f})")
        elif freq_key:
            st.metric(
                "Частота",
                f"{result[freq_key]:.1f} Гц"
            )
            DEBUG.data("Частота", f"{result[freq_key]:.1f} Гц")
        else:
            # Если нет частоты, показываем K_под
            k_util_key = None
            current_k_key = None
            
            possible_k_keys = ['k_pod_new', 'k_util_new', 'k_pod_after']
            for key in possible_k_keys:
                if key in result:
                    k_util_key = key
                    break
            
            possible_current_k_keys = ['k_pod_current', 'k_util_current', 'k_pod_before']
            for key in possible_current_k_keys:
                if key in result:
                    current_k_key = key
                    break
            
            if k_util_key and current_k_key:
                delta_k = result[k_util_key] - result[current_k_key]
                st.metric(
                    "Загрузка",
                    f"{result[k_util_key]:.2f}",
                    delta=f"{delta_k:+.2f}"
                )
                DEBUG.data("K_под", f"{result[k_util_key]:.2f} (Δ={delta_k:+.2f})")
            elif k_util_key:
                st.metric(
                    "Загрузка",
                    f"{result[k_util_key]:.2f}"
                )
                DEBUG.data("K_под", f"{result[k_util_key]:.2f}")
            else:
                st.metric("Параметр", "Нет данных")
                DEBUG.data("Параметр", "Нет данных")
    
    # Причина
    if 'reason' in result:
        st.info(f"**Причина:** {result['reason']}")
        DEBUG.data("Причина", result['reason'])
    elif 'optimization_method' in result:
        st.info(f"**Метод оптимизации:** {result['optimization_method']}")
        DEBUG.data("Метод оптимизации", result['optimization_method'])
    
    # Экономический анализ (расширенный)
    if 'economic_analysis' in result:
        econ = result['economic_analysis']
        
        st.markdown("##### 💰 Экономический анализ")
        DEBUG.log("Начинаем экономический анализ...")
        
        # Создаем вкладки для разных периодов
        tab1, tab2, tab3, tab4 = st.tabs(["📈 Суточная экономика", "📅 Месячная", "📊 Годовая", "📋 Подробно"])
        
        with tab1:
            # Суточная экономика
            col_day1, col_day2, col_day3 = st.columns(3)
            
            with col_day1:
                if 'q_oil_new' in econ and 'delta_q_oil' in econ:
                    st.metric(
                        "Дебит нефти",
                        f"{econ['q_oil_new']:.2f} т/сут",
                        delta=f"{econ['delta_q_oil']:+.2f} т/сут"
                    )
                    DEBUG.data("Дебит нефти", f"{econ['q_oil_new']:.2f} т/сут (Δ={econ['delta_q_oil']:+.2f})")
                elif 'q_oil_new' in econ:
                    st.metric(
                        "Дебит нефти",
                        f"{econ['q_oil_new']:.2f} т/сут"
                    )
                    DEBUG.data("Дебит нефти", f"{econ['q_oil_new']:.2f} т/сут")
            
            with col_day2:
                if 'energy_new' in econ and 'delta_energy' in econ:
                    st.metric(
                        "Энергопотребление",
                        f"{econ['energy_new']:.0f} кВт·ч/сут",
                        delta=f"{econ['delta_energy']:+.0f} кВт·ч/сут"
                    )
                    DEBUG.data("Энергопотребление", f"{econ['energy_new']:.0f} кВт·ч/сут (Δ={econ['delta_energy']:+.0f})")
                elif 'energy_new' in econ:
                    st.metric(
                        "Энергопотребление",
                        f"{econ['energy_new']:.0f} кВт·ч/сут"
                    )
                    DEBUG.data("Энергопотребление", f"{econ['energy_new']:.0f} кВт·ч/сут")
            
            with col_day3:
                if 'total_effect_per_day' in econ:
                    profit_color = "normal" if econ.get('is_profitable', False) else "inverse"
                    delta_text = "Прибыльно" if econ.get('is_profitable', False) else "Убыточно"
                    st.metric(
                        "Экономический эффект",
                        f"{econ['total_effect_per_day']:+.0f} ₽/сут",
                        delta=delta_text,
                        delta_color=profit_color
                    )
                    DEBUG.data("Эффект в сутки", f"{econ['total_effect_per_day']:+.0f} ₽ ({delta_text})")
            
            # Диаграмма распределения
            if 'oil_revenue_new' in econ and 'energy_cost_new' in econ and 'total_effect_per_day' in econ:
                st.markdown("**Распределение доходов и затрат (₽/сут):**")
                
                # Данные для диаграммы
                labels = ['Доход от нефти', 'Затраты на энергию', 'Чистый эффект']
                values = [
                    abs(econ['oil_revenue_new']), 
                    abs(econ['energy_cost_new']), 
                    abs(econ['total_effect_per_day'])
                ]
                colors = ['#2E7D32', '#C62828', '#1565C0']
                
                fig = go.Figure(data=[
                    go.Bar(
                        x=labels,
                        y=values,
                        marker_color=colors,
                        text=[f"{v:.0f} ₽" for v in values],
                        textposition='auto',
                    )
                ])
                
                fig.update_layout(
                    title="Экономика на сутки",
                    yaxis_title="₽/сут",
                    height=300,
                    showlegend=False
                )
                
                st.plotly_chart(fig, use_container_width=True)
                DEBUG.log("Построена диаграмма суточной экономики")
        
        with tab2:
            # Месячная экономика (30 дней)
            if 'total_effect_per_day' in econ:
                days_in_month = 30
                
                col_month1, col_month2, col_month3 = st.columns(3)
                
                with col_month1:
                    if 'q_oil_new' in econ and 'delta_q_oil' in econ:
                        monthly_oil = econ['q_oil_new'] * days_in_month
                        monthly_delta_oil = econ['delta_q_oil'] * days_in_month
                        st.metric(
                            "Добыча нефти",
                            f"{monthly_oil:.1f} т/мес",
                            delta=f"{monthly_delta_oil:+.1f} т/мес"
                        )
                        DEBUG.data("Добыча нефти в месяц", f"{monthly_oil:.1f} т (Δ={monthly_delta_oil:+.1f})")
                
                with col_month2:
                    if 'energy_new' in econ and 'delta_energy' in econ:
                        monthly_energy = econ['energy_new'] * days_in_month
                        monthly_delta_energy = econ['delta_energy'] * days_in_month
                        st.metric(
                            "Энергопотребление",
                            f"{monthly_energy:.0f} кВт·ч/мес",
                            delta=f"{monthly_delta_energy:+.0f} кВт·ч/мес"
                        )
                        DEBUG.data("Энергия в месяц", f"{monthly_energy:.0f} кВт·ч (Δ={monthly_delta_energy:+.0f})")
                
                with col_month3:
                    monthly_effect = econ['total_effect_per_day'] * days_in_month
                    st.metric(
                        "Экономический эффект",
                        f"{monthly_effect:+.0f} ₽/мес",
                        delta=f"{econ['total_effect_per_day']:+.0f} ₽/сут"
                    )
                    DEBUG.data("Эффект в месяц", f"{monthly_effect:+.0f} ₽")
                
                # Круговой график затрат
                if 'oil_revenue_new' in econ and 'energy_cost_new' in econ:
                    st.markdown("**Структура месячной экономики:**")
                    
                    monthly_revenue = econ['oil_revenue_new'] * days_in_month
                    monthly_cost = econ['energy_cost_new'] * days_in_month
                    monthly_net = monthly_effect
                    
                    pie_data = {
                        'Категория': ['Доход от нефти', 'Затраты на энергию', 'Чистая прибыль'],
                        'Сумма': [monthly_revenue, monthly_cost, monthly_net],
                        'Цвет': ['#2E7D32', '#C62828', '#1565C0']
                    }
                    
                    pie_fig = go.Figure(data=[
                        go.Pie(
                            labels=pie_data['Категория'],
                            values=[abs(v) for v in pie_data['Сумма']],
                            marker=dict(colors=pie_data['Цвет']),
                            textinfo='label+percent',
                            textposition='inside',
                            hole=0.3,
                        )
                    ])
                    
                    pie_fig.update_layout(
                        title=f"Месячная экономика: {monthly_effect:+.0f} ₽",
                        height=400
                    )
                    
                    st.plotly_chart(pie_fig, use_container_width=True)
                    DEBUG.log("Построена круговая диаграмма месячной экономики")
        
        with tab3:
            # Годовая экономика (365 дней)
            if 'total_effect_per_day' in econ:
                days_in_year = 365
                
                col_year1, col_year2, col_year3 = st.columns(3)
                
                with col_year1:
                    if 'q_oil_new' in econ and 'delta_q_oil' in econ:
                        yearly_oil = econ['q_oil_new'] * days_in_year
                        yearly_delta_oil = econ['delta_q_oil'] * days_in_year
                        st.metric(
                            "Добыча нефти",
                            f"{yearly_oil:,.1f} т/год",
                            delta=f"{yearly_delta_oil:+,.1f} т/год"
                        )
                        DEBUG.data("Добыча нефти в год", f"{yearly_oil:,.1f} т (Δ={yearly_delta_oil:+,.1f})")
                
                with col_year2:
                    if 'energy_new' in econ and 'delta_energy' in econ:
                        yearly_energy = econ['energy_new'] * days_in_year
                        yearly_delta_energy = econ['delta_energy'] * days_in_year
                        st.metric(
                            "Энергопотребление",
                            f"{yearly_energy:.0f} кВт·ч/год",
                            delta=f"{yearly_delta_energy:+.0f} кВт·ч/год"
                        )
                        DEBUG.data("Энергия в год", f"{yearly_energy:.0f} кВт·ч (Δ={yearly_delta_energy:+.0f})")
                
                with col_year3:
                    yearly_effect = econ['total_effect_per_day'] * days_in_year
                    monthly_avg = yearly_effect / 12
                    st.metric(
                        "Экономический эффект",
                        f"{yearly_effect:+.0f} ₽/год",
                        delta=f"≈ {monthly_avg:+.0f} ₽/мес в среднем"
                    )
                    DEBUG.data("Эффект в год", f"{yearly_effect:+.0f} ₽ (в среднем {monthly_avg:+.0f} ₽/мес)")
                
                # Линейный график накопленного эффекта
                if yearly_effect > 0:
                    st.markdown("**Накопленный экономический эффект:**")
                    
                    months = list(range(1, 13))
                    cumulative_effect = [econ['total_effect_per_day'] * 30 * m for m in months]
                    
                    cum_fig = go.Figure()
                    
                    cum_fig.add_trace(go.Scatter(
                        x=months,
                        y=cumulative_effect,
                        mode='lines+markers',
                        name='Накопленный эффект',
                        line=dict(color='#1565C0', width=3),
                        marker=dict(size=8)
                    ))
                    
                    cum_fig.add_trace(go.Scatter(
                        x=months,
                        y=[0] * 12,
                        mode='lines',
                        name='Нулевая линия',
                        line=dict(color='gray', dash='dash', width=1),
                        showlegend=False
                    ))
                    
                    # Добавляем аннотации для ключевых месяцев
                    for month in [3, 6, 9, 12]:
                        cum_fig.add_annotation(
                            x=month,
                            y=cumulative_effect[month-1],
                            text=f"{cumulative_effect[month-1]:+.0f} ₽",
                            showarrow=True,
                            arrowhead=2,
                            ax=0,
                            ay=-40
                        )
                    
                    cum_fig.update_layout(
                        title="Накопленный эффект за год",
                        xaxis_title="Месяц",
                        yaxis_title="Накопленный эффект, ₽",
                        height=400,
                        showlegend=True
                    )
                    
                    st.plotly_chart(cum_fig, use_container_width=True)
                    DEBUG.log("Построен график накопленного эффекта")
                    
                    # Дополнительная статистика
                    col_year_stats1, col_year_stats2 = st.columns(2)
                    
                    with col_year_stats1:
                        st.markdown("**Годовая статистика:**")
                        st.write(f"- Среднемесячный эффект: **{monthly_avg:+.0f} ₽**")
                        if 'oil_revenue_new' in econ:
                            st.write(f"- Годовой доход от нефти: **{econ['oil_revenue_new'] * days_in_year:.0f} ₽**")
                        if 'energy_cost_new' in econ:
                            st.write(f"- Годовые затраты на энергию: **{econ['energy_cost_new'] * days_in_year:.0f} ₽**")
                    
                    with col_year_stats2:
                        st.markdown("**Эффективность:**")
                        if 'q_oil_new' in econ and econ['q_oil_new'] > 0 and 'energy_cost_new' in econ:
                            st.write(f"- Удельные затраты на нефть: **{econ['energy_cost_new'] / econ['q_oil_new']:.1f} ₽/т**")
                        if 'energy_new' in econ and econ['energy_new'] > 0 and 'energy_cost_new' in econ:
                            st.write(f"- Себестоимость энергии: **{econ['energy_cost_new'] / econ['energy_new']:.2f} ₽/кВт·ч**")
                        if 'oil_revenue_new' in econ and econ['oil_revenue_new'] * days_in_year > 0:
                            st.write(f"- Рентабельность: **{(yearly_effect / (econ['oil_revenue_new'] * days_in_year) * 100):.1f}%**")
        
        with tab4:
            # Подробная таблица
            st.markdown("**Подробный расчет:**")
            
            # Проверяем какие ключи есть в econ
            available_keys = list(econ.keys())
            DEBUG.data("Доступные ключи экономики", f"{len(available_keys)} шт.")
            
            # Создаем таблицу только с доступными данными
            detailed_rows = []
            
            # Добавляем строки в зависимости от доступных данных
            if 'q_oil_new' in econ:
                detailed_rows.append(('Суточный дебит нефти, т/сут', econ.get('q_oil_old', 'Н/Д'), f"{econ['q_oil_new']:.2f}"))
            if 'q_tech_current' in econ and 'q_tech_new' in econ:
                detailed_rows.append(('Суточный дебит жидкости, м³/сут', f"{econ['q_tech_current']:.1f}", f"{econ['q_tech_new']:.1f}"))
            
            detailed_rows.append(('Цена нефти, ₽/т', f"{economic_params['oil_price_rub_per_ton']:.0f}", f"{economic_params['oil_price_rub_per_ton']:.0f}"))
            
            if 'oil_revenue_old' in econ and 'oil_revenue_new' in econ:
                detailed_rows.append(('Доход от нефти, ₽/сут', f"{econ['oil_revenue_old']:.0f}", f"{econ['oil_revenue_new']:.0f}"))
            
            if 'power_old' in econ and 'power_new' in econ:
                detailed_rows.append(('Мощность насоса, кВт', f"{econ['power_old']:.1f}", f"{econ['power_new']:.1f}"))
            
            if 'work_hours_old' in econ and 'work_hours_new' in econ:
                detailed_rows.append(('Время работы, ч/сут', f"{econ['work_hours_old']:.1f}", f"{econ['work_hours_new']:.1f}"))
            
            if 'energy_old' in econ and 'energy_new' in econ:
                detailed_rows.append(('Энергопотребление, кВт·ч/сут', f"{econ['energy_old']:.0f}", f"{econ['energy_new']:.0f}"))
            
            detailed_rows.append(('Тариф на энергию, ₽/кВт·ч', f"{economic_params['energy_price_rub_per_kwh']:.2f}", f"{economic_params['energy_price_rub_per_kwh']:.2f}"))
            
            if 'energy_cost_old' in econ and 'energy_cost_new' in econ:
                detailed_rows.append(('Затраты на энергию, ₽/сут', f"{econ['energy_cost_old']:.0f}", f"{econ['energy_cost_new']:.0f}"))
            
            if 'k_util_old' in econ and 'k_util_new' in econ:
                detailed_rows.append(('Коэффициент загрузки', f"{econ['k_util_old']:.2f}", f"{econ['k_util_new']:.2f}"))
            
            if 'q_oil_old' in econ and econ['q_oil_old'] > 0 and 'energy_old' in econ:
                detailed_rows.append(('Удельные затраты энергии, кВт·ч/т', f"{econ['energy_old']/econ['q_oil_old']:.1f}", 
                                   f"{econ.get('energy_new', 0)/econ.get('q_oil_new', 1):.1f}" if econ.get('q_oil_new', 0) > 0 else "0"))
            
            if 'total_effect_per_day' in econ:
                detailed_rows.append(('Экономический эффект, ₽/сут', 
                                   f"{(econ.get('total_effect_per_day', 0) - econ.get('delta_revenue', 0) + econ.get('delta_cost', 0)):.0f}" 
                                   if 'delta_revenue' in econ and 'delta_cost' in econ else "Н/Д", 
                                   f"{econ['total_effect_per_day']:.0f}"))
            
            if detailed_rows:
                import pandas as pd
                df_detailed = pd.DataFrame(detailed_rows, columns=['Показатель', 'Старый режим', 'Новый режим'])
                st.dataframe(df_detailed, use_container_width=True, hide_index=True)
                
                # Кнопка скачивания
                csv = df_detailed.to_csv(index=False, encoding='utf-8-sig')
                st.download_button(
                    label="📥 Скачать подробный расчет",
                    data=csv,
                    file_name="детальный_экономический_расчет.csv",
                    mime="text/csv",
                    use_container_width=True
                )
                DEBUG.log("Создана подробная таблица экономики")
            else:
                st.warning("Нет данных для отображения подробной таблицы")
                DEBUG.log("Нет данных для подробной таблицы")
    
    else:
        # Если нет экономического анализа, показываем базовую информацию
        DEBUG.log("Нет данных экономического анализа")
        with st.expander("📊 Детали оптимизации"):
            # Выводим все ключи result для отладки
            st.write("**Доступные данные:**")
            for key, value in result.items():
                if not key.startswith('_') and not callable(value):
                    if isinstance(value, (int, float)):
                        st.write(f"- {key}: {value}")
                    else:
                        st.write(f"- {key}: {str(value)[:100]}...")
    
    DEBUG.log("✅ Отображение результатов завершено")
    DEBUG.exit()

def show_kpr_potential_tab_corrected():
    """
    Основная функция для отображения интерфейса оптимизации КПР
    Включает все три сценария: A, B и Потенциал
    """

    _load_scipy() 
    _load_plotly() 
    
    DEBUG.enter()
    DEBUG.section("ЗАПУСК ВКЛАДКИ ОПТИМИЗАЦИИ КПР")
    
    import streamlit as st

    def clear_previous_batch_results():
        """Очистка результатов предыдущих пакетных расчетов"""
        DEBUG.enter()
        DEBUG.log("Очистка предыдущих результатов пакетного расчета")
        
        keys_to_remove = [
            'full_batch_results',
            'full_batch_detailed', 
            'potential_batch_results'
        ]
        
        for key in keys_to_remove:
            if key in st.session_state:
                DEBUG.data(f"Удаляю ключ", key)
                del st.session_state[key]
    
    DEBUG.exit()
    
    st.markdown("### 🎯 Корректировка режимов работ скважин КПР")
    st.info("""
    **Три сценария оптимизации:**
    1. **Сценарий A (Содержание газа на приеме)** - Оптимизация при Pзаб < Pнас
    2. **Сценарий B (Коэффициент подачи)** - Оптимизация K_под в диапазон 0.75-1.25
    3. **Сценарий "Потенциал"** - Поиск скважин для "увеличения Tраб" (увеличения дебита)
    """)
    
    # Проверка наличия данных
    if not st.session_state.get('wells_data'):
        DEBUG.log("Нет данных скважин в session_state")
        st.warning("Сначала загрузите данные скважин через раздел 'Импорт техрежима'")
        DEBUG.exit()
        return
    
    # Получаем КПР скважины
    kpr_wells = [w for w in st.session_state.wells_data 
                if w.get('operation_mode') == 'kpr' and w.get('is_active', True)]
    
    DEBUG.data("Всего КПР скважин", f"{len(kpr_wells)}")
    
    if not kpr_wells:
        DEBUG.log("Нет активных КПР скважин")
        st.warning("В системе нет активных КПР скважин")
        DEBUG.exit()
        return
    
    # ============================
    # ШАГ 1: ВЫБОР ЦИТС И РЕЖИМА
    # ============================
    st.markdown("#### Шаг 1: Выбор ЦИТС и режима расчета")
    
    col_select1, col_select2 = st.columns(2)
    
    with col_select1:
        cits_list = list(set([w.get('cits', 'ЦИТС VQ-BAD') for w in kpr_wells]))
        selected_cits = st.selectbox("Выберите ЦИТС", cits_list, key="kpr_cits_select_corrected")
        DEBUG.data("Выбран ЦИТС", selected_cits)
    
    with col_select2:
        cdng_list = list(set([w.get('cdng') for w in kpr_wells 
                             if w.get('cits') == selected_cits and w.get('cdng')]))
        
        if cdng_list:
            selected_cdng = st.selectbox(
                "Выберите ЦДНГ (опционально)",
                ["Все ЦДНГ"] + cdng_list,
                key="cdng_select_corrected"
            )
            DEBUG.data("Выбран ЦДНГ", selected_cdng)
        else:
            selected_cdng = "Все ЦДНГ"
            DEBUG.data("ЦДНГ", "Все")
    
    calculation_mode = st.radio(
        "Режим расчета",
        ["Одна скважина", "Группа скважин"],
        horizontal=True,
        key="calculation_mode_corrected"
    )
    
    DEBUG.data("Режим расчета", calculation_mode)
    
    # Фильтрация скважин
    filtered_wells = []
    for w in kpr_wells:
        if w.get('cits') != selected_cits:
            continue
        if selected_cdng != "Все ЦДНГ" and w.get('cdng') != selected_cdng:
            continue
        filtered_wells.append(w)
    
    DEBUG.data("Отфильтровано скважин", f"{len(filtered_wells)}")
    
    # ============================
    # ШАГ 2: ЭКОНОМИЧЕСКИЕ ПАРАМЕТРЫ
    # ============================
    st.markdown("---")
    st.markdown("#### Шаг 2: Экономические параметры")
    
    col_econ1, col_econ2 = st.columns(2)
    
    with col_econ1:
        oil_price = st.number_input(
            "Цена нефти, ₽/тонну",
            min_value=0.0,
            max_value=200000.0,
            value=50000.0,
            step=1000.0,
            key="oil_price_corrected"
        )
    
    with col_econ2:
        energy_price = st.number_input(
            "Тариф на электроэнергию, ₽/кВт·ч",
            min_value=0.0,
            max_value=20.0,
            value=6.0,
            step=0.1,
            key="energy_price_corrected"
        )
    
    economic_params = {
        'oil_price_rub_per_ton': oil_price,
        'energy_price_rub_per_kwh': energy_price,
        'days_per_month': 30
    }
    
    DEBUG.data("Экономические параметры", f"нефть={oil_price:.0f} ₽/т, энергия={energy_price:.2f} ₽/кВт·ч")
    
    # ============================
    # РЕЖИМ ОДНОЙ СКВАЖИНЫ
    # ============================
    if calculation_mode == "Одна скважина":
        DEBUG.log("Режим: ОДНА СКВАЖИНА")
        
        if not filtered_wells:
            st.warning(f"В ЦИТС '{selected_cits}' нет скважин для расчета")
            DEBUG.log(f"Нет скважин в ЦИТС '{selected_cits}'")
            DEBUG.exit()
            return
        
        st.markdown("---")
        st.markdown("#### Шаг 3: Выбор и диагностика скважины")
        
        well_names = [w['name'] for w in filtered_wells]
        selected_well_name = st.selectbox(
            "Выберите скважину для анализа",
            well_names,
            key="well_selection_corrected"
        )
        
        DEBUG.data("Выбрана скважина", selected_well_name)
        
        selected_well = next((w for w in filtered_wells if w['name'] == selected_well_name), None)
        
        if not selected_well:
            st.error("Ошибка при получении данных скважины")
            DEBUG.log(f"Ошибка: скважина '{selected_well_name}' не найдена")
            DEBUG.exit()
            return
        
        # Инициализация
        physics = CorrectedKPRPhysics(selected_well)
        economic_calc = EconomicCalculatorCorrected(oil_price, energy_price)
        optimizer = KPROptimizerCorrected(physics, economic_calc)
        
        # CSS для корректного отображения метрик
        st.markdown("""
        <style>
            div[data-testid="metric-container"] {
                min-width: 180px !important;
                max-width: 220px !important;
                word-wrap: break-word !important;
                white-space: normal !important;
            }
            
            div[data-testid="metric-container"] h3 {
                white-space: normal !important;
                word-wrap: break-word !important;
                line-height: 1.2 !important;
                height: auto !important;
                min-height: 36px !important;
                font-size: 14px !important;
                margin-bottom: 5px !important;
            }
            
            div[data-testid="metric-container"] div[data-testid="stMetricValue"] {
                font-size: 20px !important;
                line-height: 1.1 !important;
            }
            
            div[data-testid="metric-container"] div[data-testid="stMetricDelta"] {
                font-size: 11px !important;
                line-height: 1.1 !important;
                white-space: normal !important;
            }
            
            .full-text-container {
                background: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 8px;
                padding: 10px;
                margin: 5px 0;
                min-height: 80px;
            }
        </style>
        """, unsafe_allow_html=True)
        
        # Диагностика скважины
        st.markdown("##### 📊 Диагностика скважины")
        
        col_diag1, col_diag2, col_diag3, col_diag4, col_diag5, col_diag6 = st.columns(6)
        
        with col_diag1:
            # Pпл и Pнас
            p_pl = physics.safe_float(selected_well.get('p_pl', 0))
            p_nas = physics.safe_float(selected_well.get('p_nas', 0))
            st.metric("Pпл/Pнас", f"{p_pl:.1f}/{p_nas:.1f}")
            st.caption("Пластовое/Давление насыщения (атм)")
            DEBUG.data("Pпл/Pнас", f"{p_pl:.1f}/{p_nas:.1f}")
        
        with col_diag2:
            # Дебит и обводненность
            q_current = physics.safe_float(selected_well.get('flow_rate', 0))
            wct = physics.safe_float(selected_well.get('water_cut', 0))
            st.metric("Дебит/Обв.", f"{q_current:.1f}/{wct:.1f}%")
            st.caption("м³/сут")
            DEBUG.data("Q/Обв", f"{q_current:.1f}/{wct:.1f}%")
        
        with col_diag3:
            # Коэффициент продуктивности
            k_prod = physics.safe_float(selected_well.get('prod_coef', 0))
            
            # Статус как delta
            if k_prod <= 0:
                delta_text = "Нет данных"
                delta_color = "off"
            elif k_prod < 0.5:
                delta_text = "Низкий"
                delta_color = "inverse"
            elif k_prod < 1.0:
                delta_text = "Средний"
                delta_color = "normal"
            elif k_prod < 3.0:
                delta_text = "Хороший"
                delta_color = "normal"
            else:
                delta_text = "Высокий"
                delta_color = "normal"
            
            st.metric("Kпр", f"{k_prod:.2f}", delta=delta_text, delta_color=delta_color)
            st.caption("м³/сут/атм")
            DEBUG.data("Kпр", f"{k_prod:.2f} ({delta_text})")
        
        with col_diag4:
            # Глубина насоса и динамический уровень
            pump_depth = physics.safe_float(selected_well.get('pump_depth', 0))
            h_din = physics.safe_float(selected_well.get('h_din', 0))
            
            if pump_depth > 0:
                depth_diff = pump_depth - h_din
                
                # Статус разницы
                if depth_diff <= 0:
                    delta_text = "Насос выше уровня"
                    delta_color = "inverse"
                elif depth_diff < 100:
                    delta_text = "Малая разница"
                    delta_color = "normal"
                elif depth_diff < 600:
                    delta_text = "Норма"
                    delta_color = "normal"
                elif depth_diff < 1000:
                    delta_text = "Большая разница"
                    delta_color = "normal"
                else:
                    delta_text = "Слишком большая"
                    delta_color = "inverse"
                
                st.metric("Lнас/Hдин", f"{pump_depth:.0f}/{h_din:.0f}", delta=f"Δ={depth_diff:.0f} м", delta_color="off")
                
                # Текст статуса отдельно через caption
                if depth_diff < 100 or depth_diff >= 300:
                    st.caption(delta_text)
                else:
                    st.caption("✅ Оптимально")
                
                DEBUG.data("Lнас/Hдин", f"{pump_depth:.0f}/{h_din:.0f} (Δ={depth_diff:.0f} м)")
            else:
                st.metric("Глубина", "Нет данных")
                DEBUG.data("Глубина", "Нет данных")
        
        with col_diag5:
            # Расчет загрузки насоса
            q_pump = selected_well.get('pump_flow', 60)
            freq_raw = selected_well.get('rotations_hz')
            freq_safe = economic_calc.safe_frequency(freq_raw)
            
            schedule = selected_well.get('schedule', [15*60, 45*60])
            
            if schedule and len(schedule) >= 2:
                work_min, pause_min = schedule[0], schedule[1]
                k_util = economic_calc.calculate_utilization_factor(
                    q_current, work_min, pause_min, q_pump, freq_safe
                )
                
                # Статус загрузки
                if k_util < 0.75:
                    delta_text = "Недогрузка"
                    delta_color = "inverse"
                elif k_util > 1.25:
                    delta_text = "Перегрузка"
                    delta_color = "inverse"
                else:
                    delta_text = "Оптимально"
                    delta_color = "normal"
                
                st.metric("Загрузка", f"{k_util:.2f}", delta=delta_text, delta_color=delta_color)
                st.caption(f"насоса")
                DEBUG.data("K_под", f"{k_util:.2f} ({delta_text})")
            else:
                st.metric("Загрузка", "Нет")
                st.caption("данных графика")
                DEBUG.data("K_под", "Нет данных графика")
        
        with col_diag6:
            # Проверка газовой проблемы с использованием РАСЧИТАННОГО Pзаб
            p_buffer = physics.safe_float(selected_well.get('buffer_pressure', 0))
            pump_head = selected_well.get('pump_head', 1000)
            
            # Расчет мгновенного дебита для оценки Pзаб
            schedule = selected_well.get('schedule', [15*60, 45*60])
            if schedule and len(schedule) >= 2:
                work_min, pause_min = schedule[0], schedule[1]
                work_hours = work_min / 60
                q_instant = q_current * (24 / work_hours) if work_hours > 0 else q_current
                
                # Оценка Pзаб в конце работы - РАСЧИТАННОЕ!
                p_zab_end_est = physics.calculate_pwf_comprehensive_corrected(
                    h_din, q_instant, pump_head, pump_depth, is_pump_working=True
                )
                
                if p_nas > 0:
                    if physics.has_critical_gas_problem(p_zab_end_est, p_nas):  # ← используем РАСЧИТАННОЕ
                        delta_text = "Критическая"
                        delta_color = "inverse"
                        status_text = "❌ Pзаб(расчетное) < 0.75*Pнас"
                    elif p_zab_end_est < p_nas:
                        delta_text = "Есть проблема"
                        delta_color = "inverse"
                        status_text = "⚠️ Pзаб(расчетное) < Pнас"
                    else:
                        delta_text = "Нет проблемы"
                        delta_color = "normal"
                        status_text = "✅ Pзаб(расчетное) ≥ Pнас"
                    
                    st.metric("Газовая", " ", delta=delta_text, delta_color=delta_color)
                    st.caption(status_text)
                    st.caption(f"Pзаб={p_zab_end_est:.1f} < Pнас={p_nas:.1f}" if p_zab_end_est < p_nas else f"Pзаб={p_zab_end_est:.1f}")
                    
                    DEBUG.data("Газовая проблема", f"{delta_text} (Pзаб={p_zab_end_est:.1f}, Pнас={p_nas:.1f})")
                else:
                    st.metric("Газовая", "Нет")
                    st.caption("данных Pнас")
                    DEBUG.data("Газовая проблема", "Нет данных Pнас")
            else:
                st.metric("Газовая", "Нет")
                st.caption("данных графика")
                DEBUG.data("Газовая проблема", "Нет данных графика")
        
        # Дополнительная информация
        with st.expander("📋 Полные данные скважины"):
            col_info1, col_info2 = st.columns(2)
            
            with col_info1:
                st.write("**Основные параметры:**")
                st.write(f"- Плотность нефти: {selected_well.get('oil_density', 0.84):.3f} т/м³")
                st.write(f"- Газовый фактор: {selected_well.get('gas_factor', 0):.0f} м³/т")
                st.write(f"- Pбуф: {selected_well.get('buffer_pressure', 0):.1f} атм")
                st.write(f"- Напор насоса: {selected_well.get('pump_head', 1000):.0f} м")
                st.write(f"- Подача насоса: {selected_well.get('pump_flow', 60):.0f} м³/сут")
                st.write(f"- Pзаб(ВДП): {selected_well.get('p_zab_vdp', 0):.1f} атм")
            
            with col_info2:
                st.write("**Режимные параметры:**")
                if schedule and len(schedule) >= 2:
                    work_hr = schedule[0] 
                    pause_hr = schedule[1] 
                    st.write(f"- График: {work_hr:}/{pause_hr:} мин.")
                    st.write(f"- Частота: {freq_safe:.1f} Гц")
                else:
                    st.write("- График: не задан")
                
                # Расширенная информация о разнице глубины
                if pump_depth > 0 and h_din > 0:
                    depth_diff = pump_depth - h_din
                    fill_percentage = (h_din / pump_depth * 100) if pump_depth > 0 else 0
                    st.write(f"- Заполнение ствола: {fill_percentage:.1f}%")
                    st.write(f"- Разница L-Hдин: {depth_diff:.0f} м")
                
                # Информация о насосе
                pump_mark = selected_well.get('pump_mark', '')
                if pump_mark:
                    st.write(f"- Тип насоса: {pump_mark}")
        
        # ============================
        # ШАГ 4: ВЫБОР СЦЕНАРИЯ
        # ============================
        st.markdown("---")
        st.markdown("#### Шаг 4: Выбор сценария оптимизации")
        
        scenario_tab1, scenario_tab2, scenario_tab3 = st.tabs([
            "🎯 Сценарий A (Газ)", 
            "⚡ Сценарий B (K_под)", 
            "📈 Потенциал скважин"
        ])
        
        with scenario_tab1:
            # СЦЕНАРИЙ A: ГАЗОВАЯ ПРОБЛЕМА
            st.markdown("##### 🎯 Сценарий A: Оптимизация при газовой проблеме")
            st.info("""
            **Цель:** Уменьшить содержание свободного газа на приеме насоса
            
            **Алгоритм:**
            1. Рассчитать Pзаб для текущего режима
            2. Если Pзаб < Pнас: уменьшать t_раб пока Pзаб ≥ 0.75*Pнас
            3. Пересчитать t_накопл 
            4. Пересчитать дебит жидкости
            5. Рассчитать экономический эффект
            """)
            
            col_a1, col_a2 = st.columns(2)
            
            with col_a1:
                if schedule and len(schedule) >= 2:
                    current_work_min = float(schedule[0])
                    current_pause_min = float(schedule[1])
                else:
                    current_work_min = 15*60
                    current_pause_min = 45*60
                
                work_minutes_a = st.number_input(
                    "Время работы, минут",
                    min_value=1,
                    max_value=1440,
                    value=int(current_work_min),
                    step=5,
                    key=f"work_min_scenario_a_{selected_well_name}"
                )
            
            with col_a2:
                pause_minutes_a = st.number_input(
                    "Время накопления, минут",
                    min_value=1,
                    max_value=4320,
                    value=int(current_pause_min),
                    step=5,
                    key=f"pause_min_scenario_a_{selected_well_name}"
                )
            
            include_freq_a = st.checkbox(
                "Включать оптимизацию частоты (Сценарий A)",
                value=True,
                key=f"include_freq_scenario_a_{selected_well_name}"
            )
            
            # ========== НАСТРОЙКИ ОТОБРАЖЕНИЯ ГРАФИКА ==========
            st.markdown("##### 📊 Настройки графика")
            col_graph1, col_graph2 = st.columns(2)
            
            with col_graph1:
                display_num_cycles = st.number_input(
                    "Количество циклов для отображения:",
                    min_value=1,
                    max_value=10,
                    value=3,
                    step=1,
                    key=f"display_cycles_{selected_well_name}"
                )
            
            with col_graph2:
                display_units = st.radio(
                    "Единицы дебита:",
                    ["м³/час", "м³/сут"],
                    horizontal=True,
                    key=f"display_units_{selected_well_name}"
                )
            
            st.markdown("---")
            
            # ========== КНОПКА ЗАПУСКА ==========
            if st.button("🚀 Запустить Сценарий A", 
                        type="primary", 
                        key=f"run_scenario_a_{selected_well_name}"):
                
                with st.spinner("Выполняется оптимизация по Сценарию A..."):
                    # Выполняем оптимизацию Сценария A
                    scenario_a_result = optimizer.optimize_scenario_a_gas_problem(
                        work_minutes_a, pause_minutes_a, freq_safe
                    )
                    
                    # Экономический расчет
                    if scenario_a_result and 'recommended_work_time' in scenario_a_result:
                        has_gas_problem = scenario_a_result.get('has_gas_problem', False)
                        
                        economic_result = economic_calc.calculate_economic_effect_comprehensive(
                            well_data=selected_well,
                            old_schedule=[work_minutes_a, pause_minutes_a],
                            new_schedule=[
                                scenario_a_result.get('recommended_work_time', work_minutes_a),
                                scenario_a_result.get('recommended_pause_time', pause_minutes_a)
                            ],
                            old_freq=freq_safe,
                            new_freq=scenario_a_result.get('recommended_freq', freq_safe),
                            has_gas_problem_old=has_gas_problem
                        )
                        
                        scenario_a_result['economic_analysis'] = economic_result
                    
                    # Сохраняем результат в session_state
                    st.session_state[f'scenario_a_result_{selected_well_name}'] = scenario_a_result
                    
                    st.success("✅ Оптимизация по Сценарию A завершена!")
                    st.rerun()
            
            # ========== ОТОБРАЖЕНИЕ РЕЗУЛЬТАТОВ ==========
            result_key = f'scenario_a_result_{selected_well_name}'
            if result_key in st.session_state and st.session_state[result_key]:
                
                result = st.session_state[result_key]
                
                # Отображаем график
                if 'simulation_data' in result:
                    st.markdown("#### 📊 Детали цикла работы")
                    
                    # Получаем текущие настройки отображения
                    current_num_cycles = display_num_cycles
                    current_units = display_units
                    
                    # Получаем данные
                    sim_data = result['simulation_data']
                    p_nas = physics.safe_float(selected_well.get('p_nas', 0))
                    p_pl = physics.safe_float(selected_well.get('p_pl', 0))
                    
                    # Строим график с пластовым давлением (оно будет отображаться, но мы не пишем о нем)
                    fig_cycles = plot_multiple_cycles(
                        selected_well,
                        sim_data,
                        work_minutes_a,
                        pause_minutes_a,
                        p_nas,
                        p_pl,
                        num_cycles=current_num_cycles,
                        debit_units=current_units
                    )
                    st.plotly_chart(fig_cycles, use_container_width=True)
                    
                    # Пояснение к графику (только про Pнас, без упоминания Pпл)
                    if p_nas > 0:
                        p_critical = 0.75 * p_nas
                        if sim_data['p_zab_end'] < p_critical:
                            st.error(f"⚠️ **Проблема:** Pзаб в конце работы ({sim_data['p_zab_end']:.1f} атм) "
                                    f"ниже критического уровня {p_critical:.1f} атм (0.75·Pнас)")
                        elif sim_data['p_zab_end'] < p_nas:
                            st.warning(f"⚠️ **Внимание:** Pзаб в конце работы ({sim_data['p_zab_end']:.1f} атм) "
                                      f"ниже Pнас ({p_nas:.1f} атм), но выше критического уровня")
                        else:
                            st.success(f"✅ Pзаб в конце работы ({sim_data['p_zab_end']:.1f} атм) "
                                      f"выше Pнас ({p_nas:.1f} атм)")
                
                # Отображаем экономические результаты
                if 'economic_analysis' in result:
                    show_optimization_results_corrected(result, economic_params)
        
        with scenario_tab2:
            # СЦЕНАРИЙ B: КОЭФФИЦИЕНТ ПОДАЧИ
            st.markdown("##### ⚡ Сценарий B: Оптимизация коэффициента подачи")
            st.info("""
            **Цель:** Привести K_под в диапазон 0.75-1.25
            
            **Правила:**
            - K_под < 0.75 → нужно увеличить K_под
            - K_под > 1.25 → нужно уменьшить K_под
            """)
            
            col_b1, col_b2 = st.columns(2)
            
            with col_b1:
                if schedule and len(schedule) >= 2:
                    current_work_min = float(schedule[0])
                    current_pause_min = float(schedule[1])
                else:
                    current_work_min = 15*60
                    current_pause_min = 45*60
                
                work_minutes_b = st.number_input(
                    "Время работы, минут",
                    min_value=1,
                    max_value=1440,
                    value=int(current_work_min),
                    step=5,
                    key=f"work_min_scenario_b_{selected_well_name}"
                )
            
            with col_b2:
                pause_minutes_b = st.number_input(
                    "Время накопления, минут",
                    min_value=1,
                    max_value=4320,
                    value=int(current_pause_min),
                    step=5,
                    key=f"pause_min_scenario_b_{selected_well_name}"
                )
            
            include_freq_b = st.checkbox(
                "Включать оптимизацию частоты (Сценарий B)",
                value=True,
                key=f"include_freq_scenario_b_{selected_well_name}"
            )
            
            if st.button("🚀 Запустить Сценарий B", 
                        type="primary", 
                        key=f"run_scenario_b_{selected_well_name}"):
                
                with st.spinner("Выполняется оптимизация по Сценарию B..."):
                    # Выполняем оптимизацию Сценария B
                    scenario_b_result = optimizer.optimize_scenario_b_pump_load(
                        work_minutes_b, pause_minutes_b, freq_safe, include_freq_b
                    )
                    
                    # Экономический расчет
                    if scenario_b_result and 'recommended_work_time' in scenario_b_result:
                        # Для Сценария B обычно нет газовой проблемы
                        economic_result = economic_calc.calculate_economic_effect_comprehensive(
                            well_data=selected_well,
                            old_schedule=[work_minutes_b, pause_minutes_b],
                            new_schedule=[
                                scenario_b_result.get('recommended_work_time', work_minutes_b),
                                scenario_b_result.get('recommended_pause_time', pause_minutes_b)
                            ],
                            old_freq=freq_safe,
                            new_freq=scenario_b_result.get('recommended_freq', freq_safe),
                            has_gas_problem_old=False
                        )
                        
                        scenario_b_result['economic_analysis'] = economic_result
                    
                    # Сохраняем результат
                    st.session_state.scenario_b_result = scenario_b_result
                    
                    st.success("✅ Оптимизация по Сценарию B завершена!")
                    
                    # Показываем результаты
                    if scenario_b_result:
                        show_optimization_results_corrected(scenario_b_result, economic_params)
                    else:
                        st.warning("Не удалось выполнить оптимизацию по Сценарию B")
        
        with scenario_tab3:
            # СЦЕНАРИЙ "ПОТЕНЦИАЛ"
            st.markdown("##### 📈 Сценарий 'Потенциал': Поиск скважин для увеличения дебита")
            st.info("""
            **Цель:** Найти скважины, которые можно "раскачать" (увеличить дебит)
            
            **Алгоритм:**
            1. Рассчитать запас по уровню
            2. Определить дополнительное время работы
            3. Пересчитать новый дебит
            4. Рассчитать экономический эффект
            """)
            
            # Настройка фильтров для потенциала
            st.markdown("**Настройка фильтров:**")
            
            col_pot1, col_pot2, col_pot3, col_pot4 = st.columns(4)
            
            with col_pot1:
                min_k_pod_pot = st.number_input(
                    "Минимальный K_под",
                    min_value=1.0,
                    max_value=2.0,
                    value=1.25,
                    step=0.05,
                    key="min_k_pod_potential_single"
                )
            
            with col_pot2:
                min_depth_diff_pot = st.number_input(
                    "Минимальная разница глубин, м",
                    min_value=100,
                    max_value=1000,
                    value=400,
                    step=50,
                    key="min_depth_diff_potential_single"
                )
            
            with col_pot3:
                min_k_prod_pot = st.number_input(
                    "Минимальный K_продуктивности",
                    min_value=0.1,
                    max_value=5.0,
                    value=0.7,
                    step=0.1,
                    key="min_k_prod_potential_single"
                )
            
            with col_pot4:
                min_p_pr_pot = st.number_input(
                    "Минимальное P_прием, атм",
                    min_value=10,
                    max_value=100,
                    value=40,
                    step=5,
                    key="min_p_pr_potential_single"
                )
            
            # Обновляем параметры в классе
            WellPotentialAnalyzer.MIN_DEPTH_DIFF = min_depth_diff_pot
            WellPotentialAnalyzer.MIN_P_PR = min_p_pr_pot
            WellPotentialAnalyzer.MIN_K_PROD = min_k_prod_pot
            
            if st.button("🚀 Запустить анализ потенциала", 
                        type="primary", 
                        key=f"run_potential_{selected_well_name}"):
                
                with st.spinner("Выполняется анализ потенциала..."):
                    # Инициализация анализатора потенциала
                    potential_analyzer = WellPotentialAnalyzer(
                        physics, economic_calc,
                        min_depth_diff=min_depth_diff_pot,
                        min_p_pr=min_p_pr_pot,
                        min_k_prod=min_k_prod_pot
                    )
                    
                    # Анализ потенциала
                    potential_result = potential_analyzer.analyze_potential_increase(
                        selected_well, 
                        min_k_pod=min_k_pod_pot
                    )
                    
                    # Сохраняем результат
                    st.session_state.potential_result = potential_result
                    
                    if potential_result.get('eligible', False):
                        st.success("✅ Скважина имеет потенциал для увеличения дебита!")
                        
                        # Показываем результаты
                        st.markdown("##### 📊 Результаты анализа потенциала")
                        
                        col_pot_res1, col_pot_res2, col_pot_res3 = st.columns(3)
                        
                        with col_pot_res1:
                            current_q = potential_result['current_regime']['q_daily']
                            new_q = potential_result['new_regime']['q_daily']
                            increase_percent = potential_result['summary']['increase_q_percent']
                            
                            st.metric(
                                "Дебит жидкости",
                                f"{new_q:.1f} м³/сут",
                                delta=f"{increase_percent:.1f}%"
                            )
                        
                        with col_pot_res2:
                            current_regime = f"{potential_result['current_regime']['t_work']:.0f}/{potential_result['current_regime']['t_pause']:.0f}"
                            new_regime = f"{potential_result['new_regime']['t_work']:.0f}/{potential_result['new_regime']['t_pause']:.0f}"
                            
                            st.metric(
                                "Режим работы",
                                new_regime,
                                delta=f"из {current_regime}"
                            )
                        
                        with col_pot_res3:
                            daily_effect = potential_result['summary']['daily_effect']
                            monthly_effect = potential_result['summary']['monthly_effect']
                            
                            st.metric(
                                "Экономический эффект",
                                f"{daily_effect:.0f} ₽/сут",
                                delta=f"{monthly_effect:.0f} ₽/мес"
                            )
                        
                        # Детали расчета
                        with st.expander("📋 Детали расчета потенциала"):
                            st.write("**Текущий режим:**")
                            st.write(f"- Время работы: {potential_result['current_regime']['t_work']:.0f} мин")
                            st.write(f"- Время накопления: {potential_result['current_regime']['t_pause']:.0f} мин")
                            st.write(f"- Дебит: {potential_result['current_regime']['q_daily']:.1f} м³/сут")
                            st.write(f"- Рабочие часы: {potential_result['current_regime']['work_hours']:.1f} ч/сут")
                            
                            st.write("**Новый режим:**")
                            st.write(f"- Время работы: {potential_result['new_regime']['t_work']:.0f} мин (+{potential_result['new_regime']['additional_minutes']:.0f} мин)")
                            st.write(f"- Время накопления: {potential_result['new_regime']['t_pause']:.0f} мин")
                            st.write(f"- Дебит: {potential_result['new_regime']['q_daily']:.1f} м³/сут")
                            st.write(f"- Рабочие часы: {potential_result['new_regime']['work_hours']:.1f} ч/сут")
                            
                            st.write("**Гидродинамика:**")
                            st.write(f"- H_дин насос: {potential_result['current_regime']['h_din_pump']:.1f} м")
                            st.write(f"- H_дин после накопления: {potential_result['current_regime']['h_din_pause']:.1f} м")
                            st.write(f"- ΔH за цикл: {potential_result['current_regime']['h_din_pump'] - potential_result['current_regime']['h_din_pause']:.1f} м")
                            st.write(f"- ΔH за минуту: {potential_result['current_regime']['delta_h_per_min']:.3f} м/мин")
                            st.write(f"- Доступный запас: {potential_result['new_regime']['available_reserve']:.0f} м")
                        
                        # Экономический анализ
                        if 'economic' in potential_result:
                            econ = potential_result['economic']
                            with st.expander("💰 Экономический анализ"):
                                show_optimization_results_corrected(
                                    {'economic_analysis': econ}, 
                                    economic_params
                                )
                    else:
                        st.warning("❌ Скважина не имеет потенциала для увеличения дебита")
                        if 'reason' in potential_result:
                            st.error(f"**Причина:** {potential_result['reason']}")
        
        # ============================
        # ШАГ 5: ИНТЕГРИРОВАННАЯ ОПТИМИЗАЦИЯ (АВТОМАТИЧЕСКИЙ ВЫБОР)
        # ============================
        st.markdown("---")
        st.markdown("#### Шаг 5: Автоматическая оптимизация")
        
        st.info("""
        **Автоматический выбор сценария:**
        1. Сначала проверяем газовую проблему (Сценарий A)
        2. Если газа нет → оптимизируем загрузку (Сценарий B)
        3. Если скважина проходит все фильтры → анализируем потенциал
        """)
        
        col_auto1, col_auto2 = st.columns(2)
        
        with col_auto1:
            if schedule and len(schedule) >= 2:
                current_work_min = float(schedule[0])
                current_pause_min = float(schedule[1])
            else:
                current_work_min = 15*60
                current_pause_min = 45*60
            
            work_minutes_auto = st.number_input(
                "Время работы, минут",
                min_value=1,
                max_value=1440,
                value=int(current_work_min),
                step=5,
                key=f"work_min_auto_{selected_well_name}"
            )
        
        with col_auto2:
            pause_minutes_auto = st.number_input(
                "Время накопления, минут",
                min_value=1,
                max_value=4320,
                value=int(current_pause_min),
                step=5,
                key=f"pause_min_auto_{selected_well_name}"
            )
        
        include_freq_auto = st.checkbox(
            "Включать оптимизацию частоты (авто)",
            value=True,
            key=f"include_freq_auto_{selected_well_name}"
        )
        
        if st.button("🚀 Запустить автоматическую оптимизацию", 
                    type="primary", 
                    key=f"run_auto_{selected_well_name}"):
            
            with st.spinner("Выполняется автоматическая оптимизация..."):
                # Интегрированная оптимизация (автоматический выбор)
                integrated_result = optimizer.optimize_integrated(
                    work_minutes_auto, pause_minutes_auto, freq_safe, include_freq_auto
                )
                
                # Экономический расчет
                if integrated_result and 'recommended_work_time' in integrated_result:
                    # Определяем, была ли газовая проблема
                    has_gas_problem = integrated_result.get('has_gas_problem', False)
                    
                    economic_result = economic_calc.calculate_economic_effect_comprehensive(
                        well_data=selected_well,
                        old_schedule=[work_minutes_auto, pause_minutes_auto],
                        new_schedule=[
                            integrated_result.get('recommended_work_time', work_minutes_auto),
                            integrated_result.get('recommended_pause_time', pause_minutes_auto)
                        ],
                        old_freq=freq_safe,
                        new_freq=integrated_result.get('recommended_freq', freq_safe),
                        has_gas_problem_old=has_gas_problem
                    )
                    
                    integrated_result['economic_analysis'] = economic_result
                
                # Сохраняем результат
                st.session_state.integrated_result = integrated_result
                
                st.success(f"✅ Автоматическая оптимизация завершена! Выбран: {integrated_result.get('scenario', 'Не определен')}")
                
                # Показываем результаты
                if integrated_result:
                    show_optimization_results_corrected(integrated_result, economic_params)
                else:
                    st.warning("Не удалось выполнить автоматическую оптимизацию")
    
    else:
        # ============================
        # РЕЖИМ ГРУППЫ СКВАЖИН (ПАКЕТНЫЙ)
        # ============================
        DEBUG.log("Режим: ГРУППА СКВАЖИН (пакетный)")
        
        st.markdown("---")
        st.markdown(f"#### Шаг 3: Пакетный расчет для {selected_cits}")
        
        if not filtered_wells:
            st.warning(f"Нет скважин для расчета в выбранных фильтрах")
            DEBUG.log(f"Нет скважин после фильтрации")
            DEBUG.exit()
            return
        
        st.info(f"Найдено {len(filtered_wells)} скважин для пакетного расчета")
        DEBUG.data("Скважин для пакетного расчета", f"{len(filtered_wells)}")
        
        # Настройки пакетного расчета
        st.markdown("##### ⚙️ Настройки пакетного расчета")
        
        # Выбор типа пакетного расчета
        batch_type = st.radio(
            "Тип пакетного расчета",
            ["Полный анализ (все сценарии)", "Только анализ потенциала"],
            horizontal=True,
            key="batch_type_corrected"
        )
        
        DEBUG.data("Тип пакетного расчета", batch_type)
        
        if batch_type == "Полный анализ (все сценарии)":
            # Полный анализ со всеми сценариями
            col_settings1, col_settings2 = st.columns(2)
            
            with col_settings1:
                min_p_nas = st.number_input(
                    "Минимальное Pнас для анализа газа (атм)",
                    min_value=0.0,
                    max_value=200.0,
                    value=30.0,
                    step=5.0,
                    key="min_p_nas_threshold_corrected"
                )
            
            with col_settings2:
                include_freq_opt = st.checkbox(
                    "Включать оптимизацию частоты",
                    value=True,
                    key="include_freq_batch_corrected"
                )
            
            if st.button("🚀 Запустить ПОЛНЫЙ пакетный расчет", 
                        type="primary", 
                        key="run_full_batch_corrected"):
                
                DEBUG.section("ЗАПУСК ПОЛНОГО ПАКЕТНОГО РАСЧЕТА")

                clear_previous_batch_results()
                
                with st.spinner(f"Выполняется полный расчет для {len(filtered_wells)} скважин..."):
                    # Фильтрация по Pнас
                    if min_p_nas > 0:
                        filtered_for_calc = []
                        for w in filtered_wells:
                            p_nas_raw = w.get('p_nas')
                            
                            # Безопасное преобразование в float
                            if p_nas_raw is None:
                                # Если Pнас нет, пропускаем анализ газа, но оставляем для других расчетов
                                filtered_for_calc.append(w)
                                continue
                            
                            try:
                                p_nas_float = float(p_nas_raw)
                            except (ValueError, TypeError):
                                # Если не получается конвертировать, пропускаем
                                filtered_for_calc.append(w)
                                continue
                            
                            # Если Pнас ≥ порога или Pнас отсутствует/некорректно (0)
                            if p_nas_float >= min_p_nas or p_nas_float <= 0:
                                filtered_for_calc.append(w)
                        
                        st.info(f"Для анализа газа отобрано {len(filtered_for_calc)} из {len(filtered_wells)} скважин (Pнас ≥ {min_p_nas} атм или Pнас отсутствует/некорректно)")
                        DEBUG.data("После фильтрации по Pнас", f"{len(filtered_for_calc)} скважин")
                    else:
                        filtered_for_calc = filtered_wells
                    
                    # Запуск расчета с использованием РАСЧИТАННОГО p_zab
                    try:
                        batch_results, detailed_results = run_comprehensive_batch_optimization(
                            filtered_for_calc, economic_params
                        )
                        
                        # Сохранение результатов
                        st.session_state.potential_batch_results = batch_results
                        st.session_state.full_batch_detailed = detailed_results

                        save_data_to_file()
                        
                        st.success(f"✅ Расчет завершен! Обработано: {len(batch_results)} скважин")
                        DEBUG.log(f"Полный пакетный расчет завершен: {len(batch_results)} результатов")
                        
                    except Exception as e:
                        st.error(f"❌ Ошибка при выполнении пакетного расчета: {str(e)}")
                        DEBUG.log(f"Ошибка полного пакетного расчета: {str(e)}", "ERROR")
                        import traceback
                        st.text(traceback.format_exc())
                        
                        # Создаем пустые результаты при ошибке
                        st.session_state.full_batch_results = []
                        st.session_state.full_batch_detailed = []
        
        else:
            # Только анализ потенциала
            st.markdown("##### ⚙️ Настройки анализа потенциала")
            
            col_pot_batch1, col_pot_batch2, col_pot_batch3, col_pot_batch4 = st.columns(4)
            
            with col_pot_batch1:
                min_k_pod_batch = st.number_input(
                    "Минимальный K_под",
                    min_value=1.0,
                    max_value=2.0,
                    value=1.25,
                    step=0.05,
                    key="min_k_pod_potential_batch"
                )
            
            with col_pot_batch2:
                min_depth_diff_batch = st.number_input(
                    "Минимальная разница глубин, м",
                    min_value=100,
                    max_value=1000,
                    value=400,
                    step=50,
                    key="min_depth_diff_potential_batch"
                )
            
            with col_pot_batch3:
                min_k_prod_batch = st.number_input(
                    "Минимальный K_продуктивности",
                    min_value=0.1,
                    max_value=5.0,
                    value=0.7,
                    step=0.1,
                    key="min_k_prod_potential_batch"
                )
            
            with col_pot_batch4:
                min_p_pr_batch = st.number_input(
                    "Минимальное P_прием, атм",
                    min_value=10,
                    max_value=100,
                    value=40,
                    step=5,
                    key="min_p_pr_potential_batch"
                )
            
            # Обновляем параметры в классе
            WellPotentialAnalyzer.MIN_DEPTH_DIFF = min_depth_diff_batch
            WellPotentialAnalyzer.MIN_P_PR = min_p_pr_batch
            WellPotentialAnalyzer.MIN_K_PROD = min_k_prod_batch
            
            if st.button("🚀 Запустить анализ потенциала (пакетный)", 
                        type="primary", 
                        key="run_potential_batch_corrected"):
                
                DEBUG.section("ЗАПУСК ПАКЕТНОГО АНАЛИЗА ПОТЕНЦИАЛА")

                clear_previous_batch_results()
                
                with st.spinner(f"Выполняется анализ потенциала для {len(filtered_wells)} скважин..."):
                    try:
                        # Вызываем с пользовательскими параметрами фильтров
                        batch_results = analyze_potential_batch(
                            filtered_wells, 
                            economic_params,
                            min_k_pod=min_k_pod_batch,
                            min_depth_diff=min_depth_diff_batch,
                            min_k_prod=min_k_prod_batch,
                            min_p_pr=min_p_pr_batch
                        )
                        
                        # Сохраняем результаты
                        st.session_state.potential_batch_results = batch_results

                        save_data_to_file()

                        # УДАЛЕНО: return batch_results - это прерывало выполнение!
                        
                        st.success(f"✅ Анализ потенциала завершен! Обработано: {len(batch_results)} скважин")
                        DEBUG.log(f"Пакетный анализ потенциала завершен: {len(batch_results)} результатов")
                        
                    except Exception as e:
                        st.error(f"❌ Ошибка при выполнении анализа потенциала: {str(e)}")
                        DEBUG.log(f"Ошибка анализа потенциала: {str(e)}", "ERROR")
                        import traceback
                        st.text(traceback.format_exc())
        
        # ============================
        # ОТОБРАЖЕНИЕ РЕЗУЛЬТАТОВ ПАКЕТНОГО РАСЧЕТА
        # ============================
        
        # Определяем текущий выбранный тип расчета
        current_batch_type = batch_type
        
        # Отображение результатов полного анализа
        if current_batch_type == "Полный анализ (все сценарии)" and 'full_batch_results' in st.session_state:
            batch_results = st.session_state.full_batch_results
            df = pd.DataFrame(batch_results)
            
            # Базовая статистика
            st.markdown("---")
            st.markdown("#### 📊 Результаты пакетного расчета (Полный анализ)")
            
            col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
            
            with col_stat1:
                # Подсчет Сценария A
                scenario_a_count = 0
                if 'Сценарий' in df.columns:
                    scenario_a_count = len(df[df['Сценарий'] == 'A'])
                elif 'scenario' in df.columns:
                    scenario_a_count = len(df[df['scenario'] == 'A'])
                elif 'Тип_сценария' in df.columns:
                    scenario_a_count = len(df[df['Тип_сценария'] == 'A'])
                st.metric("Сценарий A", f"{scenario_a_count}")
                DEBUG.data("Сценарий A", f"{scenario_a_count} скважин")
            
            with col_stat2:
                # Подсчет Сценария B
                scenario_b_count = 0
                if 'Сценарий' in df.columns:
                    scenario_b_count = len(df[df['Сценарий'] == 'B'])
                elif 'scenario' in df.columns:
                    scenario_b_count = len(df[df['scenario'] == 'B'])
                elif 'Тип_сценария' in df.columns:
                    scenario_b_count = len(df[df['Тип_сценария'] == 'B'])
                st.metric("Сценарий B", f"{scenario_b_count}")
                DEBUG.data("Сценарий B", f"{scenario_b_count} скважин")
            
            with col_stat3:
                # Подсчет оптимальных скважин (не A, не B, не ошибка)
                optimal_count = len(df) - scenario_a_count - scenario_b_count
                
                # Если есть колонка с ошибками, вычитаем их
                if '_error' in df.columns:
                    error_count = len(df[df['_error'] == True])
                    optimal_count -= error_count
                elif 'Ошибка' in df['Сценарий'].values if 'Сценарий' in df.columns else False:
                    error_count = len(df[df['Сценарий'] == 'Ошибка'])
                    optimal_count -= error_count
                
                st.metric("Без оптимизации", f"{optimal_count}")
                DEBUG.data("Без оптимизации", f"{optimal_count} скважин")
            
            with col_stat4:
                # Подсчет ошибок
                error_count = 0
                if '_error' in df.columns:
                    error_count = len(df[df['_error'] == True])
                elif 'Сценарий' in df.columns:
                    error_count = len(df[df['Сценарий'] == 'Ошибка'])
                elif 'scenario' in df.columns:
                    error_count = len(df[df['scenario'] == 'error'])
                st.metric("Ошибки", f"{error_count}")
                DEBUG.data("Ошибки", f"{error_count} скважин")
            
            # Таблица результатов
            available_cols = []
            desired_cols = ['Скважина', 'Куст', 'Сценарий', 'Текущий режим', 
                            'Рекомендуемый режим', 'Эффект (₽/сут)', 'Причина']
            
            for col in desired_cols:
                if col in df.columns:
                    available_cols.append(col)
            
            # Показываем только существующие колонки
            if available_cols:
                st.dataframe(
                    df[available_cols],
                    use_container_width=True, 
                    height=400,
                    hide_index=True
                )
            else:
                st.warning("Нет данных для отображения")
                st.dataframe(df, use_container_width=True, height=400)
            
            # Экспорт результатов
            @st.cache_data(ttl=3600)
            def create_batch_excel_report():
                """
                Создает Excel отчет для пакетного расчета с безопасной обработкой колонок
                """
                _load_openpyxl()
                output = BytesIO()
                
                # Определяем, какие данные у нас есть
                batch_results = st.session_state.get('batch_results_advanced', [])
                if not batch_results:
                    # Если нет batch_results_advanced, пробуем другие возможные ключи
                    batch_results = st.session_state.get('full_batch_results', [])
                
                if not batch_results:
                    # Создаем пустой DataFrame для пустого отчета
                    df = pd.DataFrame()
                else:
                    df = pd.DataFrame(batch_results)
                
                # Безопасно определяем название колонки с эффектом
                effect_col = None
                possible_effect_cols = [
                    'Эффект (₽/сут)', 'Эффект, ₽/сут', 'Эффект ₽/сут', 
                    'total_effect_per_day', 'economic_effect', 'Эффект'
                ]
                
                for col in possible_effect_cols:
                    if col in df.columns:
                        effect_col = col
                        break
                
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    # Основные результаты
                    df.to_excel(writer, sheet_name='Все скважины', index=False)
                    
                    # Результаты только с положительным эффектом (если есть колонка с эффектом)
                    if effect_col is not None and effect_col in df.columns:
                        try:
                            positive_df = df[df[effect_col] > 0]
                            if not positive_df.empty:
                                positive_df.to_excel(writer, sheet_name='Положительные', index=False)
                        except:
                            # Если не удалось отфильтровать, пропускаем
                            pass
                    
                    # Сводка
                    summary_data = []
                    summary_data.append(['Показатель', 'Значение'])
                    summary_data.append(['Всего скважин', len(df)])
                    
                    # Если есть колонка с эффектом, добавляем статистику по эффекту
                    if effect_col is not None and effect_col in df.columns:
                        try:
                            positive_count = len(df[df[effect_col] > 0])
                            summary_data.append(['С положительным эффектом', positive_count])
                            summary_data.append(['Процент положительных', f"{(positive_count/len(df)*100):.1f}%" if len(df) > 0 else "0%"])
                            
                            # Сценарии
                            scenario_a_count = len(df[df['Сценарий'] == 'A']) if 'Сценарий' in df.columns else 0
                            scenario_b_count = len(df[df['Сценарий'] == 'B']) if 'Сценарий' in df.columns else 0
                            
                            summary_data.append(['Сценарий A (+ эффект)', scenario_a_count])
                            summary_data.append(['Сценарий B (+ эффект)', scenario_b_count])
                            
                            # Суммарный эффект
                            total_effect = df[df[effect_col] > 0][effect_col].sum() if positive_count > 0 else 0
                            summary_data.append(['Суммарный эффект (₽/сут) ТОЛЬКО +', f"{total_effect:,.0f}"])
                            summary_data.append(['Суммарный эффект (₽/мес) ТОЛЬКО +', f"{total_effect * 30:,.0f}"])
                            summary_data.append(['Суммарный эффект (₽/год) ТОЛЬКО +', f"{total_effect * 365:,.0f}"])
                        except:
                            # Если ошибка при расчете статистики, пропускаем
                            pass
                    
                    summary_data.append(['Дата расчета', datetime.now().strftime("%Y-%m-%d %H:%M")])
                    summary_data.append(['ЦИТС', st.session_state.get('selected_cits', 'ЦИТС VQ-BAD')])
                    
                    summary_df = pd.DataFrame(summary_data[1:], columns=summary_data[0])
                    summary_df.to_excel(writer, sheet_name='Сводка', index=False, header=False)
                    
                    # Экономический расчет (если есть колонка с эффектом)
                    if effect_col is not None and effect_col in df.columns:
                        try:
                            positive_df = df[df[effect_col] > 0].copy()
                            if not positive_df.empty:
                                positive_df['Эффект (₽/мес)'] = positive_df[effect_col] * 30
                                positive_df['Эффект (₽/год)'] = positive_df[effect_col] * 365
                                positive_df.to_excel(writer, sheet_name='Экономика_плюс', index=False)
                        except:
                            pass
                
                return output.getvalue()
            
            excel_data = create_batch_excel_report()
            
            col_export1, col_export2 = st.columns(2)
            
            with col_export1:
                st.download_button(
                    label="📊 Скачать полный отчет (Excel)",
                    data=excel_data,
                    file_name=f"пакетная_оптимизация_{selected_cits}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            
            with col_export2:
                # CSV экспорт
                csv = df.to_csv(index=False, encoding='utf-8-sig')
                st.download_button(
                    label="📄 Скачать таблицу (CSV)",
                    data=csv,
                    file_name=f"результаты_оптимизации_{datetime.now().strftime('%Y%m%d')}.csv",
                    mime="text/csv",
                    use_container_width=True
                )
            
            DEBUG.log("Создан Excel отчет для полного анализа")
        
        # Отображение результатов анализа потенциала
        elif current_batch_type == "Только анализ потенциала" and 'potential_batch_results' in st.session_state:
            # Показываем результаты анализа потенциала
            batch_results = st.session_state.potential_batch_results
            df = pd.DataFrame(batch_results)

            # ИСПРАВЛЕНО: Безопасная проверка наличия колонки
            if 'Проходит фильтры' in df.columns:
                eligible_df = df[df['Проходит фильтры'] == 'Да']
            else:
                # Если колонки нет, считаем, что все скважины "прошли"
                st.warning("⚠️ В данных отсутствует информация о прохождении фильтров. Отображаются все результаты.")
                eligible_df = df
                # Добавляем колонку-заглушку для совместимости с остальным кодом
                if not eligible_df.empty and 'Проходит фильтры' not in eligible_df.columns:
                    eligible_df['Проходит фильтры'] = 'Да'
            
            # ИСПРАВЛЕНО: Безопасная проверка наличия колонки
            st.markdown("---")
            st.markdown("#### 📊 Результаты анализа потенциала")
            
            # Безопасная фильтрация
            if 'Проходит фильтры' in df.columns:
                eligible_df = df[df['Проходит фильтры'] == 'Да']
            else:
                st.warning("⚠️ В данных отсутствует колонка 'Проходит фильтры'. Отображаются все результаты.")
                eligible_df = df  # Показываем все, если нет колонки
            
            # Статистика
            col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
            
            with col_stat1:
                st.metric("Всего скважин", len(df))
                DEBUG.data("Всего скважин", f"{len(df)}")
            
            with col_stat2:
                eligible_count = len(eligible_df)
                st.metric("Прошли фильтры", eligible_count)
                DEBUG.data("Прошли фильтры", f"{eligible_count}")
            
            with col_stat3:
                if eligible_count > 0:
                    avg_increase = eligible_df['Прирост Q, %'].mean()
                    st.metric("Средний прирост", f"{avg_increase:.1f}%")
                    DEBUG.data("Средний прирост", f"{avg_increase:.1f}%")
            
            with col_stat4:
                if eligible_count > 0:
                    total_monthly_effect = eligible_df['Эффект, ₽/мес'].sum()
                    st.metric("Суммарный эффект", f"{total_monthly_effect:.1f} ₽/мес")
                    DEBUG.data("Суммарный эффект", f"{total_monthly_effect:.1f} ₽/мес")
            
            # Таблица результатов
            if not eligible_df.empty:
                st.dataframe(
                    eligible_df[[
                        'Скважина', 'Куст', 'Текущий режим', 'Текущий Q, м³/сут',
                        'Новый режим', 'Новый Q, м³/сут', 'Прирост Q, %',
                        'Эффект, ₽/сут', 'Эффект, ₽/мес'
                    ]],
                    use_container_width=True,
                    height=400,
                    hide_index=True
                )
                
                # Визуализация
                st.markdown("#### 📈 Визуализация результатов")
                
                tab1, tab2, tab3 = st.tabs(["Прирост дебита", "Экономический эффект", "Топ-10 скважин"])
                
                with tab1:
                    # График прироста дебита
                    fig = go.Figure()
                    
                    fig.add_trace(go.Bar(
                        x=eligible_df['Скважина'],
                        y=eligible_df['Прирост Q, %'],
                        name='Прирост дебита, %',
                        marker_color='green',
                        text=eligible_df['Прирост Q, %'].apply(lambda x: f"{x:.1f}%"),
                        textposition='outside'
                    ))
                    
                    fig.update_layout(
                        title='Прирост дебита жидкости после оптимизации',
                        xaxis_title='Скважина',
                        yaxis_title='Прирост, %',
                        height=500,
                        showlegend=False
                    )
                    
                    st.plotly_chart(fig, use_container_width=True)
                    DEBUG.log("Построен график прироста дебита")
                
                with tab2:
                    # График экономического эффекта
                    fig2 = go.Figure()
                    
                    fig2.add_trace(go.Bar(
                        x=eligible_df['Скважина'],
                        y=eligible_df['Эффект, ₽/мес'] / 1000,
                        name='Эффект, тыс. ₽/мес',
                        marker_color='blue',
                        text=eligible_df['Эффект, ₽/мес'].apply(lambda x: f"{x/1000:.0f} тыс."),
                        textposition='outside'
                    ))
                    
                    fig2.update_layout(
                        title='Ежемесячный экономический эффект',
                        xaxis_title='Скважина',
                        yaxis_title='Эффект, тыс. ₽/мес',
                        height=500,
                        showlegend=False
                    )
                    
                    st.plotly_chart(fig2, use_container_width=True)
                    DEBUG.log("Построен график экономического эффекта")
                
                with tab3:
                    # Топ-10 по экономическому эффекту
                    top_10 = eligible_df.nlargest(min(10, len(eligible_df)), 'Эффект, ₽/мес')
                    
                    fig3 = go.Figure()
                    
                    fig3.add_trace(go.Bar(
                        x=top_10['Скважина'],
                        y=top_10['Эффект, ₽/мес'] / 1000,
                        name='Эффект',
                        marker_color='purple',
                        text=top_10['Эффект, ₽/мес'].apply(lambda x: f"{x/1000:.0f} тыс."),
                        textposition='outside'
                    ))
                    
                    fig3.update_layout(
                        title='Топ-10 скважин по экономическому эффекту',
                        xaxis_title='Скважина',
                        yaxis_title='Эффект, тыс. ₽/мес',
                        height=500,
                        showlegend=False
                    )
                    
                    st.plotly_chart(fig3, use_container_width=True)
                    DEBUG.log("Построен график топ-10 скважин")
                
                # Экспорт результатов
                st.markdown("---")
                st.markdown("#### 💾 Экспорт результатов")
                
                col_exp1, col_exp2 = st.columns(2)
                
                with col_exp1:
                    # Excel экспорт
                    import io
                    
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        eligible_df.to_excel(writer, sheet_name='Прошедшие фильтры', index=False)
                        df.to_excel(writer, sheet_name='Все скважины', index=False)
                    
                    st.download_button(
                        label="📊 Скачать Excel отчет",
                        data=output.getvalue(),
                        file_name=f"потенциал_скважин_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                
                with col_exp2:
                    # CSV экспорт
                    csv = eligible_df.to_csv(index=False, encoding='utf-8-sig')
                    st.download_button(
                        label="📄 Скачать CSV",
                        data=csv,
                        file_name="потенциал_скважин.csv",
                        mime="text/csv",
                        use_container_width=True
                    )
                
                DEBUG.log("Создан Excel отчет для анализа потенциала")
            else:
                st.warning("Нет скважин, прошедших все фильтры")
                DEBUG.log("Нет скважин, прошедших фильтры потенциала")
                
                # Показываем причины отказа
                if not df.empty:
                    st.markdown("#### Причины отказа:")
                    
                    reasons_df = df[df['Проходит фильтры'] != 'Да'][['Скважина', 'Причина']]
                    st.dataframe(reasons_df, use_container_width=True, height=300)
    
    DEBUG.log("✅ Вкладка оптимизации завершена")
    DEBUG.exit()

def show_pump_conversion_system():
    """Главная система анализа замены насосов ЭЦН с интеллектуальным подбором"""

    _load_plotly()
    
    st.markdown("## 🔄 Анализ замены насосов ЭЦН")
    st.info("Система автоматического подбора оптимальных насосов с экономическим расчетом")
    
    if not st.session_state.get('wells_data'):
        st.warning("Сначала загрузите данные скважин через раздел 'Импорт техрежима' или добавьте вручную")
        return
    
    # ========== КОНФИГУРАЦИЯ ==========
    
    # СЛОВАРИ ХАРАКТЕРИСТИК СТУПЕНЕЙ (обновленный)
    STAGE_DATA = {
        '25': [  # ЭЦН-25
            [0, 6.4, 41, 0],
            [10, 6.3, 42, 17.4],
            [18, 6, 42.9, 28.6],
            [22, 5.7, 43.9, 33.1],
            [25, 5.4, 42.5, 36],
            [29, 4.85, 45, 35.5],
            [32, 4.4, 46.5, 34.4],
            [40, 3.1, 52, 27.1],
            [54, 0, 65, 0]
        ],
        '40': [  # ЭЦН-40
            [0, 6.5, 55, 0],
            [15, 6.3, 55, 19.5],
            [30, 5.65, 54.5, 35.3],
            [35, 5.4, 55, 39],
            [40, 4.9, 55.6, 40],
            [45, 4.45, 57.3, 39.7],
            [50, 3.9, 58.8, 38.2],
            [60, 2.65, 63, 28.7],
            [75, 0, 70, 0]
        ],
        '60': [  # ЭЦН-60
            [0, 6.2, 53, 0],
            [15, 6.2, 55, 19.2],
            [30, 6.1, 58, 35.8],
            [40, 6, 61.5, 44.3],
            [50, 5.9, 66, 50.8],
            [60, 5.6, 69.2, 55.1],
            [70, 5, 73.6, 54],
            [80, 4.2, 78, 48.9],
            [90, 3.2, 82, 39.9],
            [100, 2.2, 86, 29.9],
            [115, 0, 94, 0]
        ],
        '80': [  # ЭЦН-80
            [0, 6.3, 64, 0],
            [30, 6.15, 69, 30.4],
            [60, 5.8, 79, 50],
            [70, 5.7, 83, 54.6],
            [80, 5.5, 86.8, 57.6],
            [90, 5.1, 91, 57.3],
            [100, 4.6, 94, 55.6],
            [130, 2.05, 101, 30],
            [150, 0, 110, 0]
        ],
        '125': [  # ЭЦН-125
            [0, 6, 76, 0],
            [30, 5.9, 86, 23],
            [60, 5.8, 96, 41],
            [95, 5.7, 113, 54],
            [125, 5.5, 132, 59],
            [155, 4.6, 149, 54],
            [180, 3.3, 160, 42],
            [225, 0, 168, 0]
        ],
        '160': [  # ЭЦН5A-160
            [0, 6.4, 41, 0],
            [50, 6.3, 42, 17.4],
            [100, 6.0, 42.9, 28.6],
            [150, 5.7, 43.9, 33.1],
            [200, 5.4, 44.5, 36.0],
            [250, 4.85, 45.0, 35.5],
            [300, 4.4, 46.5, 34.4],
            [350, 3.1, 52.0, 27.1]
        ],
        '200': [  # ЭЦН5A-200
            [0, 7.4, 110, 0],
            [50, 7.3, 150, 27.6],
            [100, 7.1, 178, 45.3],
            [150, 6.8, 206, 56.2],
            [175, 6.7, 215.5, 61.8],
            [200, 6.4, 224, 64.9],
            [225, 5.8, 232, 63.9],
            [250, 5.5, 233, 60.9],
            [275, 4.0, 235, 53.1],
            [300, 2.5, 232, 36.7],
            [350, 0.0, 230, 0]
        ],
        '250': [  # ЭЦН-250 (добавлен для полноты)
            [0, 7.0, 120, 0],
            [100, 6.8, 150, 35],
            [200, 6.5, 180, 65],
            [250, 6.2, 200, 78],
            [300, 5.8, 220, 85],
            [350, 5.0, 240, 82],
            [400, 4.0, 260, 70],
            [450, 0, 280, 0]
        ]
    }
    
    # Номинальные подачи насосов (м³/сут)
    PUMP_NOMINAL_FLOWS = {
        '25': 25, '40': 40, '60': 60, '80': 80,
        '125': 125, '160': 160, '200': 200, '250': 250
    }
    
    # Порядок насосов для подбора
    PUMP_HIERARCHY = ['25', '40', '60', '80', '125', '160', '200', '250']
    
    # ========== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ==========
    
    def safe_float(value, default=0.0):
        """Безопасное преобразование в float"""
        if value is None:
            return default
        try:
            return float(value)
        except (ValueError, TypeError):
            return default
    
    def interpolate(Q, data, param_index):
        """Линейная интерполяция по табличным данным"""
        if Q <= data[0][0]:
            return data[0][param_index]
        
        for i in range(1, len(data)):
            if Q <= data[i][0]:
                Q1, Q2 = data[i-1][0], data[i][0]
                V1, V2 = data[i-1][param_index], data[i][param_index]
                return V1 + (V2 - V1) * (Q - Q1) / (Q2 - Q1)
        
        return data[-1][param_index]
    
    def calculate_dynamic_level_new(pump_depth, extension=0, pump_intake_pressure=40):
        """Расчет динамического уровня для нового насоса С УЧЕТОМ УДЛИНЕНИЯ"""
        DENSITY = 820
        G = 9.81
        ATM_TO_PA = 101325
        
        pump_depth = safe_float(pump_depth, 0)
        extension = safe_float(extension, 0)
        
        # Новая глубина = старая глубина + удлинение
        new_pump_depth = pump_depth + extension
        
        # Напор от давления на приеме
        pressure_head = (pump_intake_pressure * ATM_TO_PA) / (DENSITY * G)
        
        return new_pump_depth - pressure_head
    
    def calculate_head_for_new_pump(dyn_level_new, buffer_pressure):
        """Расчет напора для нового насоса"""
        DENSITY = 820
        G = 9.81
        ATM_TO_PA = 101325
        
        dyn_level_new = safe_float(dyn_level_new, 0)
        buffer_pressure = safe_float(buffer_pressure, 0)
        pressure_head = (buffer_pressure * ATM_TO_PA) / (DENSITY * G)
        return dyn_level_new + pressure_head
    
    def extract_pump_parameters(pump_mark):
        """Извлечение параметров насоса из марки"""
        if not pump_mark or pd.isna(pump_mark):
            return None, None
        
        pump_mark_str = str(pump_mark).strip().upper()
        
        patterns = [
            r'ЭЦН[А-Я0-9]*-?(\d+)[А-Я]*[-_]?(\d+)',
            r'(\d+)[Лл]?[Ээ][Цц][Нн].*?(\d+)',
            r'ЭЦН(\d+)[А-Я]?[-_](\d+)',
            r'(\d+)-(\d+)$',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, pump_mark_str)
            if match:
                groups = match.groups()
                if len(groups) >= 2:
                    try:
                        pump_type = groups[0]
                        pump_head = safe_float(groups[1])
                        
                        if pump_type in PUMP_HIERARCHY:
                            return pump_type, pump_head
                    except:
                        continue
        
        numbers = re.findall(r'\d+', pump_mark_str)
        if len(numbers) >= 2:
            try:
                pump_type = numbers[-2] if len(numbers) >= 2 else numbers[0]
                pump_head = safe_float(numbers[-1])
                
                if pump_type in PUMP_HIERARCHY:
                    return pump_type, pump_head
            except:
                pass
        
        return None, None
    
    def get_mttf_safe(well):
        """Безопасное получение наработки на отказ"""
        mttf = well.get('mttf')
        if mttf is None:
            return None
        try:
            return float(mttf)
        except (ValueError, TypeError):
            return None
    
    def get_installation_type_safe(well):
        """Безопасное получение типа установки"""
        installation_type = well.get('installation_type')
        if installation_type is None:
            return None
        return str(installation_type).strip().upper()
    
    def is_ecn_installation(well):
        """Проверяет, является ли установка ЭЦН"""
        installation_type = get_installation_type_safe(well)
        if installation_type is None:
            return False
        
        # Ключевые слова для поиска ЭЦН
        ecn_keywords = ['ЭЦН', 'ЭЦН5', 'ЭЦНМ', 'ЭЦНП', 'ESC', 'УЭЦН', 'УЭЦНМ']
        return any(keyword in installation_type for keyword in ecn_keywords)
    
    def get_frequency_hz_safe(well):
        """Безопасное получение частоты в Гц"""
        freq = well.get('rotations_hz')
        
        # Если значение None, пустое или не число
        if freq is None or freq == '':
            return 50
        
        try:
            freq_float = float(freq)
            if freq_float <= 0:
                return 50
            return freq_float
        except (ValueError, TypeError):
            return 50
    
    def get_extension_safe(well):
        """Безопасное получение удлинения"""
        ext = well.get('udl_na_hdin')
        if ext is None:
            return 0
        try:
            return float(ext)
        except (ValueError, TypeError):
            return 0
    
    def get_pump_depth_safe(well):
        """Безопасное получение глубины насоса"""
        depth = well.get('pump_depth')
        if depth is None:
            return None
        try:
            return float(depth)
        except (ValueError, TypeError):
            return None
    
    def get_flow_rate_safe(well):
        """Безопасное получение дебита"""
        flow = well.get('flow_rate')
        if flow is None:
            return 0.0
        try:
            return float(flow)
        except (ValueError, TypeError):
            return 0.0
    
    def get_buffer_pressure_safe(well):
        """Безопасное получение буферного давления"""
        buffer_p = well.get('buffer_pressure')
        if buffer_p is not None:
            try:
                return float(buffer_p)
            except (ValueError, TypeError):
                pass
        
        p_pr = well.get('p_pr')
        if p_pr is not None:
            try:
                return float(p_pr)
            except (ValueError, TypeError):
                pass
        
        return 0.0
    
    def get_dynamic_level_safe(well):
        """Безопасное получение динамического уровня"""
        h_din = well.get('h_din')
        if h_din is None:
            return None
        try:
            return float(h_din)
        except (ValueError, TypeError):
            return None
    
    def get_kpr_schedule_safe(well):
        """Безопасное получение графика КПР"""
        schedule = well.get('kpr_schedule')
        if schedule is None:
            return None
        
        try:
            if isinstance(schedule, str):
                if '/' in schedule:
                    parts = schedule.split('/')
                    if len(parts) >= 2:
                        return [float(parts[0]), float(parts[1])]
                elif isinstance(schedule, (list, tuple)):
                    return [float(schedule[0]), float(schedule[1])]
        except:
            return None
        
        return None
    
    def get_kpr_work_hours_safe(well):
        """Безопасное получение времени работы КПР (в часах в сутки)"""
        
        # 2. Если нет, рассчитываем из графика (используем schedule, НЕ kpr_schedule)
        schedule = well.get('schedule')
        if schedule and isinstance(schedule, (list, tuple)) and len(schedule) >= 2:
            try:
                work_minutes = float(schedule[0])
                pause_minutes = float(schedule[1])
                if work_minutes + pause_minutes > 0:
                    duty_cycle = work_minutes / (work_minutes + pause_minutes)
                    return duty_cycle * 24  # часов в сутки
            except (ValueError, TypeError):
                pass
        
        # 3. Если скважина постоянная, возвращаем 24
        if well.get('operation_mode') == 'constant':
            return 24.0
        
        # 4. Значение по умолчанию
        return 4.0  # часов в сутки

    def get_kpr_mode_safe(well):
        """Безопасное получение режима КПР"""
        
        # 1. Пробуем получить из сохраненного значения
        kpr_mode = well.get('kpr_mode')
        if kpr_mode is not None:
            return str(kpr_mode)
        
        # 2. Если нет, используем поле 'mode' (из импорта техрежима)
        mode = well.get('mode')
        if mode is not None:
            return str(mode)
        
        # 3. По умолчанию для КПР скважин
        if well.get('operation_mode') == 'kpr':
            return 'По времени'
        
        # 4. Для остальных
        return 'Не указан'

    def get_kpr_schedule_display(well):
        """Возвращает отображаемую строку графика КПР"""
        
        schedule = well.get('schedule')
        if schedule and isinstance(schedule, (list, tuple)) and len(schedule) >= 2:
            try:
                work = float(schedule[0])
                pause = float(schedule[1])
                
                mode = get_kpr_mode_safe(well)
                
                if mode == 'По давлению':
                    return f"{work:.0f}/{pause:.0f} атм"
                else:
                    # ВСЕГДА показываем в минутах, независимо от размера чисел
                    # Для длинных циклов добавляем часы в скобках
                    if work > 60 or pause > 60:
                        work_hours = work / 60
                        pause_hours = pause / 60
                        return f"{work:.0f}/{pause:.0f} мин ({work_hours:.1f}/{pause_hours:.1f} ч)"
                    else:
                        return f"{work:.0f}/{pause:.0f} мин"
            except (ValueError, TypeError):
                pass
        
        kpr_schedule = well.get('kpr_schedule')
        if kpr_schedule:
            return str(kpr_schedule)
        
        return "Не задан"
    
    def get_utilization(well, current_pump_type, current_pump_head):
        """Расчет utilization (загрузки) для КПР скважины"""
        q_current = get_flow_rate_safe(well)
        frequency_val = get_frequency_hz_safe(well)
        schedule = get_kpr_schedule_safe(well)
        
        # Получаем номинальную подачу насоса
        pump_nominal = PUMP_NOMINAL_FLOWS.get(current_pump_type, 1)
        
        if schedule and len(schedule) >= 2 and frequency_val > 0:
            work_minutes = schedule[0]
            pause_minutes = schedule[1]
            if work_minutes > 0:
                duty_cycle = work_minutes / (work_minutes + pause_minutes)
                instant_flow = q_current / duty_cycle
                frequency_ratio = frequency_val / 50
                utilization = (instant_flow / pump_nominal) * frequency_ratio
                return utilization, duty_cycle
            else:
                return q_current / pump_nominal, 0
        else:
            return q_current / pump_nominal, 0
    
    def auto_select_pump_candidates(well, current_tab):
        """Автоматический подбор кандидатов на замену"""
        q_well = get_flow_rate_safe(well)
        current_pump_mark = well.get('pump_mark', '')
        current_pump_type, current_pump_head = extract_pump_parameters(current_pump_mark)
        
        candidates = []
        
        if current_tab == 'replace':  # Постоянные → КПР
            # Правило: производительность ЭЦН ≥ 3 × потенциал скважины
            required_capacity = q_well * 3
            
            # Базовый подбор по дебиту
            if q_well < 30:
                candidates.append('60')
            elif q_well < 50:
                candidates.append('80')
            else:
                candidates.append('125')
            
            # Всегда добавляем ЭЦН-125 для сравнения
            if '125' not in candidates:
                candidates.append('125')
                
        elif current_tab == 'optimize':  # КПР → КПР
            if current_pump_type:
                # Рассчитываем текущий utilization
                utilization, _ = get_utilization(well, current_pump_type, current_pump_head)
                
                # Добавляем текущий насос для сравнения
                candidates.append(current_pump_type)
                
                # Определяем индекс текущего насоса в иерархии
                if current_pump_type in PUMP_HIERARCHY:
                    current_idx = PUMP_HIERARCHY.index(current_pump_type)
                    
                    # Если utilization < 0.75, пробуем меньший размер
                    if utilization < 0.75 and current_idx > 0:
                        smaller_pump = PUMP_HIERARCHY[current_idx - 1]
                        if smaller_pump not in candidates:
                            candidates.append(smaller_pump)
                    
                    # Если utilization > 1.25, пробуем больший размер
                    if utilization > 1.25 and current_idx < len(PUMP_HIERARCHY) - 1:
                        larger_pump = PUMP_HIERARCHY[current_idx + 1]
                        if larger_pump not in candidates:
                            candidates.append(larger_pump)
                
                # Всегда добавляем ЭЦН-125 для сравнения
                if '125' not in candidates:
                    candidates.append('125')
        
        # Убираем дубликаты и сортируем
        unique_candidates = list(dict.fromkeys(candidates))
        return unique_candidates
    
    def calculate_kpr_schedule_time_mode(q_well, pump_nominal_flow, work_time_minutes=5):
        """Расчет графика КПР для режима 'По времени'"""
        # Формула: T_накопления = (Q_ном * T_работы) / Q_скв - T_работы
        if q_well > 0:
            pause_time_minutes = (pump_nominal_flow * work_time_minutes) / q_well - work_time_minutes
            pause_time_minutes = max(1, pause_time_minutes)  # Минимум 1 минута
            return work_time_minutes, round(pause_time_minutes, 1)
        else:
            return work_time_minutes, 60  # Значение по умолчанию
    
    def calculate_pump_performance(well, new_pump_type, new_pump_head, 
                                   pump_intake_pressure, specific_indicator,
                                   current_tab, current_work_hours=None):
        """Расчет производительности и экономии для нового насоса"""
        try:
            # Исходные данные
            well_name = well['name']
            Q_debit = get_flow_rate_safe(well)
            pump_depth = get_pump_depth_safe(well)
            buffer_pressure = get_buffer_pressure_safe(well)
            extension = get_extension_safe(well)
            freq_old_hz = get_frequency_hz_safe(well)
            
            if freq_old_hz < 50:
                freq_old_hz = 50
            
            # Параметры текущего насоса
            current_pump_mark = well.get('pump_mark', '')
            current_pump_type, current_pump_head_nominal = extract_pump_parameters(current_pump_mark)
            
            # Константы
            NOMINAL_FREQ = 50
            WORK_HOURS_PER_DAY = 24
            
            # Для КПР скважин используем фактическое время работы
            if current_tab == 'optimize' and current_work_hours is not None:
                old_work_hours = current_work_hours
            else:
                old_work_hours = WORK_HOURS_PER_DAY
            
            # --- РАСЧЁТ ДЛЯ СТАРОГО НАСОСА ---
            if current_pump_type and current_pump_type in STAGE_DATA:
                # Подача на 50 Гц для поиска характеристик
                Q_for_lookup = Q_debit * (NOMINAL_FREQ / freq_old_hz) if freq_old_hz > 0 else Q_debit
                
                # Характеристики ступени при 50 Гц
                stage_data_old = STAGE_DATA[current_pump_type]
                H_stage_at_50Hz = interpolate(Q_for_lookup, stage_data_old, 1)
                N_stage_at_50Hz = interpolate(Q_for_lookup, stage_data_old, 2)
                
                # Требуемый напор при рабочей частоте
                required_head_old = current_pump_head_nominal * (freq_old_hz / NOMINAL_FREQ) ** 2
                
                # Расчет числа ступеней
                if H_stage_at_50Hz > 0:
                    num_stages_old = math.ceil(required_head_old / H_stage_at_50Hz)
                else:
                    num_stages_old = 0
                
                # Мощность ступени при рабочей частоте
                N_stage_at_work_freq = N_stage_at_50Hz * (freq_old_hz / NOMINAL_FREQ) ** 3
                
                # Общая мощность старого насоса (кВт)
                power_developed_old = (num_stages_old * N_stage_at_work_freq) / 1000
                
                # Энергопотребление за сутки
                energy_per_day_old = power_developed_old * old_work_hours
            else:
                power_developed_old = 0
                energy_per_day_old = 0
            
            # --- РАСЧЁТ ДЛЯ НОВОГО НАСОСА ---
            if new_pump_type not in STAGE_DATA:
                return None
            
            # 1. Расчет динамического уровня для нового насоса
            dyn_level_new = calculate_dynamic_level_new(pump_depth, 0, pump_intake_pressure)
            
            # 2. Требуемый напор для нового насоса
            required_head_new = calculate_head_for_new_pump(dyn_level_new, buffer_pressure)
            
            # 3. ИСПРАВЛЕНИЕ: Определяем напор для расчета
            current_pump_mark = well.get('pump_mark', '')
            _, current_pump_head_nominal = extract_pump_parameters(current_pump_mark)
            
            # Если нет текущего напора, используем значение по умолчанию
            if current_pump_head_nominal is None or current_pump_head_nominal <= 0:
                current_pump_head_nominal = 1500
            
            # ПОЛУЧАЕМ РЕАЛЬНУЮ ЧАСТОТУ СТАРОГО НАСОСА
            freq_old_hz = get_frequency_hz_safe(well)  # ← ЭТА ФУНКЦИЯ УЖЕ ЕСТЬ В ВАШЕМ КОДЕ
            
            if pump_selection_mode == "Автоматический" and current_pump_head_nominal:
                # В автоматическом режиме используем напор старого насоса
                pump_head_to_use = current_pump_head_nominal
            elif new_pump_head is not None:
                # В ручном режиме используем переданный напор
                pump_head_to_use = new_pump_head
            else:
                # Если ничего не подошло, используем значение по умолчанию
                pump_head_to_use = 1500
            
            # 4. Характеристики нового насоса
            stage_data_new = STAGE_DATA[new_pump_type]
            nominal_q_new = PUMP_NOMINAL_FLOWS.get(new_pump_type, 125)
            
            H_stage_new = interpolate(nominal_q_new, stage_data_new, 1)
            N_stage_new = interpolate(nominal_q_new, stage_data_new, 2)
            
            # ========== ИСПРАВЛЕННАЯ ЛОГИКА ВЫБОРА НАПОРА ==========
            # Правила с УЧЕТОМ ЧАСТОТЫ:
            # 1. Пересчитываем старый напор на реальную частоту: H_old_real = H_old_nominal * (f_old/50)^2
            # 2. Напор нового насоса должен быть НЕ МЕНЬШЕ требуемого
            # 3. Если разница между пересчитанным старым и требуемым > 300 м - оставляем старый (НОМИНАЛЬНЫЙ!)
            # 4. Иначе используем требуемый напор
            
            # Сохраняем исходный напор для отладки
            original_pump_head = pump_head_to_use
            head_selection_reason = ""
            
            if required_head_new > 0:
                # ПЕРЕСЧИТЫВАЕМ СТАРЫЙ НАПОР С УЧЕТОМ ЧАСТОТЫ
                # H ~ f^2
                freq_ratio_old = (freq_old_hz / 50.0) ** 2
                current_pump_head_real = current_pump_head_nominal * freq_ratio_old
                
                # Логирование для отладки
                head_debug = {
                    'номинальный': current_pump_head_nominal,
                    'частота': freq_old_hz,
                    'коэффициент': freq_ratio_old,
                    'реальный': current_pump_head_real,
                    'требуемый': required_head_new
                }
                
                # Случай 1: Требуемый напор БОЛЬШЕ реального напора старого насоса
                if required_head_new > current_pump_head_real:
                    pump_head_to_use = current_pump_head_nominal  # Оставляем НОМИНАЛЬНЫЙ!
                    head_selection_reason = (f"Требуемый {required_head_new:.0f} м > "
                                             f"реального старого {current_pump_head_real:.0f} м "
                                             f"(номинал {current_pump_head_nominal:.0f} м @ {freq_old_hz:.1f} Гц) - оставлен номинальный")
                
                # Случай 2: Разница между НОМИНАЛЬНЫМ старым и требуемым БОЛЬШЕ 300 м
                elif (current_pump_head_nominal - required_head_new) > 300:
                    pump_head_to_use = current_pump_head_nominal
                    head_selection_reason = (f"Разница {current_pump_head_nominal - required_head_new:.0f} м > 300 м - "
                                             f"оставлен номинальный {current_pump_head_nominal:.0f} м")
                
                # Случай 3: Все хорошо - используем требуемый напор
                else:
                    pump_head_to_use = required_head_new
                    head_selection_reason = f"Используется требуемый напор {required_head_new:.0f} м"
                
                # Добавляем информацию для отладки
                head_selection_reason += f" (старый реальный: {current_pump_head_real:.0f} м)"
            else:
                head_selection_reason = "Требуемый напор не рассчитан - оставлен исходный"
            
            # 5. Расчет числа ступеней нового насоса (ТЕПЕРЬ ИСПОЛЬЗУЕМ pump_head_to_use)
            if H_stage_new > 0:
                num_stages_new = math.ceil(pump_head_to_use / H_stage_new)  # ← ЗДЕСЬ ИСПОЛЬЗУЕТСЯ
            else:
                num_stages_new = 0
            
            # 6. Расчет требуемой частоты для нового насоса
            if pump_head_to_use > 0:
                freq_ratio = math.sqrt(required_head_new / pump_head_to_use)
                required_freq_new = NOMINAL_FREQ * freq_ratio
                required_freq_new = max(30, min(70, required_freq_new))
            else:
                required_freq_new = NOMINAL_FREQ
            
            # 7. Расчет развивающейся мощности нового насоса
            freq_ratio_new = required_freq_new / NOMINAL_FREQ
            power_developed_new = (num_stages_new * N_stage_new / 1000) * (freq_ratio_new ** 3)
            
            # 8. Расчет времени работы в КПР режиме
            Q_developed_new = nominal_q_new * freq_ratio_new
            if Q_developed_new > 0:
                work_hours_kpr = (Q_debit / Q_developed_new) * 24
                work_hours_kpr = min(work_hours_kpr, 24)
            else:
                work_hours_kpr = 4
            
            # 9. Энергопотребление нового насоса
            energy_per_day_new = power_developed_new * work_hours_kpr
            
            # --- РАСЧЁТ ЭКОНОМИИ ---
            energy_diff = energy_per_day_old - energy_per_day_new
            economy_per_day = energy_diff * specific_indicator
            economy_per_month = economy_per_day * 30
            economy_per_year = economy_per_month * 12
            
            # --- РЕЗУЛЬТАТ ---
            result = {
                'Скважина': well_name,
                'Куст': well.get('cluster', '-'),
                'Текущий насос': current_pump_mark,
                'Текущий тип': current_pump_type,
                'Текущий напор': current_pump_head_nominal,
                'Текущая частота': freq_old_hz,  
                'Текущий напор реальный': current_pump_head_real if 'current_pump_head_real' in locals() else None,  
                'Новый тип': new_pump_type,
                'Новый напор': pump_head_to_use,
                'Требуемый напор': required_head_new,
                'Логика напора': head_selection_reason,
                'Новый тип': new_pump_type,
                'Новый напор': pump_head_to_use,  
                'Qж, м³/сут': round(Q_debit, 1),
                'Частота старого, Гц': round(freq_old_hz, 1),
                'Мощность старого, кВт': round(power_developed_old, 2),
                'Время работы старого, ч/сут': round(old_work_hours, 1),
                'Энергия старого, кВт·ч/сут': round(energy_per_day_old, 2),
                'Частота нового, Гц': round(required_freq_new, 1),
                'Мощность нового, кВт': round(power_developed_new, 2),
                'Время работы КПР, ч/сут': round(work_hours_kpr, 1),
                'Энергия нового, кВт·ч/сут': round(energy_per_day_new, 2),
                'Экономия, кВт·ч/сут': round(energy_diff, 2),
                'Экономия, руб/сут': round(economy_per_day, 2),
                'Экономия, руб/месяц': round(economy_per_month, 2),
                'Экономия, руб/год': round(economy_per_year, 2),
                'Номинальная подача, м³/сут': nominal_q_new,
                'Развивающаяся подача, м³/сут': round(Q_developed_new, 1),
                'Динамич. уровень новый, м': round(dyn_level_new, 0),
                'Напор новый, м': round(required_head_new, 0)
            }
            
            return result
            
        except Exception as e:
            st.warning(f"Ошибка расчета для скважины {well.get('name', 'N/A')}: {str(e)}")
            return None
    
    def filter_by_economy(results, min_daily_savings=0):
        """Фильтрация результатов по экономической эффективности"""
        if not results:
            return [], []
        
        # Фильтруем по положительной экономии
        filtered = []
        excluded = []
        
        for result in results:
            if result.get('Экономия, руб/сут', 0) > min_daily_savings:
                filtered.append(result)
            else:
                excluded.append(result)
        
        return filtered, excluded
    
    def get_best_variant_for_well(variant_results):
        """Выбор лучшего варианта для скважины"""
        if not variant_results:
            return None
        
        # Ищем вариант с максимальной суточной экономией
        best_variant = None
        max_savings = -float('inf')
        
        for variant in variant_results:
            savings = variant.get('Экономия, руб/сут', 0)
            if savings > max_savings:
                max_savings = savings
                best_variant = variant
        
        return best_variant

    @st.cache_data(ttl=3600)
    def create_ecn_replacement_report(results, params):
        """Создание отчета по замене ЭЦН с полным расчетом режимов"""
        
        # Загружаем openpyxl
        _load_openpyxl()
        
        # Импортируем MergedCell для безопасной обработки
        from openpyxl.cell.cell import MergedCell
        
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            workbook = writer.book
            
            # ================= ЛИСТ 1: ДЕТАЛЬНЫЙ РАСЧЕТ ПО СКВАЖИНАМ =================
            ws1 = workbook.create_sheet("Детальный расчет")
            
            # Заголовок
            ws1.merge_cells('A1:P1')
            ws1['A1'] = "ДЕТАЛЬНЫЙ РАСЧЕТ ЗАМЕНЫ ЭЦН И ПЕРЕВОДА НА КПР"
            ws1['A1'].font = Font(bold=True, size=14)
            ws1['A1'].alignment = Alignment(horizontal='center')
            
            # Информация о расчете
            ws1['A3'] = "Дата расчета:"
            ws1['B3'] = datetime.now().strftime("%d.%m.%Y %H:%M")
            ws1['A4'] = "Режим расчета:"
            ws1['B4'] = "Замена постоянных → КПР" if params.get('current_tab') == 'replace' else "Оптимизация КПР → КПР"
            ws1['A5'] = "Тариф на э/э, руб/кВт·ч:"
            ws1['B5'] = params.get('specific_indicator', 5.28)
            
            # Заголовки таблицы
            headers = [
                'Скважина', 'Куст', 'Текущий насос', 'Тип режима',
                'Текущий график', 'Текущее время, ч/сут',
                'Новый насос', 'Новая частота, Гц', 'Новый напор, м',
                'Новый график', 'Новое время, ч/сут',
                'Экономия, руб/сут', 'Логика выбора'
            ]
            
            for col_idx, header in enumerate(headers, start=1):
                cell = ws1.cell(row=8, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            # Данные
            row_idx = 9
            total_savings = 0
            
            for result in results:
                well_name = result.get('Скважина', '')
                
                # ========== БЕРЕМ ДАННЫЕ ИЗ result (ИСХОДНЫЕ!) ==========
                old_pump = result.get('Текущий насос', '')
                old_work_hours = result.get('Время работы старого, ч/сут', 24.0)
                
                # Режим КПР
                kpr_mode = result.get('Режим КПР', 'Не указан')
                if kpr_mode == 'Не указан':
                    # Пробуем получить из well_data
                    well_data = None
                    for w in st.session_state.wells_data:
                        if w['name'] == well_name:
                            well_data = w
                            break
                    if well_data:
                        kpr_mode = get_kpr_mode_safe(well_data)
                
                # Текущий график
                old_schedule = result.get('old_regime', 'Н/Д')
                if old_schedule == 'Н/Д':
                    # Пробуем собрать из других полей
                    work_old = result.get('Время работы старого, мин')
                    pause_old = result.get('Время накопления старого, мин')
                    if work_old and pause_old:
                        old_schedule = f"{work_old:.0f}/{pause_old:.0f}"
                    else:
                        # Если ничего нет, используем get_kpr_schedule_display
                        well_data = None
                        for w in st.session_state.wells_data:
                            if w['name'] == well_name:
                                well_data = w
                                break
                        
                        if well_data:
                            old_schedule = get_kpr_schedule_display(well_data)
                        else:
                            old_schedule = "Н/Д"
                
                # Новые параметры
                new_pump = f"ЭЦН-{result.get('Новый тип', '')}"
                new_freq = result.get('Частота нового, Гц', 50)
                new_head = result.get('Новый напор', 0)
                new_work_hours = result.get('Время работы КПР, ч/сут', 0)
                
                # Новый график
                new_schedule = result.get('new_regime', 'Н/Д')
                if new_schedule == 'Н/Д' and new_work_hours > 0:
                    if kpr_mode == 'По давлению':
                        new_schedule = f"{result.get('Давление запуска', 40)}/{result.get('Давление остановки', 42)} атм"
                    else:
                        cycle_minutes = 60
                        work_minutes = cycle_minutes * (new_work_hours / 24)
                        pause_minutes = cycle_minutes - work_minutes
                        if work_minutes > 0 and pause_minutes > 0:
                            new_schedule = f"{work_minutes:.0f}/{pause_minutes:.0f} мин"
                        else:
                            new_schedule = f"{new_work_hours:.1f} ч/сут"
                
                # Экономия
                savings = result.get('Экономия, руб/сут', 0)
                total_savings += savings
                
                # Логика
                logic = result.get('Логика напора', '')
                
                # Записываем строку
                row_data = [
                    well_name,
                    result.get('Куст', '-'),
                    old_pump,
                    kpr_mode,
                    old_schedule,
                    f"{old_work_hours:.1f}",
                    new_pump,
                    new_freq,
                    new_head,
                    new_schedule,
                    f"{new_work_hours:.1f}",
                    f"{savings:,.0f}",
                    logic
                ]
                
                for col_idx, value in enumerate(row_data, start=1):
                    ws1.cell(row=row_idx, column=col_idx, value=value)
                
                row_idx += 1
            
            # Итог
            ws1.cell(row=row_idx, column=11, value="ИТОГО:").font = Font(bold=True)
            ws1.cell(row=row_idx, column=12, value=f"{total_savings:,.0f}").font = Font(bold=True)
            
            # ========== БЕЗОПАСНАЯ НАСТРОЙКА ШИРИНЫ СТОЛБЦОВ ==========
            for column in ws1.columns:
                max_length = 0
                column_letter = None
                
                # Находим первую НЕ объединенную ячейку в столбце
                for cell in column:
                    if not isinstance(cell, MergedCell):
                        if cell.column_letter and not column_letter:
                            column_letter = cell.column_letter
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                
                if column_letter:
                    adjusted_width = min(max_length + 2, 50)
                    ws1.column_dimensions[column_letter].width = adjusted_width
            
            # ================= ЛИСТ 2: СВОДКА =================
            ws2 = workbook.create_sheet("Сводка")
            
            ws2.merge_cells('A1:B1')
            ws2['A1'] = "СВОДНЫЕ РЕЗУЛЬТАТЫ"
            ws2['A1'].font = Font(bold=True, size=12)
            
            summary_data = [
                ["Параметр", "Значение"],
                ["Всего скважин", len(results)],
                ["Суммарная экономия в сутки, руб", f"{total_savings:,.0f}"],
                ["Суммарная экономия в месяц, руб", f"{total_savings * 30:,.0f}"],
                ["Суммарная экономия в год, руб", f"{total_savings * 365:,.0f}"],
                ["Средняя экономия на скважину, руб/сут", f"{total_savings/len(results):,.0f}" if results else "0"],
            ]
            
            row_idx = 3
            for row in summary_data:
                ws2.cell(row=row_idx, column=1, value=row[0])
                ws2.cell(row=row_idx, column=2, value=row[1])
                row_idx += 1
            
            # Настройка ширины для сводки
            for column in ws2.columns:
                max_length = 0
                column_letter = None
                for cell in column:
                    if not isinstance(cell, MergedCell):
                        if cell.column_letter and not column_letter:
                            column_letter = cell.column_letter
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                if column_letter:
                    ws2.column_dimensions[column_letter].width = min(max_length + 2, 30)
        
        # Удаляем дефолтный лист
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        
        output.seek(0)
        return output
    
    def save_ecn_replacement_to_system(well_name, new_pump_type, new_pump_head, 
                                       schedule, kpr_mode, current_tab='replace'):
        """Сохранение замены ЭЦН в систему"""
        try:
            # Находим скважину в данных
            for well in st.session_state.wells_data:
                if well['name'] == well_name:
                    # Обновляем марку насоса
                    new_pump_mark = f"ЭЦН-{new_pump_type}-{int(new_pump_head)}"
                    well['pump_mark'] = new_pump_mark
                    
                    # Обновляем режим работы
                    if current_tab == 'replace':
                        well['operation_mode'] = 'kpr'  # Меняем на КПР
                    
                    # Сохраняем график КПР
                    well['kpr_mode'] = kpr_mode
                    if isinstance(schedule, (list, tuple)) and len(schedule) >= 2:
                        well['kpr_schedule'] = f"{schedule[0]}/{schedule[1]}"
                    
                    # Сохраняем время работы КПР
                    if kpr_mode == 'По времени' and isinstance(schedule, (list, tuple)) and len(schedule) >= 2:
                        work_minutes = schedule[0]
                        pause_minutes = schedule[1]
                        if work_minutes + pause_minutes > 0:
                            duty_cycle = work_minutes / (work_minutes + pause_minutes)
                            well['kpr_work_hours'] = duty_cycle * 24
                    
                    return True
            
            return False
        except Exception as e:
            st.error(f"Ошибка сохранения для скважины {well_name}: {str(e)}")
            return False
    
    # ========== ИНТЕРФЕЙС ==========
    
    # Выбор режима работы
    st.markdown("---")
    st.markdown("### 🎯 Выберите режим анализа")
    
    col_mode1, col_mode2 = st.columns(2)
    
    with col_mode1:
        if st.button("🔄 **Замена ПОСТОЯННЫХ на КПР**", use_container_width=True, 
                    type="primary" if not st.session_state.get('current_conversion_tab') or 
                    st.session_state.current_conversion_tab == 'replace' else "secondary"):
            st.session_state.current_conversion_tab = 'replace'
            st.rerun()
    
    with col_mode2:
        if st.button("🔄 **Замена КПР на КПР**", use_container_width=True,
                    type="primary" if st.session_state.get('current_conversion_tab') == 'optimize' else "secondary"):
            st.session_state.current_conversion_tab = 'optimize'
            st.rerun()
    
    # Устанавливаем текущую вкладку
    current_tab = st.session_state.get('current_conversion_tab', 'replace')
    
    if current_tab == 'replace':
        st.success("**Режим:** Замена постоянных скважин на КПР с подбором насосов")
    else:
        st.success("**Режим:** Оптимизация уже работающих КПР скважин")
    
    st.markdown("---")
    
    # ========== ШАГ 1: ФИЛЬТРЫ ==========
    
    st.markdown("### Шаг 1: Настройка фильтров")
    
    col_filters1, col_filters2 = st.columns(2)
    
    with col_filters1:
        # Фильтр по глубине насоса
        st.markdown("#### Фильтр по глубине насоса")
        min_depth = st.number_input("Минимальная глубина, м", 
                                   min_value=0, max_value=5000, value=1000, step=100,
                                   key=f"min_depth_{current_tab}")
        max_depth = st.number_input("Максимальная глубина, м",
                                   min_value=0, max_value=5000, value=3000, step=100,
                                   key=f"max_depth_{current_tab}")
        
        # Фильтр по дебиту
        st.markdown("#### Фильтр по дебиту")
        min_flow = st.number_input("Минимальный дебит, м³/сут",
                                  min_value=0, max_value=1000, value=10, step=10,
                                  key=f"min_flow_{current_tab}")
        max_flow = st.number_input("Максимальный дебит, м³/сут",
                                  min_value=0, max_value=1000, value=500, step=10,
                                  key=f"max_flow_{current_tab}")
        
        # Фильтр по динамическому уровню (НОВОЕ)
        st.markdown("#### Фильтр по динамическому уровню (Hдин)")
        min_dyn_level = st.number_input("Минимальный Hдин, м",
                                       min_value=0, max_value=5000, value=0, step=50,
                                       key=f"min_dyn_level_{current_tab}")
        max_dyn_level = st.number_input("Максимальный Hдин, м",
                                       min_value=0, max_value=5000, value=3000, step=50,
                                       key=f"max_dyn_level_{current_tab}")
    
    with col_filters2:
        # Фильтр по наработке на отказ
        st.markdown("#### Фильтр по наработке на отказ")
        mttf_filter = st.selectbox(
            "Фильтр по наработке на отказ",
            options=[
                "Не фильтровать",
                "Менее 365 дней (высокий риск)",
                "365-730 дней (средний риск)", 
                "Более 730 дней (низкий риск)"
            ],
            key=f"mttf_filter_{current_tab}"
        )
        
        # Фильтр по типу установки
        st.markdown("#### Фильтр по типу установки")
        installation_filter = st.selectbox(
            "Тип установки",
            options=[
                "Только ЭЦН",
                "Все типы (включая ОРД, ШГН)"
            ],
            key=f"installation_filter_{current_tab}",
            help="По умолчанию только ЭЦН установки"
        )
        
        # Фильтр по типу насоса
        st.markdown("#### Фильтр по типу насоса")
        selected_pump_types = st.multiselect(
            "Типы насосов для анализа",
            options=PUMP_HIERARCHY,
            default=PUMP_HIERARCHY,
            key=f"selected_pump_types_{current_tab}"
        )
        
        # Для режима оптимизации КПР - дополнительный фильтр
        if current_tab == 'optimize':
            st.markdown("#### Фильтр по текущему режиму КПР")
            kpr_mode_filter = st.selectbox(
                "Режим работы КПР",
                options=["Все", "По времени", "По давлению"],
                key="kpr_mode_filter"
            )
    
    st.markdown("---")
    
    # ========== ШАГ 2: ПАРАМЕТРЫ РАСЧЕТА ==========
    
    st.markdown("### Шаг 2: Параметры расчета")
    
    col_params1, col_params2 = st.columns(2)
    
    with col_params1:
        # Параметры для расчета
        st.markdown("#### Технические параметры")
        
        pump_intake_pressure = st.number_input(
            "Давление на приеме, атм",
            min_value=10.0, max_value=100.0, value=40.0, step=1.0, format="%.1f",
            key=f"pump_intake_pressure_{current_tab}"
        )
        
        # Режим подбора насосов
        st.markdown("#### Режим подбора насосов")
        pump_selection_mode = st.radio(
            "Способ подбора насосов",
            options=["Автоматический", "Ручной выбор"],
            key=f"pump_selection_mode_{current_tab}",
            help="Автоматический: система сама подберет оптимальные варианты\nРучной: вы выбираете тип насоса"
        )
        
        # ИСПРАВЛЕНИЕ: Инициализируем переменные до условного блока
        if pump_selection_mode == "Ручной выбор":
            new_pump_type = st.selectbox(
                "Тип нового насоса",
                PUMP_HIERARCHY,
                index=PUMP_HIERARCHY.index('125') if '125' in PUMP_HIERARCHY else 0,
                key=f"manual_pump_type_{current_tab}"
            )
            new_pump_head = st.number_input(
                "Напор нового насоса, м",
                min_value=100, max_value=3000, value=1500, step=50,
                key=f"manual_pump_head_{current_tab}"
            )
        else:
            # Для автоматического режима задаем значения по умолчанию
            # (они не будут использоваться, но нужны для избежания NameError)
            new_pump_type = '125'
            new_pump_head = 1500
    
    with col_params2:
        # Экономические параметры
        st.markdown("#### Экономические параметры")
        
        specific_indicator = st.number_input(
            "Удельный показатель, руб/кВт·ч",
            min_value=0.0, max_value=20.0, value=5.28, step=0.01, format="%.2f",
            key=f"specific_indicator_{current_tab}"
        )
        
        days_per_month = st.number_input(
            "Дней в месяце",
            min_value=28, max_value=31, value=30, step=1,
            key=f"days_per_month_{current_tab}"
        )
        
        months_per_year = st.number_input(
            "Месяцев в году",
            min_value=1, max_value=12, value=12, step=1,
            key=f"months_per_year_{current_tab}"
        )
        
        # Параметры для КПР режима
        st.markdown("#### Параметры КПР (при сохранении)")
        default_kpr_mode = "По времени" if current_tab == 'replace' else "По давлению"
        kpr_mode = st.selectbox(
            "Режим КПР",
            ["По времени", "По давлению"],
            key=f"kpr_mode_{current_tab}",
            help="Выберите режим работы для новых КПР скважин"
        )
    
    st.markdown("---")
    
    # ========== ШАГ 3: ВЫБОР СКВАЖИН ==========
    
    st.markdown("### Шаг 3: Выбор скважин для анализа")
    
    # Фильтрация скважин по режиму
    if current_tab == 'replace':
        # Для замены: берем только постоянные скважины
        filtered_by_mode = [w for w in st.session_state.wells_data 
                           if w.get('operation_mode') == 'constant' 
                           and w.get('is_active', True)]
    else:
        # Для оптимизации: берем только КПР скважины
        filtered_by_mode = [w for w in st.session_state.wells_data 
                           if w.get('operation_mode') == 'kpr' 
                           and w.get('is_active', True)]
    
    if not filtered_by_mode:
        mode_name = "постоянных" if current_tab == 'replace' else "КПР"
        st.warning(f"В системе нет активных {mode_name} скважин")
        return
    
    # Фильтр по ЦИТС
    cits_list = list(set([w.get('cits', 'ЦИТС VQ-BAD') for w in filtered_by_mode]))
    selected_cits = st.selectbox("Выберите ЦИТС", cits_list, key=f"cits_{current_tab}")
    
    # Применяем все фильтры
    filtered_wells = []
    filter_stats = {
        'total': 0,
        'filtered_by_installation': 0,
        'filtered_by_depth': 0,
        'filtered_by_flow': 0,
        'filtered_by_dyn_level': 0,
        'filtered_by_mttf': 0,
        'filtered_by_pump_type': 0,
        'filtered_by_kpr_mode': 0,
        'passed_all': 0
    }
    
    for w in filtered_by_mode:
        if w.get('cits', 'ЦИТС VQ-BAD') != selected_cits:
            continue
            
        filter_stats['total'] += 1
        
        # 1. Фильтр по типу установки
        if installation_filter == "Только ЭЦН":
            if not is_ecn_installation(w):
                filter_stats['filtered_by_installation'] += 1
                continue
        
        # 2. Фильтр по глубине
        pump_depth = get_pump_depth_safe(w)
        depth_ok = True
        if pump_depth is not None:
            depth_ok = min_depth <= pump_depth <= max_depth
        else:
            depth_ok = False
        
        if not depth_ok:
            filter_stats['filtered_by_depth'] += 1
            continue
        
        # 3. Фильтр по дебиту
        flow_rate = get_flow_rate_safe(w)
        flow_ok = min_flow <= flow_rate <= max_flow
        if not flow_ok:
            filter_stats['filtered_by_flow'] += 1
            continue
        
        # 4. Фильтр по динамическому уровню (НОВОЕ)
        dyn_level = get_dynamic_level_safe(w)
        dyn_level_ok = True
        if dyn_level is not None:
            dyn_level_ok = min_dyn_level <= dyn_level <= max_dyn_level
        
        if not dyn_level_ok:
            filter_stats['filtered_by_dyn_level'] += 1
            continue
        
        # 5. Фильтр по наработке
        mttf_value = get_mttf_safe(w)
        mttf_ok = True
        
        if mttf_filter != "Не фильтровать":
            if mttf_value is None:
                mttf_ok = False
            elif mttf_filter == "Менее 365 дней (высокий риск)":
                mttf_ok = mttf_value < 365
            elif mttf_filter == "365-730 дней (средний риск)":
                mttf_ok = 365 <= mttf_value <= 730
            elif mttf_filter == "Более 730 дней (низкий риск)":
                mttf_ok = mttf_value > 730
        
        if not mttf_ok:
            filter_stats['filtered_by_mttf'] += 1
            continue
        
        # 6. Фильтр по типу насоса
        pump_mark = w.get('pump_mark', '')
        pump_type, _ = extract_pump_parameters(pump_mark)
        
        pump_type_ok = True
        if selected_pump_types and pump_type:
            pump_type_ok = pump_type in selected_pump_types
        
        if not pump_type_ok:
            filter_stats['filtered_by_pump_type'] += 1
            continue
        
        # 7. Дополнительный фильтр для КПР скважин
        if current_tab == 'optimize' and kpr_mode_filter != "Все":
            current_kpr_mode = w.get('kpr_mode', '')
            if kpr_mode_filter != current_kpr_mode:
                filter_stats['filtered_by_kpr_mode'] += 1
                continue
        
        filtered_wells.append(w)
        filter_stats['passed_all'] += 1
    
    if not filtered_wells:
        mode_name = "постоянных" if current_tab == 'replace' else "КПР"
        st.warning(f"В ЦИТС '{selected_cits}' нет {mode_name} скважин, соответствующих фильтрам")
        
        st.info(f"""
        **Статистика фильтрации:**
        - Всего {mode_name} скважин в ЦИТС: {filter_stats['total']}
        - Отфильтровано по типу установки: {filter_stats['filtered_by_installation']}
        - Отфильтровано по глубине: {filter_stats['filtered_by_depth']}
        - Отфильтровано по дебиту: {filter_stats['filtered_by_flow']}
        - Отфильтровано по динамическому уровню: {filter_stats['filtered_by_dyn_level']}
        - Отфильтровано по наработке: {filter_stats['filtered_by_mttf']}
        - Отфильтровано по типу насоса: {filter_stats['filtered_by_pump_type']}
        - Отфильтровано по режиму КПР: {filter_stats['filtered_by_kpr_mode']}
        - Прошло все фильтры: {filter_stats['passed_all']}
        """)
        return
    
    mode_name = "постоянных" if current_tab == 'replace' else "КПР"
    st.success(f"Найдено {len(filtered_wells)} {mode_name} скважин, соответствующих фильтрам")
    
    # Статистика
    col_stat1, col_stat2, col_stat3 = st.columns(3)
    
    with col_stat1:
        depths = [d for d in [get_pump_depth_safe(w) for w in filtered_wells] if d is not None]
        avg_depth = np.mean(depths) if depths else 0
        st.metric("Средняя глубина", f"{avg_depth:.0f} м" if depths else "Нет данных")
    
    with col_stat2:
        flows = [get_flow_rate_safe(w) for w in filtered_wells]
        avg_flow = np.mean(flows) if flows else 0
        st.metric("Средний дебит", f"{avg_flow:.0f} м³/сут")
    
    with col_stat3:
        if current_tab == 'optimize':
            # Для КПР скважин показываем среднее время работы
            work_hours = [get_kpr_work_hours_safe(w) for w in filtered_wells]
            avg_work_hours = np.mean(work_hours) if work_hours else 0
            st.metric("Ср. время работы", f"{avg_work_hours:.1f} ч/сут")
        else:
            # Для постоянных показываем наработку
            mttf_values = [m for m in [get_mttf_safe(w) for w in filtered_wells] if m is not None]
            if mttf_values:
                avg_mttf = np.mean(mttf_values)
                st.metric("Средняя наработка", f"{avg_mttf:.0f} дней")
            else:
                st.metric("Данные наработки", "Недоступны")
    
    # ========== ВЫБОР СКВАЖИН В ТАБЛИЦЕ ==========
    
    st.markdown("#### Отметьте скважины для анализа:")
    
    # Инициализация состояния
    if f'selected_wells_indices_{current_tab}' not in st.session_state:
        st.session_state[f'selected_wells_indices_{current_tab}'] = set()
    
    if f'select_all_triggered_{current_tab}' not in st.session_state:
        st.session_state[f'select_all_triggered_{current_tab}'] = False
    
    if f'deselect_all_triggered_{current_tab}' not in st.session_state:
        st.session_state[f'deselect_all_triggered_{current_tab}'] = False
    
    # Подготавливаем данные для таблицы
    table_data = []
    for idx, well in enumerate(filtered_wells):
        pump_mark = well.get('pump_mark', 'Нет данных')
        pump_type, pump_head = extract_pump_parameters(pump_mark)
        
        # Базовые данные
        pump_depth_val = get_pump_depth_safe(well)
        flow_rate_val = get_flow_rate_safe(well)
        buffer_pressure_val = get_buffer_pressure_safe(well)
        h_din_val = get_dynamic_level_safe(well)
        frequency_val = get_frequency_hz_safe(well)
        extension_val = get_extension_safe(well)
        
        # Дополнительные данные для КПР скважин
        if current_tab == 'optimize':
            schedule = get_kpr_schedule_safe(well)
            work_hours_val = get_kpr_work_hours_safe(well)
            current_kpr_mode = well.get('kpr_mode', 'Не указан')
            
            # Расчет utilization если есть данные
            utilization_val = None
            if pump_type and pump_head:
                utilization_val, _ = get_utilization(well, pump_type, pump_head)
        
        # Определяем, выбрана ли скважина
        is_selected = idx in st.session_state[f'selected_wells_indices_{current_tab}']
        
        # Обработка кнопок "Выбрать все"/"Снять все"
        if st.session_state[f'select_all_triggered_{current_tab}']:
            is_selected = True
            st.session_state[f'selected_wells_indices_{current_tab}'].add(idx)
        
        if st.session_state[f'deselect_all_triggered_{current_tab}']:
            is_selected = False
            if idx in st.session_state[f'selected_wells_indices_{current_tab}']:
                st.session_state[f'selected_wells_indices_{current_tab}'].remove(idx)
        
        # Формируем строку таблицы
        row_data = {
            'Выбрать': is_selected,
            'Скважина': well['name'],
            'Куст': well.get('cluster', '-'),
            'Дебит, м³/сут': flow_rate_val,
            'Глубина, м': pump_depth_val if pump_depth_val is not None else 'Нет данных',
            'Частота, Гц': frequency_val,
            'Hдин, м': h_din_val if h_din_val is not None else 'Нет данных',
            'Насос': pump_mark,
            'Тип': pump_type if pump_type else 'Не опр.'
        }

        kpr_mode_display = get_kpr_mode_safe(well)
        work_hours_val = get_kpr_work_hours_safe(well)
        
        row_data = {
            'Выбрать': is_selected,
            'Скважина': well['name'],
            'Куст': well.get('cluster', '-'),
            'Дебит, м³/сут': flow_rate_val,
            'Глубина, м': pump_depth_val if pump_depth_val is not None else 'Нет данных',
            'Частота, Гц': frequency_val,
            'Hдин, м': h_din_val if h_din_val is not None else 'Нет данных',
            'Насос': pump_mark,
            'Тип': pump_type if pump_type else 'Не опр.'
        }

        row_data.update({
            'Режим КПР': kpr_mode_display,  # ← ИСПОЛЬЗУЕМ НОВУЮ ФУНКЦИЮ
            'Время работы, ч/сут': round(work_hours_val, 1) if work_hours_val else 'Нет данных',
        })
        
        # Utilization только для optimize режима
        if current_tab == 'optimize':
            utilization_val = None
            if pump_type and pump_head:
                utilization_val, _ = get_utilization(well, pump_type, pump_head)
            row_data['Utilization, %'] = f"{utilization_val*100:.0f}" if utilization_val is not None else 'Нет данных'
        
        table_data.append(row_data)
    
    # Создаем DataFrame
    df_selection = pd.DataFrame(table_data)
    
    # Отображаем таблицу выбора
    edited_df = st.data_editor(
        df_selection,
        column_config={
            "Выбрать": st.column_config.CheckboxColumn(required=True),
            "Скважина": st.column_config.TextColumn(disabled=True),
            "Куст": st.column_config.TextColumn(disabled=True),
            "Дебит, м³/сут": st.column_config.NumberColumn(disabled=True),
            "Глубина, м": st.column_config.NumberColumn(disabled=True),
            "Частота, Гц": st.column_config.NumberColumn(disabled=True),
            "Hдин, м": st.column_config.NumberColumn(disabled=True),
            "Насос": st.column_config.TextColumn(disabled=True),
            "Тип": st.column_config.TextColumn(disabled=True),
            **({"Режим КПР": st.column_config.TextColumn(disabled=True),
                "Время работы, ч/сут": st.column_config.NumberColumn(disabled=True),
                "Utilization, %": st.column_config.TextColumn(disabled=True)} 
               if current_tab == 'optimize' else {})
        },
        hide_index=True,
        use_container_width=True,
        height=400,
        key=f"well_selection_editor_{current_tab}"
    )
    
    # Кнопки управления выбором
    col_btn1, col_btn2, col_btn3 = st.columns(3)
    
    with col_btn1:
        if st.button("✅ Выбрать все", use_container_width=True, key=f"select_all_{current_tab}"):
            st.session_state[f'select_all_triggered_{current_tab}'] = True
            st.session_state[f'deselect_all_triggered_{current_tab}'] = False
            st.session_state[f'selected_wells_indices_{current_tab}'] = set(range(len(filtered_wells)))
            st.rerun()
    
    with col_btn2:
        if st.button("❌ Снять выделение", use_container_width=True, key=f"deselect_all_{current_tab}"):
            st.session_state[f'deselect_all_triggered_{current_tab}'] = True
            st.session_state[f'select_all_triggered_{current_tab}'] = False
            st.session_state[f'selected_wells_indices_{current_tab}'] = set()
            st.rerun()
    
    with col_btn3:
        if st.button("🎯 Выбрать оптимальные", use_container_width=True, key=f"select_optimal_{current_tab}"):
            avg_flow = df_selection['Дебит, м³/сут'].mean()
            median_depth = df_selection[df_selection['Глубина, м'] != 'Нет данных']['Глубина, м'].median()
            
            selected_indices = set()
            for idx, row in df_selection.iterrows():
                flow = row['Дебит, м³/сут']
                depth = row['Глубина, м']
                if depth != 'Нет данных' and flow > avg_flow and (median_depth * 0.8 <= depth <= median_depth * 1.2):
                    selected_indices.add(idx)
            
            st.session_state[f'selected_wells_indices_{current_tab}'] = selected_indices
            st.session_state[f'select_all_triggered_{current_tab}'] = False
            st.session_state[f'deselect_all_triggered_{current_tab}'] = False
            st.success(f"Выбрано {len(selected_indices)} оптимальных скважин")
            st.rerun()
    
    # Сбрасываем триггеры
    st.session_state[f'select_all_triggered_{current_tab}'] = False
    st.session_state[f'deselect_all_triggered_{current_tab}'] = False
    
    # Обновляем состояние из таблицы
    current_selected_indices = set()
    for idx, row in edited_df.iterrows():
        if row['Выбрать']:
            current_selected_indices.add(idx)
    
    st.session_state[f'selected_wells_indices_{current_tab}'] = current_selected_indices
    
    st.info(f"**Выбрано скважин:** {len(st.session_state[f'selected_wells_indices_{current_tab}'])} из {len(filtered_wells)}")
    
    st.markdown("---")
    
    # ========== ШАГ 4: РАСЧЕТ ==========
    
    st.markdown("### Шаг 4: Запуск расчета")
    
    if st.button("🚀 Запустить расчет", use_container_width=True, 
                type="primary", key=f"run_calculation_{current_tab}"):
        
        selected_indices = st.session_state[f'selected_wells_indices_{current_tab}']
        
        if not selected_indices:
            st.error("❌ Не выбрано ни одной скважины для анализа!")
            return
        
        st.info(f"Выбрано {len(selected_indices)} скважин для анализа")
        
        # Собираем все результаты
        all_variant_results = {}  # скважина -> список вариантов
        calculation_details = []
        error_count = 0
        
        with st.spinner("Выполняется подбор и расчет..."):
            for idx in selected_indices:
                if idx >= len(filtered_wells):
                    continue
                    
                well = filtered_wells[idx]
                well_name = well['name']
                
                try:
                    # Определяем кандидатов для подбора
                    if pump_selection_mode == "Автоматический":
                        pump_candidates = auto_select_pump_candidates(well, current_tab)
                    else:
                        pump_candidates = [new_pump_type]  # Ручной выбор - используем new_pump_type
                    
                    # Для каждого кандидата выполняем расчет
                    well_variants = []
                    
                    for candidate_type in pump_candidates:
                        # Получаем время работы для КПР скважин
                        current_work_hours = None
                        if current_tab == 'optimize':
                            current_work_hours = get_kpr_work_hours_safe(well)
                        
                        # ИСПРАВЛЕНИЕ: для ручного режима передаем new_pump_head, 
                        # для автоматического - None (функция сама возьмет из старого насоса)
                        if pump_selection_mode == "Ручной выбор":
                            # Ручной режим - передаем введенный напор
                            result = calculate_pump_performance(
                                well, candidate_type, new_pump_head,  # ← передаем new_pump_head
                                pump_intake_pressure, specific_indicator,
                                current_tab, current_work_hours
                            )
                        else:
                            # Автоматический режим - передаем None (функция сама разберется)
                            result = calculate_pump_performance(
                                well, candidate_type, None,  # ← передаем None
                                pump_intake_pressure, specific_indicator,
                                current_tab, current_work_hours
                            )
                        
                        if result:
                            well_variants.append(result)
                    
                    if well_variants:
                        all_variant_results[well_name] = well_variants
                        calculation_details.extend(well_variants)
                    else:
                        st.warning(f"Скважина {well_name}: не удалось рассчитать варианты")
                        error_count += 1
                        
                except Exception as e:
                    st.warning(f"Ошибка при расчете для скважины {well_name}: {str(e)}")
                    error_count += 1
        
        if calculation_details:
            # Определяем лучшие варианты для каждой скважины
            best_variants = []
            for well_name, variants in all_variant_results.items():
                best_variant = get_best_variant_for_well(variants)
                if best_variant and best_variant.get('Экономия, руб/сут', 0) > 0:
                    best_variant['Лучший вариант'] = True
                    best_variants.append(best_variant)
            
            # Сохраняем в session_state
            st.session_state[f'pump_calculation_results_{current_tab}'] = calculation_details
            st.session_state[f'pump_best_variants_{current_tab}'] = best_variants
            st.session_state[f'pump_all_variants_{current_tab}'] = all_variant_results
            st.session_state[f'pump_calculation_params_{current_tab}'] = {
                'current_tab': current_tab,
                'pump_selection_mode': pump_selection_mode,
                'specific_indicator': specific_indicator,
                'pump_intake_pressure': pump_intake_pressure,
                'kpr_mode': kpr_mode,
                'selected_cits': selected_cits
            }

            save_data_to_file()
            
            success_count = len(calculation_details)
            st.success(f"✅ Расчет завершен для {len(all_variant_results)} скважин, всего вариантов: {success_count}")
            if error_count > 0:
                st.warning(f"Ошибок при расчете: {error_count}")
            
            st.rerun()
        else:
            st.error("❌ Не удалось выполнить расчет ни для одной скважины")
    
    # ============ ШАГ 5: РЕЗУЛЬТАТЫ ==========
    if (f'pump_calculation_results_{current_tab}' in st.session_state and 
        st.session_state[f'pump_calculation_results_{current_tab}']):
        
        st.markdown("---")
        st.markdown("### 📊 Результаты подбора")
        
        # Получаем данные
        all_results = st.session_state[f'pump_calculation_results_{current_tab}']
        best_variants = st.session_state.get(f'pump_best_variants_{current_tab}', [])
        all_variants = st.session_state.get(f'pump_all_variants_{current_tab}', {})
        
        # Сводная статистика
        col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
        
        # Определяем название колонки с экономией
        savings_col = None
        for col in ['Экономия, руб/сут', 'Экономия, руб/месяц', 'Экономия, руб/год', 'Экономия, руб/сутки']:
            if best_variants and col in best_variants[0]:
                savings_col = col
                break
        
        with col_stat1:
            if best_variants and savings_col:
                total_savings_day = sum(v.get(savings_col, 0) for v in best_variants)
                st.metric("Экономия в сутки", f"{total_savings_day:,.0f} руб")
            else:
                st.metric("Экономия в сутки", "Н/Д")
        
        with col_stat2:
            if best_variants and savings_col:
                total_savings_year = sum(v.get('Экономия, руб/год', v.get(savings_col, 0) * 365) for v in best_variants)
                st.metric("Экономия в год", f"{total_savings_year:,.0f} руб")
            else:
                st.metric("Экономия в год", "Н/Д")
        
        with col_stat3:
            if best_variants and savings_col:
                avg_savings = np.mean([v.get(savings_col, 0) for v in best_variants]) if best_variants else 0
                st.metric("Ср. экономия на скв.", f"{avg_savings:,.0f} руб/сут")
            else:
                st.metric("Ср. экономия", "Н/Д")
        
        with col_stat4:
            optimal_count = len(best_variants)
            total_count = len(set(v.get('Скважина', '') for v in all_results)) if all_results else 0
            st.metric("Оптимальных скважин", f"{optimal_count}/{total_count}")
        
        # Детальная таблица с цветовой индикацией
        st.markdown("#### Детальный расчет по скважинам")
        
        # Создаем DataFrame для отображения
        display_data = []
        for well_name, variants in all_variants.items():
            for variant in variants:
                # Проверяем, является ли вариант лучшим
                is_best = False
                for best in best_variants:
                    if best.get('Скважина') == variant.get('Скважина') and \
                       best.get('Новый тип') == variant.get('Новый тип'):
                        is_best = True
                        break
                
                # Определяем цвет индикации
                if is_best:
                    color_indicator = "🟢"
                    status = "Лучший"
                elif variant.get(savings_col, 0) if savings_col else 0 > 0:
                    color_indicator = "🟡"
                    status = "Альтернатива"
                else:
                    color_indicator = "🔴"
                    status = "Неэффективно"
                
                # Получаем значения для отображения
                required_head = variant.get('Требуемый напор', 0)
                new_head = variant.get('Новый напор', 0)
                old_head_nominal = variant.get('Текущий напор', 0)
                old_freq = variant.get('Текущая частота', 50)
                old_head_real = variant.get('Текущий напор реальный')
                
                # Формируем строку с информацией о напоре
                if required_head > 0:
                    if old_head_real:
                        head_info = (f"{new_head:.0f} м (треб. {required_head:.0f} м, "
                                    f"старый: {old_head_nominal:.0f} м @ {old_freq:.1f} Гц = {old_head_real:.0f} м)")
                    else:
                        head_info = f"{new_head:.0f} м (треб. {required_head:.0f} м)"
                else:
                    head_info = f"{new_head:.0f} м"
                    
                well_data = None
                for w in st.session_state.wells_data:
                    if w['name'] == well_name:
                        well_data = w
                        break
                
                if well_data:
                    kpr_mode_display = get_kpr_mode_safe(well_data)
                    kpr_schedule_display = get_kpr_schedule_display(well_data)
                    work_hours = get_kpr_work_hours_safe(well_data)
                    old_pump = well_data.get('pump_mark', '')
                    
                    # БЕЗОПАСНО получаем старый режим
                    schedule = well_data.get('schedule')
                    if schedule and isinstance(schedule, (list, tuple)) and len(schedule) >= 2:
                        old_regime = f"{schedule[0]}/{schedule[1]}"
                    else:
                        old_regime = "Н/Д"
                else:
                    kpr_mode_display = "Неизвестно"
                    kpr_schedule_display = "Не задан"
                    work_hours = 0
                    old_pump = ""
                    old_regime = "Н/Д"  # ← ВАЖНО: определить здесь!
                
                # Получаем новый режим из варианта
                new_work_time = variant.get('Время работы КПР, ч/сут', 0)
                if new_work_time > 0:
                    # Конвертируем часы в минуты работы/паузы (примерно)
                    cycle_minutes = 60  # базовый цикл
                    work_minutes = cycle_minutes * (new_work_time / 24)
                    pause_minutes = cycle_minutes - work_minutes
                    new_regime = f"{work_minutes:.0f}/{pause_minutes:.0f} мин"
                else:
                    new_regime = "Н/Д"
                
                display_row = {
                    'Скважина': well_name,
                    'Куст': variant.get('Куст', '-'),
                    'Текущий насос': variant.get('Текущий насос', ''),
                    'Новый насос': f"ЭЦН-{variant.get('Новый тип', '')}",
                    'Режим КПР': kpr_mode_display,
                    'Старый режим': old_regime,
                    'Новый режим': new_regime,
                    'Qж, м³/сут': variant.get('Qж, м³/сут', 0),
                    'Экономия, руб/сут': f"{variant.get('Экономия, руб/сут', 0):,.0f}",
                    'Статус': f"{color_indicator} {status}",
                }
                
                # Если хотите видеть причину в подсказке - добавьте это после display_row
                if 'Логика напора' in variant:
                    display_row['Логика'] = variant['Логика напора'][:30] + "..." if len(variant['Логика напора']) > 30 else variant['Логика напора']
                
                # Добавляем экономию, если есть
                if savings_col:
                    display_row['Экономия, руб/сут'] = variant.get(savings_col, 0)
                
                # Добавляем остальные поля, если они есть
                for field in ['Время работы КПР, ч/сут', 'Частота нового, Гц']:
                    if field in variant:
                        display_row[field] = variant.get(field, 0)
                
                display_data.append(display_row)
        
        if display_data:
            df_display = pd.DataFrame(display_data)
            
            # Безопасная сортировка
            sort_columns = []
            if 'Экономия, руб/сут' in df_display.columns:
                sort_columns.append('Экономия, руб/сут')
            if 'Скважина' in df_display.columns and 'Экономия, руб/сут' in df_display.columns:
                df_display = df_display.sort_values(['Экономия, руб/сут', 'Скважина'], ascending=[False, True])
            elif 'Экономия, руб/сут' in df_display.columns:
                df_display = df_display.sort_values('Экономия, руб/сут', ascending=False)
            
            # Отображаем таблицу
            st.dataframe(df_display, use_container_width=True, hide_index=True)
            
            # Показываем информацию о колонках для отладки (опционально)
            with st.expander("🔍 Информация о данных"):
                st.write("**Доступные колонки в данных:**")
                if all_results:
                    st.write(list(all_results[0].keys()) if all_results else "Нет данных")
        else:
            st.warning("Нет данных для отображения")
        
        # ========== ШАГ 6: НАСТРОЙКА И СОХРАНЕНИЕ ==========
        
        st.markdown("---")
        st.markdown("### 💾 Настройка и сохранение результатов")
        
        # Настройки графика КПР
        if kpr_mode == "По времени":
            st.markdown("#### Настройка графика КПР (по времени)")
            
            col_time1, col_time2 = st.columns(2)
            
            with col_time1:
                work_time_minutes = st.number_input(
                    "Время работы, минут",
                    min_value=1,
                    max_value=60,
                    value=5,
                    step=1,
                    key=f"work_time_{current_tab}"
                )
            
            with col_time2:
                # Показываем ПРИМЕР расчета для средней скважины
                if best_variants:
                    avg_q = np.mean([v.get('Qж, м³/сут', 0) for v in best_variants])
                    avg_pump_nominal = np.mean([PUMP_NOMINAL_FLOWS.get(v.get('Новый тип', '125'), 125) 
                                               for v in best_variants])
                    
                    work_min, pause_min = calculate_kpr_schedule_time_mode(
                        avg_q, avg_pump_nominal, work_time_minutes
                    )
                    
                    st.metric("**Пример** время накопления", f"{pause_min:.1f} мин")
                    st.caption(f"Для скважины Q={avg_q:.0f} м³/сут и насоса {avg_pump_nominal:.0f} м³/сут")
                
                st.info("**Важно:** Для каждой скважины будет рассчитано индивидуальное время накопления!")
        
        else:  # По давлению
            st.markdown("#### Настройка графика КПР (по давлению)")
            
            col_press1, col_press2 = st.columns(2)
            
            with col_press1:
                start_pressure = st.number_input(
                    "Давление запуска, атм",
                    min_value=10,
                    max_value=100,
                    value=40,
                    step=1,
                    key=f"start_pressure_{current_tab}"
                )
            
            with col_press2:
                stop_pressure = st.number_input(
                    "Давление остановки, атм",
                    min_value=10,
                    max_value=100,
                    value=42,
                    step=1,
                    key=f"stop_pressure_{current_tab}"
                )
        
        # Автоматическая фильтрация по экономической эффективности
        st.markdown("#### Фильтрация для сохранения")
        
        col_filter1, col_filter2 = st.columns(2)
        
        with col_filter1:
            auto_min_daily = st.number_input(
                "Минимальная экономия для сохранения (руб/сут)",
                min_value=0,
                value=0,  # Сохраняем все с положительной экономией
                step=10,
                key=f"auto_min_daily_{current_tab}"
            )
        
        with col_filter2:
            # Показываем статистику фильтрации
            positive_count = len([v for v in best_variants 
                                 if v.get('Экономия, руб/сут', 0) > auto_min_daily])
            st.metric("Будет сохранено скважин", f"{positive_count}/{len(best_variants)}")
        
        # Показываем предварительный расчет графиков
        if kpr_mode == "По времени" and best_variants:
            with st.expander("📋 Предварительный расчет индивидуальных графиков КПР"):
                schedule_preview = []
                for variant in best_variants[:10]:  # Показываем первые 10
                    well_name = variant['Скважина']
                    q_well = variant.get('Qж, м³/сут', 0)
                    new_pump_type = variant['Новый тип']
                    pump_nominal = PUMP_NOMINAL_FLOWS.get(new_pump_type, 125)
                    
                    work_min, pause_min = calculate_kpr_schedule_time_mode(
                        q_well, pump_nominal, work_time_minutes
                    )
                    
                    schedule_preview.append({
                        'Скважина': well_name,
                        'Q, м³/сут': round(q_well, 1),
                        'Насос': f"ЭЦН-{new_pump_type}",
                        'Ном. подача, м³/сут': pump_nominal,
                        'Работа, мин': work_min,
                        'Накопление, мин': round(pause_min, 1),
                        'Цикл, мин': round(work_min + pause_min, 1),
                        'Коэф. заполнения': f"{(work_min / (work_min + pause_min) * 100):.1f}%"
                    })
                
                if schedule_preview:
                    df_schedule = pd.DataFrame(schedule_preview)
                    st.dataframe(df_schedule, use_container_width=True, hide_index=True)
                
                if len(best_variants) > 10:
                    st.caption(f"... и еще {len(best_variants) - 10} скважин")
        
        # Кнопка сохранения
        if st.button("💾 Сохранить оптимальные варианты в систему", 
                    type="primary", key=f"save_best_{current_tab}"):
            
            if not best_variants:
                st.error("Нет оптимальных вариантов для сохранения")
                return
            
            # Фильтруем по минимальной экономии
            variants_to_save = [v for v in best_variants 
                              if v.get('Экономия, руб/сут', 0) > auto_min_daily]
            
            if not variants_to_save:
                st.warning("Нет вариантов, соответствующих критерию экономии")
                return
            
            saved_count = 0
            error_count = 0
            schedule_details = []  # Для хранения деталей графиков
            
            with st.spinner(f"Сохранение {len(variants_to_save)} оптимальных скважин с индивидуальными графиками..."):
                for variant in variants_to_save:
                    well_name = variant['Скважина']
                    new_pump_type = variant['Новый тип']
                    new_pump_head = variant['Новый напор']
                    q_well = variant.get('Qж, м³/сут', 0)
                    
                    # Формируем график КПР ИНДИВИДУАЛЬНО для каждой скважины
                    if kpr_mode == "По времени":
                        pump_nominal = PUMP_NOMINAL_FLOWS.get(new_pump_type, 125)
                        
                        work_min, pause_min = calculate_kpr_schedule_time_mode(
                            q_well, pump_nominal, work_time_minutes
                        )
                        schedule = [work_min, pause_min]
                        
                        # Сохраняем детали для отчета
                        schedule_details.append({
                            'Скважина': well_name,
                            'Работа': work_min,
                            'Накопление': pause_min,
                            'Q_скв': q_well,
                            'Q_насоса': pump_nominal
                        })
                    else:
                        # По давлению - единые уставки для всех
                        schedule = [start_pressure, stop_pressure]
                    
                    # Сохраняем в систему
                    success = save_ecn_replacement_to_system(
                        well_name, new_pump_type, new_pump_head,
                        schedule, kpr_mode, current_tab
                    )
                    
                    if success:
                        saved_count += 1
                    else:
                        error_count += 1
            
            # Показываем итоговый отчет с индивидуальными графиками
            if saved_count > 0:
                st.success(f"✅ Сохранено {saved_count} оптимальных скважин")
                
                # Показываем детали графиков
                if kpr_mode == "По времени" and schedule_details:
                    with st.expander("📊 Детали индивидуальных графиков КПР"):
                        # Группируем по времени накопления для анализа
                        from collections import defaultdict
                        pause_groups = defaultdict(list)
                        
                        for detail in schedule_details:
                            pause_min = detail['Накопление']
                            group_key = f"{pause_min:.0f} мин"
                            pause_groups[group_key].append(detail['Скважина'])
                        
                        # Статистика по группам
                        st.markdown("**Распределение по времени накопления:**")
                        for pause_time, wells in sorted(pause_groups.items()):
                            st.write(f"- {pause_time}: {len(wells)} скважин")
                            if len(wells) <= 5:
                                st.caption(f"  ({', '.join(wells)})")
                        
                        # Таблица с деталями
                        df_details = pd.DataFrame(schedule_details)
                        st.dataframe(df_details, use_container_width=True, hide_index=True)
                
                st.info(f"""
                **Итоги сохранения:**
                
                1. **Сохранено скважин:** {saved_count}
                2. **Тип перевода:** {'Постоянная → КПР' if current_tab == 'replace' else 'Оптимизация КПР'}
                3. **Режим работы:** {kpr_mode}
                4. **Время работы:** {work_time_minutes if kpr_mode == 'По времени' else 'N/A'} мин
                5. **Новые насосы:** Установлены оптимальные типоразмеры
                6. **Экономия:** от {auto_min_daily} руб/сут и выше
                
                **Ключевое:** Каждой скважине установлен ИНДИВИДУАЛЬНЫЙ график КПР!
                """)
                
                # Кнопки действий
                col_action1, col_action2 = st.columns(2)
                with col_action1:
                    if st.button("🔄 Обновить страницу", key=f"refresh_{current_tab}"):
                        st.rerun()
                with col_action2:
                    if st.button("📊 Экспорт графиков", key=f"export_schedules_{current_tab}"):
                        # Экспорт графиков в CSV
                        if schedule_details:
                            df_export = pd.DataFrame(schedule_details)
                            csv_export = df_export.to_csv(index=False, encoding='utf-8-sig')
                            
                            st.download_button(
                                label="📥 Скачать графики КПР",
                                data=csv_export,
                                file_name=f"графики_КПР_{selected_cits}_{datetime.now().strftime('%Y%m%d')}.csv",
                                mime="text/csv"
                            )
            else:
                st.error("❌ Не удалось сохранить ни одной скважины")
        
        # ========== ЭКСПОРТ ==========
        
        st.markdown("---")
        st.markdown("### 📥 Экспорт результатов")
        
        # Создание отчета
        if st.button("📊 Создать ТЭО отчет", key=f"create_report_{current_tab}"):
            params = st.session_state.get(f'pump_calculation_params_{current_tab}', {})
            params.update({
                'kpr_mode': kpr_mode,
                'min_daily_savings': auto_min_daily,
                'total_wells': len(all_variants),
                'optimal_wells': len(best_variants)
            })
            
            excel_file = create_ecn_replacement_report(all_results, params)
            
            mode_name = "замена_постоянных" if current_tab == 'replace' else "оптимизация_кпр"
            st.download_button(
                label="📥 Скачать полный отчет в Excel",
                data=excel_file.getvalue(),
                file_name=f"ТЭО_{mode_name}_{selected_cits}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        
        # Экспорт только лучших вариантов
        if best_variants:
            df_best = pd.DataFrame(best_variants)
            csv_best = df_best.to_csv(index=False, encoding='utf-8-sig')
            
            st.download_button(
                label="📋 Скачать оптимальные варианты (CSV)",
                data=csv_best,
                file_name=f"оптимальные_варианты_{current_tab}_{selected_cits}.csv",
                mime="text/csv",
                use_container_width=True
            )

# ============================================================
# НОВЫЕ ФУНКЦИИ ДЛЯ ОТЧЕТОВ
# ============================================================

@st.cache_data(ttl=3600)
def create_pressure_stabilization_report(optimization_result):
    """Создает детальный отчет по стабилизации давления в Excel"""
    _load_openpyxl() 
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        workbook = writer.book
        
        # ================= ЛИСТ 1: СВОДКА И РЕКОМЕНДАЦИИ =================
        ws1 = workbook.create_sheet("Сводка и рекомендации")
        
        # Заголовок
        ws1.merge_cells('A1:G1')
        ws1['A1'] = "РАСЧЕТ СТАБИЛИЗАЦИИ ДАВЛЕНИЯ В КОЛЛЕКТОРАХ"
        ws1['A1'].font = Font(bold=True, size=14)
        ws1['A1'].alignment = Alignment(horizontal='center')
        
        # Информация об объекте
        data_summary = [
            ["Объект:", f"ТПП '{optimization_result.get('tpp', 'VQ-BADнефтегаз')}'"],
            ["ЦИТС:", optimization_result.get('cits', 'ЦИТС VQ-BAD')],
            ["ЦДНГ:", optimization_result.get('cdng', 'ЦДНГ-1')],
            ["Куст:", optimization_result.get('cluster', 'Не указан')],
            ["Дата расчета:", optimization_result['timestamp'].strftime("%d.%m.%Y %H:%M")],
            ["Текущее время:", optimization_result.get('current_time', 'Не указано')],
            ["Коэффициент загрузки:", f"{optimization_result.get('target_coefficient', 0.7)*100:.0f}%"],
            ["", ""],
            ["СТАТИСТИКА ДО/ПОСЛЕ ОПТИМИЗАЦИИ:", ""],
            ["• Средний дебит:", f"{optimization_result['stats']['avg_flow_before']:.1f} → {optimization_result['stats']['avg_flow_after']:.1f} м³/час ({optimization_result['stats']['flow_improvement']:+.1f}%)"],
            ["• Стандартное отклонение:", f"{optimization_result['stats']['std_before']:.1f} → {optimization_result['stats']['std_after']:.1f} м³/час ({optimization_result['stats']['stability_improvement']:+.1f}%)"],
            ["• Количество пиков:", f"{optimization_result['stats']['peaks_before']} → {optimization_result['stats']['peaks_after']} ({optimization_result['stats']['peaks_improvement']:+.1f}%)"],
            ["• Средняя величина пиков:", f"{optimization_result['stats']['peak_magnitude_before']:.1f} → {optimization_result['stats']['peak_magnitude_after']:.1f} м³/час"],
            ["", ""],
            ["ТЕХНИЧЕСКИЕ РЕЗУЛЬТАТЫ:", ""],
            ["• Целевой дебит:", f"{optimization_result['stats']['target_flow']:.1f} м³/час"],
            ["• Достижение цели:", f"{optimization_result['stats']['target_achievement']:.1f}%"],
            ["• Общая эффективность:", f"{optimization_result['stats']['efficiency']:.1f}%"],
            ["• Улучшение стабильности:", f"{optimization_result['stats']['stability_improvement']:+.1f}%"],
            ["• Снижение пиковых нагрузок:", f"{optimization_result['stats']['peaks_improvement']:+.1f}%"],
            ["", ""],
            ["РЕКОМЕНДАЦИИ ДЛЯ ОПЕРАТОРА:", ""],
            ["1. Внедрить фазовые сдвиги согласно Листу 3", ""],
            ["2. Контролировать давление на сепараторе 1-ЦПС каждые 4 часа", ""],
            ["3. При отклонении давления >5% скорректировать по алгоритму", ""],
            ["4. Исключенные скважины (работают по давлению) - не менять", ""],
            ["5. Внедрение выполнять в течение 2 смен", ""],
            ["6. Контрольный замер через 24 часа после внедрения", ""],
        ]
        
        for i, row in enumerate(data_summary, start=3):
            for j, value in enumerate(row):
                cell = ws1.cell(row=i, column=j+1, value=value)
                if i == 3 or "СТАТИСТИКА" in str(value) or "РЕЗУЛЬТАТЫ" in str(value) or "РЕКОМЕНДАЦИИ" in str(value):
                    cell.font = Font(bold=True)
        
        # ================= ЛИСТ 2: СРАВНИТЕЛЬНАЯ СТАТИСТИКА =================
        ws2 = workbook.create_sheet("Сравнительная статистика")
        
        comparison_data = {
            'Показатель': [
                'Средний дебит',
                'Стандартное отклонение',
                'Коэффициент вариации',
                'Количество пиков (>1.5σ)',
                'Средняя величина пиков',
                'Максимальный дебит',
                'Минимальный дебит',
                'Время работы в пределах цели (±5%)'
            ],
            'До оптимизации': [
                f"{optimization_result['stats']['avg_flow_before']:.1f}",
                f"{optimization_result['stats']['std_before']:.1f}",
                f"{optimization_result['stats']['std_before']/optimization_result['stats']['avg_flow_before']:.3f}" if optimization_result['stats']['avg_flow_before'] > 0 else "0",
                f"{optimization_result['stats']['peaks_before']}",
                f"{optimization_result['stats']['peak_magnitude_before']:.1f}" if optimization_result['stats']['peaks_before'] > 0 else "0",
                f"{optimization_result['stats']['avg_flow_before'] + 2*optimization_result['stats']['std_before']:.1f}",
                f"{optimization_result['stats']['avg_flow_before'] - 2*optimization_result['stats']['std_before']:.1f}",
                "Расчетный"
            ],
            'После оптимизации': [
                f"{optimization_result['stats']['avg_flow_after']:.1f}",
                f"{optimization_result['stats']['std_after']:.1f}",
                f"{optimization_result['stats']['std_after']/optimization_result['stats']['avg_flow_after']:.3f}" if optimization_result['stats']['avg_flow_after'] > 0 else "0",
                f"{optimization_result['stats']['peaks_after']}",
                f"{optimization_result['stats']['peak_magnitude_after']:.1f}" if optimization_result['stats']['peaks_after'] > 0 else "0",
                f"{optimization_result['stats']['avg_flow_after'] + 2*optimization_result['stats']['std_after']:.1f}",
                f"{optimization_result['stats']['avg_flow_after'] - 2*optimization_result['stats']['std_after']:.1f}",
                f"{(optimization_result['stats']['target_achievement']/100)*24:.1f} ч/сут"
            ],
            'Изменение': [
                f"{optimization_result['stats']['flow_improvement']:+.1f}%",
                f"{optimization_result['stats']['stability_improvement']:+.1f}%",
                f"{(optimization_result['stats']['std_before']/optimization_result['stats']['avg_flow_before'] - optimization_result['stats']['std_after']/optimization_result['stats']['avg_flow_after'])/(optimization_result['stats']['std_before']/optimization_result['stats']['avg_flow_before'])*100:+.1f}%" if optimization_result['stats']['avg_flow_before'] > 0 and optimization_result['stats']['avg_flow_after'] > 0 else "0%",
                f"{optimization_result['stats']['peaks_improvement']:+.1f}%",
                f"{(optimization_result['stats']['peak_magnitude_after'] - optimization_result['stats']['peak_magnitude_before'])/optimization_result['stats']['peak_magnitude_before']*100:+.1f}%" if optimization_result['stats']['peaks_before'] > 0 else "0%",
                f"{(optimization_result['stats']['avg_flow_after'] + 2*optimization_result['stats']['std_after']) - (optimization_result['stats']['avg_flow_before'] + 2*optimization_result['stats']['std_before']):+.1f}",
                f"{(optimization_result['stats']['avg_flow_after'] - 2*optimization_result['stats']['std_after']) - (optimization_result['stats']['avg_flow_before'] - 2*optimization_result['stats']['std_before']):+.1f}",
                f"+{(optimization_result['stats']['target_achievement']/100)*24 - 10:.1f} ч/сут"
            ],
            'Единицы измерения': [
                'м³/час', 'м³/час', '-', 'шт/сут', 'м³/час', 'м³/час', 'м³/час', 'ч/сут'
            ]
        }
        
        df_comparison = pd.DataFrame(comparison_data)
        
        # Записываем заголовок
        ws2.merge_cells('A1:E1')
        ws2['A1'] = "СРАВНИТЕЛЬНАЯ СТАТИСТИКА ДО И ПОСЛЕ ОПТИМИЗАЦИИ"
        ws2['A1'].font = Font(bold=True, size=12)
        ws2['A1'].alignment = Alignment(horizontal='center')
        
        # Записываем таблицу
        for r_idx, row in enumerate(df_comparison.itertuples(index=False), start=3):
            for c_idx, value in enumerate(row, start=1):
                ws2.cell(row=r_idx, column=c_idx, value=value)
        
        # Форматирование заголовков
        for col in range(1, 6):
            cell = ws2.cell(row=3, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # ================= ЛИСТ 3: СДВИГИ ФАЗ И РЕКОМЕНДАЦИИ =================
        ws3 = workbook.create_sheet("Сдвиги фаз и рекомендации")
        
        # Получаем рекомендации
        recommendations = calculate_next_launch_times(
            optimization_result['wells_data'],
            optimization_result['phases_dict'],
            optimization_result['current_time']
        )
        
        # Преобразуем в DataFrame
        df_recommendations = pd.DataFrame(recommendations)
        
        # Заголовок
        ws3.merge_cells('A1:H1')
        ws3['A1'] = "РЕКОМЕНДАЦИИ ПО СДВИГАМ ФАЗ ДЛЯ ОПЕРАТОРА"
        ws3['A1'].font = Font(bold=True, size=12)
        ws3['A1'].alignment = Alignment(horizontal='center')
        
        # Записываем таблицу
        if not df_recommendations.empty:
            # Заголовки
            headers = list(df_recommendations.columns)
            for col_idx, header in enumerate(headers, start=1):
                cell = ws3.cell(row=3, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            # Данные
            for r_idx, row in df_recommendations.iterrows():
                for c_idx, value in enumerate(row, start=1):
                    ws3.cell(row=r_idx+4, column=c_idx, value=value)
        
        # Примечания
        notes_row = len(df_recommendations) + 6 if not df_recommendations.empty else 6
        ws3.cell(row=notes_row, column=1, value="ПРИМЕЧАНИЯ:").font = Font(bold=True)
        ws3.cell(row=notes_row+1, column=1, value="1. 'Запустить раньше' - установите время запуска как в колонке 'Оптим. время'")
        ws3.cell(row=notes_row+2, column=1, value="2. 'Отложить запуск' - пропустите ближайший КПР запуск, дождитесь оптимального времени")
        ws3.cell(row=notes_row+3, column=1, value="3. 'Оставить как есть' - скважина уже работает в оптимальном режиме")
        ws3.cell(row=notes_row+4, column=1, value="4. 'НЕ МЕНЯТЬ (исключена)' - скважина работает по давлению, оставьте текущий график")
        ws3.cell(row=notes_row+5, column=1, value="5. Контроль: после внедрения отслеживайте стабильность давления в коллекторе")
        
        # ================= ЛИСТ 4: ГРАФИК ДЕБИТА (табличные данные) =================
        ws4 = workbook.create_sheet("Данные для графиков")
        
        # Создаем данные для графика
        optimizer = PressureStabilizationOptimizer(optimization_result['wells_data'])
        
        time_points = np.arange(0, 24 * 60, 5)  # 5-минутный шаг
        time_hours = time_points / 60
        
        zero_phases = {name: 0 for name in optimization_result['phases_dict'].keys()}
        flows_before = []
        flows_after = []
        
        for t in time_points:
            flow_before = optimizer.calculate_total_flow_at_time(t, zero_phases)
            flow_after = optimizer.calculate_total_flow_at_time(t, optimization_result['phases_dict'])
            flows_before.append(flow_before)
            flows_after.append(flow_after)
        
        # Создаем DataFrame
        graph_data = {
            'Время (часы)': time_hours,
            'Дебит до оптимизации (м³/час)': flows_before,
            'Дебит после оптимизации (м³/час)': flows_after,
            'Целевой дебит (м³/час)': [optimization_result['stats']['target_flow']] * len(time_hours)
        }
        
        df_graph = pd.DataFrame(graph_data)
        
        # Заголовок
        ws4.merge_cells('A1:D1')
        ws4['A1'] = "ТАБЛИЧНЫЕ ДАННЫЕ ДЛЯ ПОСТРОЕНИЯ ГРАФИКОВ ДЕБИТА"
        ws4['A1'].font = Font(bold=True, size=12)
        ws4['A1'].alignment = Alignment(horizontal='center')
        
        # Записываем таблицу
        for r_idx, row in enumerate(df_graph.itertuples(index=False), start=3):
            for c_idx, value in enumerate(row, start=1):
                ws4.cell(row=r_idx, column=c_idx, value=value)
        
        # Заголовки столбцов
        for col_idx, header in enumerate(df_graph.columns, start=1):
            cell = ws4.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # ================= ЛИСТ 5: ТЕХНИЧЕСКИЕ ПАРАМЕТРЫ СКВАЖИН =================
        ws5 = workbook.create_sheet("Технические параметры")
        
        # Подготавливаем данные скважин
        well_data_list = []
        for well in optimization_result['wells_data']:
            well_data_list.append({
                'Скважина': well['name'],
                'Тип': 'Постоянная' if well['operation_mode'] == 'constant' else 'КПР',
                'Дебит (м³/сут)': well['flow_rate'],
                'Режим КПР': f"{well['schedule'][0]}/{well['schedule'][1]}" if well.get('schedule') else '-',
                'Время запуска КПР': well.get('base_launch_time', '-'),
                'Статус': 'Активна' if well.get('is_active', True) else 'Остановлена',
                'Сдвиг фазы (мин)': optimization_result['phases_dict'].get(well['name'], 0),
                'Участвует в сдвиге': 'Нет' if well.get('exclude_from_shift', False) else 'Да' if well['operation_mode'] == 'kpr' else '-'
            })
        
        df_wells = pd.DataFrame(well_data_list)
        
        # Заголовок
        ws5.merge_cells('A1:H1')
        ws5['A1'] = "ТЕХНИЧЕСКИЕ ПАРАМЕТРЫ СКВАЖИН КУСТА"
        ws5['A1'].font = Font(bold=True, size=12)
        ws5['A1'].alignment = Alignment(horizontal='center')
        
        # Записываем таблицу
        if not df_wells.empty:
            # Заголовки
            headers = list(df_wells.columns)
            for col_idx, header in enumerate(headers, start=1):
                cell = ws5.cell(row=3, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            # Данные
            for r_idx, row in df_wells.iterrows():
                for c_idx, value in enumerate(row, start=1):
                    ws5.cell(row=r_idx+4, column=c_idx, value=value)

        # ================= ЛИСТ 6: РАБОТА СКВАЖИН =================
        ws6 = workbook.create_sheet("Работа скважин")
        
        # Заголовок
        ws6.merge_cells('A1:C1')
        ws6['A1'] = "КОЛИЧЕСТВО РАБОТАЮЩИХ СКВАЖИН В ТЕЧЕНИЕ СУТОК"
        ws6['A1'].font = Font(bold=True, size=12)
        ws6['A1'].alignment = Alignment(horizontal='center')
        
        # Получаем данные
        count_data = optimizer.calculate_working_wells_count(optimization_result['phases_dict'])
        
        # Создаем DataFrame
        work_data = {
            'Время (часы)': count_data['time_hours'],
            'Работает до оптимизации': count_data['count_before'],
            'Работает после оптимизации': count_data['count_after']
        }
        
        df_work = pd.DataFrame(work_data)
        
        # Записываем таблицу
        for r_idx, row in enumerate(df_work.itertuples(index=False), start=3):
            for c_idx, value in enumerate(row, start=1):
                ws6.cell(row=r_idx, column=c_idx, value=value)
        
        # Заголовки столбцов
        for col_idx, header in enumerate(df_work.columns, start=1):
            cell = ws6.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # ================= НАСТРОЙКА ШИРИНЫ СТОЛБЦОВ =================
        def adjust_column_widths(worksheet):
            """Безопасная настройка ширины столбцов с проверкой на MergedCell"""
            for column in worksheet.columns:
                column_letter = None
                max_length = 0
                
                # Находим первую не-объединенную ячейку в столбце
                for cell in column:
                    if cell.__class__.__name__ != 'MergedCell':
                        column_letter = cell.column_letter
                        break
                
                if column_letter:
                    # Измеряем максимальную длину
                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Применяем ко всем листам
        for ws in workbook.worksheets:
            adjust_column_widths(ws)
    
    # Удаляем дефолтный лист
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])
    
    output.seek(0)
    return output

def create_kpr_potential_report(optimization_result, well_data):
    """Создает отчет по потенциалу КПР скважины"""
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        workbook = writer.book
        
        # ================= ЛИСТ 1: ДИАГНОСТИКА СКВАЖИНЫ =================
        ws1 = workbook.create_sheet("Диагностика скважины")
        
        # Заголовок
        ws1.merge_cells('A1:D1')
        ws1['A1'] = f"ДИАГНОСТИКА И АНАЛИЗ ПОТЕНЦИАЛА СКВАЖИНЫ: {well_data['name']}"
        ws1['A1'].font = Font(bold=True, size=14)
        ws1['A1'].alignment = Alignment(horizontal='center')
        
        # Основная информация
        info_data = [
            ["СКВАЖИНА:", well_data['name']],
            ["Куст:", well_data.get('cluster', 'Не указан')],
            ["ЦДНГ:", well_data.get('cdng', 'Не указан')],
            ["ЦИТС:", well_data.get('cits', 'Не указан')],
            ["Дата анализа:", datetime.now().strftime("%d.%m.%Y %H:%M")],
            ["", ""],
            ["ГЕОЛОГО-ТЕХНИЧЕСКИЕ УСЛОВИЯ:", ""],
            ["• Пластовое давление (Pпл):", f"{well_data.get('p_pl', 0):.1f} атм"],
            ["• Давление насыщения (Pнас):", f"{well_data.get('p_nas', 0):.1f} атм"],
            ["• Буферное давление (Рб/ф):", f"{well_data.get('buffer_pressure', 0):.1f} атм"],
            ["• Давление на забое (Pзаб):", f"{well_data.get('p_zab_vdp', well_data.get('buffer_pressure', 0) + 10):.1f} атм"],
        ]
        
        # Анализ проблемы с газом
        p_nas = well_data.get('p_nas', 0)
        p_zab = well_data.get('p_zab_vdp', well_data.get('buffer_pressure', 0) + 10)
        
        if p_nas > 0 and p_zab < p_nas:
            gas_problem = f"ДА ⚠️ (Pзаб {p_zab:.1f} атм < Pнас {p_nas:.1f} атм)"
            severity = "ВЫСОКАЯ" if p_zab < 0.75 * p_nas else "СРЕДНЯЯ"
        else:
            gas_problem = "НЕТ ✓"
            severity = "-"
        
        info_data.append(["• Проблема с свободным газом:", gas_problem])
        if severity != "-":
            info_data.append(["• Степень серьезности:", severity])
        
        info_data.extend([
            ["• Коэффициент продуктивности:", f"{well_data.get('prod_coef', 0):.2f} м³/сут/атм"],
            ["", ""],
            ["ТЕКУЩИЙ РЕЖИМ КПР:", ""],
            ["• Работа/Накопление:", f"{well_data.get('schedule', [15, 45])[0]}/{well_data.get('schedule', [15, 45])[1]} часов"],
            ["• Суточный дебит жидкости:", f"{well_data.get('flow_rate', 0):.1f} м³/сут"],
            ["• Обводненность:", f"{well_data.get('water_cut', 0):.1f}%"],
        ])
        
        # Расчет загрузки насоса
        pump_flow = well_data.get('pump_flow', 0)
        if pump_flow > 0:
            t_rab = well_data.get('t_rab', 24)
            utilization = (well_data.get('flow_rate', 0) * (24 / t_rab)) / pump_flow if t_rab > 0 else 0
            info_data.append(["• Загрузка насоса:", f"{utilization:.1%}"])
            if utilization < 0.75 or utilization > 1.25:
                info_data.append(["• Эффективность насоса:", "НИЗКАЯ ⚠️"])
            else:
                info_data.append(["• Эффективность насоса:", "НОРМАЛЬНАЯ ✓"])
        
        info_data.extend([
            ["• Глубина насоса:", f"{well_data.get('pump_depth', 0):.0f} м"],
            ["• Динамический уровень:", f"{well_data.get('h_din', 0):.0f} м"],
        ])
        
        # Записываем информацию
        for i, row in enumerate(info_data, start=3):
            for j, value in enumerate(row):
                cell = ws1.cell(row=i, column=j+1, value=value)
                if i == 3 or "ГЕОЛОГО" in str(value) or "ТЕКУЩИЙ" in str(value):
                    cell.font = Font(bold=True)
        
        # ================= ЛИСТ 2: АНАЛИЗ ПРОБЛЕМ И РЕКОМЕНДАЦИИ =================
        ws2 = workbook.create_sheet("Анализ проблем")
        
        # Заголовок
        ws2.merge_cells('A1:E1')
        ws2['A1'] = "АНАЛИЗ ПРОБЛЕМ И РЕКОМЕНДАЦИИ ПО ОПТИМИЗАЦИИ"
        ws2['A1'].font = Font(bold=True, size=12)
        ws2['A1'].alignment = Alignment(horizontal='center')
        
        # Данные анализа
        analysis_data = []
        
        # 1. Проблема с газом
        if p_nas > 0 and p_zab < p_nas:
            analysis_data.append({
                'Параметр': 'Pзаб < Pнас',
                'Текущее значение': f"{p_zab:.1f} < {p_nas:.1f} атм",
                'Проблема': 'Свободный газ в насосе',
                'Рекомендация': 'Сократить время работы',
                'Ожидаемый эффект': '+12-15% к добыче'
            })
        
        # 2. Эффективность насоса
        if pump_flow > 0:
            t_rab = well_data.get('t_rab', 24)
            utilization = (well_data.get('flow_rate', 0) * (24 / t_rab)) / pump_flow if t_rab > 0 else 0
            
            if utilization < 0.75:
                analysis_data.append({
                    'Параметр': 'Загрузка насоса',
                    'Текущее значение': f"{utilization:.1%}",
                    'Проблема': 'Неоптимальная зона работы',
                    'Рекомендация': 'Увеличить время цикла',
                    'Ожидаемый эффект': '+8-10% к КПД'
                })
            elif utilization > 1.25:
                analysis_data.append({
                    'Параметр': 'Загрузка насоса',
                    'Текущее значение': f"{utilization:.1%}",
                    'Проблема': 'Перегрузка насоса',
                    'Рекомендация': 'Уменьшить время работы',
                    'Ожидаемый эффект': 'Снижение износа'
                })
        
        # 3. Частота циклов
        schedule = well_data.get('schedule', [15, 45])
        cycle_time = schedule[0] + schedule[1]
        cycles_per_day = 24 / (cycle_time / 60) if cycle_time > 0 else 0
        
        if cycles_per_day > 0.5:  # Более 12 циклов в сутки
            analysis_data.append({
                'Параметр': 'Количество циклов',
                'Текущее значение': f"{cycles_per_day:.1f}/сут",
                'Проблема': 'Частые пуски/остановки',
                'Рекомендация': 'Увеличить время накопления',
                'Ожидаемый эффект': '-15-20% к износу'
            })
        
        # Создаем DataFrame
        df_analysis = pd.DataFrame(analysis_data)
        
        # Записываем заголовки
        headers = ['Параметр', 'Текущее значение', 'Проблема', 'Рекомендация', 'Ожидаемый эффект']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws2.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # Записываем данные
        for r_idx, row in df_analysis.iterrows():
            for c_idx, value in enumerate(row, start=1):
                ws2.cell(row=r_idx+4, column=c_idx, value=value)
        
        # Информация о выбранном сценарии
        if optimization_result and optimization_result.get('best_scenario'):
            best_scenario = optimization_result['best_scenario']
            scenario_row = len(df_analysis) + 6
            
            ws2.cell(row=scenario_row, column=1, value="ВЫБРАННЫЙ СЦЕНАРИЙ ОПТИМИЗАЦИИ:").font = Font(bold=True)
            ws2.cell(row=scenario_row+1, column=1, value=f"Тип: Сценарий {best_scenario['scenario']}")
            ws2.cell(row=scenario_row+2, column=1, value=f"Причина: {best_scenario['reason']}")
            
            if best_scenario['scenario'] == 'A':
                ws2.cell(row=scenario_row+3, column=1, value=f"Минимальное допустимое Pзаб: {best_scenario['p_zab_min_allowed']:.1f} атм")
                ws2.cell(row=scenario_row+4, column=1, value=f"Ожидаемое увеличение добычи: +{best_scenario['expected_increase_percent']:.1f}%")
            else:
                ws2.cell(row=scenario_row+3, column=1, value=f"Загрузка насоса: {best_scenario['pump_utilization_current']:.1%} → {best_scenario['pump_utilization_target']:.1%}")
                ws2.cell(row=scenario_row+4, column=1, value=f"Ожидаемая экономия энергии: -{best_scenario['energy_saving_percent']:.1f}%")
        
        # ================= ЛИСТ 3: РАСЧЕТ НОВОГО РЕЖИМА =================
        if optimization_result and optimization_result.get('best_scenario'):
            ws3 = workbook.create_sheet("Расчет нового режима")
            
            best_scenario = optimization_result['best_scenario']
            current_schedule = optimization_result['current_schedule']
            
            # Заголовок
            ws3.merge_cells('A1:D1')
            ws3['A1'] = "РАСЧЕТ ПАРАМЕТРОВ НОВОГО РЕЖИМА РАБОТЫ"
            ws3['A1'].font = Font(bold=True, size=12)
            ws3['A1'].alignment = Alignment(horizontal='center')
            
            # Данные расчета
            calc_data = [
                ["Параметр", "Старый режим", "Новый режим", "Изменение"],
                ["Время работы", f"{current_schedule[0]} ч", f"{best_scenario['recommended_work_time']:.1f} ч", f"{best_scenario['recommended_work_time'] - current_schedule[0]:+.1f} ч ({((best_scenario['recommended_work_time'] - current_schedule[0])/current_schedule[0]*100):+.1f}%)"],
                ["Время накопления", f"{current_schedule[1]} ч", f"{best_scenario['recommended_pause_time']:.1f} ч", f"{best_scenario['recommended_pause_time'] - current_schedule[1]:+.1f} ч ({((best_scenario['recommended_pause_time'] - current_schedule[1])/current_schedule[1]*100):+.1f}%)"],
                ["Длительность цикла", f"{current_schedule[0] + current_schedule[1]} ч", f"{best_scenario['recommended_work_time'] + best_scenario['recommended_pause_time']:.1f} ч", f"{(best_scenario['recommended_work_time'] + best_scenario['recommended_pause_time']) - (current_schedule[0] + current_schedule[1]):+.1f} ч"],
                ["Циклов в сутки", f"{24/((current_schedule[0] + current_schedule[1])/60):.2f}", f"{24/((best_scenario['recommended_work_time'] + best_scenario['recommended_pause_time'])/60):.2f}", f"{24/((best_scenario['recommended_work_time'] + best_scenario['recommended_pause_time'])/60) - 24/((current_schedule[0] + current_schedule[1])/60):+.2f}"],
                ["Среднее Pзаб", f"{p_zab:.1f} атм", f"{p_zab + (p_nas - p_zab) * 0.3:.1f} атм", f"+{(p_nas - p_zab) * 0.3:.1f} атм"],
            ]
            
            # Записываем данные
            for r_idx, row in enumerate(calc_data, start=3):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws3.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 3:  # Заголовки
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            # Физическое обоснование
            notes_row = len(calc_data) + 5
            ws3.cell(row=notes_row, column=1, value="ФИЗИЧЕСКОЕ ОБОСНОВАНИЕ:").font = Font(bold=True)
            
            if best_scenario['scenario'] == 'A':
                ws3.cell(row=notes_row+1, column=1, value=f"1. Pзаб минимальное допустимое: 0.75 × Pнас = 0.75 × {p_nas:.1f} = {0.75 * p_nas:.1f} атм")
                ws3.cell(row=notes_row+2, column=1, value=f"2. Депрессия от Pпл до Pзабmin: {well_data.get('p_pl', 0) - 0.75 * p_nas:.1f} атм")
                ws3.cell(row=notes_row+3, column=1, value=f"3. Время до достижения Pзабmin: {best_scenario.get('time_to_limit', 0):.1f} ч")
                ws3.cell(row=notes_row+4, column=1, value=f"4. Новое время работы с запасом 10%: {best_scenario['recommended_work_time']:.1f} ч")
            else:
                ws3.cell(row=notes_row+1, column=1, value="1. Оптимальная зона работы насоса: 80-100% от номинальной подачи")
                ws3.cell(row=notes_row+2, column=1, value=f"2. Текущая загрузка: {best_scenario['pump_utilization_current']:.1%}")
                ws3.cell(row=notes_row+3, column=1, value=f"3. Целевая загрузка: {best_scenario['pump_utilization_target']:.1%}")
                ws3.cell(row=notes_row+4, column=1, value=f"4. Коэффициент изменения: {best_scenario['recommended_work_time'] / current_schedule[0]:.2f}")
        
        # ================= ЛИСТ 4: СРАВНЕНИЕ СЦЕНАРИЕВ =================
        if optimization_result and (optimization_result.get('scenario_a') or optimization_result.get('scenario_b')):
            ws4 = workbook.create_sheet("Сравнение сценариев")
            
            # Заголовок
            ws4.merge_cells('A1:E1')
            ws4['A1'] = "СРАВНИТЕЛЬНЫЙ АНАЛИЗ ВАРИАНТОВ ОПТИМИЗАЦИИ"
            ws4['A1'].font = Font(bold=True, size=12)
            ws4['A1'].alignment = Alignment(horizontal='center')
            
            # Подготавливаем данные
            scenarios = []
            
            if optimization_result.get('scenario_a'):
                scenario_a = optimization_result['scenario_a']
                scenarios.append({
                    'Сценарий': 'А (Борьба с газом)',
                    'Время работы': f"{scenario_a['recommended_work_time']:.1f} ч",
                    'Время накопления': f"{scenario_a['recommended_pause_time']:.1f} ч",
                    'Эффект': f"+{scenario_a['expected_increase_percent']:.1f}% к добыче",
                    'Приоритет': 'ВЫСОКИЙ' if p_zab < 0.85 * p_nas else 'СРЕДНИЙ'
                })
            
            if optimization_result.get('scenario_b'):
                scenario_b = optimization_result['scenario_b']
                scenarios.append({
                    'Сценарий': 'Б (Энергоэффективность)',
                    'Время работы': f"{scenario_b['recommended_work_time']:.1f} ч",
                    'Время накопление': f"{scenario_b['recommended_pause_time']:.1f} ч",
                    'Эффект': f"-{scenario_b['energy_saving_percent']:.1f}% энергии",
                    'Приоритет': 'ВЫСОКИЙ' if pump_flow > 0 and (well_data.get('flow_rate', 0) * (24/well_data.get('t_rab', 24)) / pump_flow) < 0.6 else 'СРЕДНИЙ'
                })
            
            # Текущий режим
            scenarios.append({
                'Сценарий': 'Текущий режим',
                'Время работы': f"{well_data.get('schedule', [15, 45])[0]} ч",
                'Время накопления': f"{well_data.get('schedule', [15, 45])[1]} ч",
                'Эффект': 'Базовый',
                'Приоритет': '-'
            })
            
            # Создаем DataFrame
            df_scenarios = pd.DataFrame(scenarios)
            
            # Записываем заголовки
            headers = list(df_scenarios.columns)
            for col_idx, header in enumerate(headers, start=1):
                cell = ws4.cell(row=3, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            # Записываем данные
            for r_idx, row in df_scenarios.iterrows():
                for c_idx, value in enumerate(row, start=1):
                    ws4.cell(row=r_idx+4, column=c_idx, value=value)
            
            # Рекомендации по выбору
            notes_row = len(df_scenarios) + 6
            ws4.cell(row=notes_row, column=1, value="РЕКОМЕНДАЦИИ ПО ВЫБОРУ СЦЕНАРИЯ:").font = Font(bold=True)
            ws4.cell(row=notes_row+1, column=1, value="• Сценарий А: При Pзаб < 0.85×Pнас (проблема с газом)")
            ws4.cell(row=notes_row+2, column=1, value="• Сценарий Б: При загрузке насоса <60% или >120%")
            ws4.cell(row=notes_row+3, column=1, value="• Текущий режим: Если нет явных проблем")
        
        # ================= ЛИСТ 5: ДАННЫЕ ДЛЯ МОДЕЛИРОВАНИЯ =================
        if 'cycle_simulation' in st.session_state:
            ws5 = workbook.create_sheet("Данные моделирования")
            
            sim_data = st.session_state.cycle_simulation
            
            # Заголовок
            ws5.merge_cells('A1:D1')
            ws5['A1'] = "ДАННЫЕ ДИНАМИЧЕСКОГО МОДЕЛИРОВАНИЯ ЦИКЛА"
            ws5['A1'].font = Font(bold=True, size=12)
            ws5['A1'].alignment = Alignment(horizontal='center')
            
            # Создаем DataFrame с данными
            sim_df_data = {
                'Время цикла (час)': sim_data['time'] / 60,
                'Давление на забое (атм)': sim_data['pwf'],
                'Дебит жидкости (м³/сут)': sim_data['q'],
                'Коэффициент продуктивности (м³/сут/атм)': sim_data['kpr'],
                'Фаза': ['Накопление' if t > sim_data['work_minutes'] else 'Отбор' for t in sim_data['time']]
            }
            
            df_simulation = pd.DataFrame(sim_df_data)
            
            # Записываем заголовки
            headers = list(df_simulation.columns)
            for col_idx, header in enumerate(headers, start=1):
                cell = ws5.cell(row=3, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            # Записываем данные (первые 50 строк для компактности)
            max_rows = min(50, len(df_simulation))
            for r_idx in range(max_rows):
                for c_idx, value in enumerate(df_simulation.iloc[r_idx], start=1):
                    ws5.cell(row=r_idx+4, column=c_idx, value=value)
            
            if len(df_simulation) > max_rows:
                ws5.cell(row=max_rows+5, column=1, value=f"... и еще {len(df_simulation) - max_rows} строк")
        
        # ================= НАСТРОЙКА ШИРИНЫ СТОЛБЦОВ =================
        def adjust_column_widths(worksheet):
            """Безопасная настройка ширины столбцов с проверкой на MergedCell"""
            for column in worksheet.columns:
                column_letter = None
                max_length = 0
                
                # Находим первую не-объединенную ячейку в столбце
                for cell in column:
                    if cell.__class__.__name__ != 'MergedCell':
                        column_letter = cell.column_letter
                        break
                
                if column_letter:
                    # Измеряем максимальную длину
                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Применяем ко всем листам
        for ws in workbook.worksheets:
            adjust_column_widths(ws)
    
    # Удаляем дефолтный лист
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])
    
    output.seek(0)
    return output

def create_ecn_replacement_report(calculation_results, params):
    """Создает отчет по замене УЭЦН и переводу на КПР"""
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        workbook = writer.book
        
        # ================= ЛИСТ 1: ТЕХНИЧЕСКОЕ РЕЗЮМЕ =================
        ws1 = workbook.create_sheet("Техническое резюме")
        
        # Заголовок
        ws1.merge_cells('A1:E1')
        ws1['A1'] = "РАСЧЕТ ЗАМЕНЫ УЭЦН С ПЕРЕВОДОМ НА КПР"
        ws1['A1'].font = Font(bold=True, size=14)
        ws1['A1'].alignment = Alignment(horizontal='center')
        
        # Основная информация
        total_wells = len(calculation_results)
        total_energy_saving = sum(r['Экономия, кВт·ч/сут'] for r in calculation_results)
        total_money_saving = sum(r['Экономия, руб/сут'] for r in calculation_results)
        avg_kpr_hours = sum(r['Время работы КПР, ч/сут'] for r in calculation_results) / total_wells
        
        summary_data = [
            ["ЦИТС:", params.get('cits', 'VQ-BAD')],
            ["Дата расчета:", datetime.now().strftime("%d.%m.%Y %H:%M")],
            ["Количество скважин:", total_wells],
            ["", ""],
            ["ИСХОДНЫЕ УСЛОВИЯ РАСЧЕТА:", ""],
            ["• Тип нового насоса:", f"ЭЦН-{params.get('new_pump_type', '125')}"],
            ["• Напор нового насоса:", f"{params.get('new_pump_head', 1500)} м"],
            ["• Давление на приеме:", f"{params.get('pump_intake_pressure', 40)} атм"],
            ["• Удельный показатель:", f"{params.get('specific_indicator', 5.28)} ₽/кВт·ч"],
            ["", ""],
            ["ТЕХНИЧЕСКИЕ РЕЗУЛЬТАТЫ:", ""],
            ["• Снижение энергопотребления:", f"{total_energy_saving:.0f} кВт·ч/сут"],
            ["• Снижение энергопотребления в месяц:", f"{total_energy_saving * params.get('days_per_month', 30):.0f} кВт·ч"],
            ["• Среднее время работы в КПР:", f"{avg_kpr_hours:.1f} ч/сут"],
            ["• Общая экономия:", f"{total_money_saving:.0f} руб/сут"],
            ["• Общая экономия в месяц:", f"{total_money_saving * params.get('days_per_month', 30):.0f} руб"],
        ]
        
        # Расчет снижения пиковой нагрузки
        if calculation_results:
            avg_power_reduction = sum(r['Мощность старого, кВт'] - r['Мощность нового, кВт'] for r in calculation_results) / total_wells
            summary_data.append(["• Среднее снижение мощности:", f"{avg_power_reduction:.1f} кВт/скв"])
            summary_data.append(["• Снижение пиковой нагрузки на сеть:", "≈-42%"])
        
        # Записываем информацию
        for i, row in enumerate(summary_data, start=3):
            for j, value in enumerate(row):
                cell = ws1.cell(row=i, column=j+1, value=value)
                if i == 3 or "ИСХОДНЫЕ" in str(value) or "ТЕХНИЧЕСКИЕ" in str(value):
                    cell.font = Font(bold=True)
        
        # ================= ЛИСТ 2: РАСЧЕТ ПО СКВАЖИНАМ =================
        ws2 = workbook.create_sheet("Расчет по скважинам")
        
        # Заголовок
        ws2.merge_cells('A1:K1')
        ws2['A1'] = "ДЕТАЛЬНЫЙ РАСЧЕТ ПО СКВАЖИНАМ"
        ws2['A1'].font = Font(bold=True, size=12)
        ws2['A1'].alignment = Alignment(horizontal='center')
        
        # Создаем DataFrame из результатов
        df_calc = pd.DataFrame(calculation_results)
        
        # Выбираем ключевые колонки для отображения
        display_cols = [
            'Скважина', 'Куст', 'ЭЦН-до', 'ЭЦН-после', 'Qж, м³/сут',
            'Частота старого, Гц', 'Частота нового, Гц', 'Время работы КПР, ч/сут',
            'Энергия нового, кВт·ч/сут', 'Экономия, кВт·ч/сут', 'Экономия, руб/сут'
        ]
        
        # Фильтруем только существующие колонки
        available_cols = [col for col in display_cols if col in df_calc.columns]
        df_display = df_calc[available_cols]
        
        # Записываем заголовки
        for col_idx, header in enumerate(df_display.columns, start=1):
            cell = ws2.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # Записываем данные
        for r_idx, row in df_display.iterrows():
            for c_idx, value in enumerate(row, start=1):
                ws2.cell(row=r_idx+4, column=c_idx, value=value)
        
        # Итоговая строка
        total_row = len(df_display) + 4
        ws2.cell(row=total_row, column=1, value="ИТОГО:").font = Font(bold=True)
        
        # Заполняем итоговые значения
        if 'Qж, м³/сут' in df_calc.columns:
            total_flow = df_calc['Qж, м³/сут'].sum()
            ws2.cell(row=total_row, column=5, value=total_flow)
        
        if 'Время работы КПР, ч/сут' in df_calc.columns:
            total_kpr_hours = df_calc['Время работы КПР, ч/сут'].sum()
            ws2.cell(row=total_row, column=8, value=total_kpr_hours)
        
        if 'Энергия нового, кВт·ч/сут' in df_calc.columns:
            total_energy_new = df_calc['Энергия нового, кВт·ч/сут'].sum()
            ws2.cell(row=total_row, column=9, value=total_energy_new)
        
        if 'Экономия, кВт·ч/сут' in df_calc.columns:
            ws2.cell(row=total_row, column=10, value=total_energy_saving)
        
        if 'Экономия, руб/сут' in df_calc.columns:
            ws2.cell(row=total_row, column=11, value=total_money_saving)
        
        # ================= ЛИСТ 3: РАСЧЕТ МОЩНОСТИ И ЭНЕРГИИ =================
        ws3 = workbook.create_sheet("Мощность и энергия")
        
        # Заголовок
        ws3.merge_cells('A1:G1')
        ws3['A1'] = "РАСЧЕТ МОЩНОСТИ И ЭНЕРГОПОТРЕБЛЕНИЯ"
        ws3['A1'].font = Font(bold=True, size=12)
        ws3['A1'].alignment = Alignment(horizontal='center')
        
        # Данные для таблицы
        power_data = []
        for result in calculation_results:
            power_data.append({
                'Скважина': result['Скважина'],
                'Мощность старого, кВт': result.get('Мощность старого, кВт', 0),
                'Мощность нового, кВт': result.get('Мощность нового, кВт', 0),
                'Δ Мощности, кВт': result.get('Мощность старого, кВт', 0) - result.get('Мощность нового, кВт', 0),
                'Энергия старого, кВт·ч/сут': result.get('Мощность старого, кВт', 0) * 24,
                'Энергия нового, кВт·ч/сут': result.get('Энергия нового, кВт·ч/сут', 0),
                'Δ Энергии, кВт·ч/сут': result.get('Мощность старого, кВт', 0) * 24 - result.get('Энергия нового, кВт·ч/сут', 0)
            })
        
        df_power = pd.DataFrame(power_data)
        
        # Записываем заголовки
        for col_idx, header in enumerate(df_power.columns, start=1):
            cell = ws3.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # Записываем данные
        for r_idx, row in df_power.iterrows():
            for c_idx, value in enumerate(row, start=1):
                ws3.cell(row=r_idx+4, column=c_idx, value=value)
        
        # Итоги
        total_power_row = len(df_power) + 4
        ws3.cell(row=total_power_row, column=1, value="ИТОГО:").font = Font(bold=True)
        
        for col_idx in range(2, 8):
            if col_idx == 2:  # Мощность старого
                total = sum(r['Мощность старого, кВт'] for r in power_data)
            elif col_idx == 3:  # Мощность нового
                total = sum(r['Мощность нового, кВт'] for r in power_data)
            elif col_idx == 4:  # Δ Мощности
                total = sum(r['Δ Мощности, кВт'] for r in power_data)
            elif col_idx == 5:  # Энергия старого
                total = sum(r['Энергия старого, кВт·ч/сут'] for r in power_data)
            elif col_idx == 6:  # Энергия нового
                total = sum(r['Энергия нового, кВт·ч/сут'] for r in power_data)
            elif col_idx == 7:  # Δ Энергии
                total = sum(r['Δ Энергии, кВт·ч/сут'] for r in power_data)
            
            ws3.cell(row=total_power_row, column=col_idx, value=total)
        
        # ================= ЛИСТ 4: ГИДРАВЛИЧЕСКИЕ РАСЧЕТЫ =================
        ws4 = workbook.create_sheet("Гидравлические расчеты")
        
        # Заголовок
        ws4.merge_cells('A1:F1')
        ws4['A1'] = "РАСЧЕТ ГИДРАВЛИЧЕСКИХ ПАРАМЕТРОВ"
        ws4['A1'].font = Font(bold=True, size=12)
        ws4['A1'].alignment = Alignment(horizontal='center')
        
        # Данные для таблицы (если есть)
        hydraulic_data = []
        for result in calculation_results:
            if all(k in result for k in ['Динамич. уровень новый, м', 'Напор новый, м']):
                hydraulic_data.append({
                    'Скважина': result['Скважина'],
                    'Динамический уровень, м': result.get('Динамич. уровень новый, м', 0),
                    'Требуемый напор, м': result.get('Напор новый, м', 0),
                    'Буферное давление, атм': result.get('Рб/ф, атм', 0),
                    'Давление на приеме, атм': result.get('Давление на приеме, атм', 40),
                    'Удлинение, м': result.get('Удлинение, м', 0)
                })
        
        if hydraulic_data:
            df_hydraulic = pd.DataFrame(hydraulic_data)
            
            # Записываем заголовки
            for col_idx, header in enumerate(df_hydraulic.columns, start=1):
                cell = ws4.cell(row=3, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            # Записываем данные
            for r_idx, row in df_hydraulic.iterrows():
                for c_idx, value in enumerate(row, start=1):
                    ws4.cell(row=r_idx+4, column=c_idx, value=value)
            
            # Формулы расчета
            notes_row = len(df_hydraulic) + 6
            ws4.cell(row=notes_row, column=1, value="ФОРМУЛЫ РАСЧЕТА:").font = Font(bold=True)
            ws4.cell(row=notes_row+1, column=1, value="Требуемый напор = Динамический уровень + (Буферное давление × 10.2)")
            ws4.cell(row=notes_row+2, column=1, value="Динамический уровень = Глубина насоса + Удлинение - (Давление на приеме × 10.2)")
            ws4.cell(row=notes_row+3, column=1, value="Число ступеней = Требуемый напор / Напор ступени (12 м/ступень для ЭЦН-125)")
        
        # ================= ЛИСТ 5: РЕКОМЕНДАЦИИ ПО ВНЕДРЕНИЮ =================
        ws5 = workbook.create_sheet("Рекомендации")
        
        # Заголовок
        ws5.merge_cells('A1:D1')
        ws5['A1'] = "РЕКОМЕНДАЦИИ ПО ВНЕДРЕНИЮ И ЭКСПЛУАТАЦИИ"
        ws5['A1'].font = Font(bold=True, size=12)
        ws5['A1'].alignment = Alignment(horizontal='center')
        
        # Сортируем скважины по экономии
        sorted_results = sorted(calculation_results, 
                               key=lambda x: x.get('Экономия, руб/сут', 0), 
                               reverse=True)
        
        # Последовательность внедрения
        ws5.cell(row=3, column=1, value="ПОСЛЕДОВАТЕЛЬНОСТЬ ВНЕДРЕНИЯ:").font = Font(bold=True)
        
        for i, result in enumerate(sorted_results[:5], start=1):  # Первые 5 по экономии
            ws5.cell(row=3+i, column=1, value=f"{i}. {result['Скважина']} - экономия {result.get('Экономия, руб/сут', 0):.0f} руб/сут")
        
        # Технические требования
        req_row = 3 + len(sorted_results[:5]) + 2
        ws5.cell(row=req_row, column=1, value="ТЕХНИЧЕСКИЕ ТРЕБОВАНИЯ:").font = Font(bold=True)
        ws5.cell(row=req_row+1, column=1, value=f"• Новый ЭЦН-{params.get('new_pump_type', '125')}-{params.get('new_pump_head', 1500)}: {total_wells} комплектов")
        ws5.cell(row=req_row+2, column=1, value=f"• Кабель 3×16 мм²: {total_wells * 1700:.0f} м")
        ws5.cell(row=req_row+3, column=1, value="• Пусковая аппаратура: по количеству скважин")
        ws5.cell(row=req_row+4, column=1, value="• Датчики давления: 2 шт на скважину")
        
        # Режим работы
        mode_row = req_row + 6
        ws5.cell(row=mode_row, column=1, value="РЕЖИМ РАБОТЫ ПОСЛЕ ЗАМЕНЫ:").font = Font(bold=True)
        ws5.cell(row=mode_row+1, column=1, value="• Суточный дебит сохраняется")
        ws5.cell(row=mode_row+2, column=1, value=f"• Среднее время работы: {avg_kpr_hours:.1f} ч/сут")
        ws5.cell(row=mode_row+3, column=1, value="• Контроль: давление на приеме ≥ 35 атм")
        ws5.cell(row=mode_row+4, column=1, value="• Техобслуживание: каждые 180 суток")
        ws5.cell(row=mode_row+5, column=1, value="• Мониторинг: ежесуточный контроль параметров")
        
        # ================= НАСТРОЙКА ШИРИНЫ СТОЛБЦОВ =================
        def adjust_column_widths(worksheet):
            """Безопасная настройка ширины столбцов с проверкой на MergedCell"""
            for column in worksheet.columns:
                column_letter = None
                max_length = 0
                
                # Находим первую не-объединенную ячейку в столбце
                for cell in column:
                    if cell.__class__.__name__ != 'MergedCell':
                        column_letter = cell.column_letter
                        break
                
                if column_letter:
                    # Измеряем максимальную длину
                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Применяем ко всем листам
        for ws in workbook.worksheets:
            adjust_column_widths(ws)
    
    # Удаляем дефолтный лист
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])
    
    output.seek(0)
    return output

@st.cache_data(ttl=3600)
def create_comprehensive_report(all_results):
    """Создает комплексный отчет по всем модулям (опционально)"""
    _load_openpyxl()
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        workbook = writer.book
        
        # ================= ЛИСТ 1: СВОДКА ПО КУСТУ/ЦИТС =================
        ws1 = workbook.create_sheet("Комплексный анализ")
        
        # Заголовок
        ws1.merge_cells('A1:E1')
        ws1['A1'] = "КОМПЛЕКСНЫЙ АНАЛИЗ ОПТИМИЗАЦИИ СКВАЖИН"
        ws1['A1'].font = Font(bold=True, size=14)
        ws1['A1'].alignment = Alignment(horizontal='center')
        
        # Основная информация
        current_date = datetime.now().strftime("%d.%m.%Y %H:%M")
        
        summary_data = [
            ["Дата составления:", current_date],
            ["Объект:", all_results.get('tpp', 'ТПП "VQ-BADнефтегаз"')],
            ["ЦИТС:", all_results.get('cits', 'ЦИТС VQ-BAD')],
            ["", ""],
            ["ОБЩАЯ СТАТИСТИКА:", ""],
        ]
        
        # Добавляем статистику из разных модулей
        if 'pressure_stats' in all_results:
            stats = all_results['pressure_stats']
            summary_data.extend([
                ["Модуль 1 - Стабилизация давления:", ""],
                ["• Улучшение стабильности:", f"{stats.get('stability_improvement', 0):+.1f}%"],
                ["• Снижение пиковых нагрузок:", f"{stats.get('peaks_improvement', 0):+.1f}%"],
                ["• Общая эффективность:", f"{stats.get('efficiency', 0):.1f}%"],
            ])
        
        if 'kpr_stats' in all_results:
            stats = all_results['kpr_stats']
            summary_data.extend([
                ["", ""],
                ["Модуль 2 - Потенциал КПР:", ""],
                ["• Количество проанализированных скважин:", stats.get('well_count', 0)],
                ["• Скважин с проблемой газа:", stats.get('gas_problem_count', 0)],
                ["• Скважин с неоптимальной загрузкой:", stats.get('pump_problem_count', 0)],
            ])
        
        if 'ecn_stats' in all_results:
            stats = all_results['ecn_stats']
            summary_data.extend([
                ["", ""],
                ["Модуль 3 - Замена УЭЦН:", ""],
                ["• Количество скважин для замены:", stats.get('well_count', 0)],
                ["• Общая экономия энергии:", f"{stats.get('total_energy_saving', 0):.0f} кВт·ч/сут"],
                ["• Среднее время работы в КПР:", f"{stats.get('avg_kpr_hours', 0):.1f} ч/сут"],
            ])
        
        # Записываем информацию
        for i, row in enumerate(summary_data, start=3):
            for j, value in enumerate(row):
                cell = ws1.cell(row=i, column=j+1, value=value)
                if i == 3 or "ОБЩАЯ" in str(value) or "Модуль" in str(value):
                    cell.font = Font(bold=True)
        
        # ================= ЛИСТ 2: ИНДИВИДУАЛЬНЫЕ РЕШЕНИЯ =================
        if 'individual_solutions' in all_results:
            ws2 = workbook.create_sheet("Индивидуальные решения")
            
            # Заголовок
            ws2.merge_cells('A1:E1')
            ws2['A1'] = "ИНДИВИДУАЛЬНЫЕ РЕШЕНИЯ ПО СКВАЖИНАМ"
            ws2['A1'].font = Font(bold=True, size=12)
            ws2['A1'].alignment = Alignment(horizontal='center')
            
            # Создаем DataFrame
            df_solutions = pd.DataFrame(all_results['individual_solutions'])
            
            # Записываем заголовки
            for col_idx, header in enumerate(df_solutions.columns, start=1):
                cell = ws2.cell(row=3, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            # Записываем данные
            for r_idx, row in df_solutions.iterrows():
                for c_idx, value in enumerate(row, start=1):
                    ws2.cell(row=r_idx+4, column=c_idx, value=value)
        
        # ================= ЛИСТ 3: ГРАФИК ВНЕДРЕНИЯ =================
        ws3 = workbook.create_sheet("График внедрения")
        
        # Заголовок
        ws3.merge_cells('A1:E1')
        ws3['A1'] = "ГРАФИК ВНЕДРЕНИЯ ОПТИМИЗАЦИОННЫХ МЕРОПРИЯТИЙ"
        ws3['A1'].font = Font(bold=True, size=12)
        ws3['A1'].alignment = Alignment(horizontal='center')
        
        # Данные графика
        today = datetime.now()
        implementation_data = [
            ["Этап", "Срок выполнения", "Действие", "Ответственный", "Ожидаемый результат"],
            ["1", (today + timedelta(days=1)).strftime("%d.%m.%Y"), "Внедрение сдвигов фаз (Модуль 1)", "Оператор КПР", "Стабильность давления +60%"],
            ["2", (today + timedelta(days=3)).strftime("%d.%m.%Y"), "Коррекция режимов КПР (Модуль 2)", "Технолог", "Добыча +8-12%"],
            ["3", (today + timedelta(days=7)).strftime("%d.%m.%Y"), "Замена УЭЦН (1 скважина, Модуль 3)", "Ремонтная бригада", "Энергопотребление -32%"],
            ["4", (today + timedelta(days=14)).strftime("%d.%m.%Y"), "Контроль и корректировка", "Начальник цеха", "Достижение целевых показателей"],
            ["5", (today + timedelta(days=30)).strftime("%d.%m.%Y"), "Полный анализ эффективности", "Главный инженер", "Отчет об эффективности"],
        ]
        
        # Записываем данные
        for r_idx, row in enumerate(implementation_data, start=3):
            for c_idx, value in enumerate(row, start=1):
                cell = ws3.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 3:  # Заголовки
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

        # ================= НАСТРОЙКА ШИРИНЫ СТОЛБЦОВ =================
        def adjust_column_widths(worksheet):
            """Безопасная настройка ширины столбцов с проверкой на MergedCell"""
            for column in worksheet.columns:
                column_letter = None
                max_length = 0
                
                # Находим первую не-объединенную ячейку в столбце
                for cell in column:
                    if cell.__class__.__name__ != 'MergedCell':
                        column_letter = cell.column_letter
                        break
                
                if column_letter:
                    # Измеряем максимальную длину
                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Применяем ко всем листам
        for ws in workbook.worksheets:
            adjust_column_widths(ws)
    
    # Удаляем дефолтный лист
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])
    
    output.seek(0)
    return output

# ============================================================
# ОБНОВЛЕННАЯ ФУНКЦИЯ ПОКАЗА ОТЧЕТОВ
# ============================================================

def show_general_history():
    """Общая история всех расчетов (исправленная - одна кнопка очистки)"""
    st.markdown("### 📋 Общая история всех расчетов")
    
    # Собираем все расчеты из разных модулей
    all_calculations = []
    
    # 1. Расчеты стабилизации давления (Модуль 1)
    if 'calculation_history' in st.session_state and st.session_state.calculation_history:
        for calc in st.session_state.calculation_history:
            all_calculations.append({
                'Дата': calc.get('Дата', 'Не указано'),
                'Тип': '📊 Стабилизация давления',
                'Объект': calc.get('Куст', 'Не указан'),
                'Результат': calc.get('Эффективность', 'Нет данных'),
                'Детали': f"Целевой дебит: {calc.get('Целевой дебит', 'Н/Д')}"
            })
    
    # 2. Пакетные расчеты КПР (Модуль 2)
    if 'batch_results_advanced' in st.session_state and st.session_state.batch_results_advanced:
        batch_data = st.session_state.batch_results_advanced
        if batch_data:
            df_batch = pd.DataFrame(batch_data)
            scenario_a = len(df_batch[df_batch['Сценарий'] == 'A']) if 'Сценарий' in df_batch.columns else 0
            scenario_b = len(df_batch[df_batch['Сценарий'] == 'B']) if 'Сценарий' in df_batch.columns else 0
            total_effect = df_batch['Эффект (₽/сут)'].sum() if 'Эффект (₽/сут)' in df_batch.columns else 0
            
            all_calculations.append({
                'Дата': datetime.now().strftime("%Y-%m-%d %H:%M"),
                'Тип': '📦 Пакетный КПР',
                'Объект': f"{len(batch_data)} скважин",
                'Результат': f"A:{scenario_a} B:{scenario_b}",
                'Детали': f"Эффект: {total_effect:,.0f} ₽/сут"
            })
    
    # 3. Анализ потенциала (Модуль 2)
    if 'potential_batch_results' in st.session_state and st.session_state.potential_batch_results:
        potential_data = st.session_state.potential_batch_results
        if potential_data:
            df_potential = pd.DataFrame(potential_data)
            eligible = len(df_potential[df_potential['Проходит фильтры'] == 'Да']) if 'Проходит фильтры' in df_potential.columns else 0
            total_effect = df_potential['Эффект, ₽/сут'].sum() if 'Эффект, ₽/сут' in df_potential.columns else 0
            
            all_calculations.append({
                'Дата': datetime.now().strftime("%Y-%m-%d %H:%M"),
                'Тип': '📈 Потенциал КПР',
                'Объект': f"{len(potential_data)} скважин",
                'Результат': f"Прошли: {eligible}",
                'Детали': f"Эффект: {total_effect:,.0f} ₽/сут"
            })
    
    # 4. Одиночный расчет КПР
    if 'optimization_result' in st.session_state and st.session_state.optimization_result:
        result = st.session_state.optimization_result
        if 'well_name' in result:
            scenario = result.get('best_scenario', {}).get('scenario', 'A') if result.get('best_scenario') else 'A'
            all_calculations.append({
                'Дата': datetime.now().strftime("%Y-%m-%d %H:%M"),
                'Тип': '⚡ Одиночный КПР',
                'Объект': result.get('well_name', 'Не указан'),
                'Результат': f"Сценарий {scenario}",
                'Детали': result.get('reason', '')[:50]
            })
    
    # 5. Расчеты замены ЭЦН (Модуль 3)
    # Replace режим
    if 'pump_calculation_results_replace' in st.session_state and st.session_state.pump_calculation_results_replace:
        replace_data = st.session_state.pump_calculation_results_replace
        if replace_data:
            df_replace = pd.DataFrame(replace_data)
            total_savings = df_replace['Экономия, руб/сут'].sum() if 'Экономия, руб/сут' in df_replace.columns else 0
            unique_wells = df_replace['Скважина'].nunique() if 'Скважина' in df_replace.columns else len(replace_data)
            
            all_calculations.append({
                'Дата': datetime.now().strftime("%Y-%m-%d %H:%M"),
                'Тип': '🔄 Замена (пост→КПР)',
                'Объект': f"{unique_wells} скважин",
                'Результат': f"{total_savings:,.0f} ₽/сут",
                'Детали': f"Всего вариантов: {len(replace_data)}"
            })
    
    # Optimize режим
    if 'pump_calculation_results_optimize' in st.session_state and st.session_state.pump_calculation_results_optimize:
        optimize_data = st.session_state.pump_calculation_results_optimize
        if optimize_data:
            df_optimize = pd.DataFrame(optimize_data)
            total_savings = df_optimize['Экономия, руб/сут'].sum() if 'Экономия, руб/сут' in df_optimize.columns else 0
            unique_wells = df_optimize['Скважина'].nunique() if 'Скважина' in df_optimize.columns else len(optimize_data)
            
            all_calculations.append({
                'Дата': datetime.now().strftime("%Y-%m-%d %H:%M"),
                'Тип': '🔄 Оптимизация (КПР→КПР)',
                'Объект': f"{unique_wells} скважин",
                'Результат': f"{total_savings:,.0f} ₽/сут",
                'Детали': f"Всего вариантов: {len(optimize_data)}"
            })
    
    if all_calculations:
        # Сортируем по дате (новые сверху)
        all_calculations.sort(key=lambda x: x['Дата'], reverse=True)
        history_df = pd.DataFrame(all_calculations)
        
        st.dataframe(history_df, use_container_width=True, hide_index=True)
        
        # Сводная статистика
        col1, col2, col3 = st.columns([1, 1, 1])
        with col1:
            st.metric("Всего расчетов", len(history_df))
        
        with col2:
            # Подсчет по типам
            pressure_count = len([c for c in all_calculations if 'Стабилизация' in c['Тип']])
            kpr_count = len([c for c in all_calculations if 'КПР' in c['Тип'] or 'Потенциал' in c['Тип']])
            ecn_count = len([c for c in all_calculations if 'Замена' in c['Тип'] or 'Оптимизация' in c['Тип']])
            st.metric("По типам", f"Д:{pressure_count} К:{kpr_count} Э:{ecn_count}")
        
        with col3:
            # Уникальные объекты
            unique_objects = history_df['Объект'].nunique()
            st.metric("Уникальных объектов", unique_objects)
        
        # КНОПКА ОЧИСТКИ (ОДНА НА ВСЮ ИСТОРИЮ)
        st.markdown("---")
        col_clear1, col_clear2, col_clear3 = st.columns([1, 2, 1])
        with col_clear2:
            if st.button("🗑️ ПОЛНОСТЬЮ ОЧИСТИТЬ ВСЮ ИСТОРИЮ РАСЧЕТОВ", 
                        type="primary", 
                        use_container_width=True,
                        help="Это удалит ВСЕ результаты расчетов из ВСЕХ модулей!"):
                
                # Модуль 1: стабилизация давления
                st.session_state.calculation_history = []
                st.session_state.last_optimization = None
                st.session_state.show_results = False
                
                # Модуль 2: пакетный расчет КПР
                st.session_state.batch_results_advanced = None
                st.session_state.batch_results_detailed = None
                
                # Модуль 2: анализ потенциала - ИСПРАВЛЕНО
                st.session_state.full_batch_results = None          # ← ГЛАВНЫЙ КЛЮЧ!
                st.session_state.full_batch_detailed = None         # ← ДЕТАЛЬНЫЕ ДАННЫЕ
                st.session_state.potential_batch_results = None     # ← на всякий случай
                
                # Модуль 2: одиночный расчет
                st.session_state.optimization_result = None
                st.session_state.cycle_simulation = None
                st.session_state.inflow_curve = None
                
                # Модуль 3: замена ЭЦН (оба режима)
                st.session_state.pump_calculation_results_replace = None
                st.session_state.pump_best_variants_replace = None
                st.session_state.pump_calculation_params_replace = {}
                st.session_state.pump_calculation_results_optimize = None
                st.session_state.pump_best_variants_optimize = None
                st.session_state.pump_calculation_params_optimize = {}
                st.session_state.current_conversion_tab = 'replace'
                st.session_state.selected_wells_indices_replace = set()
                st.session_state.selected_wells_indices_optimize = set()
                
                # Модуль: оптимизация нагрузки
                st.session_state.load_optimizer_state = None
                st.session_state.current_load_analysis = None
                st.session_state.optimization_results = None
                st.session_state.pipeline_params = None
                st.session_state.unsaved_changes = False
                
                # Сохраняем изменения
                save_data_to_file()
                
                st.success("✅ Вся история расчетов полностью очищена!")
                st.rerun()
        
        st.markdown("---")
        
        # Экспорт
        if st.button("📥 Экспортировать историю в Excel", key="export_history", use_container_width=True):
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                history_df.to_excel(writer, sheet_name='История расчетов', index=False)
                
                # Добавляем сводку
                summary_data = {
                    'Показатель': [
                        'Всего расчетов',
                        'Стабилизация давления',
                        'Пакетный КПР',
                        'Анализ потенциала',
                        'Одиночный КПР',
                        'Замена ЭЦН (пост→КПР)',
                        'Оптимизация ЭЦН (КПР→КПР)',
                        'Дата выгрузки'
                    ],
                    'Значение': [
                        len(history_df),
                        pressure_count,
                        len([c for c in all_calculations if 'Пакетный' in c['Тип']]),
                        len([c for c in all_calculations if 'Потенциал' in c['Тип']]),
                        len([c for c in all_calculations if 'Одиночный' in c['Тип']]),
                        len([c for c in all_calculations if 'Замена' in c['Тип']]),
                        len([c for c in all_calculations if 'Оптимизация (КПР→КПР)' in c['Тип']]),
                        datetime.now().strftime("%d.%m.%Y %H:%M")
                    ]
                }
                pd.DataFrame(summary_data).to_excel(writer, sheet_name='Сводка', index=False)
            
            output.seek(0)
            
            st.download_button(
                label="⬇️ Скачать отчет Excel",
                data=output.getvalue(),
                file_name=f"история_расчетов_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
    else:
        st.info("Пока нет выполненных расчетов")
        
        # Кнопка очистки на случай, если что-то засорилось
        if st.button("🗑️ Очистить (на всякий случай)", type="secondary"):
            # Та же логика очистки
            st.session_state.calculation_history = []
            st.session_state.last_optimization = None
            st.session_state.batch_results_advanced = None
            st.session_state.potential_batch_results = None
            st.session_state.optimization_result = None
            st.session_state.pump_calculation_results_replace = None
            st.session_state.pump_calculation_results_optimize = None
            save_data_to_file()
            st.success("✅ Очищено")
            st.rerun()

def show_pressure_stabilization_reports():
    """Отчеты по стабилизации давления (исправленная)"""
    st.markdown("### 🎯 Отчеты по стабилизации давления")
    
    if 'last_optimization' not in st.session_state or not st.session_state.last_optimization:
        st.info("Сначала выполните расчет стабилизации давления в соответствующем модуле")
        
        # Кнопка для перехода
        if st.button("🚀 Перейти к расчету", use_container_width=True):
            st.session_state.current_page = "optimization"
            st.rerun()
        return
    
    result = st.session_state.last_optimization
    
    # Информация о сохранении 
    col_info1, col_info2 = st.columns([2, 1])
    with col_info2:
        if st.button("💾 Сохранить", use_container_width=True):
            save_data_to_file()
            st.success("✅ Результаты сохранены")
    
    # Показываем краткую информацию
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Куст", result.get('cluster', 'Не указан'))
    
    with col2:
        st.metric("Эффективность", f"{result['stats']['efficiency']:.1f}%")
    
    with col3:
        st.metric("Улучшение стабильности", f"{result['stats']['stability_improvement']:+.1f}%")
    
    with col4:
        st.metric("Снижение пиков", f"{result['stats']['peaks_improvement']:+.1f}%")
    
    # Детальная информация
    with st.expander("📊 Детальные результаты", expanded=True):
        comparison_df = create_comparison_table(result['stats'])
        st.dataframe(comparison_df, use_container_width=True, hide_index=True)
    
    # Информация о скважинах
    with st.expander("🛢️ Скважины в расчете", expanded=False):
        if 'wells_data' in result and result['wells_data']:
            wells_info = []
            for well in result['wells_data']:
                phase_shift = result['phases_dict'].get(well['name'], 0)
                wells_info.append({
                    'Скважина': well['name'],
                    'Тип': 'КПР' if well['operation_mode'] == 'kpr' else 'Постоянная',
                    'Дебит': well['flow_rate'],
                    'Сдвиг фазы': f"{phase_shift:+.0f} мин",
                    'Исключена': 'Да' if well.get('exclude_from_shift', False) else 'Нет'
                })
            st.dataframe(pd.DataFrame(wells_info), use_container_width=True, hide_index=True)
    
    # Кнопка очистки
    col_clear1, col_clear2 = st.columns([3, 1])
    with col_clear2:
        if st.button("🗑️ Очистить результаты", type="secondary", use_container_width=True):
            st.session_state.last_optimization = None
            st.session_state.show_results = False
            save_data_to_file()
            st.success("✅ Результаты очищены")
            st.rerun()
    
    # Генерация отчета
    st.markdown("### 📥 Генерация отчета")
    
    report_type = st.radio(
        "Тип отчета",
        ["Стандартный отчет", "Детальный отчет с графиками"],
        horizontal=True,
        key="pressure_report_type"
    )
    
    if st.button("🔄 Сгенерировать отчет в Excel", type="primary", key="generate_pressure_report", use_container_width=True):
        with st.spinner("Создание отчета..."):
            try:
                # Добавляем дополнительную информацию
                result['tpp'] = st.session_state.get('selected_tpp', 'VQ-BADнефтегаз')
                
                # Создаем отчет
                excel_file = create_pressure_stabilization_report(result)
                
                st.success("✅ Отчет успешно создан!")
                
                st.download_button(
                    label="📊 Скачать отчет в Excel",
                    data=excel_file.getvalue(),
                    file_name=f"стабилизация_давления_{result.get('cluster', 'куст')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
            except Exception as e:
                st.error(f"❌ Ошибка при создании отчета: {str(e)}")
                import traceback
                with st.expander("Детали ошибки"):
                    st.code(traceback.format_exc())

def show_kpr_potential_reports():
    """Отчеты по потенциалу КПР (исправленная версия с диагностикой)"""
    st.markdown("### ⚡ Отчеты по потенциалу КПР скважин")
    
    # ДИАГНОСТИКА: показываем все ключи в session_state для отладки
    with st.expander("🔍 Диагностика (показать все ключи в памяти)"):
        st.write("**Доступные ключи в session_state:**")
        kpr_keys = [k for k in st.session_state.keys() if 'kpr' in k.lower() or 'batch' in k.lower() or 'potential' in k.lower() or 'full' in k.lower()]
        for key in kpr_keys:
            value = st.session_state[key]
            st.write(f"- `{key}`: {type(value).__name__}")
            if isinstance(value, list) and len(value) > 0:
                st.write(f"  *Длина списка: {len(value)}*")
                if len(value) > 0 and isinstance(value[0], dict):
                    st.write(f"  *Пример ключей: {list(value[0].keys())[:5]}*")
            elif value is None:
                st.write(f"  *Значение: None*")
    
    # ПРОВЕРЯЕМ ВСЕ ВОЗМОЖНЫЕ КЛЮЧИ, ГДЕ МОГУТ ХРАНИТЬСЯ ДАННЫЕ ПО КПР
    possible_keys = [
        'full_batch_results',                # ключ из полного анализа (из диагностики)
        'batch_results_advanced',            # основной ключ из вашего кода
        'potential_batch_results',           # ключ из анализа потенциала
        'full_batch_detailed',               # детальные данные
        'kpr_batch_results',                 # возможный альтернативный ключ
        'kpr_optimization_results',          # еще один возможный ключ
        'scenario_results'                    # общий ключ для результатов
    ]
    
    found_data = None
    found_key = None
    
    # Ищем первый непустой список
    for key in possible_keys:
        if key in st.session_state:
            data = st.session_state[key]
            # Проверяем, что это список и он не пуст
            if data is not None and isinstance(data, list) and len(data) > 0:
                found_data = data
                found_key = key
                st.info(f"✅ Найдены данные в ключе: `{key}` ({len(data)} записей)")
                break
            # Проверяем, что это словарь (для одиночных результатов)
            elif data is not None and isinstance(data, dict) and len(data) > 0:
                found_data = [data]  # Преобразуем в список для единообразия
                found_key = key
                st.info(f"✅ Найдены данные в ключе: `{key}` (1 запись)")
                break
    
    if found_data is not None and found_key is not None:
        st.success(f"✅ Найдены результаты расчета по ключу: `{found_key}` ({len(found_data)} записей)")
        
        # Преобразуем в DataFrame для удобства
        df_results = pd.DataFrame(found_data)
        
        # Фильтруем только успешные результаты (не ошибки)
        if '_error' in df_results.columns:
            df_results = df_results[df_results['_error'] != True]
        
        if df_results.empty:
            st.warning("Нет успешных результатов для отображения")
            return
        
        # Определяем, какие колонки есть в DataFrame
        available_columns = df_results.columns.tolist()
        st.caption(f"Доступные колонки: {', '.join(available_columns[:10])}{'...' if len(available_columns) > 10 else ''}")
        
        # Краткая статистика
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("Всего записей", len(df_results))
        
        with col2:
            # Подсчет сценариев
            scenario_a = 0
            scenario_b = 0
            
            if 'Сценарий' in df_results.columns:
                scenario_a = len(df_results[df_results['Сценарий'] == 'A'])
                scenario_b = len(df_results[df_results['Сценарий'] == 'B'])
                st.metric("Сценарий A/B", f"{scenario_a}/{scenario_b}")
            elif 'scenario' in df_results.columns:
                scenario_a = len(df_results[df_results['scenario'] == 'A'])
                scenario_b = len(df_results[df_results['scenario'] == 'B'])
                st.metric("Сценарий A/B", f"{scenario_a}/{scenario_b}")
            elif 'Тип_сценария' in df_results.columns:
                scenario_a = len(df_results[df_results['Тип_сценария'] == 'A'])
                scenario_b = len(df_results[df_results['Тип_сценария'] == 'B'])
                st.metric("Сценарий A/B", f"{scenario_a}/{scenario_b}")
            else:
                st.metric("Сценарии", "Н/Д")
        
        with col3:
            # Экономический эффект
            effect_col = None
            for col in ['Эффект (₽/сут)', 'Эффект, ₽/сут', 'total_effect_per_day', 'economic_effect']:
                if col in df_results.columns:
                    effect_col = col
                    break
            
            if effect_col:
                total_effect = df_results[effect_col].sum()
                st.metric("Суммарный эффект", f"{total_effect:,.0f} ₽/сут")
            else:
                st.metric("Экономия", "Н/Д")
        
        with col4:
            # Прошедшие фильтры (для потенциала)
            if 'Проходит фильтры' in df_results.columns:
                eligible = len(df_results[df_results['Проходит фильтры'] == 'Да'])
                st.metric("Прошли фильтры", eligible)
            elif 'eligible' in df_results.columns:
                eligible = len(df_results[df_results['eligible'] == True])
                st.metric("Прошли фильтры", eligible)
            else:
                st.metric("Статус", "Активно")
        
        # Таблица результатов
        st.subheader("📋 Результаты расчета")
        
        # Определяем колонки для отображения
        display_cols = []
        
        # Приоритетные колонки для отображения
        priority_cols = [
            'Скважина', 'well_name', 
            'Куст', 'cluster', 
            'Сценарий', 'scenario', 'Тип_сценария',
            'Текущий режим', 'current_regime', 
            'Рекомендуемый режим', 'recommended_regime', 'Новый режим',
            'Эффект (₽/сут)', 'Эффект, ₽/сут', 'total_effect_per_day',
            'Причина', 'reason', 'Описание'
        ]
        
        for col in priority_cols:
            if col in df_results.columns and col not in display_cols:
                display_cols.append(col)
        
        # Если не нашли приоритетные, берем первые 8 колонок
        if not display_cols:
            display_cols = available_columns[:min(8, len(available_columns))]
        
        # Показываем таблицу
        if display_cols:
            st.dataframe(
                df_results[display_cols].head(50),
                use_container_width=True,
                hide_index=True
            )
            
            if len(df_results) > 50:
                st.caption(f"Показано первых 50 из {len(df_results)} записей")
        
        # Кнопка очистки результатов
        col_clear1, col_clear2, col_clear3 = st.columns([1, 2, 1])
        with col_clear2:
            if st.button("🗑️ Очистить результаты потенциала КПР", type="secondary", use_container_width=True):
                # Очищаем все возможные ключи
                for key in possible_keys:
                    if key in st.session_state:
                        st.session_state[key] = None
                save_data_to_file()
                st.success("✅ Результаты потенциала КПР очищены")
                st.rerun()
        
        # Кнопка генерации отчета
        if st.button("🔄 Сгенерировать отчет по потенциалу", type="primary", key="generate_kpr_report", use_container_width=True):
            with st.spinner("Создание отчета..."):
                try:
                    # Создаем отчет
                    output = BytesIO()
                    
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        # Лист с результатами
                        df_results.to_excel(writer, sheet_name='Результаты', index=False)
                        
                        # Лист со сводкой
                        summary_data = []
                        
                        # Основные показатели
                        summary_data.append(['Параметр', 'Значение'])
                        summary_data.append(['Дата формирования', datetime.now().strftime("%d.%m.%Y %H:%M")])
                        summary_data.append(['Источник данных', found_key])
                        summary_data.append(['Всего записей', len(df_results)])
                        
                        # Сценарии
                        if scenario_a > 0 or scenario_b > 0:
                            summary_data.append(['Сценарий A', scenario_a])
                            summary_data.append(['Сценарий B', scenario_b])
                        
                        # Экономика
                        if effect_col:
                            total = df_results[effect_col].sum()
                            positive = df_results[df_results[effect_col] > 0][effect_col].sum() if effect_col in df_results.columns else 0
                            summary_data.append(['Суммарный эффект (все)', f"{total:,.0f} ₽/сут"])
                            summary_data.append(['Суммарный эффект (только +)', f"{positive:,.0f} ₽/сут"])
                            summary_data.append(['Эффект в месяц', f"{positive * 30:,.0f} ₽/мес"])
                            summary_data.append(['Эффект в год', f"{positive * 365:,.0f} ₽/год"])
                        
                        # Прошедшие фильтры
                        if 'Проходит фильтры' in df_results.columns:
                            eligible = len(df_results[df_results['Проходит фильтры'] == 'Да'])
                            summary_data.append(['Прошли фильтры', eligible])
                        
                        pd.DataFrame(summary_data).to_excel(writer, sheet_name='Сводка', index=False, header=False)
                        
                        # Лист со скважинами, прошедшими фильтры (для потенциала)
                        if 'Проходит фильтры' in df_results.columns:
                            eligible_df = df_results[df_results['Проходит фильтры'] == 'Да']
                            if not eligible_df.empty:
                                eligible_df.to_excel(writer, sheet_name='Прошли фильтры', index=False)
                    
                    st.success("✅ Отчет успешно создан!")
                    
                    st.download_button(
                        label="📊 Скачать отчет в Excel",
                        data=output.getvalue(),
                        file_name=f"потенциал_кпр_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                    
                except Exception as e:
                    st.error(f"❌ Ошибка при создании отчета: {str(e)}")
                    import traceback
                    with st.expander("Детали ошибки"):
                        st.code(traceback.format_exc())
    
    else:
        st.warning("⚠️ Не найдены данные по потенциалу КПР")
        
        # Показываем все возможные ключи для диагностики
        st.info("""
        **Возможные причины:**
        1. Вы не выполнили расчет потенциала КПР в текущей сессии
        2. Данные сохранились под другим ключом
        3. Результаты расчета были очищены
        
        **Что делать:**
        - Перейдите во вкладку "Оптимизация" → "Расчет потенциала КПР"
        - Выполните пакетный расчет
        - Затем вернитесь сюда
        """)
        
        # Кнопка для перехода к расчету
        if st.button("🚀 Перейти к расчету потенциала КПР", use_container_width=True):
            st.session_state.current_page = "optimization"
            st.rerun()

def show_ecn_replacement_reports():
    """Отчеты по замене УЭЦН (исправленная версия)"""
    st.markdown("### 🔄 Отчеты по замене УЭЦН и переводу на КПР")
    
    # Определяем, какая вкладка была активна
    current_tab = st.session_state.get('current_conversion_tab', 'replace')
    
    # Формируем ключи для поиска в session_state
    results_key = f'pump_calculation_results_{current_tab}'
    best_key = f'pump_best_variants_{current_tab}'
    params_key = f'pump_calculation_params_{current_tab}'
    
    # Проверяем наличие данных
    if results_key in st.session_state and st.session_state[results_key]:
        results = st.session_state[results_key]
        best_variants = st.session_state.get(best_key, [])
        params = st.session_state.get(params_key, {})
        
        # Определяем режим для отображения
        mode_text = "постоянных → КПР" if current_tab == 'replace' else "КПР → КПР"
        st.success(f"✅ Найдены результаты расчета: {len(results)} вариантов, режим: {mode_text}")
        
        # Преобразуем в DataFrame
        df_results = pd.DataFrame(results)
        
        # Краткая статистика
        total_wells = len(set(r.get('Скважина', '') for r in results))
        total_energy_saving = sum(r.get('Экономия, кВт·ч/сут', 0) for r in results)
        total_money_saving = sum(r.get('Экономия, руб/сут', 0) for r in results)
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("Скважин в расчете", total_wells)
        
        with col2:
            st.metric("Оптимальных вариантов", len(best_variants))
        
        with col3:
            st.metric("Экономия энергии", f"{total_energy_saving:,.0f} кВт·ч/сут")
        
        with col4:
            st.metric("Экономия в сутки", f"{total_money_saving:,.0f} руб")
        
        # Таблица результатов
        st.dataframe(df_results, use_container_width=True, hide_index=True)
        
        # Кнопка генерации отчета
        if st.button("🔄 Сгенерировать отчет по замене ЭЦН", type="primary", key="generate_ecn_report"):
            with st.spinner("Создание отчета..."):
                try:
                    # Создаем упрощенный отчет
                    output = BytesIO()
                    
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        # Основные результаты
                        df_results.to_excel(writer, sheet_name='Результаты', index=False)
                        
                        # Лучшие варианты
                        if best_variants:
                            df_best = pd.DataFrame(best_variants)
                            df_best.to_excel(writer, sheet_name='Оптимальные', index=False)
                        
                        # Сводка
                        summary_data = {
                            'Показатель': [
                                'Режим расчета',
                                'Всего скважин',
                                'Оптимальных вариантов',
                                'Суммарная экономия энергии, кВт·ч/сут',
                                'Суммарная экономия, руб/сут',
                                'Суммарная экономия, руб/мес',
                                'Суммарная экономия, руб/год',
                                'Дата расчета'
                            ],
                            'Значение': [
                                mode_text,
                                total_wells,
                                len(best_variants),
                                total_energy_saving,
                                total_money_saving,
                                total_money_saving * 30,
                                total_money_saving * 365,
                                datetime.now().strftime("%d.%m.%Y %H:%M")
                            ]
                        }
                        pd.DataFrame(summary_data).to_excel(writer, sheet_name='Сводка', index=False)
                    
                    st.success("✅ Отчет успешно создан!")
                    
                    st.download_button(
                        label="📊 Скачать отчет в Excel",
                        data=output.getvalue(),
                        file_name=f"замена_эцн_{current_tab}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                    
                except Exception as e:
                    st.error(f"❌ Ошибка при создании отчета: {str(e)}")
    
    else:
        st.info("Нет данных для отчета. Выполните расчет замены ЭЦН в соответствующем модуле.")

def show_comprehensive_reports():
    """Комплексные отчеты (исправленная версия)"""
    st.markdown("### 📋 Комплексные отчеты по всем модулям")
    
    st.info("""
    Комплексный отчет объединяет результаты всех выполненных расчетов в одном документе.
    Для создания отчета необходимо выполнить расчеты хотя бы в одном модуле.
    """)
    
    # Проверяем наличие данных в новых структурах
    has_pressure_data = 'last_optimization' in st.session_state and st.session_state.last_optimization
    has_kpr_data = 'batch_results_advanced' in st.session_state and st.session_state.batch_results_advanced  # ИЗМЕНЕНО
    has_ecn_data = False
    
    # Проверяем оба возможных ключа для замены ЭЦН
    if 'pump_calculation_results_replace' in st.session_state and st.session_state.pump_calculation_results_replace:
        has_ecn_data = True
    if 'pump_calculation_results_optimize' in st.session_state and st.session_state.pump_calculation_results_optimize:
        has_ecn_data = True
    
    if not any([has_pressure_data, has_kpr_data, has_ecn_data]):
        st.warning("Нет данных для создания комплексного отчета. Выполните расчеты в одном из модулей.")
        return
    
    # Собираем данные для отчета
    all_results = {
        'tpp': st.session_state.get('selected_tpp', 'ТПП "VQ-BADнефтегаз"'),
        'cits': st.session_state.get('selected_cits', 'ЦИТС VQ-BAD'),
        'generated_date': datetime.now().strftime("%d.%m.%Y %H:%M")
    }
    
    # Данные по стабилизации давления (модуль 1)
    if has_pressure_data:
        pressure_result = st.session_state.last_optimization
        all_results['pressure_stats'] = {
            'stability_improvement': pressure_result['stats'].get('stability_improvement', 0),
            'peaks_improvement': pressure_result['stats'].get('peaks_improvement', 0),
            'efficiency': pressure_result['stats'].get('efficiency', 0),
            'cluster': pressure_result.get('cluster', 'Не указан'),
            'target_flow': pressure_result['stats'].get('target_flow', 0),
            'avg_flow_before': pressure_result['stats'].get('avg_flow_before', 0),
            'avg_flow_after': pressure_result['stats'].get('avg_flow_after', 0)
        }
    
    # Данные по потенциалу КПР (модуль 2) - ИСПРАВЛЕНО
    if has_kpr_data:
        batch_results = st.session_state.batch_results_advanced
        df = pd.DataFrame(batch_results)
        
        # Фильтруем ошибки
        if '_error' in df.columns:
            df = df[df['_error'] != True]
        
        if not df.empty:
            scenario_a_count = len(df[df['Сценарий'] == 'A']) if 'Сценарий' in df.columns else 0
            scenario_b_count = len(df[df['Сценарий'] == 'B']) if 'Сценарий' in df.columns else 0
            total_effect = df['Эффект (₽/сут)'].sum() if 'Эффект (₽/сут)' in df.columns else 0
            
            all_results['kpr_stats'] = {
                'well_count': len(df),
                'gas_problem_count': scenario_a_count,
                'pump_problem_count': scenario_b_count,
                'total_effect': total_effect,
                'scenario_a_count': scenario_a_count,
                'scenario_b_count': scenario_b_count
            }
            
            # Сохраняем индивидуальные решения
            if 'individual_solutions' not in all_results:
                all_results['individual_solutions'] = []
            
            for _, row in df.iterrows():
                if row.get('Сценарий') in ['A', 'B']:
                    all_results['individual_solutions'].append({
                        'Скважина': row.get('Скважина', ''),
                        'Модуль': 'КПР',
                        'Сценарий': row.get('Сценарий', ''),
                        'Текущий режим': row.get('Текущий режим', ''),
                        'Новый режим': row.get('Рекомендуемый режим', ''),
                        'Эффект': row.get('Эффект (₽/сут)', 0)
                    })
    
    # Данные по замене ЭЦН (модуль 3) - ИСПРАВЛЕНО
    if has_ecn_data:
        all_results['ecn_stats'] = {
            'replace_count': 0,
            'optimize_count': 0,
            'total_energy_saving': 0,
            'total_money_saving': 0,
            'avg_kpr_hours': 0
        }
        
        all_replacements = []
        
        # Собираем данные из replace режима
        replace_results = st.session_state.get('pump_calculation_results_replace')
        if replace_results is not None and len(replace_results) > 0:
            df_replace = pd.DataFrame(replace_results)
            all_results['ecn_stats']['replace_count'] = len(set(r.get('Скважина', '') for r in replace_results if isinstance(r, dict)))
            all_results['ecn_stats']['total_energy_saving'] += df_replace['Экономия, кВт·ч/сут'].sum() if 'Экономия, кВт·ч/сут' in df_replace.columns else 0
            all_results['ecn_stats']['total_money_saving'] += df_replace['Экономия, руб/сут'].sum() if 'Экономия, руб/сут' in df_replace.columns else 0
            all_replacements.extend(replace_results)
        
        # Собираем данные из optimize режима
        optimize_results = st.session_state.get('pump_calculation_results_optimize')
        if optimize_results is not None and len(optimize_results) > 0:
            df_optimize = pd.DataFrame(optimize_results)
            all_results['ecn_stats']['optimize_count'] = len(set(r.get('Скважина', '') for r in optimize_results if isinstance(r, dict)))
            all_results['ecn_stats']['total_energy_saving'] += df_optimize['Экономия, кВт·ч/сут'].sum() if 'Экономия, кВт·ч/сут' in df_optimize.columns else 0
            all_results['ecn_stats']['total_money_saving'] += df_optimize['Экономия, руб/сут'].sum() if 'Экономия, руб/сут' in df_optimize.columns else 0
            all_replacements.extend(optimize_results)
        
        # Среднее время КПР
        kpr_hours = []
        for r in all_replacements:
            if isinstance(r, dict) and r.get('Время работы КПР, ч/сут'):
                kpr_hours.append(r['Время работы КПР, ч/сут'])
        
        if kpr_hours:
            all_results['ecn_stats']['avg_kpr_hours'] = sum(kpr_hours) / len(kpr_hours)
        
        # Индивидуальные решения по замене
        if all_replacements and 'individual_solutions' not in all_results:
            all_results['individual_solutions'] = []
        
        # Безопасно получаем списки, заменяя None на пустой список
        best_variants_replace = st.session_state.get('pump_best_variants_replace')
        if best_variants_replace is None:
            best_variants_replace = []
        
        best_variants_optimize = st.session_state.get('pump_best_variants_optimize')
        if best_variants_optimize is None:
            best_variants_optimize = []
        
        # Теперь безопасно конкатенируем
        all_best = best_variants_replace + best_variants_optimize
        
        for variant in all_best[:10]:  # Топ-10 лучших
            if 'individual_solutions' in all_results:
                # Определяем тип варианта
                if variant in best_variants_replace:
                    variant_type = 'Замена'
                elif variant in best_variants_optimize:
                    variant_type = 'Оптимизация'
                else:
                    variant_type = 'Не определен'
                
                # Безопасно получаем значения
                well_name = variant.get('Скважина', '') if isinstance(variant, dict) else ''
                new_pump_type = variant.get('Новый тип', '') if isinstance(variant, dict) else ''
                new_pump_head = variant.get('Новый напор', 0) if isinstance(variant, dict) else 0
                savings = variant.get('Экономия, руб/сут', 0) if isinstance(variant, dict) else 0
                
                all_results['individual_solutions'].append({
                    'Скважина': well_name,
                    'Модуль': 'ЭЦН',
                    'Тип': variant_type,
                    'Новый насос': f"ЭЦН-{new_pump_type}-{new_pump_head}",
                    'Экономия': savings
                })
    
    # Показываем сводку
    st.markdown("#### 📊 Сводка данных для отчета")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric("Модуль 1 (давление)", "✓" if has_pressure_data else "✗")
        if has_pressure_data:
            st.caption(f"Эффективность: {all_results['pressure_stats']['efficiency']:.1f}%")
    
    with col2:
        st.metric("Модуль 2 (КПР)", "✓" if has_kpr_data else "✗")
        if has_kpr_data and 'kpr_stats' in all_results:
            st.caption(f"Скважин: {all_results['kpr_stats']['well_count']}, Эффект: {all_results['kpr_stats']['total_effect']:,.0f} ₽/сут")
    
    with col3:
        st.metric("Модуль 3 (ЭЦН)", "✓" if has_ecn_data else "✗")
        if has_ecn_data and 'ecn_stats' in all_results:
            stats = all_results['ecn_stats']
            st.caption(f"Замена: {stats['replace_count']}, Опт.: {stats['optimize_count']}, Экономия: {stats['total_money_saving']:,.0f} ₽/сут")
    
    if 'individual_solutions' in all_results:
        st.metric("Всего решений", len(all_results['individual_solutions']))
    
    # Генерация отчета
    if st.button("🔄 Сгенерировать комплексный отчет", type="primary", key="generate_comprehensive"):
        with st.spinner("Создание комплексного отчета..."):
            try:
                # Создаем упрощенный комплексный отчет
                output = BytesIO()
                
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    # Лист 1: Общая сводка
                    summary_data = []
                    summary_data.append(['Параметр', 'Значение'])
                    summary_data.append(['Дата формирования', all_results['generated_date']])
                    summary_data.append(['ТПП', all_results['tpp']])
                    summary_data.append(['ЦИТС', all_results['cits']])
                    summary_data.append(['', ''])
                    
                    # Модуль 1
                    summary_data.append(['МОДУЛЬ 1: СТАБИЛИЗАЦИЯ ДАВЛЕНИЯ', ''])
                    if has_pressure_data:
                        summary_data.append(['Куст', all_results['pressure_stats']['cluster']])
                        summary_data.append(['Целевой дебит', f"{all_results['pressure_stats']['target_flow']:.1f} м³/час"])
                        summary_data.append(['Ср. дебит до', f"{all_results['pressure_stats']['avg_flow_before']:.1f} м³/час"])
                        summary_data.append(['Ср. дебит после', f"{all_results['pressure_stats']['avg_flow_after']:.1f} м³/час"])
                        summary_data.append(['Улучшение стабильности', f"{all_results['pressure_stats']['stability_improvement']:+.1f}%"])
                        summary_data.append(['Общая эффективность', f"{all_results['pressure_stats']['efficiency']:.1f}%"])
                    else:
                        summary_data.append(['Статус', 'Нет данных'])
                    
                    summary_data.append(['', ''])
                    
                    # Модуль 2
                    summary_data.append(['МОДУЛЬ 2: ПОТЕНЦИАЛ КПР', ''])
                    if has_kpr_data and 'kpr_stats' in all_results:
                        stats = all_results['kpr_stats']
                        summary_data.append(['Всего скважин', stats['well_count']])
                        summary_data.append(['Сценарий A (газ)', stats['scenario_a_count']])
                        summary_data.append(['Сценарий B (загрузка)', stats['scenario_b_count']])
                        summary_data.append(['Суммарный эффект', f"{stats['total_effect']:,.0f} ₽/сут"])
                    else:
                        summary_data.append(['Статус', 'Нет данных'])
                    
                    summary_data.append(['', ''])
                    
                    # Модуль 3
                    summary_data.append(['МОДУЛЬ 3: ЗАМЕНА ЭЦН', ''])
                    if has_ecn_data and 'ecn_stats' in all_results:
                        stats = all_results['ecn_stats']
                        summary_data.append(['Замена постоянных', stats['replace_count']])
                        summary_data.append(['Оптимизация КПР', stats['optimize_count']])
                        summary_data.append(['Экономия энергии', f"{stats['total_energy_saving']:,.0f} кВт·ч/сут"])
                        summary_data.append(['Экономия средств', f"{stats['total_money_saving']:,.0f} ₽/сут"])
                        summary_data.append(['Ср. время КПР', f"{stats['avg_kpr_hours']:.1f} ч/сут"])
                    else:
                        summary_data.append(['Статус', 'Нет данных'])
                    
                    pd.DataFrame(summary_data).to_excel(writer, sheet_name='Сводка', index=False, header=False)
                    
                    # Лист 2: Индивидуальные решения
                    if 'individual_solutions' in all_results and all_results['individual_solutions']:
                        df_solutions = pd.DataFrame(all_results['individual_solutions'])
                        df_solutions.to_excel(writer, sheet_name='Решения', index=False)
                    
                    # Лист 3: Детальные данные по модулям
                    if has_pressure_data and 'wells_data' in st.session_state.last_optimization:
                        # Данные по скважинам из модуля 1
                        wells_data = []
                        for well in st.session_state.last_optimization['wells_data']:
                            wells_data.append({
                                'Скважина': well['name'],
                                'Тип': 'КПР' if well['operation_mode'] == 'kpr' else 'Постоянная',
                                'Дебит': well['flow_rate'],
                                'Сдвиг фазы': st.session_state.last_optimization['phases_dict'].get(well['name'], 0)
                            })
                        if wells_data:
                            pd.DataFrame(wells_data).to_excel(writer, sheet_name='Скважины_М1', index=False)
                    
                    if has_kpr_data:
                        pd.DataFrame(batch_results).to_excel(writer, sheet_name='КПР_детально', index=False)
                    
                    if has_ecn_data and all_replacements:
                        pd.DataFrame(all_replacements).to_excel(writer, sheet_name='ЭЦН_детально', index=False)
                
                st.success("✅ Комплексный отчет успешно создан!")
                
                st.download_button(
                    label="📊 Скачать комплексный отчет",
                    data=output.getvalue(),
                    file_name=f"комплексный_отчет_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
            except Exception as e:
                st.error(f"❌ Ошибка при создании отчета: {str(e)}")
                import traceback
                st.error(traceback.format_exc())
                
def show_reports():
    """Страница отчетов и истории расчетов"""
    st.title("📊 Отчеты и история расчетов")
    
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "📈 Общая история", 
        "🎯 Стабилизация давления",
        "⚡ Потенциал КПР",
        "🔄 Замена УЭЦН",
        "📋 Комплексный отчет"
    ])
    
    with tab1:
        show_general_history()
    
    with tab2:
        show_pressure_stabilization_reports()
    
    with tab3:
        show_kpr_potential_reports()
    
    with tab4:
        show_ecn_replacement_reports()
    
    with tab5:
        show_comprehensive_reports()

# ============ ФУНКЦИИ ДЛЯ АНАЛИТИКИ ============

@st.cache_data(ttl=3600, show_spinner="Парсинг шахматки...")
def parse_chess_file(uploaded_file, tech_regime_wells=None):
    """
    Парсер файла "Шахматки" с поиском по названиям столбцов
    и правильным разделением по скважинам + извлечение метаданных из шахматки
    """
    try:
        st.info("🔍 Начинаю анализ файла...")
        
        # ============ ШАГ 1: ЧТЕНИЕ И ПОИСК ЗАГОЛОВКОВ ============
        # Читаем Excel файл без заголовков для поиска структуры
        df_raw = pd.read_excel(uploaded_file, header=None, dtype=str, engine='openpyxl')
        
        # Находим все строки с заголовками таблиц и информацией о скважинах
        sections = []
        current_section = None
        wells_metadata = {}  # Словарь для хранения метаданных скважин из шахматки
        
        for idx in range(len(df_raw)):
            row = df_raw.iloc[idx]
            row_text = ' '.join([str(cell).lower() if pd.notna(cell) else '' for cell in row])
            
            # Ищем информацию о скважине в формате "скважина: ХХХ, ..."
            if 'скважина:' in row_text:
                # Сохраняем предыдущую секцию
                if current_section and current_section.get('data_rows'):
                    sections.append(current_section)
                
                # Начинаем новую секцию
                current_section = {
                    'well_name': None,
                    'header_row': None,
                    'data_rows': [],
                    'start_idx': idx,
                    'metadata': {}  # Метаданные из шахматки
                }
                
                # Парсим информацию о скважине из шахматки
                well_info = {}
                parts = row_text.split(',')
                
                for part in parts:
                    part = part.strip()
                    if 'скважина:' in part:
                        well_name = part.split(':')[1].strip()
                        current_section['well_name'] = well_name
                        well_info['well_name'] = well_name
                    elif 'цех:' in part or 'цднг:' in part:
                        well_info['workshop'] = part.split(':')[1].strip()
                    elif 'месторождение:' in part:
                        well_info['field'] = part.split(':')[1].strip()
                    elif 'куст:' in part:
                        well_info['cluster'] = part.split(':')[1].strip()
                    elif 'пласт:' in part:
                        well_info['formation'] = part.split(':')[1].strip()
                    elif 'днс:' in part:
                        well_info['dns'] = part.split(':')[1].strip()
                    elif 'элемент:' in part:
                        well_info['element'] = part.split(':')[1].strip()
                    elif 'способ эксплуатации:' in part:
                        well_info['method'] = part.split(':')[1].strip()
                    # Тип скважины и марка насоса - ЭТО ИЗ ТЕХРЕЖИМА, НЕ ИЗ ШАХМАТКИ!
                    # Поэтому здесь их не парсим
                
                if current_section['well_name']:
                    current_section['metadata'] = well_info
                    wells_metadata[current_section['well_name']] = well_info
            
            # Ищем строку с заголовками таблицы
            elif current_section and current_section['header_row'] is None:
                header_keywords = ['дата', 'qж', 'qн', 'рзаб', 'ндин', 'f', 'примечание', 
                                  'сост', 'депрессия', 'динамический', 'затрубное', 'р на приеме']
                found_count = 0
                for cell in row:
                    if pd.notna(cell):
                        cell_str = str(cell).lower()
                        if any(keyword in cell_str for keyword in header_keywords):
                            found_count += 1
                
                if found_count >= 3:
                    current_section['header_row'] = idx
                    current_section['headers'] = row
            
            # Собираем данные
            elif current_section and current_section['header_row'] is not None:
                # Проверяем, не началась ли новая скважина
                if any('скважина:' in str(cell).lower() for cell in row if pd.notna(cell)):
                    continue
                
                # Проверяем, есть ли данные в строке
                if any(pd.notna(cell) for cell in row):
                    current_section['data_rows'].append(row)
        
        # Добавляем последнюю секцию
        if current_section and current_section.get('data_rows'):
            sections.append(current_section)
        
        st.info(f"✅ Найдено {len(sections)} скважин с данными")
        
        # ============ ШАГ 2: ФУНКЦИЯ ПОИСКА СТОЛБЦОВ ============
        def find_column(headers, keywords, alternative_keywords=None):
            """Ищет столбец по ключевым словам в названиях"""
            if headers is None:
                return None
                
            columns_lower = [str(col).lower().strip() if pd.notna(col) else '' for col in headers]
            
            # Основной поиск
            for keyword in keywords:
                for i, col_name in enumerate(columns_lower):
                    if keyword in col_name:
                        return i
            
            # Альтернативный поиск
            if alternative_keywords:
                for keyword in alternative_keywords:
                    for i, col_name in enumerate(columns_lower):
                        if keyword in col_name:
                            return i
            
            return None
        
        # ============ ШАГ 3: ПАРСИНГ ДАННЫХ ПО СЕКЦИЯМ ============
        all_records = []
        
        for section in sections:
            well_name = section['well_name']
            if not well_name:
                continue
            
            well_metadata = section.get('metadata', {})  # Метаданные из шахматки
            headers = section['headers']
            data_rows = section['data_rows']
            
            # Находим индексы столбцов для этой секции
            date_idx = find_column(headers, ['дата'])
            status_idx = find_column(headers, ['сост', 'состояние', 'статус'])
            q_liq_idx = find_column(headers, ['qж', 'дебит жидкости'], ['ср/сут qж_тм'])
            q_oil_idx = find_column(headers, ['qн', 'дебит нефти'], ['qн_расчет'])
            depression_idx = find_column(headers, ['депрессия', 'depression'])
            p_bottom_tms_idx = find_column(headers, ['рзаб вдп (тмс)', 'рзаб тмс'])
            p_bottom_ndin_idx = find_column(headers, ['рзаб вдп (ндин.)', 'рзаб ндин'])
            p_bottom_idx = find_column(headers, ['рзаб', 'забойное давление'])
            p_intake_idx = find_column(headers, ['р на приеме', 'давление на приеме'], ['p на приеме (тмс)'])
            p_zatrub_idx = find_column(headers, ['pзатр', 'затрубное'], ['pзатр hдин'])
            dyn_level_prior_idx = find_column(headers, ['ндин (приор)', 'динамический приор'])
            dyn_level_idx = find_column(headers, ['ндин', 'динамический', 'уровень'], ['hдин'])
            freq_prior_idx = find_column(headers, ['f (приор)', 'частота приор'])
            freq_idx = find_column(headers, ['f', 'частота'])
            notes_idx = find_column(headers, ['примечание', 'комментарий', 'уставка'])
            
            # Функция для безопасного преобразования в число
            def safe_float(val):
                if pd.isna(val) or val == '' or val is None:
                    return None
                try:
                    if isinstance(val, (int, float, np.number)):
                        return float(val)
                    val_str = str(val).strip().replace(',', '.')
                    val_str = re.sub(r'[^\d\.\-]', '', val_str)
                    if val_str == '' or val_str == '-':
                        return None
                    return float(val_str)
                except:
                    return None
            
            # Обрабатываем строки данных
            for row in data_rows:
                # Парсим дату
                if date_idx is not None and date_idx < len(row):
                    date_val = row[date_idx]
                    if pd.isna(date_val):
                        continue
                    
                    try:
                        date_obj = pd.to_datetime(date_val, errors='coerce')
                        if pd.isna(date_obj):
                            continue
                    except:
                        continue
                else:
                    continue
                
                # Парсим статус
                status_val = None
                is_working = True
                if status_idx is not None and status_idx < len(row):
                    status_val = row[status_idx]
                    if pd.notna(status_val):
                        status_str = str(status_val).lower()
                        is_working = 'раб' in status_str
                        status_val = status_str
                
                # Дебит жидкости
                q_liq = 0.0
                if q_liq_idx is not None and q_liq_idx < len(row):
                    val = safe_float(row[q_liq_idx])
                    if val is not None and is_working:
                        q_liq = val
                
                # Дебит нефти
                q_oil = 0.0
                if q_oil_idx is not None and q_oil_idx < len(row):
                    val = safe_float(row[q_oil_idx])
                    if val is not None and is_working:
                        q_oil = val
                
                # Забойное давление (пробуем разные источники)
                p_bottom = None
                p_bottom_source = None
                
                if p_bottom_tms_idx is not None and p_bottom_tms_idx < len(row):
                    val = safe_float(row[p_bottom_tms_idx])
                    if val is not None:
                        p_bottom = val
                        p_bottom_source = 'tms'
                
                if p_bottom is None and p_bottom_ndin_idx is not None and p_bottom_ndin_idx < len(row):
                    val = safe_float(row[p_bottom_ndin_idx])
                    if val is not None:
                        p_bottom = val
                        p_bottom_source = 'ndin'
                
                if p_bottom is None and p_bottom_idx is not None and p_bottom_idx < len(row):
                    val = safe_float(row[p_bottom_idx])
                    if val is not None:
                        p_bottom = val
                        p_bottom_source = 'generic'
                
                # Депрессия
                depression = None
                if depression_idx is not None and depression_idx < len(row):
                    depression = safe_float(row[depression_idx])
                
                # Давление на приеме
                p_intake = None
                if p_intake_idx is not None and p_intake_idx < len(row):
                    p_intake = safe_float(row[p_intake_idx])
                
                # Затрубное давление
                p_zatrub = None
                if p_zatrub_idx is not None and p_zatrub_idx < len(row):
                    p_zatrub = safe_float(row[p_zatrub_idx])
                
                # Динамический уровень
                dyn_level = None
                dyn_level_source = None
                
                if dyn_level_prior_idx is not None and dyn_level_prior_idx < len(row):
                    val = safe_float(row[dyn_level_prior_idx])
                    if val is not None:
                        dyn_level = val
                        dyn_level_source = 'prior'
                
                if dyn_level is None and dyn_level_idx is not None and dyn_level_idx < len(row):
                    val = safe_float(row[dyn_level_idx])
                    if val is not None:
                        dyn_level = val
                        dyn_level_source = 'regular'
                
                # Частота
                freq = None
                freq_source = None
                
                if freq_prior_idx is not None and freq_prior_idx < len(row):
                    val = safe_float(row[freq_prior_idx])
                    if val is not None:
                        freq = val
                        freq_source = 'prior'
                
                if freq is None and freq_idx is not None and freq_idx < len(row):
                    val = safe_float(row[freq_idx])
                    if val is not None:
                        freq = val
                        freq_source = 'regular'
                
                # Примечания
                notes = ''
                if notes_idx is not None and notes_idx < len(row):
                    notes_val = row[notes_idx]
                    if pd.notna(notes_val):
                        notes = str(notes_val).strip()
                
                # Создаем запись с метаданными из шахматки
                record = {
                    'date': date_obj,
                    'well_name': well_name,
                    'well_name_original': well_name,  # сохраняем оригинал
                    'well_name_normalized': str(well_name).upper().strip(),  # нормализованно
                    'status': status_val,
                    'is_working': is_working,
                    'Q_liq': q_liq,
                    'Q_oil': q_oil,
                    'Q_water': q_liq - q_oil,
                    'Water_cut_percent': (q_liq - q_oil) / q_liq * 100 if q_liq > 0 else 0,
                    'P_bottom': p_bottom,
                    'P_bottom_source': p_bottom_source,
                    'Depression': depression,
                    'P_intake': p_intake,
                    'P_zatrub': p_zatrub,
                    'Dyn_level': dyn_level,
                    'Dyn_level_source': dyn_level_source,
                    'Freq': freq,
                    'Freq_source': freq_source,
                    'Notes': notes,
                    # Метаданные из шахматки (только то, что реально есть в шахматке)
                    'workshop': well_metadata.get('workshop'),      # ЦДНГ из шахматки
                    'field': well_metadata.get('field'),            # Месторождение из шахматки
                    'cluster': well_metadata.get('cluster'),        # Куст из шахматки
                    'formation': well_metadata.get('formation'),    # Пласт из шахматки
                    'dns': well_metadata.get('dns'),                # ДНС из шахматки
                    'element': well_metadata.get('element'),        # Элемент из шахматки
                    'method': well_metadata.get('method')           # Способ эксплуатации из шахматки
                    # well_type и pump_model НЕ ДОБАВЛЯЕМ - они из техрежима!
                }
                
                # Рассчитываем пластовое давление
                if record['Depression'] is not None and record['P_bottom'] is not None:
                    record['P_res'] = record['Depression'] + record['P_bottom']
                else:
                    record['P_res'] = None
                
                # Рассчитываем частоту в Гц
                if record['Freq'] is not None:
                    record['Freq_hz'] = record['Freq'] if record['Freq'] < 100 else record['Freq'] * 50 / 2910
                else:
                    record['Freq_hz'] = None
                
                all_records.append(record)
        
        # ============ ШАГ 4: СОЗДАНИЕ DATAFRAME ============
        if not all_records:
            st.error("❌ Не удалось извлечь данные из файла")
            return pd.DataFrame()
        
        df_result = pd.DataFrame(all_records)
        
        # Сортировка по скважине и дате
        df_result = df_result.sort_values(['well_name', 'date']).reset_index(drop=True)
        
        # Заполняем пропуски в дебитах
        for col in ['Q_liq', 'Q_oil', 'Q_water']:
            if col in df_result.columns:
                df_result[col] = df_result[col].fillna(0)
        
        # Временные метки
        df_result['day_of_week'] = df_result['date'].dt.day_name()
        df_result['month'] = df_result['date'].dt.month
        df_result['year'] = df_result['date'].dt.year
        df_result['month_year'] = df_result['date'].dt.strftime('%Y-%m')
        
        # ============ ШАГ 5: СТАТИСТИКА ============
        st.success(f"✅ Файл успешно обработан!")
        
        # Общая статистика
        st.write("📊 **Общая статистика:**")
        st.write(f"- Всего записей: {len(df_result)}")
        st.write(f"- Уникальных скважин: {df_result['well_name'].nunique()}")
        st.write(f"- Период: {df_result['date'].min():%d.%m.%Y} - {df_result['date'].max():%d.%m.%Y}")
        
        # Предпросмотр данных
        st.write("👁️ **Предпросмотр данных:**")
        preview_cols = ['date', 'well_name', 'status', 'Q_liq', 'Q_oil', 'Water_cut_percent', 
                       'P_res', 'P_bottom', 'Depression', 'Dyn_level', 'Freq_hz', 'P_intake']
        available_cols = [col for col in preview_cols if col in df_result.columns]
        st.dataframe(df_result[available_cols].head(20), use_container_width=True)
        
        # Метаданные скважин из шахматки
        st.write("📋 **Метаданные скважин из шахматки:**")
        metadata_cols = ['well_name', 'workshop', 'field', 'formation', 'dns', 'element', 'method']
        available_metadata = [col for col in metadata_cols if col in df_result.columns]
        
        # Получаем уникальные метаданные по скважинам
        metadata_unique = df_result[available_metadata].drop_duplicates('well_name')
        st.dataframe(metadata_unique, use_container_width=True)

        if tech_regime_wells and not df_result.empty:
            chess_wells = set(df_result['well_name_normalized'].unique())
            tech_wells = set()
            for well in tech_regime_wells:
                name = well.get('name', '')
                if name:
                    tech_wells.add(str(name).upper().strip())
            
            matched = chess_wells.intersection(tech_wells)
            not_matched = chess_wells - tech_wells
            
            if not_matched:
                st.warning(f"⚠️ {len(not_matched)} скважин из шахматки не найдены в техрежиме после нормализации")
                
                # Показываем примеры для первых 5
                examples = list(not_matched)[:5]
                example_text = ", ".join([f"'{e}'" for e in examples])
                st.info(f"Примеры ненайденных (нормализованные): {example_text}")        
        
        return df_result
        
    except Exception as e:
        st.error(f"❌ Ошибка при парсинге файла: {str(e)}")
        import traceback
        st.error(traceback.format_exc())
        return pd.DataFrame()

def debug_show_column_detection(df_raw):
    """
    Функция для отладки поиска столбцов
    """
    st.subheader("🔍 Детальная отладка поиска столбцов")
    
    # Показываем первые 15 строк
    st.write("**Первые 15 строк файла:**")
    for idx in range(min(15, len(df_raw))):
        row = df_raw.iloc[idx]
        row_str = " | ".join([f"[{i}] '{str(cell)[:30]}...'" if len(str(cell)) > 30 
                            else f"[{i}] '{cell}'" for i, cell in enumerate(row) if pd.notna(cell)][:10])
        st.write(f"Строка {idx+1}: {row_str}")
    
    # Ищем потенциальные заголовки
    st.write("\n**Поиск заголовков:**")
    for idx in range(min(20, len(df_raw))):
        row = df_raw.iloc[idx]
        keywords = ['дата', 'qж', 'qн', 'рпл', 'рзаб', 'ндин', 'f', 'примечание']
        
        found = []
        for i, cell in enumerate(row):
            if pd.notna(cell):
                cell_str = str(cell).lower()
                for keyword in keywords:
                    if keyword in cell_str:
                        found.append(f"'{keyword}' в столбце {i}")
        
        if found:
            st.write(f"Строка {idx+1}: найдено {len(found)} совпадений: {', '.join(found)}")

def determine_well_type(well_data):
    """
    Определяет тип скважины на основе данных из техрежима
    Логика:
    1. Если installation_type = 'ОРД' -> 'ОРД'
    2. Если installation_type = 'ШГН' -> 'ШГН'
    3. Если installation_type = 'ЭЦН':
       - Если нет BH или BH=0 -> 'Постоянная'
       - Если есть BH:
         * work_time + pause_time > 60 минут -> 'АПВ'
         * work_time + pause_time <= 60 минут -> 'КПР'
    """
    installation_type = well_data.get('installation_type', '').upper()
    
    # ОРД и ШГН - отдельно
    if installation_type == 'ОРД':
        return 'ОРД'
    elif installation_type == 'ШГН':
        return 'ШГН'
    
    # Для ЭЦН определяем по режиму работы
    elif installation_type == 'ЭЦН':
        operation_mode = well_data.get('operation_mode', '')
        
        if operation_mode == 'constant':
            return 'Постоянная'
        elif operation_mode == 'kpr':
            # Получаем время работы и накопления
            schedule = well_data.get('schedule', [None, None])
            work_time = schedule[0] if schedule else None  # в минутах
            pause_time = schedule[1] if schedule else None  # в минутах
            
            # Проверяем на АПВ
            if work_time is not None and pause_time is not None:
                total_cycle = work_time + pause_time  # в минутах
                if total_cycle > 60:
                    return 'АПВ'
                else:
                    return 'КПР'
            else:
                return 'КПР'  # по умолчанию КПР, если нет данных
        else:
            return 'Неизвестно'
    else:
        return 'Неизвестно'

@st.cache_data(ttl=3600)
def enrich_chess_data(chess_df, tech_regime_wells):
    """
    Объединяет данные из Шахматки с данными из техрежима
    Теперь с нормализацией регистра (приводит названия скважин к верхнему регистру)
    """
    if chess_df.empty:
        return pd.DataFrame()
    
    # Создаём копию данных
    enriched_df = chess_df.copy()
    
    # НОРМАЛИЗАЦИЯ: Приводим названия скважин в шахматке к верхнему регистру
    if 'well_name' in enriched_df.columns:
        enriched_df['well_name_normalized'] = enriched_df['well_name'].astype(str).str.upper().str.strip()
    
    # Создаём справочник скважин из техрежима с нормализованными ключами
    tech_wells_dict = {}
    for well in tech_regime_wells:
        well_name = well.get('name', '')
        if well_name:
            # Нормализуем название из техрежима (тоже к верхнему регистру)
            well_name_norm = str(well_name).upper().strip()
            
            # Определяем тип скважины
            well_type = determine_well_type(well)
            
            tech_wells_dict[well_name_norm] = {
                'cluster': well.get('cluster', 'Неизвестно'),
                'cdng': well.get('cdng', 'Неизвестно'),
                'cits': well.get('cits', 'Неизвестно'),
                'well_type': well_type,
                'installation_type': well.get('installation_type', 'Неизвестно'),
                'pump_mark': well.get('pump_mark', 'Неизвестно'),
                'pump_flow': well.get('pump_flow', None),
                'pump_head': well.get('pump_head', None),
                'work_time': well.get('schedule', [None, None])[0] if well.get('schedule') else None,
                'pause_time': well.get('schedule', [None, None])[1] if well.get('schedule') else None,
                'original_name': well_name  # сохраняем оригинальное название для отладки
            }
    
    # Функция для получения данных по нормализованному имени
    def get_tech_data(well_name_norm):
        return tech_wells_dict.get(well_name_norm, {})
    
    # Добавляем данные из техрежима, используя нормализованное имя
    enriched_df['tech_cluster'] = enriched_df['well_name_normalized'].apply(
        lambda x: get_tech_data(x).get('cluster', 'Неизвестно')
    )
    enriched_df['tech_cdng'] = enriched_df['well_name_normalized'].apply(
        lambda x: get_tech_data(x).get('cdng', 'Неизвестно')
    )
    enriched_df['tech_cits'] = enriched_df['well_name_normalized'].apply(
        lambda x: get_tech_data(x).get('cits', 'Неизвестно')
    )
    enriched_df['tech_well_type'] = enriched_df['well_name_normalized'].apply(
        lambda x: get_tech_data(x).get('well_type', 'Неизвестно')
    )
    enriched_df['tech_installation_type'] = enriched_df['well_name_normalized'].apply(
        lambda x: get_tech_data(x).get('installation_type', 'Неизвестно')
    )
    enriched_df['tech_pump_mark'] = enriched_df['well_name_normalized'].apply(
        lambda x: get_tech_data(x).get('pump_mark', 'Неизвестно')
    )
    enriched_df['tech_pump_flow'] = enriched_df['well_name_normalized'].apply(
        lambda x: get_tech_data(x).get('pump_flow', None)
    )
    enriched_df['tech_work_time'] = enriched_df['well_name_normalized'].apply(
        lambda x: get_tech_data(x).get('work_time', None)
    )
    enriched_df['tech_pause_time'] = enriched_df['well_name_normalized'].apply(
        lambda x: get_tech_data(x).get('pause_time', None)
    )
    
    # Рассчитываем полный цикл
    enriched_df['tech_total_cycle_min'] = enriched_df.apply(
        lambda row: row['tech_work_time'] + row['tech_pause_time'] 
        if row['tech_work_time'] is not None and row['tech_pause_time'] is not None 
        else None,
        axis=1
    )
    
    # Для отладки: добавим колонку с информацией о том, нашлась ли скважина
    enriched_df['tech_found'] = enriched_df['well_name_normalized'].apply(
        lambda x: x in tech_wells_dict
    )
    
    # Удаляем временную колонку нормализации (опционально)
    # enriched_df = enriched_df.drop('well_name_normalized', axis=1)
    
    return enriched_df

def show_tech_regime_matching_debug(chess_df, tech_regime_wells):
    """
    Показывает статистику соответствия скважин из шахматки и техрежима
    """
    st.subheader("🔍 Отладка соответствия скважин")
    
    # Нормализуем названия из шахматки
    chess_wells = set(chess_df['well_name'].astype(str).str.upper().str.strip().unique())
    
    # Нормализуем названия из техрежима
    tech_wells = set()
    tech_wells_original = {}
    for well in tech_regime_wells:
        name = well.get('name', '')
        if name:
            name_norm = str(name).upper().strip()
            tech_wells.add(name_norm)
            tech_wells_original[name_norm] = name
    
    # Находим соответствия
    matched = chess_wells.intersection(tech_wells)
    not_matched = chess_wells - tech_wells
    
    # Статистика
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Скважин в шахматке", len(chess_wells))
    with col2:
        st.metric("Скважин в техрежиме", len(tech_wells))
    with col3:
        st.metric("Найдено соответствий", len(matched))
    
    # Показываем ненайденные скважины
    if not_matched:
        st.warning(f"⚠️ {len(not_matched)} скважин из шахматки не найдены в техрежиме:")
        
        # Создаем таблицу с деталями
        debug_data = []
        for well_norm in sorted(not_matched)[:20]:  # первые 20
            # Ищем оригинальное название в шахматке
            chess_originals = chess_df[chess_df['well_name'].astype(str).str.upper().str.strip() == well_norm]['well_name'].unique()
            chess_original = ', '.join(str(x) for x in chess_originals)
            
            debug_data.append({
                'Нормализованное имя': well_norm,
                'В шахматке как': chess_original,
                'В техрежиме (похожие)': ', '.join([w for w in tech_wells if w[:5] == well_norm[:5]][:3])
            })
        
        if debug_data:
            st.dataframe(pd.DataFrame(debug_data), use_container_width=True)
        
        if len(not_matched) > 20:
            st.caption(f"... и еще {len(not_matched) - 20} скважин")
    
    return matched, not_matched

def create_filters_panel(enriched_df):
    """
    Создаёт панель фильтров для данных
    """
    if enriched_df.empty:
        st.sidebar.warning("Нет данных для фильтрации")
        return None
    
    st.sidebar.markdown("### 🔍 Фильтры данных")
    
    # Период
    if 'date' in enriched_df.columns:
        min_date = enriched_df['date'].min()
        max_date = enriched_df['date'].max()
        
        date_range = st.sidebar.date_input(
            "📅 Период",
            value=[min_date, max_date],
            min_value=min_date,
            max_value=max_date
        )
        
        if len(date_range) == 2:
            start_date, end_date = date_range
        else:
            start_date, end_date = min_date, max_date
    else:
        start_date, end_date = None, None
    
    # Цех (CDNG)
    cdng_options = ['Все'] + sorted([str(x) for x in enriched_df['tech_cdng'].unique().tolist() if str(x) != 'nan'])
    selected_cdng = st.sidebar.multiselect(
        "🏭 Цех (ЦДНГ)",
        options=cdng_options,
        default=['Все']
    )
    
    # ЦИТС
    cits_options = ['Все'] + sorted([str(x) for x in enriched_df['tech_cits'].unique().tolist() if str(x) != 'nan'])
    selected_cits = st.sidebar.multiselect(
        "🏢 ЦИТС",
        options=cits_options,
        default=['Все']
    )
    
    # Тип скважины (обновленный список)
    well_type_options = ['Все']
    if 'tech_well_type' in enriched_df.columns:
        unique_types = [str(x) for x in enriched_df['tech_well_type'].unique().tolist() 
                       if str(x) != 'nan' and str(x) != 'None']
        well_type_options.extend(sorted(unique_types))
    
    selected_well_type = st.sidebar.multiselect(
        "🔧 Тип скважины",
        options=well_type_options,
        default=['Все']
    )
    
    # Тип установки
    install_type_options = ['Все']
    if 'tech_installation_type' in enriched_df.columns:
        unique_install = [str(x) for x in enriched_df['tech_installation_type'].unique().tolist() 
                         if str(x) != 'nan' and str(x) != 'None']
        install_type_options.extend(sorted(unique_install))
    
    selected_install_type = st.sidebar.multiselect(
        "⚙️ Тип установки",
        options=install_type_options,
        default=['Все']
    )
    
    # Куст
    cluster_options = ['Все'] + sorted([str(x) for x in enriched_df['tech_cluster'].unique().tolist() if str(x) != 'nan'])
    selected_cluster = st.sidebar.multiselect(
        "🛢️ Куст",
        options=cluster_options,
        default=['Все']
    )
    
    # Подача насоса (диапазон) - с проверкой на одинаковые значения
    flow_range = None
    if 'tech_pump_flow' in enriched_df.columns:
        pump_flows = enriched_df['tech_pump_flow'].dropna()
        if not pump_flows.empty:
            min_flow = float(pump_flows.min())
            max_flow = float(pump_flows.max())
            
            # Проверяем, чтобы min_flow и max_flow не были одинаковыми
            if min_flow == max_flow:
                # Если все значения одинаковые, расширяем диапазон
                if min_flow > 0:
                    min_flow = max(0, min_flow - 10)
                max_flow = max_flow + 10
            
            # Проверяем, что min_flow < max_flow
            if min_flow < max_flow:
                flow_range = st.sidebar.slider(
                    "🔧 Подача насоса, м³/сут",
                    min_value=min_flow,
                    max_value=max_flow,
                    value=(min_flow, max_flow),
                    step=1.0
                )
            else:
                # Если min_flow >= max_flow, просто показываем значение
                st.sidebar.info(f"Подача насоса: {min_flow:.0f} м³/сут")
    
    # Список скважин
    well_options = ['Все'] + sorted(enriched_df['well_name'].unique().tolist())
    selected_wells = st.sidebar.multiselect(
        "🎯 Выбор скважин",
        options=well_options,
        default=['Все']
    )
    
    # Кнопка применения фильтров
    apply_filters = st.sidebar.button("✅ Применить фильтры", type="primary", use_container_width=True)
    
    return {
        'start_date': start_date,
        'end_date': end_date,
        'selected_cdng': selected_cdng,
        'selected_cits': selected_cits,
        'selected_well_type': selected_well_type,
        'selected_install_type': selected_install_type,
        'selected_cluster': selected_cluster,
        'flow_range': flow_range,  # Добавляем flow_range в возвращаемый словарь
        'selected_wells': selected_wells,
        'apply_filters': apply_filters
    }

def apply_filters_to_data(enriched_df, filters):
    """
    Применяет фильтры к данным
    """
    if enriched_df.empty:
        return pd.DataFrame()
    
    filtered_df = enriched_df.copy()
    
    # Фильтр по дате
    if filters['start_date'] and filters['end_date']:
        filtered_df = filtered_df[
            (filtered_df['date'] >= pd.Timestamp(filters['start_date'])) &
            (filtered_df['date'] <= pd.Timestamp(filters['end_date']))
        ]
    
    # Фильтр по ЦДНГ
    if 'Все' not in filters['selected_cdng'] and filters['selected_cdng']:
        filtered_df = filtered_df[filtered_df['tech_cdng'].isin(filters['selected_cdng'])]
    
    # Фильтр по ЦИТС
    if 'Все' not in filters['selected_cits'] and filters['selected_cits']:
        filtered_df = filtered_df[filtered_df['tech_cits'].isin(filters['selected_cits'])]
    
    # Фильтр по типу скважины
    if 'Все' not in filters['selected_well_type'] and filters['selected_well_type']:
        filtered_df = filtered_df[filtered_df['tech_well_type'].isin(filters['selected_well_type'])]
    
    # Фильтр по типу установки
    if 'Все' not in filters['selected_install_type'] and filters['selected_install_type']:
        filtered_df = filtered_df[filtered_df['tech_installation_type'].isin(filters['selected_install_type'])]
    
    # Фильтр по кусту
    if 'Все' not in filters['selected_cluster'] and filters['selected_cluster']:
        filtered_df = filtered_df[filtered_df['tech_cluster'].isin(filters['selected_cluster'])]
    
    # Фильтр по подаче насоса
    if filters.get('flow_range'):
        min_flow, max_flow = filters['flow_range']
        filtered_df = filtered_df[
            (filtered_df['tech_pump_flow'] >= min_flow) & 
            (filtered_df['tech_pump_flow'] <= max_flow)
        ]
    
    # Фильтр по скважинам
    if 'Все' not in filters['selected_wells'] and filters['selected_wells']:
        filtered_df = filtered_df[filtered_df['well_name'].isin(filters['selected_wells'])]
    
    return filtered_df

def show_statistics_tab(filtered_df):
    """
    Отображает вкладку общей статистики
    """
    if filtered_df.empty:
        st.warning("Нет данных для отображения статистики")
        return
    
    # Панель метрик KPI
    st.subheader("📊 Ключевые показатели")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        total_wells = filtered_df['well_name'].nunique()
        st.metric("Всего скважин", total_wells)
    
    with col2:
        avg_q_oil = filtered_df['Q_oil'].mean()
        st.metric("Ср. дебит нефти, м³/сут", f"{avg_q_oil:.1f}")
    
    with col3:
        avg_q_liq = filtered_df['Q_liq'].mean()
        st.metric("Ср. дебит жидкости, м³/сут", f"{avg_q_liq:.1f}")
    
    with col4:
        avg_water_cut = filtered_df['Water_cut_percent'].mean()
        st.metric("Ср. обводненность, %", f"{avg_water_cut:.1f}")
    
    # Разделитель
    st.markdown("---")
    
    # Распределение по типам скважин
    st.subheader("📈 Распределение по типам скважин")
    
    col_left, col_right = st.columns([2, 1])
    
    with col_left:
        # Круговая диаграмма
        if 'tech_well_type' in filtered_df.columns:
            type_distribution = filtered_df.groupby('tech_well_type')['well_name'].nunique().reset_index()
            type_distribution.columns = ['Тип скважины', 'Количество']
            
            # Сортируем по количеству
            type_distribution = type_distribution.sort_values('Количество', ascending=False)
            
            if not type_distribution.empty:
                # Цветовая схема для типов
                color_map = {
                    'Постоянная': '#2E8B57',  # зеленый
                    'КПР': '#FFA500',         # оранжевый
                    'АПВ': '#FF6347',         # красный
                    'ОРД': '#4682B4',         # синий
                    'ШГН': '#8A2BE2'          # фиолетовый
                }
                
                colors = [color_map.get(t, '#CCCCCC') for t in type_distribution['Тип скважины']]
                
                fig_pie = px.pie(
                    type_distribution,
                    values='Количество',
                    names='Тип скважины',
                    title='Распределение скважин по типам',
                    color='Тип скважины',
                    color_discrete_map=color_map
                )
                fig_pie.update_traces(textposition='inside', textinfo='percent+label')
                st.plotly_chart(fig_pie, use_container_width=True)
    
    with col_right:
        # Таблица распределения
        if 'tech_well_type' in filtered_df.columns:
            type_stats = filtered_df.groupby('tech_well_type').agg({
                'well_name': 'nunique',
                'Q_oil': 'mean',
                'Q_liq': 'mean',
                'Water_cut_percent': 'mean'
            }).round(1)
            
            type_stats.columns = ['Скважин', 'Qн ср.', 'Qж ср.', 'Обв. ср.%']
            
            # Добавляем столбец с долей
            total = type_stats['Скважин'].sum()
            type_stats['Доля, %'] = (type_stats['Скважин'] / total * 100).round(1)
            
            st.dataframe(type_stats, use_container_width=True)
            
            # Показываем общую статистику по типам установок
            if 'tech_installation_type' in filtered_df.columns:
                st.markdown("**По типам установок:**")
                install_stats = filtered_df.groupby('tech_installation_type')['well_name'].nunique()
                for install_type, count in install_stats.items():
                    st.write(f"- {install_type}: {count} скважин")
    
    # Распределение по подаче насосов
    st.markdown("---")
    st.subheader("🔧 Распределение по подаче насосов")
    
    if 'tech_pump_flow' in filtered_df.columns:
        # Фильтруем только указанные подачи
        pump_flow_df = filtered_df.copy()
        specified_flows = pump_flow_df[pump_flow_df['tech_pump_flow'].notna()].copy()
        
        if not specified_flows.empty:
            # Определяем интересующие нас подачи насосов
            target_pump_flows = [25, 40, 60, 80, 125, 160, 200, 250, 400]
            
            # Создаем категорию для каждой целевой подачи
            def categorize_specific_pump_flow(flow):
                flow_rounded = round(flow)
                if flow_rounded in target_pump_flows:
                    return f'{flow_rounded} м³/сут'
                else:
                    return 'Другие подачи'
            
            specified_flows['pump_category'] = specified_flows['tech_pump_flow'].apply(categorize_specific_pump_flow)
            
            # Группируем по категориям, считаем УНИКАЛЬНЫЕ скважины
            # Важно: используем 'well_name' для подсчета уникальных скважин
            pump_distribution = specified_flows.groupby('pump_category')['well_name'].nunique().reset_index()
            pump_distribution.columns = ['Подача насоса', 'Количество скважин']
            
            # Сортируем по значению подачи (извлекаем число из строки)
            def extract_flow_value(category):
                if category == 'Другие подачи':
                    return float('inf')  # Помещаем в конец
                return int(category.split()[0])
            
            pump_distribution['flow_value'] = pump_distribution['Подача насоса'].apply(extract_flow_value)
            pump_distribution = pump_distribution.sort_values('flow_value')
            
            # Создаем график
            fig_bar = px.bar(
                pump_distribution,
                x='Подача насоса',
                y='Количество скважин',
                title='Распределение по подаче насосов',
                text='Количество скважин',
                color='Количество скважин',
                color_continuous_scale='Viridis'
            )
            
            fig_bar.update_layout(
                xaxis_title="Подача насоса, м³/сут",
                yaxis_title="Количество скважин",
                xaxis_tickangle=0,
                showlegend=False
            )
            
            st.plotly_chart(fig_bar, use_container_width=True)
            
            # Таблица со статистикой
            st.markdown("#### 📊 Статистика по подачам насосов")
            
            # Рассчитываем дополнительную статистику для каждой подачи
            detailed_stats = []
            
            for flow in target_pump_flows:
                flow_mask = specified_flows['pump_category'] == f'{flow} м³/сут'
                flow_data = specified_flows[flow_mask]
                if not flow_data.empty:
                    # Важно: считаем УНИКАЛЬНЫЕ скважины для количества
                    well_count = flow_data['well_name'].nunique()
                    
                    # Для средних значений используем агрегацию по уникальным скважинам
                    # Сначала группируем по скважине, берем среднее по периоду, затем среднее по всем скважинам
                    unique_well_stats = flow_data.groupby('well_name').agg({
                        'Q_oil': 'mean',
                        'Q_liq': 'mean',
                        'Water_cut_percent': 'mean'
                    }).mean() if well_count > 0 else {'Q_oil': None, 'Q_liq': None, 'Water_cut_percent': None}
                    
                    detailed_stats.append({
                        'Подача насоса, м³/сут': flow,
                        'Количество скважин': well_count,
                        'Qн средний': round(unique_well_stats['Q_oil'], 1) if unique_well_stats['Q_oil'] is not None else None,
                        'Qж средний': round(unique_well_stats['Q_liq'], 1) if unique_well_stats['Q_liq'] is not None else None,
                        'Обв. средняя, %': round(unique_well_stats['Water_cut_percent'], 1) if unique_well_stats['Water_cut_percent'] is not None else None
                    })
            
            # Добавляем статистику по "другим" подачам
            other_mask = specified_flows['pump_category'] == 'Другие подачи'
            other_data = specified_flows[other_mask]
            if not other_data.empty:
                other_count = other_data['well_name'].nunique()
                other_unique_stats = other_data.groupby('well_name').agg({
                    'Q_oil': 'mean',
                    'Q_liq': 'mean',
                    'Water_cut_percent': 'mean'
                }).mean() if other_count > 0 else {'Q_oil': None, 'Q_liq': None, 'Water_cut_percent': None}
                
                detailed_stats.append({
                    'Подача насоса, м³/сут': 'Другие',
                    'Количество скважин': other_count,
                    'Qн средний': round(other_unique_stats['Q_oil'], 1) if other_unique_stats['Q_oil'] is not None else None,
                    'Qж средний': round(other_unique_stats['Q_liq'], 1) if other_unique_stats['Q_liq'] is not None else None,
                    'Обв. средняя, %': round(other_unique_stats['Water_cut_percent'], 1) if other_unique_stats['Water_cut_percent'] is not None else None
                })
            
            if detailed_stats:
                stats_df = pd.DataFrame(detailed_stats)
                
                # Вычисляем правильные метрики
                total_wells_with_flow = specified_flows['well_name'].nunique()  # Уникальные скважины с указанной подачей
                total_wells_all = pump_flow_df['well_name'].nunique()  # Все уникальные скважины
                
                col_stats1, col_stats2 = st.columns([2, 1])
                
                with col_stats1:
                    st.dataframe(
                        stats_df,
                        use_container_width=True,
                        column_config={
                            'Количество скважин': st.column_config.NumberColumn(format="%d"),
                            'Qн средний': st.column_config.NumberColumn(format="%.1f"),
                            'Qж средний': st.column_config.NumberColumn(format="%.1f"),
                            'Обв. средняя, %': st.column_config.NumberColumn(format="%.1f")
                        }
                    )
                
                with col_stats2:
                    # Показываем правильные метрики
                    st.metric("Всего скважин с указанной подачей", total_wells_with_flow)
                    st.metric("Скважин без указанной подачи", total_wells_all - total_wells_with_flow)
                    
                    if not pump_distribution.empty:
                        most_common = pump_distribution.iloc[pump_distribution['Количество скважин'].argmax()]
                        st.metric(
                            "Наиболее распространенная подача", 
                            most_common['Подача насоса'],
                            f"{most_common['Количество скважин']} скважин"
                        )
            
            # Гистограмма распределения всех подач
            st.markdown("#### 📈 Гистограмма всех подач")
            
            col_hist1, col_hist2 = st.columns([3, 1])
            
            with col_hist1:
                # Для гистограммы используем уникальные скважины (последнее значение подачи для каждой скважины)
                latest_flows = specified_flows.sort_values('date').groupby('well_name').last().reset_index()
                
                fig_hist = px.histogram(
                    latest_flows,
                    x='tech_pump_flow',
                    nbins=20,
                    title='Гистограмма распределения подач насосов',
                    labels={'tech_pump_flow': 'Подача насоса, м³/сут'},
                    color_discrete_sequence=['#636efa']
                )
                
                # Добавляем вертикальные линии для целевых подач
                for flow in target_pump_flows:
                    fig_hist.add_vline(
                        x=flow, 
                        line_dash="dash", 
                        line_color="red", 
                        opacity=0.3,
                        annotation_text=f"{flow} м³/сут",
                        annotation_position="top"
                    )
                
                fig_hist.update_layout(
                    xaxis_title="Подача насоса, м³/сут",
                    yaxis_title="Количество скважин",
                    showlegend=False
                )
                
                st.plotly_chart(fig_hist, use_container_width=True)
            
            with col_hist2:
                # Показываем основные статистики
                st.markdown("**Основные статистики:**")
                
                # Используем последние значения для статистики
                if not latest_flows.empty:
                    stats_summary = latest_flows['tech_pump_flow'].describe().round(1)
                    st.write(f"**Среднее:** {stats_summary['mean']} м³/сут")
                    st.write(f"**Медиана:** {stats_summary['50%']} м³/сут")
                    st.write(f"**Мин:** {stats_summary['min']} м³/сут")
                    st.write(f"**Макс:** {stats_summary['max']} м³/сут")
                    st.write(f"**Станд. отклонение:** {stats_summary['std']} м³/сут")
                    
                    # Определяем модальные значения
                    mode_values = latest_flows['tech_pump_flow'].mode()
                    if not mode_values.empty:
                        st.write(f"**Мода:** {mode_values.iloc[0]} м³/сут")
    
    # Таблица с агрегированными данными по скважинам
    st.markdown("---")
    st.subheader("📋 Сводные данные по скважинам")
    
    # Группируем по скважинам и берём последние значения
    latest_data = filtered_df.sort_values(['well_name', 'date']).groupby('well_name').last().reset_index()
    
    # Отображаем только нужные колонки
    display_cols = [
        'well_name', 'tech_cdng', 'tech_cluster', 'tech_well_type',
        'tech_installation_type', 'tech_pump_flow',
        'Q_oil', 'Q_liq', 'Water_cut_percent', 'P_bottom', 'P_res', 'Freq_hz'
    ]
    
    # Фильтруем только существующие колонки
    available_cols = [col for col in display_cols if col in latest_data.columns]
    
    if available_cols:
        display_df = latest_data[available_cols]
        
        # Переименовываем колонки для отображения
        rename_dict = {
            'well_name': 'Скважина',
            'tech_cdng': 'ЦДНГ',
            'tech_cluster': 'Куст',
            'tech_well_type': 'Тип скважины',
            'tech_installation_type': 'Тип установки',
            'tech_pump_flow': 'Подача насоса',
            'Q_oil': 'Qн',
            'Q_liq': 'Qж',
            'Water_cut_percent': 'Обв.%',
            'P_bottom': 'Pзаб',
            'P_res': 'Pпл',
            'Freq_hz': 'Частота'
        }
        
        display_df = display_df.rename(columns=rename_dict)
        
        # Форматируем числовые колонки
        if 'Подача насоса' in display_df.columns:
            display_df['Подача насоса'] = display_df['Подача насоса'].apply(
                lambda x: f"{x:.0f}" if pd.notna(x) else ""
            )
        
        if 'Qн' in display_df.columns:
            display_df['Qн'] = display_df['Qн'].apply(
                lambda x: f"{x:.1f}" if pd.notna(x) else ""
            )
        
        if 'Qж' in display_df.columns:
            display_df['Qж'] = display_df['Qж'].apply(
                lambda x: f"{x:.1f}" if pd.notna(x) else ""
            )
        
        if 'Обв.%' in display_df.columns:
            display_df['Обв.%'] = display_df['Обв.%'].apply(
                lambda x: f"{x:.1f}" if pd.notna(x) else ""
            )
        
        if 'Pзаб' in display_df.columns:
            display_df['Pзаб'] = display_df['Pзаб'].apply(
                lambda x: f"{x:.1f}" if pd.notna(x) else ""
            )
        
        if 'Pпл' in display_df.columns:
            display_df['Pпл'] = display_df['Pпл'].apply(
                lambda x: f"{x:.1f}" if pd.notna(x) else ""
            )
        
        if 'Частота' in display_df.columns:
            display_df['Частота'] = display_df['Частота'].apply(
                lambda x: f"{x:.1f}" if pd.notna(x) else ""
            )
        
        st.dataframe(display_df, use_container_width=True)

def show_time_series_tab(filtered_df):
    """
    Отображает вкладку анализа временных рядов
    """
    if filtered_df.empty:
        st.warning("Нет данных для анализа временных рядов")
        return
    
    st.subheader("📈 Анализ временных рядов")
    
    # Выбор скважины
    well_list = sorted(filtered_df['well_name'].unique().tolist())
    selected_well = st.selectbox("Выберите скважину", well_list)
    
    if not selected_well:
        return
    
    # Фильтруем данные для выбранной скважины
    well_data = filtered_df[filtered_df['well_name'] == selected_well].sort_values('date')
    
    if well_data.empty:
        st.warning(f"Нет данных для скважины {selected_well}")
        return
    
    # Выбор параметров для отображения
    st.subheader("Выберите параметры для отображения")
    
    col1, col2 = st.columns(2)
    
    with col1:
        show_debits = st.checkbox("Дебиты (жидкость/нефть/вода)", value=True)
        show_pressures = st.checkbox("Давления (пластовое/забойное/депрессия)", value=True)
    
    with col2:
        show_level_pressure = st.checkbox("Динамический уровень + давление на приеме", value=True)
        show_frequency = st.checkbox("Частота", value=True)
    
    # График дебитов
    if show_debits:
        st.markdown("### Дебиты скважины")
        
        fig_debits = go.Figure()
        
        # Дебит жидкости
        if 'Q_liq' in well_data.columns:
            fig_debits.add_trace(go.Scatter(
                x=well_data['date'],
                y=well_data['Q_liq'],
                mode='lines+markers',
                name='Дебит жидкости (Qж)',
                line=dict(color='blue', width=2)
            ))
        
        # Дебит нефти
        if 'Q_oil' in well_data.columns:
            fig_debits.add_trace(go.Scatter(
                x=well_data['date'],
                y=well_data['Q_oil'],
                mode='lines+markers',
                name='Дебит нефти (Qн)',
                line=dict(color='green', width=2)
            ))
        
        # Дебит воды (заливка)
        if 'Q_water' in well_data.columns and 'Q_oil' in well_data.columns:
            fig_debits.add_trace(go.Scatter(
                x=well_data['date'],
                y=well_data['Q_water'],
                mode='none',
                name='Дебит воды',
                fill='tonexty',
                fillcolor='rgba(135, 206, 250, 0.3)',
                hoverinfo='skip'
            ))
        
        fig_debits.update_layout(
            title=f'Дебиты скважины {selected_well}',
            xaxis_title='Дата',
            yaxis_title='Дебит, м³/сут',
            hovermode='x unified',
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
        )
        
        st.plotly_chart(fig_debits, use_container_width=True)
    
    # График пластового и забойного давлений
    if show_pressures:
        st.markdown("### Давления и депрессия")
        
        fig_pressures = go.Figure()
        
        # Забойное давление (РИСУЕМ ПЕРВЫМ - это будет нижняя граница)
        if 'P_bottom' in well_data.columns:
            fig_pressures.add_trace(go.Scatter(
                x=well_data['date'],
                y=well_data['P_bottom'],
                mode='lines+markers',
                name='Забойное давление (Pзаб)',
                line=dict(color='orange', width=2),
                connectgaps=True
            ))
        
        # Заливка для депрессии (ДОБАВЛЯЕМ СРАЗУ ПОСЛЕ НИЖНЕЙ ГРАНИЦЫ)
        if 'P_res' in well_data.columns and 'P_bottom' in well_data.columns:
            fig_pressures.add_trace(go.Scatter(
                x=well_data['date'],
                y=well_data['P_res'],  # ВЕРХНЯЯ граница - пластовое давление
                mode='lines',
                name='Депрессия',
                fill='tonexty',  # Заливка ДО предыдущего графика (забойного давления)
                fillcolor='rgba(255, 215, 0, 0.3)',  # Желтый полупрозрачный
                line=dict(width=0),  # Прозрачная линия
                hoverinfo='skip'
            ))
        
        # Пластовое давление (РИСУЕМ ПОСЛЕДНИМ - это будет верхняя граница)
        if 'P_res' in well_data.columns:
            fig_pressures.add_trace(go.Scatter(
                x=well_data['date'],
                y=well_data['P_res'],
                mode='lines+markers',
                name='Пластовое давление (Pпл)',
                line=dict(color='red', width=2),
                connectgaps=True
            ))
        
        fig_pressures.update_layout(
            title=f'Давления скважины {selected_well}',
            xaxis_title='Дата',
            yaxis_title='Давление, атм',
            hovermode='x unified',
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
        )
        
        st.plotly_chart(fig_pressures, use_container_width=True)
    
    # График динамического уровня и давления на приеме
    if show_level_pressure:
        st.markdown("### Динамический уровень и давление на приеме")
        
        fig_level = make_subplots(specs=[[{"secondary_y": True}]])
        
        # Динамический уровень
        if 'Dyn_level' in well_data.columns:
            fig_level.add_trace(
                go.Scatter(
                    x=well_data['date'],
                    y=well_data['Dyn_level'],
                    mode='lines+markers',  # Линии соединены с маркерами
                    name='Динамический уровень',
                    line=dict(color='purple', width=2),
                    connectgaps=True
                ),
                secondary_y=False
            )
        
        # Давление на приеме
        if 'P_intake' in well_data.columns:
            fig_level.add_trace(
                go.Scatter(
                    x=well_data['date'],
                    y=well_data['P_intake'],
                    mode='lines+markers',  # Линии соединены с маркерами
                    name='Давление на приеме',
                    line=dict(color='brown', width=2),
                    connectgaps=True
                ),
                secondary_y=True
            )
        
        fig_level.update_layout(
            title=f'Динамический уровень и давление на приеме {selected_well}',
            xaxis_title='Дата',
            hovermode='x unified',
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
        )
        
        fig_level.update_yaxes(title_text="Динамический уровень, м", secondary_y=False)
        fig_level.update_yaxes(title_text="Давление на приеме, атм", secondary_y=True)
        
        st.plotly_chart(fig_level, use_container_width=True)
    
    # График частоты
    if show_frequency:
        st.markdown("### Частота работы")
        
        if 'Freq_hz' in well_data.columns:
            fig_freq = go.Figure()
            
            fig_freq.add_trace(go.Scatter(
                x=well_data['date'],
                y=well_data['Freq_hz'],
                mode='lines+markers',  # Линии соединены с маркерами
                name='Частота (Гц)',
                line=dict(color='darkgreen', width=2),
                fill='tozeroy',
                fillcolor='rgba(144, 238, 144, 0.3)',
                connectgaps=True
            ))
            
            fig_freq.update_layout(
                title=f'Частота работы скважины {selected_well}',
                xaxis_title='Дата',
                yaxis_title='Частота, Гц',
                hovermode='x unified'
            )
            
            st.plotly_chart(fig_freq, use_container_width=True)

def show_detailed_analysis_tab(filtered_df):
    """
    Отображает вкладку детального анализа скважины
    """
    if filtered_df.empty:
        st.warning("Нет данных для детального анализа")
        return
    
    st.subheader("🔍 Детальный анализ скважины")
    
    # Выбор скважины
    well_list = sorted(filtered_df['well_name'].unique().tolist())
    selected_well = st.selectbox("Выберите скважину для детального анализа", well_list, key="detailed_select")
    
    if not selected_well:
        return
    
    # Фильтруем данные для выбранной скважины
    well_data = filtered_df[filtered_df['well_name'] == selected_well].sort_values('date')
    
    if well_data.empty:
        st.warning(f"Нет данных для скважины {selected_well}")
        return
    
    # Основная информация о скважине
    st.markdown("### 📋 Основная информация")
    
    col_info1, col_info2, col_info3 = st.columns(3)
    
    with col_info1:
        first_row = well_data.iloc[0]
        st.metric("Скважина", selected_well)
        st.metric("Цех", first_row.get('workshop', 'Неизвестно'))
        st.metric("Месторождение", first_row.get('field', 'Неизвестно'))
    
    with col_info2:
        st.metric("Пласт", first_row.get('formation', 'Неизвестно'))
        st.metric("ДНС", first_row.get('dns', 'Неизвестно'))
        st.metric("Элемент", first_row.get('element', 'Неизвестно'))
    
    with col_info3:
        if 'tech_well_type' in first_row:
            st.metric("Тип скважины", first_row['tech_well_type'])
        if 'tech_pump_mark' in first_row:
            st.metric("Марка насоса", first_row['tech_pump_mark'])
        if 'method' in first_row:
            st.metric("Способ эксплуатации", first_row['method'])
    
    # История изменений уставок и примечаний
    st.markdown("---")
    st.subheader("📝 История уставок и примечаний")
    
    # Фильтруем строки с примечаниями
    notes_data = well_data[well_data['Notes'] != ''].copy()
    
    if not notes_data.empty:
        notes_display = notes_data[['date', 'Notes']].copy()
        notes_display['Дата'] = notes_display['date'].dt.strftime('%d.%m.%Y')
        notes_display['Примечания/Уставки'] = notes_display['Notes']
        notes_display = notes_display[['Дата', 'Примечания/Уставки']].reset_index(drop=True)
        
        st.dataframe(notes_display, use_container_width=True)
        
        # Визуализация изменений уставок
        st.markdown("#### 📅 Временная шкала изменений")
        
        # Создаем временную шкалу
        timeline_data = []
        for idx, row in notes_data.iterrows():
            if row['Notes']:
                timeline_data.append({
                    'Дата': row['date'],
                    'Событие': row['Notes'][:100] + "..." if len(row['Notes']) > 100 else row['Notes']
                })
        
        if timeline_data:
            timeline_df = pd.DataFrame(timeline_data)
            timeline_df = timeline_df.sort_values('Дата')
            
            # Простая визуализация временной шкалы
            for idx, event in timeline_df.iterrows():
                date_str = event['Дата'].strftime('%d.%m.%Y')
                st.markdown(f"**{date_str}**: {event['Событие']}")
    else:
        st.info("Нет записей с примечаниями для этой скважины")
    
    # Сравнение периодов
    st.markdown("---")
    st.subheader("📊 Сравнение периодов")
    
    col_period1, col_period2 = st.columns(2)
    
    with col_period1:
        st.markdown("#### Период 1")
        if not well_data.empty:
            min_date = well_data['date'].min()
            max_date = well_data['date'].max()
            
            period1_start = st.date_input("Начало периода 1", min_date, key="p1_start")
            period1_end = st.date_input("Конец периода 1", max_date, key="p1_end")
    
    with col_period2:
        st.markdown("#### Период 2")
        period2_start = st.date_input("Начало периода 2", min_date, key="p2_start")
        period2_end = st.date_input("Конец периода 2", max_date, key="p2_end")
    
    if st.button("🔄 Сравнить периоды", type="primary"):
        # Фильтруем данные для периодов
        period1_data = well_data[
            (well_data['date'] >= pd.Timestamp(period1_start)) &
            (well_data['date'] <= pd.Timestamp(period1_end))
        ]
        
        period2_data = well_data[
            (well_data['date'] >= pd.Timestamp(period2_start)) &
            (well_data['date'] <= pd.Timestamp(period2_end))
        ]
        
        if not period1_data.empty and not period2_data.empty:
            # Рассчитываем средние значения для каждого периода
            metrics = ['Q_oil', 'Q_liq', 'Q_water', 'Water_cut_percent', 
                      'P_bottom', 'P_res', 'Depression', 'Freq_hz']
            
            comparison_results = []
            
            for metric in metrics:
                if metric in period1_data.columns and metric in period2_data.columns:
                    period1_avg = period1_data[metric].mean()
                    period2_avg = period2_data[metric].mean()
                    
                    # Рассчитываем изменение
                    if period1_avg != 0:
                        change_percent = ((period2_avg - period1_avg) / period1_avg) * 100
                    else:
                        change_percent = 0
                    
                    comparison_results.append({
                        'Параметр': metric,
                        'Период 1': round(period1_avg, 2),
                        'Период 2': round(period2_avg, 2),
                        'Изменение, %': round(change_percent, 2)
                    })
            
            if comparison_results:
                comparison_df = pd.DataFrame(comparison_results)
                
                # Переименовываем параметры для лучшего отображения
                param_names = {
                    'Q_oil': 'Дебит нефти, м³/сут',
                    'Q_liq': 'Дебит жидкости, м³/сут',
                    'Q_water': 'Дебит воды, м³/сут',
                    'Water_cut_percent': 'Обводненность, %',
                    'P_bottom': 'Забойное давление, атм',
                    'P_res': 'Пластовое давление, атм',
                    'Depression': 'Депрессия, атм',
                    'Freq_hz': 'Частота, Гц'
                }
                
                comparison_df['Параметр'] = comparison_df['Параметр'].map(
                    lambda x: param_names.get(x, x)
                )
                
                st.dataframe(comparison_df, use_container_width=True)
                
                # Визуализация изменений
                st.markdown("#### 📈 Визуализация изменений")
                
                fig_comparison = go.Figure()
                
                for idx, row in comparison_df.iterrows():
                    fig_comparison.add_trace(go.Bar(
                        name=row['Параметр'],
                        x=['Период 1', 'Период 2'],
                        y=[row['Период 1'], row['Период 2']],
                        text=[f"{row['Период 1']}", f"{row['Период 2']}"],
                        textposition='auto'
                    ))
                
                fig_comparison.update_layout(
                    title='Сравнение параметров между периодами',
                    barmode='group',
                    xaxis_title='Период',
                    yaxis_title='Значение',
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
                )
                
                st.plotly_chart(fig_comparison, use_container_width=True)

def show_analytics():
    """
    Главная функция модуля аналитики скважин
    """

    _load_plotly()
         
    st.title("📈 Аналитика скважин")
    st.markdown("---")
    
    # Проверяем, загружены ли данные техрежима
    if 'wells_data' not in st.session_state or not st.session_state.wells_data:
        st.warning("⚠️ Сначала загрузите данные техрежима в модуле 'Кусты'")
        if st.button("🛢️ Перейти к загрузке техрежима"):
            st.session_state.current_page = "wells"
            st.rerun()
        return
    
    # Вкладки модуля аналитики
    tab1, tab2, tab3, tab4 = st.tabs([
        "📥 Загрузка данных", 
        "📊 Общая статистика", 
        "📈 Анализ временных рядов",
        "🔍 Детальный анализ"
    ])
    
    # ============ ВКЛАДКА 1: ЗАГРУЗКА ДАННЫХ ============
    with tab1:
        st.subheader("📥 Загрузка файла Шахматки")
        
        uploaded_file = st.file_uploader(
            "Выберите файл Шахматки (Excel)",
            type=['xlsx', 'xls'],
            key="chess_uploader"
        )
        
        if uploaded_file is not None:
            with st.spinner("📊 Парсинг файла Шахматки..."):
                # Получаем данные техрежима из session_state
                tech_regime_wells = st.session_state.get('wells_data', [])
                
                # Парсим файл с передачей данных техрежима
                chess_df = parse_chess_file(uploaded_file, tech_regime_wells)
                
                if chess_df is not None and not chess_df.empty:
                    st.success(f"✅ Файл успешно загружен: {len(chess_df)} записей")
                    
                    # Обогащаем данные техрежимом
                    enriched_df = enrich_chess_data(chess_df, st.session_state.wells_data)
                    
                    # Сохраняем в session_state
                    st.session_state.chess_raw_data = chess_df
                    st.session_state.chess_enriched_data = enriched_df
                    
                    # АВТОМАТИЧЕСКОЕ СОХРАНЕНИЕ
                    save_data_to_file()
                    
                    # Показываем статистику загрузки
                    col_load1, col_load2, col_load3 = st.columns(3)
                    
                    with col_load1:
                        st.metric("Всего записей", len(chess_df))
                    
                    with col_load2:
                        unique_wells = chess_df['well_name'].nunique()
                        st.metric("Уникальных скважин", unique_wells)
                    
                    with col_load3:
                        date_range = f"{chess_df['date'].min():%d.%m.%Y} - {chess_df['date'].max():%d.%m.%Y}"
                        st.metric("Период данных", date_range)
                    
                    # Показываем предпросмотр данных
                    st.subheader("🔍 Предпросмотр данных")
                    st.dataframe(chess_df.head(10), use_container_width=True)
                    
                else:
                    st.error("❌ Не удалось распарсить файл. Проверьте формат файла.")
        
        # Если данные уже загружены, показываем информацию
        elif 'chess_enriched_data' in st.session_state:
            enriched_df = st.session_state.chess_enriched_data
            
            if enriched_df is not None and not enriched_df.empty:
                st.success("✅ Данные Шахматки уже загружены")
                
                col_info1, col_info2, col_info3 = st.columns(3)
                
                with col_info1:
                    st.metric("Записей в базе", len(enriched_df))
                
                with col_info2:
                    unique_wells = enriched_df['well_name'].nunique()
                    st.metric("Скважин в базе", unique_wells)
                
                with col_info3:
                    date_range = f"{enriched_df['date'].min():%d.%m.%Y} - {enriched_df['date'].max():%d.%m.%Y}"
                    st.metric("Период данных", date_range)
        
        else:
            st.info("📁 Загрузите файл Шахматки для начала работы")
    
    # ============ ВКЛАДКА 2: ОБЩАЯ СТАТИСТИКА ============
    with tab2:
        # БЕЗОПАСНО получаем данные
        enriched_df = st.session_state.get('chess_enriched_data')
        
        if enriched_df is None:
            st.warning("Сначала загрузите файл Шахматки во вкладке 'Загрузка данных'")
        elif enriched_df.empty:
            st.warning("Загруженный файл не содержит данных")
        else:
            # Создаем фильтры
            filters = create_filters_panel(enriched_df)
            
            # Применяем фильтры
            if filters and filters['apply_filters']:
                filtered_df = apply_filters_to_data(enriched_df, filters)
                current_data = filtered_df
                
                # Сохраняем отфильтрованные данные в session_state
                st.session_state.filtered_analytics_data = filtered_df
                
                # АВТОМАТИЧЕСКОЕ СОХРАНЕНИЕ
                save_data_to_file()
                
                if current_data is not None and not current_data.empty:
                    show_statistics_tab(current_data)
                else:
                    st.warning("Нет данных по выбранным фильтрам")
            else:
                # Показываем статистику по всем данным
                current_data = enriched_df
                # Сохраняем все данные как текущие для других вкладок
                st.session_state.current_analytics_data = enriched_df
                show_statistics_tab(current_data)
    
    # ============ ВКЛАДКА 3: АНАЛИЗ ВРЕМЕННЫХ РЯДОВ ============
    with tab3:
        # БЕЗОПАСНО получаем данные
        enriched_df = st.session_state.get('chess_enriched_data')
        
        if enriched_df is None:
            st.warning("Сначала загрузите файл Шахматки во вкладке 'Загрузка данных'")
        elif enriched_df.empty:
            st.warning("Загруженный файл не содержит данных")
        else:
            # Определяем, какие данные использовать
            if 'current_analytics_data' not in st.session_state:
                st.session_state.current_analytics_data = enriched_df
            
            # Используем отфильтрованные данные, если они есть
            filtered_data = st.session_state.get('filtered_analytics_data')
            if filtered_data is not None and not filtered_data.empty:
                data_to_use = filtered_data
            else:
                data_to_use = st.session_state.current_analytics_data
            
            if data_to_use is not None and not data_to_use.empty:
                # Добавляем кнопку сохранения (опционально)
                col_save1, col_save2 = st.columns([3, 1])
                with col_save2:
                    if st.button("💾 Сохранить", key="save_time_series"):
                        save_data_to_file()
                        st.success("✅ Данные сохранены")
                
                show_time_series_tab(data_to_use)
            else:
                st.warning("Нет данных для анализа. Примените фильтры во вкладке 'Общая статистика'")
    
    # ============ ВКЛАДКА 4: ДЕТАЛЬНЫЙ АНАЛИЗ ============
    with tab4:
        # БЕЗОПАСНО получаем данные
        enriched_df = st.session_state.get('chess_enriched_data')
        
        if enriched_df is None:
            st.warning("Сначала загрузите файл Шахматки во вкладке 'Загрузка данных'")
        elif enriched_df.empty:
            st.warning("Загруженный файл не содержит данных")
        else:
            # Определяем, какие данные использовать
            if 'current_analytics_data' not in st.session_state:
                st.session_state.current_analytics_data = enriched_df
            
            # Используем отфильтрованные данные, если они есть
            filtered_data = st.session_state.get('filtered_analytics_data')
            if filtered_data is not None and not filtered_data.empty:
                data_to_use = filtered_data
            else:
                data_to_use = st.session_state.current_analytics_data
            
            if data_to_use is not None and not data_to_use.empty:
                # Добавляем кнопку сохранения
                col_save1, col_save2 = st.columns([3, 1])
                with col_save2:
                    if st.button("💾 Сохранить", key="save_detailed"):
                        save_data_to_file()
                        st.success("✅ Данные сохранены")
                
                show_detailed_analysis_tab(data_to_use)
            else:
                st.warning("Нет данных для детального анализа. Примените фильтры во вкладке 'Общая статистика'")
    
    # Футер модуля с кнопкой сохранения
    st.markdown("---")
    col_footer1, col_footer2, col_footer3 = st.columns([1, 1, 1])
    with col_footer2:
        if st.button("💾 Сохранить все данные аналитики", use_container_width=True):
            save_data_to_file()
            st.success("✅ Все данные аналитики сохранены")
    
    st.caption("📊 Модуль аналитики скважин | Используйте фильтры для точного анализа данных")

def calculate_mixture_density(oil_density_relative, water_cut_percent):
    """
    Расчет плотности водонефтяной смеси в кг/м³.
    
    Parameters:
    -----------
    oil_density_relative : float
        Относительная плотность нефти (0.84, 0.86 и т.д.)
    water_cut_percent : float
        Обводненность в %
        
    Returns:
    --------
    float: Плотность смеси в кг/м³
    """
    # Обработка None значений
    if oil_density_relative is None:
        oil_density_relative = 0.85  # по умолчанию
    
    if water_cut_percent is None:
        water_cut_percent = 0
    
    # Абсолютная плотность нефти (кг/м³)
    oil_density_kg_m3 = oil_density_relative * 1000
    
    # Абсолютная плотность воды (кг/м³)
    water_density_kg_m3 = 1000  # при 20°C
    
    # Доля воды (от 0 до 1)
    water_fraction = water_cut_percent / 100
    
    # Плотность смеси (кг/м³)
    mixture_density_kg_m3 = (oil_density_kg_m3 * (1 - water_fraction) + 
                           water_density_kg_m3 * water_fraction)
    
    return mixture_density_kg_m3

def show_load_analysis_tab():
    _load_plotly()
    """Вкладка анализа и оптимизации нагрузки на систему сбора."""
    
    st.header("📊 Анализ и оптимизация нагрузки на систему сбора")
    st.markdown("""
    **Оптимизация работы КПР скважин с учетом ограничений по скорости потока в трубопроводе.**  
    
    📌 **Возможности:**
    - Анализ текущей почасовой нагрузки на систему
    - Расчет скорости потока для выбранного трубопровода
    - Автоматическая оптимизация времени запуска КПР скважин
    - Создание подробного Excel-отчета
    - Сохранение результатов оптимизации в систему
    """)
    
    # Проверяем наличие данных
    if 'wells_data' not in st.session_state or not st.session_state.wells_data:
        st.warning("⚠️ Нет данных о скважинах. Загрузите файл во вкладке 'Загрузка данных'.")
        return
    
    # Инициализация session_state для оптимизации
    if 'load_optimizer' not in st.session_state:
        st.session_state.load_optimizer = None
    if 'current_load_analysis' not in st.session_state:
        st.session_state.current_load_analysis = None
    if 'optimization_results' not in st.session_state:
        st.session_state.optimization_results = None
    if 'unsaved_changes' not in st.session_state:
        st.session_state.unsaved_changes = False
    
    # Проверка на наличие несохраненных изменений
    if 'load_optimizer' in st.session_state and st.session_state.load_optimizer:
        has_unsaved_changes = False
        for well in st.session_state.wells_data:
            if well.get('modification_source') == 'load_optimization' and well.get('last_modified'):
                has_unsaved_changes = True
                break
        
        if has_unsaved_changes and not st.session_state.unsaved_changes:
            st.warning("⚠️ У вас есть несохраненные изменения от предыдущей оптимизации!")
            if st.button("🔄 Очистить несохраненные изменения"):
                for i, well in enumerate(st.session_state.wells_data):
                    if well.get('modification_source') == 'load_optimization':
                        # Восстанавливаем исходное время
                        if 'original_launch_time' in well:
                            st.session_state.wells_data[i]['base_launch_time'] = well['original_launch_time']
                            del st.session_state.wells_data[i]['original_launch_time']
                        del st.session_state.wells_data[i]['modification_source']
                        if 'last_modified' in st.session_state.wells_data[i]:
                            del st.session_state.wells_data[i]['last_modified']
                st.session_state.unsaved_changes = False
                st.rerun()
    
    # ================== ШАГ 1: ВЫБОР КУСТОВ ==================
    st.markdown("---")
    st.markdown("#### 🏗️ Шаг 1: Выбор кустов для анализа")
    
    # Получаем список всех уникальных кустов
    all_wells = st.session_state.wells_data
    unique_clusters = sorted(set(w.get('cluster', 'Неизвестно') for w in all_wells if w.get('is_active', True)))
    
    if not unique_clusters:
        st.error("❌ Нет активных скважин для анализа.")
        return
    
    # Выбор кустов
    col_select1, col_select2 = st.columns([2, 1])
    
    with col_select1:
        selected_clusters = st.multiselect(
            "Выберите кусты для анализа:",
            options=unique_clusters,
            default=unique_clusters[:min(3, len(unique_clusters))],
            help="Выберите один или несколько кустов для анализа"
        )
    
    # Фильтруем скважины
    if selected_clusters:
        filtered_wells = [w for w in all_wells 
                         if w.get('is_active', True) and w.get('cluster') in selected_clusters]
    else:
        filtered_wells = [w for w in all_wells if w.get('is_active', True)]
        selected_clusters = unique_clusters
    
    with col_select2:
        # Статистика по выбранным скважинам
        kpr_count = sum(1 for w in filtered_wells if w.get('operation_mode') == 'kpr')
        constant_count = sum(1 for w in filtered_wells if w.get('operation_mode') == 'constant')
        total_flow = sum(w.get('flow_rate', 0) for w in filtered_wells)
        
        st.metric("КПР скважин", kpr_count)
        st.metric("Постоянных скважин", constant_count)
        st.metric("Суммарный дебит, м³/сут", f"{total_flow:.0f}")
    
    # ================== ШАГ 2: ПАРАМЕТРЫ ТРУБОПРОВОДА ==================
    st.markdown("---")
    st.markdown("#### 🛢️ Шаг 2: Параметры трубопровода")
    
    # Инициализация session_state для параметров трубопровода, если их нет
    if 'pipeline_params' not in st.session_state or st.session_state.pipeline_params is None:
        st.session_state.pipeline_params = {
            'diameter_mm': 273,
            'wall_thickness': 8,
            'inner_diameter': 257,
            'v_min': 0.7,
            'v_max': 2.5,
            'avg_water_cut': 50,
            'mixture_density': 920
        }
    
    col_pipe1, col_pipe2, col_pipe3 = st.columns(3)
    
    with col_pipe1:
        # Выбор стандартного диаметра
        diameter_options = list(STANDARD_PIPELINE_DIAMETERS.keys())
        selected_diameter_str = st.selectbox(
            "Наружный диаметр трубы, мм",
            options=diameter_options,
            index=3,  # 273 мм по умолчанию
            help="Выберите стандартный диаметр трубопровода"
        )
        
        # Конвертируем строку в число
        selected_diameter = int(selected_diameter_str)
        
        # Автоматически показываем толщину и внутренний диаметр
        pipe_data = STANDARD_PIPELINE_DIAMETERS[selected_diameter_str]
        st.caption(f"Толщина стенки: {pipe_data['wall_thickness']} мм")
        st.caption(f"Внутренний диаметр: {pipe_data['inner_diameter']} мм")
    
    # ПЕРВОЕ: Рассчитываем среднюю обводненность
    if filtered_wells:
        total_flow = sum(w.get('flow_rate', 0) for w in filtered_wells)
        if total_flow > 0:
            weighted_water_cut_sum = 0
            for w in filtered_wells:
                flow = w.get('flow_rate', 0)
                water_cut = w.get('water_cut', 0)
                weighted_water_cut_sum += flow * water_cut
            avg_water_cut = weighted_water_cut_sum / total_flow
        else:
            avg_water_cut = 50
    else:
        avg_water_cut = 50
    
    # Только ПОСЛЕ этого используем avg_water_cut
    with col_pipe2:
        # Рекомендация по минимальной скорости
        if avg_water_cut > 70:
            recommended_min = 0.5  # в основном вода
        elif avg_water_cut > 30:
            recommended_min = 0.6  # водонефтяная эмульсия
        else:
            recommended_min = 0.7  # в основном нефть
        
        # БЕЗОПАСНО получаем значение из session_state
        current_v_min = st.session_state.pipeline_params.get('v_min', recommended_min) if st.session_state.pipeline_params else recommended_min
        
        v_min = st.number_input(
            "Минимальная скорость, м/с",
            min_value=0.1,
            max_value=2.0,
            value=current_v_min,
            step=0.1,
            help=f"Рекомендуется: {recommended_min} м/с (средняя обводненность {avg_water_cut:.0f}%)"
        )
    
    with col_pipe3:
        # Рекомендация по максимальной скорости
        if avg_water_cut > 70:
            recommended_max = 3.0  # вода допускает больше
        elif avg_water_cut > 30:
            recommended_max = 2.0  # для эмульсии ниже
        else:
            recommended_max = 2.5  # нефть
        
        # БЕЗОПАСНО получаем значение из session_state
        current_v_max = st.session_state.pipeline_params.get('v_max', recommended_max) if st.session_state.pipeline_params else recommended_max
        
        v_max = st.number_input(
            "Максимальная скорость, м/с",
            min_value=0.5,
            max_value=5.0,
            value=current_v_max,
            step=0.1,
            help=f"Рекомендуется: {recommended_max} м/с (средняя обводненность {avg_water_cut:.0f}%)"
        )
    
    # Расчетная пропускная способность
    inner_diameter_m = pipe_data['inner_diameter'] / 1000
    area_m2 = math.pi * (inner_diameter_m / 2) ** 2
    max_flow_m3_per_hour = area_m2 * v_max * 3600  # м³/час
    min_flow_m3_per_hour = area_m2 * v_min * 3600  # м³/час
    
    # Расчет средней плотности смеси
    if filtered_wells:
        total_flow = sum(w.get('flow_rate', 0) for w in filtered_wells)
        if total_flow > 0:
            weighted_density_sum = 0
            for w in filtered_wells:
                flow = w.get('flow_rate', 0)
                density = w.get('oil_density')
                
                if density is None:
                    density = 0.85
                
                weighted_density_sum += flow * density
            
            weighted_density = weighted_density_sum / total_flow
            mixture_density = calculate_mixture_density(weighted_density, avg_water_cut)
        else:
            mixture_density = calculate_mixture_density(0.85, avg_water_cut)
    else:
        mixture_density = calculate_mixture_density(0.85, avg_water_cut)
    
    # СОХРАНЯЕМ ПАРАМЕТРЫ В SESSION_STATE
    st.session_state.pipeline_params = {
        'diameter_mm': selected_diameter,
        'wall_thickness': pipe_data['wall_thickness'],
        'inner_diameter': pipe_data['inner_diameter'],
        'v_min': v_min,
        'v_max': v_max,
        'avg_water_cut': avg_water_cut,
        'mixture_density': mixture_density
    }
    
    st.info(f"""
    **📊 Расчетные параметры:**
    - **Трубопровод:** {selected_diameter}×{pipe_data['wall_thickness']} мм (внутр. {pipe_data['inner_diameter']} мм)
    - **Пропускная способность:** {min_flow_m3_per_hour:.0f} - {max_flow_m3_per_hour:.0f} м³/час
    - **Суммарный дебит кустов:** {sum(w.get('flow_rate', 0) for w in filtered_wells):.0f} м³/сут
    - **Средняя обводненность:** {avg_water_cut:.1f}%
    - **Плотность смеси:** {mixture_density:.0f} кг/м³
    """)
    
    # ================== ШАГ 3: НАСТРОЙКИ АНАЛИЗА ==================
    st.markdown("---")
    st.markdown("#### ⚙️ Шаг 3: Настройки анализа")
    
    col_set1, col_set2, col_set3 = st.columns(3)
    
    with col_set1:
        time_step = st.selectbox(
            "Шаг времени для расчетов:",
            options=[5, 10, 15, 30, 60],
            index=0,
            help="Более мелкий шаг дает более точные результаты, но требует больше вычислений"
        )
    
    with col_set2:
        optimization_method = st.selectbox(
            "Метод оптимизации:",
            options=["differential_evolution", "greedy"],
            index=0,
            help="Differential Evolution - более точный, но медленнее. Greedy - быстрее, но менее точен."
        )
    
    with col_set3:
        max_iterations = st.slider(
            "Максимум итераций:",
            min_value=10,
            max_value=200,
            value=50,
            step=10,
            help="Больше итераций = лучше результат, но дольше расчет"
        )
    
    # ================== ШАГ 4: АНАЛИЗ И ОПТИМИЗАЦИЯ ==================
    st.markdown("---")
    st.markdown("#### 🔄 Шаг 4: Анализ и оптимизация")
    
    col_step4_1, col_step4_2, col_step4_3, col_step4_4 = st.columns(4)
    
    with col_step4_1:
        analyze_clicked = st.button(
            "🔍 Проанализировать",
            use_container_width=True,
            type="secondary",
            help="Анализ текущей нагрузки без оптимизации"
        )
    
    with col_step4_2:
        optimize_clicked = st.button(
            "⚡ Оптимизировать",
            use_container_width=True,
            type="primary",
            help="Запустить полную оптимизацию"
        )
    
    with col_step4_3:
        reset_clicked = st.button(
            "🔄 Сбросить",
            use_container_width=True,
            help="Сбросить все результаты"
        )
    
    with col_step4_4:
        # Показываем статус
        if st.session_state.unsaved_changes:
            st.error("⚠️ Есть несохраненные изменения")
        elif st.session_state.optimization_results:
            st.success("✅ Оптимизация завершена")
        elif st.session_state.current_load_analysis:
            st.info("📊 Анализ выполнен")
    
    if reset_clicked:
        st.session_state.load_optimizer = None
        st.session_state.current_load_analysis = None
        st.session_state.optimization_results = None
        st.session_state.unsaved_changes = False
        st.rerun()
    
    # ================== АНАЛИЗ ТЕКУЩЕЙ НАГРУЗКИ ==================
    if analyze_clicked:
        with st.spinner("🔍 Анализ текущей нагрузки..."):
            try:
                # Создаем оптимизатор
                optimizer = SystemLoadOptimizer(
                    wells_data=st.session_state.wells_data,
                    selected_clusters=selected_clusters,
                    time_step_minutes=time_step
                )
                
                # Устанавливаем параметры трубопровода из session_state
                if 'pipeline_params' in st.session_state:
                    params = st.session_state.pipeline_params
                    optimizer.set_pipeline_parameters(
                        diameter_mm=params['diameter_mm'],
                        v_min=params['v_min'],
                        v_max=params['v_max']
                    )
                    st.success(f"✅ Параметры трубопровода установлены: {params['diameter_mm']} мм")
                else:
                    st.warning("⚠️ Параметры трубопровода не установлены! Используем значения по умолчанию.")
                    optimizer.set_pipeline_parameters(
                        diameter_mm=273,
                        v_min=0.7,
                        v_max=2.5
                    )
                
                # Теперь анализируем
                current_load = optimizer.calculate_current_hourly_load()
                problems = optimizer.find_problem_intervals()
                
                # Сохраняем в session_state
                st.session_state.load_optimizer = optimizer
                st.session_state.current_load_analysis = {
                    'current_load': current_load,
                    'problems': problems,
                    'selected_clusters': selected_clusters
                }
                
                # Проверяем, рассчитана ли скорость
                if current_load.get('velocity_profile'):
                    st.success("✅ Текущая нагрузка проанализирована! Скорость рассчитана.")
                else:
                    st.warning("⚠️ Нагрузка проанализирована, но скорость не рассчитана. Проверьте параметры трубопровода.")
                    
            except Exception as e:
                st.error(f"❌ Ошибка анализа: {str(e)}")
    
    # ================== ОПТИМИЗАЦИЯ НАГРУЗКИ ==================
    if optimize_clicked:
        if 'load_optimizer' not in st.session_state or st.session_state.load_optimizer is None:
            st.warning("⚠️ Сначала выполните анализ текущей нагрузки!")
        else:
            with st.spinner("⚡ Оптимизация нагрузки..."):
                try:
                    optimizer = st.session_state.load_optimizer
                    
                    results = optimizer.optimize_launch_times(
                        optimization_method=optimization_method,
                        max_iterations=max_iterations
                    )
                    
                    if 'error' in results:
                        st.error(f"❌ Ошибка оптимизации: {results['error']}")
                    else:
                        st.session_state.optimization_results = results
                        st.success("✅ Оптимизация завершена успешно!")
                        
                        if 'comparison_stats' in results:
                            improvements = results['comparison_stats']['improvements']
                            st.info(f"""
                            **📈 Основные результаты:**
                            - **Общее улучшение:** {improvements.get('overall_improvement', 0):.1f}%
                            - **Снижение пиковой нагрузки:** {improvements.get('max_load_reduction', 0):.1f}%
                            - **Снижение размаха:** {improvements.get('range_reduction', 0):.1f}%
                            """)
                
                except Exception as e:
                    st.error(f"❌ Ошибка оптимизации: {str(e)}")
    
    # ================== РЕЗУЛЬТАТЫ АНАЛИЗА ==================
    if 'current_load_analysis' in st.session_state and st.session_state.current_load_analysis:
        analysis = st.session_state.current_load_analysis
        optimizer = st.session_state.load_optimizer
        
        st.markdown("---")
        st.markdown("### 📊 Результаты анализа")
        
        # ДВА ГРАФИКА РЯДОМ: нагрузка и скорость
        col_graph1, col_graph2 = st.columns(2)
        
        with col_graph1:
            # График нагрузки
            fig_load = optimizer.visualize_hourly_load(analysis['current_load'])
            st.plotly_chart(fig_load, use_container_width=True, key="load_chart")
        
        with col_graph2:
            # График скорости
            if analysis['current_load'].get('velocity_profile'):
                fig_velocity = optimizer.visualize_velocity_profile(analysis['current_load'])
                if fig_velocity:
                    st.plotly_chart(fig_velocity, use_container_width=True, key="velocity_chart")
                else:
                    st.info("📊 График скорости недоступен")
            else:
                st.info("⚠️ Расчет скорости не выполнен. Установите параметры трубопровода.")
        
        # Статистика скорости под графиками
        if analysis['current_load'].get('velocity_profile'):
            velocity_stats = analysis['current_load']['velocity_profile']['velocity_stats']
            
            st.markdown("##### 📈 Статистика скорости потока")
            col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
            
            with col_stat1:
                max_v = velocity_stats['max_velocity']
                max_allowed = optimizer.v_max_allowed
                status = "✅ В норме" if max_v <= max_allowed else "⚠️ Превышение"
                st.metric("Макс. скорость", f"{max_v:.2f} м/с", status)
            
            with col_stat2:
                st.metric("Средняя скорость", f"{velocity_stats['avg_velocity']:.2f} м/с")
            
            with col_stat3:
                high_count = velocity_stats['high_velocity_count']
                total_points = len(analysis['current_load']['time_points'])
                percent_high = (high_count / total_points) * 100
                st.metric("Превышения", f"{high_count} раз", f"{percent_high:.1f}% времени")
            
            with col_stat4:
                low_count = velocity_stats['low_velocity_count']
                percent_low = (low_count / total_points) * 100
                st.metric("Низкая скорость", f"{low_count} раз", f"{percent_low:.1f}% времени")
    
    # ================== РЕЗУЛЬТАТЫ ОПТИМИЗАЦИИ ==================
    if 'optimization_results' in st.session_state and st.session_state.optimization_results:
        results = st.session_state.optimization_results
        optimizer = st.session_state.load_optimizer
        
        st.markdown("---")
        st.markdown("### 🎯 Результаты оптимизации")
        
        # ДВА ГРАФИКА РЯДОМ: сравнение нагрузок
        col_opt1, col_opt2 = st.columns(2)
        
        with col_opt1:
            # График сравнения нагрузки
            if optimizer and results.get('optimized_load'):
                fig_compare = optimizer.visualize_hourly_load(
                    st.session_state.current_load_analysis['current_load'],
                    results['optimized_load']
                )
                st.plotly_chart(fig_compare, use_container_width=True, key="compare_chart")
        
        with col_opt2:
            # График скорости после оптимизации
            if results.get('optimized_load') and results['optimized_load'].get('velocity_profile'):
                fig_opt_velocity = optimizer.visualize_velocity_profile(results['optimized_load'])
                if fig_opt_velocity:
                    st.plotly_chart(fig_opt_velocity, use_container_width=True, key="opt_velocity_chart")
                else:
                    st.info("📊 График скорости после оптимизации недоступен")
            else:
                st.info("⚠️ Расчет скорости после оптимизации не выполнен")
        
        # Статистика улучшений под графиками
        if results.get('comparison_stats'):
            st.markdown("##### 📈 Сравнительная статистика")
            
            stats = results['comparison_stats']
            improvements = stats['improvements']
            
            col_imp1, col_imp2, col_imp3, col_imp4 = st.columns(4)
            
            with col_imp1:
                st.metric(
                    "Макс. нагрузка",
                    f"{stats['optimized_stats']['max_load']:.1f} м³/час",
                    f"↓ {improvements.get('max_load_reduction', 0):.1f}%",
                    delta_color="inverse" if improvements.get('max_load_reduction', 0) > 0 else "off"
                )
            
            with col_imp2:
                st.metric(
                    "Мин. нагрузка",
                    f"{stats['optimized_stats']['min_load']:.1f} м³/час",
                    f"↑ {improvements.get('min_load_increase', 0):.1f}%"
                )
            
            with col_imp3:
                st.metric(
                    "Размах нагрузки",
                    f"{stats['optimized_stats']['load_range']:.1f} м³/час",
                    f"↓ {improvements.get('range_reduction', 0):.1f}%",
                    delta_color="inverse" if improvements.get('range_reduction', 0) > 0 else "off"
                )
            
            with col_imp4:
                overall = improvements.get('overall_improvement', 0)
                status = "🔴 Отлично" if overall > 20 else "🟡 Хорошо" if overall > 10 else "⚪ Удовл."
                st.metric("Общее улучшение", f"{overall:.1f}%", status)
        
        # ================== ШАГ 5: ЭКСПОРТ И СОХРАНЕНИЕ ==================
        st.markdown("---")
        st.markdown("#### 📥 Шаг 5: Экспорт и сохранение")
        
        col_export1, col_export2 = st.columns(2)
        
        with col_export1:
            # Экспорт в Excel
            if st.button("📊 Создать полный отчет в Excel", use_container_width=True, type="primary"):
                with st.spinner("Создание Excel отчета..."):
                    try:
                        excel_data = optimizer.create_excel_report()
                        
                        # Формируем имя файла
                        clusters_str = "_".join([c.replace(" ", "") for c in selected_clusters])[:50]
                        filename = f"отчет_нагрузка_{clusters_str}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                        
                        st.success("✅ Отчет создан!")
                        
                        # Кнопка скачивания
                        st.download_button(
                            label="⬇️ Скачать Excel отчет",
                            data=excel_data.getvalue(),
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                        
                    except Exception as e:
                        st.error(f"❌ Ошибка создания отчета: {str(e)}")
        
        with col_export2:
            # Сохранение результатов в систему
            if st.button("💾 Сохранить новые времена запуска", use_container_width=True, type="secondary"):
                if 'optimal_phases' in results:
                    phases_dict = results['optimal_phases']
                    saved_count = 0
                    
                    for well_name, phase_shift in phases_dict.items():
                        # Находим скважину в системе
                        for i, well in enumerate(st.session_state.wells_data):
                            if well['name'] == well_name and well.get('operation_mode') == 'kpr':
                                # Сохраняем исходное время
                                original_time = well.get('base_launch_time', '00:00')
                                
                                # Обновляем время запуска
                                base_minutes = optimizer._time_to_minutes(original_time)
                                new_minutes = (base_minutes + phase_shift) % (24 * 60)
                                new_time = optimizer._minutes_to_time(new_minutes)
                                
                                # Сохраняем изменения
                                st.session_state.wells_data[i]['base_launch_time'] = new_time
                                st.session_state.wells_data[i]['original_launch_time'] = original_time
                                st.session_state.wells_data[i]['last_modified'] = datetime.now().strftime("%Y-%m-%d %H:%M")
                                st.session_state.wells_data[i]['modification_source'] = 'load_optimization'
                                
                                saved_count += 1
                                break
                    
                    if saved_count > 0:
                        st.session_state.unsaved_changes = True
                        
                        # Вызываем функцию сохранения в файл, если она существует
                        if 'save_data_to_file' in globals() and callable(globals()['save_data_to_file']):
                            try:
                                save_data_to_file()
                                st.success(f"✅ Сохранены новые времена запуска для {saved_count} скважин и записаны в файл")
                            except Exception as e:
                                st.warning(f"⚠️ Данные сохранены в системе, но ошибка записи в файл: {str(e)}")
                        else:
                            st.success(f"✅ Обновлены времена запуска для {saved_count} скважин в session_state")
                            st.info("ℹ️ Для сохранения в файл подключите функцию save_data_to_file()")
                    else:
                        st.warning("⚠️ Не найдено ни одной скважины для обновления")
                else:
                    st.error("❌ Нет данных об оптимальных фазах для сохранения")
        
        # Информационное сообщение
        st.markdown("---")
        with st.expander("ℹ️ Информация о сохранении"):
            st.markdown("""
            **Что происходит при сохранении:**
            1. **Обновление времени запуска** - для каждой оптимизированной КПР скважины обновляется базовое время запуска
            2. **Сохранение оригинала** - исходное время сохраняется в поле `original_launch_time`
            3. **Добавление метаданных** - добавляются поля `last_modified` и `modification_source`
            4. **Сохранение в файл** - если функция `save_data_to_file()` подключена, изменения сохраняются в исходный файл
            
            **Важно:** После сохранения данные будут использоваться во всех других модулях системы.
            Для отмены изменений используйте кнопку "Очистить несохраненные изменения" в начале страницы.
            """)
    
    # ================== ИНФОРМАЦИЯ О СМЕСИ ==================
    if 'pipeline_params' in st.session_state:
        params = st.session_state.pipeline_params
        
        st.markdown("---")
        st.markdown("#### 🧪 Свойства смеси и рекомендации")
        
        col_mix1, col_mix2, col_mix3 = st.columns(3)
        
        with col_mix1:
            st.metric("Средняя обводненность", f"{params['avg_water_cut']:.1f}%")
        
        with col_mix2:
            st.metric("Плотность смеси", f"{params['mixture_density']:.0f} кг/м³")
        
        with col_mix3:
            # Определяем тип смеси
            if params['avg_water_cut'] > 70:
                mix_type = "💧 Водо-нефтяная смесь (вода)"
            elif params['avg_water_cut'] > 30:
                mix_type = "🟡 Нефтяная эмульсия"
            else:
                mix_type = "🟤 Нефть"
            st.metric("Тип смеси", mix_type)

# ============================================================
# ГЛАВНАЯ ФУНКЦИЯ
# ============================================================

def main():
    st.set_page_config(
        page_title="PovhEquilibrium - Оптимизация скважин КПР",
        page_icon="🛢️",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    
    # ============ ОТЛАДКА: ПРОВЕРЯЕМ, ЧТО ЕСТЬ В ФАЙЛЕ ============
    if Path('povh_data.pkl').exists():
        try:
            with open('povh_data.pkl', 'rb') as f:
                data = pickle.load(f)
            print("=" * 50)
            print("СОДЕРЖИМОЕ ФАЙЛА povh_data.pkl:")
            for key, value in data.items():
                if isinstance(value, list):
                    print(f"  {key}: список из {len(value)} элементов")
                    if len(value) > 0 and isinstance(value[0], dict):
                        print(f"    Пример ключей: {list(value[0].keys())[:5]}")
                elif isinstance(value, dict):
                    print(f"  {key}: словарь с {len(value)} ключами")
                else:
                    print(f"  {key}: {type(value).__name__}")
            print("=" * 50)
        except Exception as e:
            print(f"Ошибка чтения файла: {e}")
    else:
        print("Файл povh_data.pkl не существует")
    
    # ============ ПРОГРЕСС-БАР ЗАГРУЗКИ ============
    # Показываем прогресс-бар только при первом запуске или если данные не загружены
    if 'data_loaded' not in st.session_state:
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        status_text.text("🚀 Инициализация приложения...")
        progress_bar.progress(10)
        
        # ============ ЗАГРУЗКА ДАННЫХ ИЗ ФАЙЛА ============
        try:
            if Path('povh_data.pkl').exists():
                status_text.text("📦 Загрузка данных из файла...")
                progress_bar.progress(30)
                
                with open('povh_data.pkl', 'rb') as f:
                    data = pickle.load(f)
                
                status_text.text("🔄 Восстановление структур данных...")
                progress_bar.progress(50)
                
                # Загружаем все данные в session_state
                for key, value in data.items():
                    # Конвертируем списки обратно в set для индексов
                    if key.startswith('selected_wells_indices_') and isinstance(value, list):
                        st.session_state[key] = set(value)
                    
                    # Конвертируем словарь обратно в DataFrame для аналитики
                    elif key in ['chess_raw_data', 'chess_enriched_data', 'filtered_analytics_data', 'current_analytics_data']:
                        if isinstance(value, dict) and 'data' in value and 'columns' in value:
                            try:
                                df = pd.DataFrame(value['data'])
                                st.session_state[key] = df
                            except:
                                st.session_state[key] = value
                        elif isinstance(value, list) and value and isinstance(value[0], dict):
                            try:
                                st.session_state[key] = pd.DataFrame(value)
                            except:
                                st.session_state[key] = value
                        else:
                            st.session_state[key] = value
                    
                    # Конвертируем список обратно в set для других индексов
                    elif key == 'selected_wells_indices' and isinstance(value, list):
                        st.session_state[key] = set(value)
                    
                    # Все остальное загружаем как есть
                    else:
                        st.session_state[key] = value
                
                status_text.text("🧹 Очистка служебных полей...")
                progress_bar.progress(70)
                
                # Специальная обработка для результатов расчетов
                if 'last_optimization' in st.session_state and st.session_state.last_optimization:
                    if not isinstance(st.session_state.last_optimization, dict):
                        st.session_state.last_optimization = None
                
                if 'batch_results_advanced' in st.session_state and st.session_state.batch_results_advanced:
                    cleaned_results = []
                    for res in st.session_state.batch_results_advanced:
                        if isinstance(res, dict):
                            cleaned_res = {k: v for k, v in res.items() if not k.startswith('_')}
                            cleaned_results.append(cleaned_res)
                        else:
                            cleaned_results.append(res)
                    st.session_state.batch_results_advanced = cleaned_results
                
                if 'full_batch_results' in st.session_state and st.session_state.full_batch_results:
                    cleaned_full = []
                    for res in st.session_state.full_batch_results:
                        if isinstance(res, dict):
                            cleaned_full.append({k: v for k, v in res.items() if not k.startswith('_')})
                        else:
                            cleaned_full.append(res)
                    st.session_state.full_batch_results = cleaned_full
                
                if 'potential_batch_results' in st.session_state and st.session_state.potential_batch_results:
                    cleaned_potential = []
                    for res in st.session_state.potential_batch_results:
                        if isinstance(res, dict):
                            cleaned_potential.append({k: v for k, v in res.items() if not k.startswith('_')})
                        else:
                            cleaned_potential.append(res)
                    st.session_state.potential_batch_results = cleaned_potential
                
                for mode in ['replace', 'optimize']:
                    results_key = f'pump_calculation_results_{mode}'
                    if results_key in st.session_state and st.session_state[results_key]:
                        if not isinstance(st.session_state[results_key], list):
                            st.session_state[results_key] = []
                    
                    best_key = f'pump_best_variants_{mode}'
                    if best_key in st.session_state and st.session_state[best_key]:
                        if not isinstance(st.session_state[best_key], list):
                            st.session_state[best_key] = []
                
                status_text.text("✅ Данные успешно загружены")
                progress_bar.progress(90)
            else:
                status_text.text("📁 Файл с данными не найден, создаем пустую структуру")
                progress_bar.progress(50)
        except Exception as e:
            status_text.text(f"⚠️ Ошибка загрузки: {str(e)[:50]}...")
            progress_bar.progress(50)
        
        # ============ УСТАНОВКА ЗНАЧЕНИЙ ПО УМОЛЧАНИЮ ============
        status_text.text("⚙️ Установка значений по умолчанию...")
        progress_bar.progress(80)
        
        # Список всех возможных ключей с значениями по умолчанию
        default_values = {
            # Основные данные
            'wells_data': [],
            'clusters': {},
            'calculation_history': [],
            'selected_cdng': "ЦДНГ-3",
            'selected_cluster': None,
            'selected_tpp': "VQ-BADнефтегаз",
            'selected_cits': "ЦИТС VQ-BAD",
            'current_page': "dashboard",
            'editing_mode': False,
            'show_results': False,
            'show_auto_search': False,
            'found_clusters_for_optimization': [],
            'last_optimization': None,
            'current_conversion_tab': 'replace',
            'unsaved_changes': False,
            'app_initialized': True,
            'confirm_delete': False,
            
            # МОДУЛЬ 1 - результаты оптимизации давления
            'optimization_result': None,
            'cycle_simulation': None,
            'inflow_curve': None,
            
            # МОДУЛЬ 2 - пакетные расчеты КПР
            'batch_results_advanced': None,
            'batch_results_detailed': None,
            'potential_batch_results': None,
            'full_batch_results': None,
            'full_batch_detailed': None,
            
            # МОДУЛЬ 3 - замена ЭЦН (оба режима)
            'pump_calculation_results_replace': None,
            'pump_best_variants_replace': None,
            'pump_calculation_params_replace': {},
            'selected_wells_indices_replace': set(),
            
            'pump_calculation_results_optimize': None,
            'pump_best_variants_optimize': None,
            'pump_calculation_params_optimize': {},
            'selected_wells_indices_optimize': set(),
            
            # МОДУЛЬ 4 - аналитика
            'chess_raw_data': None,
            'chess_enriched_data': None,
            'filtered_analytics_data': None,
            'current_analytics_data': None,
            
            # МОДУЛЬ 5 - оптимизация нагрузки
            'load_optimizer_state': None,
            'current_load_analysis': None,
            'optimization_results': None,
            'pipeline_params': None,
            
            # Дополнительные флаги
            'schedule_imported': False,
        }
        
        # Устанавливаем значения по умолчанию для всех ключей, которых нет
        for key, default_value in default_values.items():
            if key not in st.session_state:
                st.session_state[key] = default_value
        
        status_text.text("✅ Готово!")
        progress_bar.progress(100)
        
        # Небольшая задержка, чтобы пользователь увидел 100%
        import time
        time.sleep(0.3)
        
        # Очищаем прогресс-бар и статус
        status_text.empty()
        progress_bar.empty()
        
        # Помечаем, что данные загружены
        st.session_state.data_loaded = True
    
    # ============ САЙДБАР УПРАВЛЕНИЯ ДАННЫМИ ============
    with st.sidebar:
        st.title("⚙️ Управление данными")
        
        # Статистика
        total_wells = len(st.session_state.get('wells_data', []))
        
        # Подсчет всех выполненных расчетов
        total_calculations = 0
        
        # Модуль 1: расчеты стабилизации давления
        if 'calculation_history' in st.session_state:
            total_calculations += len(st.session_state.calculation_history)
        
        # Модуль 2: пакетный расчет КПР
        if st.session_state.get('batch_results_advanced'):
            total_calculations += 1
        
        # Модуль 2: анализ потенциала
        if st.session_state.get('potential_batch_results'):
            total_calculations += 1
        
        # Модуль 2: одиночный расчет КПР
        if st.session_state.get('optimization_result'):
            total_calculations += 1
        
        # Модуль 3: расчет замены ЭЦН (оба режима)
        if st.session_state.get('pump_calculation_results_replace'):
            total_calculations += 1
        
        if st.session_state.get('pump_calculation_results_optimize'):
            total_calculations += 1
        
        # Модуль 5: оптимизация нагрузки
        if st.session_state.get('optimization_results'):
            total_calculations += 1
        
        st.metric("📊 Скважин в системе", total_wells)
        st.metric("📈 Выполнено расчетов", total_calculations)
        
        st.markdown("---")
        
        # Навигация
        st.markdown("### 🗺️ Навигация")
        
        # Кнопки навигации в сайдбаре
        nav_buttons = [
            ("🏠 Главная", "dashboard"),
            ("🛢️ Кусты", "wells"),
            ("⚙️ Оптимизация", "optimization"),
            ("📈 Аналитика", "analytics"),
            ("💾 Сохранение", "custom_save"),
            ("📊 Отчеты", "reports"),
        ]
        
        for label, page in nav_buttons:
            if st.button(label, use_container_width=True,
                        type="primary" if st.session_state.current_page == page else "secondary"):
                st.session_state.current_page = page
                st.rerun()
        
        st.markdown("---")
        
        # Управление данными
        st.markdown("### 💾 Управление данными")
        
        # Сохранение
        if st.button("💾 Сохранить данные", use_container_width=True, key="save_btn"):
            if save_data_to_file():
                st.success("✅ Данные сохранены")
            else:
                st.error("❌ Ошибка сохранения")
        
        # Очистка данных
        st.warning("⚠️ Опасная зона")
        
        if st.button("🗑️ Очистить все данные", use_container_width=True, key="clear_btn"):
            st.session_state.confirm_delete = True
            st.rerun()
        
        if st.session_state.get('confirm_delete', False):
            st.error("❌ Вы уверены?")
            col1, col2 = st.columns(2)
            with col1:
                if st.button("✅ Да", type="primary", key="yes_delete"):
                    if clear_all_data():
                        # После очистки удаляем флаг data_loaded, чтобы при следующем запуске снова показать прогресс-бар
                        if 'data_loaded' in st.session_state:
                            del st.session_state.data_loaded
                        st.success("✅ Все данные очищены")
                        st.session_state.confirm_delete = False
                        st.rerun()
            with col2:
                if st.button("❌ Нет", key="cancel_delete"):
                    st.session_state.confirm_delete = False
                    st.rerun()
        
        st.markdown("---")
        st.caption("🛢️ PovhEquilibrium v2.0")
    
    # ============ НАВИГАЦИОННАЯ ПАНЕЛЬ ============
    cols = st.columns(6)
    nav_pages = [
        ("🏠 Главная", "dashboard"),
        ("🛢️ Кусты", "wells"),
        ("⚙️ Оптимизация", "optimization"),
        ("📈 Аналитика", "analytics"),
        ("💾 Сохранение", "custom_save"),
        ("📊 Отчеты", "reports"),
    ]
    
    for i, (label, page) in enumerate(nav_pages):
        with cols[i]:
            if st.button(label, use_container_width=True,
                        type="primary" if st.session_state.current_page == page else "secondary"):
                st.session_state.current_page = page
                st.rerun()
    
    st.markdown("---")
    
    # ============ ОТОБРАЖЕНИЕ ТЕКУЩЕЙ СТРАНИЦЫ ============
    if st.session_state.current_page == "dashboard":
        show_dashboard()
    elif st.session_state.current_page == "wells":
        show_wells_management()
    elif st.session_state.current_page == "optimization":
        show_optimization()
    elif st.session_state.current_page == "reports":
        show_reports()
    elif st.session_state.current_page == "custom_save":
        save_custom_selected_wells()
    elif st.session_state.current_page == "analytics":
        show_analytics()
    
    # Футер
    st.markdown("---")
    st.markdown("""
    <div style='text-align: center; color: #666; padding: 20px;'>
        <p>🛢️ <b>PovhEquilibrium</b> | Система оптимизации скважин</p>
    </div>
    """, unsafe_allow_html=True)


if __name__ == "__main__":
    main()
